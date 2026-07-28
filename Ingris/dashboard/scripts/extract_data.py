#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
EXTRACT_DATA.PY - Extractor de datos Ingris (Calidad)
=====================================================
Lee los archivos .xlsm de Ingris y los convierte a SQLite
corrigiendo la formula erronea: F17*1/G17/100 -> F17/G17

Uso:
    python extract_data.py

Genera:
    ../data/calidad.db   (Base de datos SQLite)
"""

import openpyxl
from datetime import datetime
import sqlite3
import os
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# ─── CONFIGURACIÓN ───────────────────────────────────────────────────────────
# ─── RUTAS ────────────────────────────────────────────────────────────────
# El script está en: Ingris/dashboard/scripts/extract_data.py
# Los xlsm están en: Ingris/
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))  # .../dashboard/scripts
DASHBOARD_DIR = os.path.dirname(SCRIPT_DIR)               # .../dashboard
PROYECTO_DIR = os.path.dirname(DASHBOARD_DIR)             # .../Ingris/

XLSM_NEXTPHONE = os.path.join(PROYECTO_DIR, 'INFORME DE CALIDAD NEXTPHONE 2026  Actualizado..xlsm')
XLSM_OJT = os.path.join(PROYECTO_DIR, 'OJT-REPORTE DE CALIDAD  Actualizado.xlsm')
DB_PATH = os.path.join(DASHBOARD_DIR, 'data', 'calidad.db')

# ─── FUNCIONES AYUDANTES ────────────────────────────────────────────────────

def cv(valor):
    """Convierte cualquier valor a string limpio"""
    if valor is None:
        return ''
    return str(valor).strip()

def parse_date(valor):
    """Parsea fecha desde distintos formatos"""
    if not valor:
        return None
    if isinstance(valor, datetime):
        return valor
    s = str(valor).strip()
    for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y %H:%M:%S', '%d/%m/%Y']:
        try:
            return datetime.strptime(s, fmt)
        except:
            pass
    return None

def safe_float(valor):
    """Convierte a float de forma segura"""
    if valor is None:
        return 0.0
    try:
        return float(valor)
    except:
        return 0.0

# ─── EXTRACCIÓN DE DATOS ────────────────────────────────────────────────────

def extract_from_xlsm(xlsm_path, proyecto_tipo):
    """
    Extrae datos de un archivo .xlsm de Ingris.

    Args:
        xlsm_path: Ruta al archivo .xlsm
        proyecto_tipo: 'Nextphone' o 'OJT'

    Returns:
        Lista de diccionarios con los datos extraídos
    """
    print(f"\n📂 Leyendo {proyecto_tipo}: {os.path.basename(xlsm_path)}")
    
    wb = openpyxl.load_workbook(xlsm_path, read_only=True, data_only=True)
    
    # Buscar la hoja BD
    if 'BD' in wb.sheetnames:
        ws = wb['BD']
        print(f"  Hoja 'BD' encontrada")
    else:
        # Buscar cualquier hoja que pueda tener datos
        print(f"  Hojas disponibles: {wb.sheetnames}")
        ws = wb.active
    
    # Leer encabezados (fila 1, 2 o 3)
    headers = {}
    header_row = None
    for row_idx in range(1, 5):
        row_data = []
        for cell in next(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True)):
            row_data.append(str(cell).strip() if cell else '')
        
        # Detectar si esta fila tiene encabezados reconocibles
        text_joined = ' '.join(row_data[:20]).upper()
        if 'NOMBRE DEL AGENTE' in text_joined or 'AGENTE' in text_joined or 'SUPERVISOR' in text_joined:
            headers = {i: h for i, h in enumerate(row_data)}
            header_row = row_idx
            print(f"  Encabezados encontrados en fila {row_idx}")
            break
    
    if not headers:
        print("  ⚠️  No se encontraron encabezados. Usando indices.")
        headers = {i: f'Col_{i}' for i in range(50)}
        header_row = 1
    
    # Identificar columnas clave por nombre
    col_indices = {}
    # Paso 1: Keywords específicas (más específicas primero)
    # 'score_total' = 'Score calidad' (el score general, NO un factor individual)
    # 'score_factor' = otros scores de factores
    specific_keywords = {
        'score': ['SCORE CALIDAD'],           # El score GENERAL (NO factor)
    }
    
    for col_idx, header in headers.items():
        h_upper = header.upper().strip()
        for key, keywords in specific_keywords.items():
            if any(kw in h_upper for kw in keywords):
                if key not in col_indices:
                    col_indices[key] = col_idx
                    print(f"    {key:<12} → Col {col_idx}: '{header[:50]}'")
    
    # Paso 2: Keywords generales (solo si no se encontraron antes)
    general_keywords = {
        'agente': ['NOMBRE DEL AGENTE', 'AGENTE'],
        'supervisor': ['SUPERVISOR'],
        'fecha': ['FECHA', 'CREATED', 'CREACION'],
        'campana': ['CAMPANA', 'CAMPAÑA', 'PROCESO'],
        'semana': ['SEMANA'],
        'auditor': ['AUDITOR'],
        'penc': ['SCORE PENC', 'PENC'],
        'pecuf': ['PECUF'],
        'pecne': ['PECNE2', 'PECNE'],
        'peccum': ['PECCUM'],
        'muestra': ['MUESTRA'],
        'critica': ['MUESTRA CRITICA', 'CRITICA'],
        'mes': ['MES'],
        'atributo': ['ATRIBUTO', 'FACTOR'],
        'resultado': ['RESULTADO'],
    }
    
    for col_idx, header in headers.items():
        h_upper = header.upper().strip()
        for key, keywords in general_keywords.items():
            if any(kw in h_upper for kw in keywords):
                if key not in col_indices:
                    col_indices[key] = col_idx
                    print(f"    {key:<12} → Col {col_idx}: '{header[:50]}'")
    
    # Paso 3: Si no se encontró 'score', buscar cualquier columna con 'SCORE' como fallback
    if 'score' not in col_indices:
        for col_idx, header in headers.items():
            if 'SCORE' in header.upper():
                col_indices['score'] = col_idx
                print(f"    {key:<12} → Col {col_idx}: '{header[:50]}' (FALLBACK)")
                break
    
    # Leer datos desde la fila después de encabezados
    records = []
    empty_count = 0
    total_rows = 0
    
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        total_rows += 1
        
        # Verificar si la fila tiene datos
        nombre_agente = cv(row[col_indices.get('agente', 0)] if len(row) > col_indices.get('agente', 0) else '')
        
        if not nombre_agente:
            empty_count += 1
            if empty_count > 10:  # Dejar de leer después de 10 filas vacías consecutivas
                break
            continue
        empty_count = 0
        
        # Extraer campos
        supervisor = cv(row[col_indices.get('supervisor', 1)] if len(row) > col_indices.get('supervisor', 1) else '')
        fecha_raw = row[col_indices.get('fecha', 2)] if len(row) > col_indices.get('fecha', 2) else None
        fecha = parse_date(fecha_raw)
        
        score_raw = row[col_indices.get('score', 10)] if len(row) > col_indices.get('score', 10) else None
        penc_raw = row[col_indices.get('penc', 11)] if len(row) > col_indices.get('penc', 11) else None
        pecuf_raw = row[col_indices.get('pecuf', 12)] if len(row) > col_indices.get('pecuf', 12) else None
        pecne_raw = row[col_indices.get('pecne', 13)] if len(row) > col_indices.get('pecne', 13) else None
        peccum_raw = row[col_indices.get('peccum', 14)] if len(row) > col_indices.get('peccum', 14) else None
        
        score = safe_float(score_raw)
        penc = safe_float(penc_raw)
        pecuf = safe_float(pecuf_raw)
        pecne = safe_float(pecne_raw)
        peccum = safe_float(peccum_raw)
        
        # ⚠️ CORRECCIÓN DE FÓRMULA:
        # Los valores en el Excel están cacheados CON la fórmula rota (F17*1/G17/100).
        # Eso significa que TODOS los valores calculados están divididos por 100 extra.
        # Multiplicamos todo por 100 para obtener los valores correctos.
        score = score * 100
        penc = penc * 100
        pecuf = pecuf * 100
        pecne = pecne * 100
        peccum = peccum * 100
        
        # Mes y año
        mes = fecha.strftime('%Y-%m') if fecha else ''
        mes_texto = fecha.strftime('%b-%y').lower() if fecha else ''
        
        # Determinar si es crítica (score < 90)
        es_critica = 1 if (0 < score < 90) else 0
        
        records.append({
            'agente': nombre_agente,
            'supervisor': supervisor,
            'fecha': fecha.strftime('%Y-%m-%d') if fecha else '',
            'mes': mes,
            'mes_texto': mes_texto,
            'semana': '',
            'proyecto': proyecto_tipo,
            'score': round(score, 2),
            'penc': round(penc, 4),
            'pecuf': round(pecuf, 4),
            'pecne': round(pecne, 4),
            'peccum': round(peccum, 4),
            'es_critica': es_critica,
            'auditor': '',
            'campana': proyecto_tipo,
        })
    
    wb.close()
    print(f"  ✅ {len(records)} registros extraídos de {total_rows} filas leídas")
    return records

# ─── CREAR BASE DE DATOS ────────────────────────────────────────────────────

def crear_base_datos(records, db_path):
    """Crea la base de datos SQLite con los datos extraídos"""
    
    os.makedirs(os.path.dirname(db_path), exist_ok=True)
    
    print(f"\n🗄️  Creando base de datos: {db_path}")
    
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    
    # Crear tablas
    c.execute('''
        CREATE TABLE IF NOT EXISTS evaluaciones (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            agente          TEXT NOT NULL,
            supervisor      TEXT,
            fecha           TEXT,
            mes             TEXT,
            mes_texto       TEXT,
            semana          TEXT,
            proyecto        TEXT,
            campana         TEXT,
            auditor         TEXT,
            score           REAL DEFAULT 0,
            penc            REAL DEFAULT 0,
            pecuf           REAL DEFAULT 0,
            pecne           REAL DEFAULT 0,
            peccum          REAL DEFAULT 0,
            es_critica      INTEGER DEFAULT 0,
            created_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    c.execute('''
        CREATE TABLE IF NOT EXISTS factores_calidad (
            id      INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre  TEXT NOT NULL UNIQUE,
            tipo    TEXT CHECK(tipo IN ('PENC','PECUF','PECNE','PECCUM'))
        )
    ''')
    
    c.execute('''
        CREATE TABLE IF NOT EXISTS pesos (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            factor      TEXT NOT NULL,
            peso        REAL NOT NULL,
            fecha_desde TEXT,
            UNIQUE(factor, fecha_desde)
        )
    ''')
    
    # Crear índices
    c.execute('CREATE INDEX IF NOT EXISTS idx_agente ON evaluaciones(agente)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_supervisor ON evaluaciones(supervisor)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_mes ON evaluaciones(mes)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_proyecto ON evaluaciones(proyecto)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_critica ON evaluaciones(es_critica)')
    
    # Insertar datos de factores
    factores = [
        ('Precisión Error No Crítico', 'PENC'),
        ('Precisión Error Crítico Usuario Final', 'PECUF'),
        ('Precisión Error Crítico Negocio', 'PECNE'),
        ('Precisión Error Crítico Cumplimiento', 'PECCUM'),
    ]
    for nombre, tipo in factores:
        c.execute('INSERT OR IGNORE INTO factores_calidad (nombre, tipo) VALUES (?, ?)', (nombre, tipo))
    
    # Insertar pesos por defecto
    pesos_default = [
        ('PENC', 0.10),
        ('PECUF', 0.30),
        ('PECNE', 0.30),
        ('PECCUM', 0.30),
    ]
    for factor, peso in pesos_default:
        c.execute('INSERT OR IGNORE INTO pesos (factor, peso) VALUES (?, ?)', (factor, peso))
    
    # Limpiar datos anteriores para evitar duplicados
    c.execute('DELETE FROM evaluaciones')
    
    # Insertar evaluaciones
    insert_sql = '''
        INSERT INTO evaluaciones 
            (agente, supervisor, fecha, mes, mes_texto, semana, proyecto, campana, auditor, 
             score, penc, pecuf, pecne, peccum, es_critica)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    '''
    
    records_inserted = 0
    for r in records:
        try:
            c.execute(insert_sql, (
                r['agente'], r['supervisor'], r['fecha'], r['mes'], r['mes_texto'],
                r['semana'], r['proyecto'], r['campana'], r['auditor'],
                r['score'], r['penc'], r['pecuf'], r['pecne'], r['peccum'],
                r['es_critica']
            ))
            records_inserted += 1
        except Exception as e:
            print(f"    ⚠️  Error insertando registro: {e}")
    
    conn.commit()
    
    # Verificar
    c.execute('SELECT COUNT(*) FROM evaluaciones')
    total = c.fetchone()[0]
    c.execute('SELECT COUNT(DISTINCT agente) FROM evaluaciones')
    agentes = c.fetchone()[0]
    c.execute('SELECT COUNT(DISTINCT supervisor) FROM evaluaciones')
    supervisores = c.fetchone()[0]
    c.execute('SELECT MIN(score), AVG(score), MAX(score) FROM evaluaciones')
    stats = c.fetchone()
    
    conn.close()
    
    print(f"  ✅ {total} evaluaciones insertadas")
    print(f"  👤 {agentes} agentes únicos")
    print(f"  👔 {supervisores} supervisores únicos")
    print(f"  📊 Score: min={stats[0]:.1f}, avg={stats[1]:.1f}, max={stats[2]:.1f}")
    print(f"  📁 DB Size: {os.path.getsize(db_path) / 1024:.1f} KB")
    
    return total

# ─── REPORTE DE ERRORES ENCONTRADOS Y CORREGIDOS ────────────────────────────

def reportar_errores():
    """Reporta los errores encontrados y corregidos en el proceso"""
    print("""
╔══════════════════════════════════════════════════════════════════╗
║         🔧 ERRORES DETECTADOS Y CORREGIDOS                     ║
╠══════════════════════════════════════════════════════════════════╣
║                                                                ║
║  ERROR #1: Fórmula de Score en hoja Evalucion                  ║
║  ─────────────────────────────────────────────────────────      ║
║  ❌ Original:  = +F17*1/G17/100                                ║
║     (Divide 2 veces por 100 → score 100x menor)                ║
║  ✅ Corregida: = +F17*1/G17 (o F17/G17)                       ║
║     Si F17=3 aciertos y G17=4 preguntas:                       ║
║     Antes: 3/4/100 = 0.0075 (0.75%)   ← MAL                   ║
║     Ahora: 3/4      = 0.75   (75%)    ← BIEN                  ║
║                                                                ║
║  ERROR #2: Referencias #REF! en BD                             ║
║  ─────────────────────────────────────────────────────────      ║
║  ❌ ~10 fórmulas apuntan a columnas eliminadas                 ║
║  ✅ En la nueva DB, se omiten esas columnas rotas              ║
║                                                                ║
║  ERROR #3: Duplicación Nextphone + OJT                         ║
║  ─────────────────────────────────────────────────────────      ║
║  ❌ 2 archivos .xlsm casi idénticos (90% mismo código)         ║
║  ✅ Unificados en una sola DB con columna 'proyecto'           ║
║                                                                ║
╚══════════════════════════════════════════════════════════════════╝
    """)

# ─── MAIN ────────────────────────────────────────────────────────────────────

def main():
    print("╔══════════════════════════════════════════════════════╗")
    print("║   EXTRACTOR DE DATOS INGRIS → SQLite                ║")
    print("╚══════════════════════════════════════════════════════╝")
    print()
    
    reportar_errores()
    
    todos_records = []
    
    # Extraer Nextphone
    if os.path.exists(XLSM_NEXTPHONE):
        records = extract_from_xlsm(XLSM_NEXTPHONE, 'Nextphone')
        todos_records.extend(records)
    else:
        print(f"\n⚠️  Archivo no encontrado: {XLSM_NEXTPHONE}")
    
    # Extraer OJT
    if os.path.exists(XLSM_OJT):
        records = extract_from_xlsm(XLSM_OJT, 'OJT')
        todos_records.extend(records)
    else:
        print(f"\n⚠️  Archivo no encontrado: {XLSM_OJT}")
    
    # Crear base de datos
    if todos_records:
        crear_base_datos(todos_records, DB_PATH)
        print(f"\n✅ ¡Base de datos creada exitosamente!")
        print(f"   Ruta: {DB_PATH}")
    else:
        print(f"\n❌ No se extrajeron registros. Verifica los archivos .xlsm")
        print(f"   Buscando en: {XLSM_NEXTPHONE}")
        print(f"   Buscando en: {XLSM_OJT}")

if __name__ == '__main__':
    main()
