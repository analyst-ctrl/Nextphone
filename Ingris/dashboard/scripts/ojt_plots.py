#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Genera gráficos OJT en PNG desde el archivo .xlsm
"""

import openpyxl
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
import os

XLSM = r"C:\xampp\htdocs\Chamba Panama\Ingris\OJT-REPORTE DE CALIDAD  Actualizado.xlsm"
OUT_DIR = r"C:\xampp\htdocs\Chamba Panama\Ingris\dashboard\assets"

wb = openpyxl.load_workbook(XLSM, data_only=True, keep_vba=False)

plt.rcParams.update({
    'font.family': 'sans-serif',
    'font.sans-serif': ['Segoe UI', 'DejaVu Sans'],
    'axes.edgecolor': '#2a2d3e',
    'axes.labelcolor': '#8b8fa3',
    'axes.titlecolor': '#e8eaf0',
    'xtick.color': '#8b8fa3',
    'ytick.color': '#8b8fa3',
    'text.color': '#e8eaf0',
    'figure.facecolor': '#0f1117',
    'axes.facecolor': '#1a1d27',
    'legend.facecolor': '#1e2130',
    'legend.edgecolor': '#2a2d3e',
    'legend.labelcolor': '#e8eaf0',
})

def save(fig, name):
    path = os.path.join(OUT_DIR, name)
    fig.savefig(path, dpi=150, bbox_inches='tight', facecolor='#0f1117')
    print(f"  Saved: {path}")
    plt.close(fig)

# ─── DATA ───────────────────────────────────────────────────
ws = wb['Graficos']

score_jun = ws.cell(14, 7).value  # G14
score_jul = ws.cell(15, 7).value  # G15
score_total = ws.cell(16, 7).value

penc = ws.cell(16, 3).value
pecuf = ws.cell(16, 4).value
pecne = ws.cell(16, 5).value
peccum = ws.cell(16, 6).value

muestras = ws.cell(5, 8).value  # H5
agentes_mon = ws.cell(5, 2).value  # B5
hc_activo = ws.cell(5, 11).value  # K5 = HC!H2
cumplimiento = ws.cell(6, 11).value  # K6

# Agent data
ws_ag = wb['Resultado por agente']
agentes = []
for r in range(11, 17):
    name = ws_ag.cell(r, 2).value
    score = ws_ag.cell(r, 11).value
    cuartil = ws_ag.cell(r, 12).value
    cumple = ws_ag.cell(r, 13).value
    if name:
        agentes.append({'name': name, 'score': score, 'cuartil': cuartil, 'cumple': cumple})

# Atributos data
ws_at = wb['Resultado por atributo']
atributos = []
for r in range(15, 41):
    name = ws_at.cell(r, 2).value
    val = ws_at.cell(r, 3).value
    if name and val is not None:
        atributos.append({'name': name.strip(), 'value': val})

wb.close()

print("=== GENERANDO PLOTS OJT ===")

# ─── 1. SCORE GENERAL OJT ──────────────────────────────────
fig, ax = plt.subplots(figsize=(8, 5))
bars = ax.bar(['Jun-26', 'Jul-26'], [score_jun*100, score_jul*100],
              color=['#4f8cff', '#a78bfa'], width=0.5, edgecolor='none', zorder=3)
for bar, val in zip(bars, [score_jun*100, score_jul*100]):
    ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 1,
            f'{val:.1f}%', ha='center', va='bottom', fontweight='bold',
            fontsize=14, color='#e8eaf0')
ax.axhline(y=90, color='#f87171', linestyle='--', linewidth=1.5, alpha=0.7, label='Umbral crítico (90%)')
ax.axhline(y=score_total*100, color='#fbbf24', linestyle=':', linewidth=1, alpha=0.5, label=f'Promedio total ({score_total*100:.1f}%)')
ax.set_ylim(0, 100)
ax.set_title('Score General OJT - Nuevo Ingreso', fontsize=16, fontweight='bold', pad=15)
ax.set_ylabel('Score (%)', fontsize=12)
ax.yaxis.set_major_formatter(mticker.FormatStrFormatter('%.0f%%'))
ax.legend(framealpha=0.3, fontsize=10)
ax.grid(axis='y', alpha=0.15, zorder=0)
save(fig, 'ojt_score_general.png')

# ─── 2. PRECISION POR COMPONENTE ──────────────────────────
fig, ax = plt.subplots(figsize=(8, 5))
labels = ['PENC\n(No Crítico)', 'PECUF\n(Usuario Final)', 'PECNE\n(Negocio)', 'PECCUM\n(Cumplimiento)']
vals = [penc*100, pecuf*100, pecne*100, peccum*100]
colors = ['#4f8cff', '#a78bfa', '#fbbf24', '#f87171']
bars = ax.bar(labels, vals, color=colors, width=0.6, edgecolor='none', zorder=3)
for bar, val in zip(bars, vals):
    ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 1,
            f'{val:.1f}%', ha='center', va='bottom', fontweight='bold', fontsize=12, color='#e8eaf0')
ax.set_ylim(0, 100)
ax.set_title('Precisión por Tipo de Error', fontsize=16, fontweight='bold', pad=15)
ax.set_ylabel('Precisión (%)', fontsize=12)
ax.yaxis.set_major_formatter(mticker.FormatStrFormatter('%.0f%%'))
ax.grid(axis='y', alpha=0.15, zorder=0)
save(fig, 'ojt_precision_componentes.png')

# ─── 3. EVOLUCION MENSUAL POR COMPONENTE ──────────────────
fig, ax = plt.subplots(figsize=(9, 5))
x = np.arange(2)
width = 0.18
meses = ['Jun-26', 'Jul-26']
penc_vals = [ws.cell(14, 3).value*100, ws.cell(15, 3).value*100]
pecuf_vals = [ws.cell(14, 4).value*100, ws.cell(15, 4).value*100]
pecne_vals = [ws.cell(14, 5).value*100, ws.cell(15, 5).value*100]
peccum_vals = [ws.cell(14, 6).value*100, ws.cell(15, 6).value*100]
score_vals = [ws.cell(14, 7).value*100, ws.cell(15, 7).value*100]

ax.bar(x - width*2, penc_vals, width, label='PENC', color='#4f8cff', zorder=3)
ax.bar(x - width, pecuf_vals, width, label='PECUF', color='#a78bfa', zorder=3)
ax.bar(x, pecne_vals, width, label='PECNE', color='#fbbf24', zorder=3)
ax.bar(x + width, peccum_vals, width, label='PECCUM', color='#f87171', zorder=3)
ax.bar(x + width*2, score_vals, width, label='Score', color='#34d399', zorder=3)
ax.set_xticks(x)
ax.set_xticklabels(meses, fontsize=12)
ax.set_ylim(0, 110)
ax.set_title('Evolución Mensual por Componente', fontsize=16, fontweight='bold', pad=15)
ax.set_ylabel('%', fontsize=12)
ax.yaxis.set_major_formatter(mticker.FormatStrFormatter('%.0f%%'))
ax.legend(framealpha=0.3, fontsize=10, ncol=3)
ax.grid(axis='y', alpha=0.15, zorder=0)
save(fig, 'ojt_evolucion_mensual.png')

# ─── 4. AGENTES - SCORE INDIVIDUAL ─────────────────────────
fig, ax = plt.subplots(figsize=(10, 5))
names = [a['name'].split()[0] for a in agentes]
scores = [a['score']*100 for a in agentes]
cuartil_colors = {'Q1': '#34d399', 'Q2': '#4f8cff', 'Q3': '#fbbf24', 'Q4': '#f87171'}
bar_colors = [cuartil_colors.get(a['cuartil'], '#8b8fa3') for a in agentes]
bars = ax.barh(names, scores, color=bar_colors, edgecolor='none', zorder=3, height=0.6)
for bar, val, a in zip(bars, scores, agentes):
    label = f'{val:.1f}% | {a["cuartil"]} | {a["cumple"]}'
    ax.text(bar.get_width() + 0.5, bar.get_y() + bar.get_height()/2,
            label, va='center', fontsize=9, color='#8b8fa3')
ax.axvline(x=90, color='#f87171', linestyle='--', linewidth=1, alpha=0.5, label='Umbral crítico')
ax.set_xlim(0, 110)
ax.set_title('Score por Agente OJT', fontsize=16, fontweight='bold', pad=15)
ax.set_xlabel('Score (%)', fontsize=12)
ax.xaxis.set_major_formatter(mticker.FormatStrFormatter('%.0f%%'))
# Legend for quartiles
from matplotlib.patches import Patch
legend_elements = [Patch(facecolor='#34d399', label='Q1 (Mejor)'),
                   Patch(facecolor='#f87171', label='Q4 (Peor)'),
                   Patch(facecolor='none', edgecolor='#f87171', linestyle='--', label='Umbral 90%')]
ax.legend(handles=legend_elements, framealpha=0.3, fontsize=9)
ax.grid(axis='x', alpha=0.15, zorder=0)
ax.invert_yaxis()
save(fig, 'ojt_agentes.png')

# ─── 5. KPI CARD (text summary) ───────────────────────────
fig, ax = plt.subplots(figsize=(8, 4))
ax.axis('off')
score_color = '#34d399' if score_total*100 >= 85 else '#fbbf24' if score_total*100 >= 70 else '#f87171'
text_content = f"""
╔══════════════════════════════════════════════════════╗
║           RESUMEN OJT - NUEVO INGRESO               ║
╠══════════════════════════════════════════════════════╣
║                                                     ║
║   Score General:       {score_total*100:5.1f}%                        ║
║   Evaluaciones:        {int(muestras):3d}                           ║
║   Agentes Monit.:      {int(agentes_mon):3d}                           ║
║   HC Activo:           {int(hc_activo):3d}                           ║
║   % Cumplimiento:      {cumplimiento*100:5.1f}%                        ║
║                                                     ║
║   PENC:  {penc*100:5.1f}%     PECUF: {pecuf*100:5.1f}%                ║
║   PECNE: {pecne*100:5.1f}%     PECCUM:{peccum*100:5.1f}%                ║
║                                                     ║
║   Cumplen: {sum(1 for a in agentes if a['cumple']=='Cumple')} de {len(agentes)} agentes              ║
║   En Q1:  {sum(1 for a in agentes if a['cuartil']=='Q1')}   |   En Q4:  {sum(1 for a in agentes if a['cuartil']=='Q4')}                ║
╚══════════════════════════════════════════════════════╝
"""
ax.text(0.5, 0.5, text_content, transform=ax.transAxes, fontfamily='monospace',
        fontsize=10, verticalalignment='center', horizontalalignment='center',
        color='#e8eaf0', linespacing=1.5)
ax.set_title('Resumen OJT - Nuevo Ingreso', fontsize=16, fontweight='bold', pad=10, color='#e8eaf0')
save(fig, 'ojt_resumen.png')

print("\nDone! All plots saved to:", OUT_DIR)
