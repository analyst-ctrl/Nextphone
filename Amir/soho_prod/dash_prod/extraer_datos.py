# -*- coding: utf-8 -*-
"""
===============================================================================
  PIPELINE RENOVACION SOHO -> dash_prod
  Gerente: Amir Josue Rodriguez Chavarria
  Campaña: RENOVACION (equipo Marinel Moreno)

  Lectura:
    - extrainfo/HC*.xlsx          -> plantillas de equipo (3 equipos)
    - Reporte_soho_junio_25 (1).xlsb -> Vici (marcaciones) + Ventas + Facturas
    - bases/Base para campaña cross-sell movil Nextphone.xlsx -> venta cruzada

  Genera:
    - data/renovacion_data.json  -> datos para el dashboard
    - index.html                 -> dashboard auto-contenido (HTML+JS+Chart.js)
      con navegación por meses (TODOS / ENERO..JUNIO)

  Uso:
    python dash_prod/extraer_datos.py
===============================================================================
"""
import glob, json, re, os, datetime
from collections import Counter, defaultdict
from datetime import date, timedelta

BASE = os.path.dirname(os.path.abspath(__file__))
PROD = os.path.dirname(BASE)

def norm(s):
    if s is None: return ''
    s = str(s).strip()
    # Arreglar mojibake: texto UTF-8 leído como cp1252 (ej: 'Ñ' -> 'Ã'+'\u2018')
    try:
        s = s.encode('cp1252').decode('utf-8')
    except Exception:
        pass
    s = s.upper()
    s = re.sub(r'\s+', ' ', s)
    for a, b in [('Á','A'),('É','E'),('Í','I'),('Ó','O'),('Ú','U'),('Ñ','N'),('Ü','U')]:
        s = s.replace(a, b)
    return s

def to_num(v):
    if v is None: return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(',', '.').replace(' ', ''))
    except Exception:
        return 0.0

def cuenta12(v):
    """Normaliza una cuenta a 12 dígitos (o los últimos 8 como fallback)."""
    if v is None: return ''
    s = str(v).replace('.0', '').strip()
    s = re.sub(r'\D', '', s)
    if len(s) >= 12: return s[-12:]
    if len(s) >= 8: return s[-8:]
    return ''


def tokens(name):
    return set(norm(name).split())

def match_member(name, roster):
    """Devuelve el miembro del roster que coincide con el nombre.

    Requisitos (para evitar falsos positivos por apellidos compartidos):
      - El PRIMER token del nombre (nombre de pila) debe estar en el miembro
      - Al menos 2 tokens en común
    """
    n = norm(name)
    if not n: return None
    toks = n.split()
    if not toks: return None
    first = toks[0]  # nombre de pila (orden real, no por longitud)
    t = set(toks)
    best, best_score = None, 0
    for member in roster:
        mt = tokens(member)
        if first not in mt:
            continue
        score = len(t & mt)
        if score >= 2 and score > best_score:
            best, best_score = member, score
    return best

def get_col(idx, *nombres):
    """Devuelve el índice de la primera columna cuyo nombre normalizado coincida."""
    for n in nombres:
        k = norm(n)
        if k in idx:
            return idx[k]
    return None

# =============================================================================
# 0. MESES
# =============================================================================
MESES_ORDEN = ['ENERO','FEBRERO','MARZO','ABRIL','MAYO','JUNIO']
MES_ABREV = {'ENE':'ENERO','FEB':'FEBRERO','MAR':'MARZO','ABR':'ABRIL',
             'MAY':'MAYO','JUN':'JUNIO','JUL':'JULIO','AGO':'AGOSTO',
             'SEP':'SEPTIEMBRE','OCT':'OCTUBRE','NOV':'NOVIEMBRE','DIC':'DICIEMBRE'}
MES_NUM = {f'{i:02d}': m for i, m in enumerate(MESES_ORDEN, 1)}

def mes_de(v):
    """Convierte una fecha/mes (texto o serial Excel) a nombre de mes en MESES_ORDEN.
    Devuelve '' si está fuera de los meses del reporte (enero-junio)."""
    if v is None: return ''
    if isinstance(v, (int, float)):
        try:
            d = date(1899, 12, 30) + timedelta(days=int(v))
            if 1 <= d.month <= 6:
                return MESES_ORDEN[d.month - 1]
            return ''
        except Exception:
            return ''
    s = norm(v)
    # nombre completo del mes: 'junio', '12 DE ENERO'
    for m in MESES_ORDEN:
        if m in s:
            return m
    # abreviatura de 3 letras: 'may/26', '01/jun', '24/jun'
    ab = re.search(r'\b([A-Z]{3})\b', s)
    if ab and ab.group(1) in MES_ABREV:
        return MES_ABREV[ab.group(1)]
    # fecha ISO: '2026-01-02 00:00:00'
    num = re.search(r'(?:^|\D)(\d{1,2})(?=\D|$)', s)
    if num and num.group(1).zfill(2) in MES_NUM:
        return MES_NUM[num.group(1).zfill(2)]
    return ''

# =============================================================================
# 1. PLANTILLAS HC -> EQUIPOS
# =============================================================================
import openpyxl

EQUIPOS_RAW = [
    # (archivo, supervisor, nombre_campana)
    ('HC AGOSTO.xlsx',                'MARINEL MORENO',      'RENOVACION'),
    ('HC Formato solicitado (2).xlsx','JULIO CESAR ARAUZ',   'SOHO'),
]

equipos = {}
for archivo, sup, campana in EQUIPOS_RAW:
    f = glob.glob(os.path.join(PROD, 'extrainfo', archivo))
    if not f:
        print('HC no encontrado:', archivo); continue
    wb = openpyxl.load_workbook(f[0], read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    hdr = {norm(h): i for i, h in enumerate(header or [])}
    col_nombre = get_col(hdr, 'Nombre', 'Nombre del Agente', 'Agente')
    col_sup = get_col(hdr, 'Supervisor')
    col_canal = get_col(hdr, 'Canal')
    miembros = []
    for r in rows:
        if r is None or all(v is None for v in r): continue
        nombre = r[col_nombre] if col_nombre is not None and col_nombre < len(r) else None
        if not nombre:
            continue
        sup_col = r[col_sup] if col_sup is not None and col_sup < len(r) else None
        canal_col = r[col_canal] if col_canal is not None and col_canal < len(r) else None
        # Filtrar por canal: el equipo RENOVACION pertenece al canal SOHO
        if not (canal_col and norm(canal_col).startswith('SOHO')):
            continue
        # El supervisor como persona solo cuenta si además está listado como integrante;
        # en HC AGOSTO Marinel está listada como coordinadora (canal SOHO) -> se incluye
        miembros.append(str(nombre).strip())
    wb.close()
    # Eliminar duplicados conservando el orden
    unicos = sorted(set(miembros), key=lambda x: norm(x))
    equipos[campana] = {
        'supervisor': sup,
        'archivo': archivo,
        'miembros': unicos,
    }
    print(f'Equipo {campana} ({sup}): {len(unicos)} integrantes')

roster = {c: [norm(m) for m in e['miembros']] for c, e in equipos.items()}

# -----------------------------------------------------------------------------
# ASIGNACIONES MANUALES (agentes que no estan en el HC pero pertenecen a un
# equipo, confirmado por la columna Coordinador en 'lineas facturadas'):
#   - Coordinador AMIR JOSE RODRIGUEZ -> equipo RENOVACION (Marinel Moreno)
#   - Coordinador JULIO CESAR ARAUZ   -> equipo SOHO
# -----------------------------------------------------------------------------
OVERRIDES = [
    # (aliases_posibles, campana, nombre_canonico_unificado)
    (['MILENA MOJICA', 'LILIAN MILENA MOJICA DE LAS SALAS'],
     'RENOVACION', 'LILIAN MILENA MOJICA DE LAS SALAS'),
    (['LIZETH GONZALEZ', 'LIZETH MILENA GONZALEZ RODRIGUEZ'],
     'RENOVACION', 'LIZETH MILENA GONZALEZ RODRIGUEZ'),
]

def match_override(nombre, aliases):
    """Match de override igual de estricto que match_member:
    el PRIMER token (nombre de pila) debe estar en el alias y >=2 tokens en común.
    Evita falsos positivos tipo 'DANNA DE LAS SALAS' vs 'LILIAN MILENA MOJICA DE LAS SALAS'."""
    n = norm(nombre)
    if not n: return False
    toks = n.split()
    first = toks[0]
    t = set(toks)
    for al in aliases:
        at = tokens(al)
        if first in at and len(t & at) >= 2:
            return True
    return False

def equipo_de(nombre):
    """Devuelve (campana, miembro_normalizado) para un nombre de agente."""
    n = norm(nombre) if nombre else ''
    # 1) Overrides manuales (match estricto: primer token + >=2 tokens comunes)
    for aliases, camp, canonico in OVERRIDES:
        if match_override(nombre, aliases):
            return camp, norm(canonico)
    # 2) Match contra el roster de cada equipo
    for campana, r in roster.items():
        m = match_member(nombre, r)
        if m:
            return campana, m
    return None, n if n else 'SIN AGENTE'

# =============================================================================
# 2. VICI - MARCACIONES
# =============================================================================
import pyxlsb

xlsb = glob.glob(os.path.join(PROD, 'Reporte_soho_junio_25 (1).xlsb'))[0]
wb = pyxlsb.open_workbook(xlsb)

def read_sheet(sn):
    with wb.get_sheet(sn) as ws:
        header = None
        rows = []
        for r in ws.rows():
            vals = [c.v for c in r]
            if header is None:
                header = vals; continue
            rows.append(vals)
    return header, rows

vici_header, vici_rows = read_sheet('Vici')
vi = {norm(h): i for i, h in enumerate(vici_header)}
c_full = get_col(vi, 'full_name')
c_status = get_col(vi, 'status_name')
c_mes = get_col(vi, 'Mes')
c_phone = get_col(vi, 'phone_number_dialed')
c_seg = get_col(vi, 'length_in_sec')
c_src = get_col(vi, 'source_id', 'lead_id', 'account')
print('Vici:', len(vici_rows), 'llamadas')

# Definición alineada con el consolidado previo: no-contacto = No contesta + Buzón/Contestador
STATUS_NO_CONTACTO = {'NO CONTESTA', 'LLAMADA NO CONTESTA', 'CONTESTADOR AUTOMATICO',
                      'BUZON DE VOZ'}

llamadas = []
for v in vici_rows:
    nombre = v[c_full] if c_full is not None and c_full < len(v) else None
    status = norm(v[c_status]) if c_status is not None and c_status < len(v) else ''
    mes = norm(v[c_mes]) if c_mes is not None and c_mes < len(v) else ''
    phone = str(v[c_phone]).replace('.0','') if c_phone is not None and c_phone < len(v) and v[c_phone] is not None else ''
    seg = to_num(v[c_seg]) if c_seg is not None and c_seg < len(v) else 0
    src = cuenta12(v[c_src]) if c_src is not None and c_src < len(v) and v[c_src] is not None else ''
    campana, miembro = equipo_de(nombre)
    llamadas.append({
        'nombre': norm(nombre), 'campana': campana, 'miembro': miembro,
        'status': status, 'mes': mes, 'phone': phone, 'seg': seg, 'src': src,
    })

def stats_vici(rows):
    total = len(rows)
    no_contesta = sum(1 for r in rows if r['status'] in {'NO CONTESTA','LLAMADA NO CONTESTA'})
    contactados = sum(1 for r in rows if r['status'] and r['status'] not in STATUS_NO_CONTACTO)
    phones = set(r['phone'] for r in rows if r['phone'])
    seg_total = sum(r['seg'] for r in rows)
    meses = Counter(r['mes'] for r in rows)
    return {
        'llamadas': total,
        'no_contesta': no_contesta,
        'contactados': contactados,
        'tasa_contacto': round(contactados / total * 100, 1) if total else 0,
        'numeros_unicos': len(phones),
        'intentos_prom': round(total / len(phones), 2) if phones else 0,
        'duracion_prom_s': round(seg_total / total, 0) if total else 0,
        'por_mes': {m: meses.get(m, 0) for m in MESES_ORDEN},
    }

def vici_por_mes(rows):
    """Métricas de marcación desglosadas por mes (enero-junio)."""
    d = {m: {'llamadas': 0, 'contactados': 0, 'no_contesta': 0, 'numeros': 0,
             'seg': 0, 'duracion_prom_s': 0.0}
         for m in MESES_ORDEN}
    for r in rows:
        m = r['mes'] or ''
        if m not in d: continue
        d[m]['llamadas'] += 1
        d[m]['seg'] += r['seg']
        if r['status'] and r['status'] not in STATUS_NO_CONTACTO:
            d[m]['contactados'] += 1
        if r['status'] in {'NO CONTESTA','LLAMADA NO CONTESTA'}:
            d[m]['no_contesta'] += 1
    for m in MESES_ORDEN:
        d[m]['numeros'] = len(set(r['phone'] for r in rows
                                  if (r['mes'] or '') == m and r['phone']))
        d[m]['duracion_prom_s'] = round(d[m]['seg'] / d[m]['llamadas'], 0) if d[m]['llamadas'] else 0.0
    return d

# Global + por campaña
global_vici = stats_vici(llamadas)
global_por_mes = {m: stats_vici([r for r in llamadas if (r['mes'] or '') == m])
                  for m in MESES_ORDEN}
vici_por_campana = {}
for c in equipos:
    rows_c = [r for r in llamadas if r['campana'] == c]
    vici_por_campana[c] = stats_vici(rows_c)

# Por campaña y mes (para el selector de campaña del dashboard)
vici_campana_por_mes = {}
for c in equipos:
    vici_campana_por_mes[c] = {
        m: stats_vici([r for r in llamadas if r['campana'] == c and (r['mes'] or '') == m])
        for m in MESES_ORDEN
    }

# Por agente: clave (campana, miembro_normalizado) — unifica nombres cortos y completos
vici_por_agente = defaultdict(list)
for r in llamadas:
    vici_por_agente[(r['campana'] or 'SIN EQUIPO', r['miembro'] or 'SIN AGENTE')].append(r)

# Tipificaciones
tipif_global = Counter(r['status'] or '(SIN TIPIFICACION)' for r in llamadas)
tipif_renov = Counter(r['status'] or '(SIN TIPIFICACION)' for r in llamadas if r['campana'] == 'RENOVACION')
tipif_glob_por_mes = {}
tipif_renov_por_mes = {}
for m in MESES_ORDEN:
    tipif_glob_por_mes[m] = Counter(r['status'] or '(SIN TIPIFICACION)'
                                    for r in llamadas if (r['mes'] or '') == m).most_common()
    tipif_renov_por_mes[m] = Counter(r['status'] or '(SIN TIPIFICACION)'
                                     for r in llamadas if r['campana'] == 'RENOVACION' and (r['mes'] or '') == m).most_common()

# =============================================================================
# 3. VENTAS
# =============================================================================
def load_ventas_moviles():
    header, rows = read_sheet('Ventas Moviles')
    idx = {norm(h): i for i, h in enumerate(header)}
    c_ag = get_col(idx, 'Agentes', 'AGENTE', 'Vendedor')
    c_tram = get_col(idx, 'Tramite')
    c_mrc = get_col(idx, 'MRC Actual')
    c_mrc_ant = get_col(idx, 'MRC Anterior')
    c_mes = get_col(idx, 'Mes', 'MES')
    c_fecha = get_col(idx, 'Fecha')
    out = []
    for v in rows:
        ag = v[c_ag] if c_ag is not None and c_ag < len(v) else None
        campana, miembro = equipo_de(ag)
        mes = ''
        if c_mes is not None and c_mes < len(v):
            mes = mes_de(v[c_mes])
        if not mes and c_fecha is not None and c_fecha < len(v):
            mes = mes_de(v[c_fecha])
        out.append({'nombre': norm(ag), 'campana': campana, 'miembro': miembro, 'mes': mes,
                    'tramite': norm(v[c_tram]) if c_tram is not None and c_tram < len(v) else '',
                    'mrc': to_num(v[c_mrc]) if c_mrc is not None and c_mrc < len(v) else 0,
                    'mrc_ant': to_num(v[c_mrc_ant]) if c_mrc_ant is not None and c_mrc_ant < len(v) else 0})
    return out

def load_ventas_fijo():
    header, rows = read_sheet('Ventas Fijo')
    idx = {norm(h): i for i, h in enumerate(header)}
    c_ven = get_col(idx, 'Vendedor', 'VENDEDOR')
    c_rgu = get_col(idx, 'Rgu')
    c_mrc = get_col(idx, 'Mrc')
    c_st = get_col(idx, 'Status Final')
    c_mes = get_col(idx, 'Mes Reg', 'Mes', 'MES')
    out = []
    for v in rows:
        ven = v[c_ven] if c_ven is not None and c_ven < len(v) else None
        campana, miembro = equipo_de(ven)
        mes = mes_de(v[c_mes]) if c_mes is not None and c_mes < len(v) else ''
        out.append({'nombre': norm(ven), 'campana': campana, 'miembro': miembro, 'mes': mes,
                    'rgu': to_num(v[c_rgu]) if c_rgu is not None and c_rgu < len(v) else 0,
                    'mrc': to_num(v[c_mrc]) if c_mrc is not None and c_mrc < len(v) else 0,
                    'status': norm(v[c_st]) if c_st is not None and c_st < len(v) else ''})
    return out

def load_lineas_facturadas():
    header, rows = read_sheet('lineas facturadas')
    idx = {norm(h): i for i, h in enumerate(header)}
    c_ven = get_col(idx, 'Vendedor')
    c_tip = get_col(idx, 'Tipoventa', 'Tipo de Venta')
    c_val = get_col(idx, 'Valor plan')
    c_cant = get_col(idx, 'Cantidad lineas')
    c_conv = get_col(idx, 'Venta convergente')
    c_mes = get_col(idx, 'Fecha fact', 'Fecha Fact', 'Fecha')
    out = []
    for v in rows:
        ven = v[c_ven] if c_ven is not None and c_ven < len(v) else None
        campana, miembro = equipo_de(ven)
        mes = mes_de(v[c_mes]) if c_mes is not None and c_mes < len(v) else ''
        out.append({'nombre': norm(ven), 'campana': campana, 'miembro': miembro, 'mes': mes,
                    'tipoventa': norm(v[c_tip]) if c_tip is not None and c_tip < len(v) else '',
                    'valor': to_num(v[c_val]) if c_val is not None and c_val < len(v) else 0,
                    'cant': to_num(v[c_cant]) if c_cant is not None and c_cant < len(v) else 0,
                    'convergente': norm(v[c_conv]) if c_conv is not None and c_conv < len(v) else ''})
    return out

ventas_moviles = load_ventas_moviles()
ventas_fijo = load_ventas_fijo()
lineas_fact = load_lineas_facturadas()
print('Ventas moviles:', len(ventas_moviles), '| Ventas fijo:', len(ventas_fijo),
      '| Lineas facturadas:', len(lineas_fact))
print('  Ventas moviles por mes:', {m: sum(1 for s in ventas_moviles if s['mes'] == m) for m in MESES_ORDEN})
print('  Ventas fijo por mes:   ', {m: sum(1 for s in ventas_fijo if s['mes'] == m) for m in MESES_ORDEN})
print('  Lineas por mes:        ', {m: sum(1 for s in lineas_fact if s['mes'] == m) for m in MESES_ORDEN})

def ventas_por_agente(sales):
    d = defaultdict(lambda: {'n': 0, 'mrc': 0.0, 'rgu': 0.0, 'valor': 0.0,
                             'tramite': Counter(), 'tipoventa': Counter(), 'status': Counter()})
    for s in sales:
        k = (s['campana'] or 'SIN EQUIPO', s['miembro'] or 'SIN AGENTE')
        a = d[k]
        a['n'] += 1
        a['mrc'] += s.get('mrc', 0) or 0
        # RGU = columna Rgu (Ventas Fijo) o Cantidad lineas (lineas facturadas)
        a['rgu'] += (s.get('rgu', 0) or 0) + (s.get('cant', 0) or 0)
        a['valor'] += s.get('valor', 0) or 0
        if s.get('tramite'): a['tramite'][s['tramite']] += 1
        if s.get('tipoventa'): a['tipoventa'][s['tipoventa']] += 1
        if s.get('status'): a['status'][s['status']] += 1
    return d

# Clasificación de trámites/tipos de venta (definición de Amir):
#   RENOVACION -> renovación (MRC anterior/actual solo de este grupo)
#   LINEA NUEVA + PORTABILIDAD + PORTABILIDAD EXTERNA -> VENTA CRUZADA
#   PORTABILIDAD INTERNA NO cuenta (no es renovación ni cruzada)
TRAMITE_RENOV = {'RENOVACION'}
TRAMITE_CRUZ = {'LINEA NUEVA', 'PORTABILIDAD', 'PORTABILIDAD EXTERNA'}
TIPOVENTA_RENOV = {'RENOVACION'}
TIPOVENTA_CRUZ = {'LINEA NUEVA', 'PORTABILIDAD EXTERNA', 'PROTABILIDAD EXTERNA'}  # typo 'PROTABILIDAD' en los datos

def ventas_por_mes_agente():
    """ventas por agente y mes: {clave: {mes: {movil,fijo,lineas,mrc,rgu,tramite,tipoventa,
    renov,cruz,mrc_ant_renov,mrc_act_renov,mrc_cruz,rgu_renov,rgu_cruz}}}"""
    d = defaultdict(lambda: {m: {'movil': 0, 'fijo': 0, 'lineas': 0, 'mrc': 0.0, 'rgu': 0.0,
                                 'tramite': {}, 'tipoventa': {},
                                 'renov': 0, 'cruz': 0,
                                 'mrc_ant_renov': 0.0, 'mrc_act_renov': 0.0, 'mrc_cruz': 0.0,
                                 'rgu_renov': 0.0, 'rgu_cruz': 0.0}
                             for m in MESES_ORDEN})
    for sales, tipo in [(ventas_moviles, 'movil'), (ventas_fijo, 'fijo'), (lineas_fact, 'lineas')]:
        for s in sales:
            k = (s['campana'] or 'SIN EQUIPO', s['miembro'] or 'SIN AGENTE')
            m = s.get('mes') or ''
            if m not in d[k]: continue
            v = d[k][m]
            v[tipo] += 1
            v['mrc'] += (s.get('mrc') or 0) + (s.get('valor') or 0)
            v['rgu'] += (s.get('rgu') or 0) + (s.get('cant') or 0)
            # VENTAS = trámites de la hoja 'Ventas Moviles' (una fila = una venta).
            # RGU = líneas de la hoja 'lineas facturadas' (cantidad de líneas).
            # Fuentes separadas: NO mezclar trámites con líneas en la misma métrica.
            if tipo == 'movil' and s.get('tramite'):
                v['tramite'][s['tramite']] = v['tramite'].get(s['tramite'], 0) + 1
                if s['tramite'] in TRAMITE_RENOV:
                    v['renov'] += 1
                    v['mrc_ant_renov'] += s.get('mrc_ant') or 0
                    v['mrc_act_renov'] += s.get('mrc') or 0
                elif s['tramite'] in TRAMITE_CRUZ:
                    v['cruz'] += 1
                    v['mrc_cruz'] += s.get('mrc') or 0
            if tipo == 'lineas' and s.get('tipoventa'):
                v['tipoventa'][s['tipoventa']] = v['tipoventa'].get(s['tipoventa'], 0) + 1
                if s['tipoventa'] in TIPOVENTA_RENOV:
                    v['rgu_renov'] += s.get('cant') or 0
                elif s['tipoventa'] in TIPOVENTA_CRUZ:
                    v['rgu_cruz'] += s.get('cant') or 0
    return d

vm_por = ventas_por_agente(ventas_moviles)
vf_por = ventas_por_agente(ventas_fijo)
lf_por = ventas_por_agente(lineas_fact)
vpm = ventas_por_mes_agente()

# =============================================================================
# 4. VENTA CRUZADA (cross-sell base)
# =============================================================================
cs = glob.glob(os.path.join(PROD, 'bases', 'Base para campaña cross-sell movil Nextphone.xlsx'))
cross_por_agente = Counter()
cross_mes = defaultdict(Counter)   # miembro -> Counter(mes)
if cs:
    cwb = openpyxl.load_workbook(cs[0], read_only=True, data_only=True)
    for sn in ['Base para gestionar CC ', 'con contacto telefonos']:
        if sn not in cwb.sheetnames: continue
        ws = cwb[sn]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        hdr = {norm(h): i for i, h in enumerate(header or [])}
        col = get_col(hdr, 'Ejecutivo', 'AGENTE', 'Agentes', 'Ejecutiva')
        col_fecha = get_col(hdr, 'FECHA', 'Fecha')
        if col is None: continue
        for r in rows:
            if r is None: continue
            ag = r[col] if col < len(r) else None
            if ag:
                _, miembro = equipo_de(ag)
                cross_por_agente[(norm(ag), miembro)] += 1
                mes = mes_de(r[col_fecha]) if col_fecha is not None and col_fecha < len(r) else ''
                if mes in MESES_ORDEN:
                    cross_mes[miembro][mes] += 1
    cwb.close()
print('Gestiones cross-sell:', sum(cross_por_agente.values()))
print('  Cross-sell por mes:', {m: sum(cm.get(m, 0) for cm in cross_mes.values()) for m in MESES_ORDEN})

# =============================================================================
# 4c. WINBACK - base de clientes a recuperar (campaña de Julio)
# =============================================================================
winback = {'detalle': [], 'por_ejecutiva': [], 'por_razon': [], 'resumen': {}}
wb_glob = glob.glob(os.path.join(PROD, 'bases', 'Winback SOHO*.xlsx'))
if wb_glob:
    wbw = openpyxl.load_workbook(wb_glob[0], read_only=True, data_only=True)
    wws = wbw['BASE 1.3'] if 'BASE 1.3' in wbw.sheetnames else wbw[wbw.sheetnames[0]]
    rows = wws.iter_rows(values_only=True)
    header = next(rows, None)
    hdr = {norm(h): i for i, h in enumerate(header or [])}
    c_eje = get_col(hdr, 'Ejecutivas', 'Ejecutivo')
    c_cli = get_col(hdr, 'CLIENTE', 'Cliente')
    c_plan = get_col(hdr, 'DESCRIPCIÓN DE PLAN', 'Descripcion de plan', 'Plan')
    c_raz = get_col(hdr, 'RAZÓN DE SALIDA', 'Razon de salida', 'Razón de salida')
    c_rgu = get_col(hdr, 'RGU')
    c_mrc = get_col(hdr, 'MRC Acum', 'MRC', 'MRC acumulado')
    c_cont = get_col(hdr, 'Persona de contacto')
    c_corr = get_col(hdr, 'Correo')
    c_t1 = get_col(hdr, 'Contacto 1')
    c_t2 = get_col(hdr, 'Contacto 2')
    for r in rows:
        if r is None:
            continue
        def g(c): return r[c] if c is not None and c < len(r) else None
        eje = norm(g(c_eje)) if g(c_eje) else 'SIN EJECUTIVA'
        razon = norm(g(c_raz)) if g(c_raz) else 'SIN RAZON'
        winback['detalle'].append({
            'e': eje, 'cli': str(g(c_cli) or '').strip(),
            'plan': str(g(c_plan) or '').strip(),
            'razon': razon, 'rgu': to_num(g(c_rgu)), 'mrc': to_num(g(c_mrc)),
            'cont': str(g(c_cont) or '').strip(), 'correo': str(g(c_corr) or '').strip(),
            't1': str(g(c_t1) or '').strip(), 't2': str(g(c_t2) or '').strip(),
        })
    wbw.close()
    por_eje = defaultdict(lambda: {'clientes': 0, 'rgu': 0.0, 'mrc': 0.0})
    por_raz = defaultdict(lambda: {'clientes': 0, 'rgu': 0.0, 'mrc': 0.0})
    for d in winback['detalle']:
        por_eje[d['e']]['clientes'] += 1
        por_eje[d['e']]['rgu'] += d['rgu']
        por_eje[d['e']]['mrc'] += d['mrc']
        por_raz[d['razon']]['clientes'] += 1
        por_raz[d['razon']]['rgu'] += d['rgu']
        por_raz[d['razon']]['mrc'] += d['mrc']
    winback['por_ejecutiva'] = sorted(
        [{'e': k, 'clientes': v['clientes'], 'rgu': v['rgu'], 'mrc': v['mrc']}
         for k, v in por_eje.items()], key=lambda x: -x['clientes'])
    winback['por_razon'] = sorted(
        [{'r': k, 'clientes': v['clientes'], 'rgu': v['rgu'], 'mrc': v['mrc']}
         for k, v in por_raz.items()], key=lambda x: -x['clientes'])
    winback['resumen'] = {
        'clientes': len(winback['detalle']),
        'rgu': round(sum(d['rgu'] for d in winback['detalle']), 1),
        'mrc': round(sum(d['mrc'] for d in winback['detalle']), 2),
        'razones': len(winback['por_razon']),
    }
print('Winback:', winback['resumen'])

# =============================================================================
# 4d. CROSS-SELL - detalle de gestiones (campaña de Julio)
# =============================================================================
cross_data = {'detalle': [], 'por_ejecutiva': [], 'resumen': {}}
if cs:
    cwb2 = openpyxl.load_workbook(cs[0], read_only=True, data_only=True)
    sin_tel_base = 0
    for sn in ['Base para gestionar CC ', 'con contacto telefonos']:
        if sn not in cwb2.sheetnames:
            continue
        ws = cwb2[sn]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        hdr = {norm(h): i for i, h in enumerate(header or [])}
        c_eje = get_col(hdr, 'Ejecutivo', 'AGENTE', 'Agentes', 'Ejecutiva')
        c_fecha = get_col(hdr, 'FECHA', 'Fecha')
        c_cuenta = get_col(hdr, 'CUENTA', 'Cuenta')
        c_cust = get_col(hdr, 'CUST_NO')
        c_cli = get_col(hdr, 'ALIAS', 'NOMBRE_CIS', 'Nombre')
        c_prov = get_col(hdr, 'PROVINCIA', 'Provincia')
        c_tel = get_col(hdr, 'contact_mobile', 'Contacto 1', 'Telefono')
        c_est = get_col(hdr, 'Estatus', 'Status', 'Estado')
        c_com = get_col(hdr, 'COMENTARIOS', 'Comentarios')
        for r in rows:
            if r is None:
                continue
            def g(c): return r[c] if c is not None and c < len(r) else None
            eje = norm(g(c_eje)) if g(c_eje) else 'SIN EJECUTIVA'
            tel = str(g(c_tel) or '').replace('.0', '').strip()
            cross_data['detalle'].append({
                'fecha': str(g(c_fecha) or '').strip(),
                'cuenta': str(g(c_cuenta) or '').replace('.0', '').strip(),
                'cust': str(g(c_cust) or '').replace('.0', '').strip(),
                'cli': str(g(c_cli) or '').strip(),
                'prov': str(g(c_prov) or '').strip(),
                'tel': tel,
                'e': eje,
                'est': str(g(c_est) or '').strip(),
                'com': str(g(c_com) or '').strip(),
            })
    if 'sin contactos telefonicos' in cwb2.sheetnames:
        ws = cwb2['sin contactos telefonicos']
        for _ in ws.iter_rows(min_row=2, values_only=True):
            sin_tel_base += 1
    cwb2.close()
    por_eje_c = Counter(d['e'] for d in cross_data['detalle'])
    cross_data['por_ejecutiva'] = sorted(
        [{'e': k, 'gestiones': v} for k, v in por_eje_c.items()],
        key=lambda x: -x['gestiones'])
    cross_data['resumen'] = {
        'gestiones': len(cross_data['detalle']),
        'con_tel': sum(1 for d in cross_data['detalle'] if d['tel'] and d['tel'] != '#N/A'),
        'sin_tel_base': sin_tel_base,
        'ejecutivas': len(por_eje_c),
    }
print('Cross-sell detalle:', cross_data['resumen'])

# =============================================================================
# 4b. RECORRIDO DE LA BASE (bases de ViciDial vs marcaciones RENOVACION)
# =============================================================================
# Llave de cruce: TELÉFONO (últimos 8 dígitos). Las bases descargadas de
# ViciDial (Base de datos Nextphone + Base Móvil sin fijo) se cruzan contra las
# marcaciones de los agentes RENOVACION para ver cuánto se recorrió de la base:
# marcados, contactados, sin tocar, % de avance, por agente y por mes.
BASES_RECORRIDO = [
    # (patron, nombre_corto, col_cuenta12, col_tel, col_nombre, col_ejecutiva)
    ('Base de datos Nextphone 14072026*.xlsx', 'NEXTPHONE', 4, 10, 5, 0),
    ('Base Móvil sin fijo 02072026 ren*.xlsx', 'MOVIL',     3, 9,  4, None),
]

def tel8(v):
    """Teléfono a 8 dígitos: últimos 8 dígitos (o todos si hay menos).
    Equivale a la antigua phone8 pero sin anteponer '6' a los de 7 dígitos
    (los fijos de Vici no deben convertirse en móviles para el match)."""
    s = str(v).replace('.0', '') if v is not None else ''
    s = re.sub(r'\D', '', s)
    return s[-8:] if s else ''

base_rows = []
for patron, corto, c_cuenta, c_tel, c_nom, c_eje in BASES_RECORRIDO:
    f = glob.glob(os.path.join(PROD, 'bases', patron))
    if not f:
        print('Base no encontrada:', patron); continue
    bwb = openpyxl.load_workbook(f[0], read_only=True, data_only=True)
    bws = bwb[bwb.sheetnames[0]]
    seen = set()
    for r in bws.iter_rows(values_only=True):
        if r is None: continue
        tel = tel8(r[c_tel]) if c_tel is not None and c_tel < len(r) else ''
        if not tel or tel in seen: continue
        seen.add(tel)
        cuenta = cuenta12(r[c_cuenta]) if c_cuenta is not None and c_cuenta < len(r) else ''
        nombre = norm(r[c_nom]) if c_nom is not None and c_nom < len(r) and r[c_nom] else ''
        ejecutiva = norm(r[c_eje]) if c_eje is not None and c_eje < len(r) and r[c_eje] else ''
        if ejecutiva in ('MOVIL', 'SIN FIJO'):
            ejecutiva = ''
        base_rows.append({'base': corto, 'cuenta': cuenta, 'tel': tel,
                          'nombre': nombre, 'ejecutiva': ejecutiva})
    bwb.close()
print('Bases de recorrido:', len(base_rows), 'telefonos unicos')

# Cruce contra las marcaciones de la campaña RENOVACION
renov_llam = [r for r in llamadas if r['campana'] == 'RENOVACION']
por_tel = defaultdict(list)
for r in renov_llam:
    t = tel8(r['phone'])
    if t:
        por_tel[t].append(r)

base_tels = {r['tel'] for r in base_rows}
marcados_tels = set(por_tel) & base_tels
contactados_tels = set()
for t in marcados_tels:
    if any(r['status'] and r['status'] not in STATUS_NO_CONTACTO for r in por_tel[t]):
        contactados_tels.add(t)

rec_por_mes = {}
for m in MESES_ORDEN:
    rec_por_mes[m] = {
        'marcados': len({t for t in marcados_tels
                         if any((r['mes'] or '') == m for r in por_tel[t])}),
        'contactados': len({t for t in contactados_tels
                            if any((r['mes'] or '') == m and r['status'] not in STATUS_NO_CONTACTO
                                   for r in por_tel[t])}),
    }

rec_por_agente = defaultdict(lambda: {'tels': set(), 'llamadas': 0, 'contactados': set()})
for t in marcados_tels:
    for r in por_tel[t]:
        a = rec_por_agente[r['miembro'] or 'SIN AGENTE']
        a['tels'].add(t)
        a['llamadas'] += 1
        if r['status'] and r['status'] not in STATUS_NO_CONTACTO:
            a['contactados'].add(t)

rec_por_base = {}
for b in ['NEXTPHONE', 'MOVIL']:
    bt = {r['tel'] for r in base_rows if r['base'] == b}
    rec_por_base[b] = {'total': len(bt), 'marcados': len(bt & marcados_tels)}

recorrido_base = {
    'llave': 'telefono (8 digitos)',
    'total': len(base_rows),
    'marcados': len(marcados_tels),
    'sin_tocar': len(base_rows) - len(marcados_tels),
    'pct_recorrido': round(len(marcados_tels) / len(base_rows) * 100, 1) if base_rows else 0,
    'contactados': len(contactados_tels),
    'intentos_prom': round(sum(len(por_tel[t]) for t in marcados_tels) / len(marcados_tels), 2) if marcados_tels else 0,
    'por_mes': rec_por_mes,
    'por_base': rec_por_base,
    'por_agente': [{'agente': k, 'telefonos': len(v['tels']), 'llamadas': v['llamadas'],
                    'contactados': len(v['contactados'])}
                   for k, v in rec_por_agente.items()],
    'detalle': sorted(
        [{'b': r['base'], 'c': r['cuenta'], 't': r['tel'], 'n': r['nombre'],
          'e': r['ejecutiva'],
          'm': 1 if r['tel'] in marcados_tels else 0,
          'co': 1 if r['tel'] in contactados_tels else 0,
          'll': len(por_tel[r['tel']]) if r['tel'] in por_tel else 0,
          'ag': por_tel[r['tel']][-1]['miembro'] if r['tel'] in por_tel and por_tel[r['tel']] else '',
          'ms': [m for m in MESES_ORDEN
                 if any((rr['mes'] or '') == m for rr in por_tel.get(r['tel'], []))]}
         for r in base_rows],
        key=lambda x: -x['m']),
}
print('Recorrido base: %d/%d marcados (%.1f%%)' % (
    len(marcados_tels), len(base_rows),
    round(len(marcados_tels) / len(base_rows) * 100, 1) if base_rows else 0))

# =============================================================================
# 5. ARMAR DATOS POR AGENTE (ranking top performance)
# =============================================================================
agentes = {}
all_keys = set(vici_por_agente.keys()) | set(vm_por.keys()) | set(vf_por.keys()) | set(lf_por.keys())

for k in all_keys:
    campana, miembro = k
    v = vici_por_agente.get(k, [])
    vs = stats_vici(v)
    vm_ag = vici_por_mes(v)
    vm_s = vm_por.get(k)
    vf_s = vf_por.get(k)
    lf_s = lf_por.get(k)
    ventas = (vm_s['n'] if vm_s else 0) + (vf_s['n'] if vf_s else 0) + (lf_s['n'] if lf_s else 0)
    rgu = (vf_s['rgu'] if vf_s else 0) + (lf_s['rgu'] if lf_s else 0)
    mrc_total = (vm_s['mrc'] if vm_s else 0) + (vf_s['mrc'] if vf_s else 0) + (lf_s['valor'] if lf_s else 0)
    cruzada = sum(n for (n_, m_), n in cross_por_agente.items() if m_ == miembro)
    vpk = vpm[k]
    agentes[k] = {
        'agente': miembro,
        'campana': campana,
        'llamadas': vs['llamadas'],
        'no_contesta': vs['no_contesta'],
        'contactados': vs['contactados'],
        'tasa_contacto': vs['tasa_contacto'],
        'numeros': vs['numeros_unicos'],
        'intentos_prom': vs['intentos_prom'],
        'duracion_prom_s': vs['duracion_prom_s'],
        'ventas_movil': vm_s['n'] if vm_s else 0,
        'ventas_fijo': vf_s['n'] if vf_s else 0,
        'lineas_fact': lf_s['n'] if lf_s else 0,
        'ventas_total': ventas,
        'rgu': round(rgu, 1),
        'mrc': round(mrc_total, 2),
        'venta_cruzada': sum(vpk[m]['cruz'] for m in MESES_ORDEN),
        'gestiones_cross': cruzada,
        'ventas_renov': sum(vpk[m]['renov'] for m in MESES_ORDEN),
        'ventas_cruz': sum(vpk[m]['cruz'] for m in MESES_ORDEN),
        'mrc_ant_renov': round(sum(vpk[m]['mrc_ant_renov'] for m in MESES_ORDEN), 2),
        'mrc_act_renov': round(sum(vpk[m]['mrc_act_renov'] for m in MESES_ORDEN), 2),
        'mrc_cruz': round(sum(vpk[m]['mrc_cruz'] for m in MESES_ORDEN), 2),
        'mrc_prom': 0.0,
        'rgu_renov': int(round(sum(vpk[m]['rgu_renov'] for m in MESES_ORDEN))),
        'rgu_cruz': int(round(sum(vpk[m]['rgu_cruz'] for m in MESES_ORDEN))),
        'tramites': dict(vm_s['tramite']) if vm_s else {},
        'tipoventa': dict(lf_s['tipoventa']) if lf_s else {},
        'status_fijo': dict(vf_s['status']) if vf_s else {},
        'por_mes': vs['por_mes'],
        'contactados_por_mes': {m: vm_ag[m]['contactados'] for m in MESES_ORDEN},
        'no_contesta_por_mes': {m: vm_ag[m]['no_contesta'] for m in MESES_ORDEN},
        'numeros_por_mes': {m: vm_ag[m]['numeros'] for m in MESES_ORDEN},
        'duracion_por_mes': {m: vm_ag[m]['duracion_prom_s'] for m in MESES_ORDEN},
        'ventas_por_mes': vpm[k],
        'cruzada_por_mes': {m: cross_mes[miembro].get(m, 0) for m in MESES_ORDEN},
        # Eficiencia (calculada al final para reflejar merges)
        'ef_ventas_x_llamada': 0.0,
        'ef_conv_pct': 0.0,
        'ef_mrc_x_venta': 0.0,
        'en_roster': False,
        'sin_actividad': False,
    }

# --- Merge de agentes SIN EQUIPO con formatos de nombre distintos ---
# Ej: 'MILENA MOJICA' y 'LILIAN MILENA MOJICA DE LAS SALAS' son la misma persona.
STOP = {'DE', 'DEL', 'LA', 'LAS', 'LOS', 'EL', 'Y', 'SAN', 'SANTA', 'MC', 'MA', 'VON'}

def sig_tokens(nombre):
    return {t for t in norm(nombre).split() if t not in STOP}

sin_keys = [k for k, a in agentes.items() if a['campana'] == 'SIN EQUIPO']
merged = set()
for i, k1 in enumerate(sin_keys):
    if k1 in merged: continue
    t1 = sig_tokens(k1[1])
    grupo = [k1]
    for k2 in sin_keys[i+1:]:
        if k2 in merged: continue
        t2 = sig_tokens(k2[1])
        if len(t1 & t2) >= 2:  # comparten al menos 2 apellidos significativos
            grupo.append(k2)
    if len(grupo) > 1:
        base = agentes[k1]
        for k2 in grupo[1:]:
            b = agentes[k2]
            for campo in ['llamadas','no_contesta','contactados','numeros','ventas_movil',
                          'ventas_fijo','lineas_fact','ventas_total','venta_cruzada',
                          'gestiones_cross','ventas_renov','ventas_cruz']:
                base[campo] += b[campo]
            base['rgu'] += b['rgu']
            base['mrc'] += b['mrc']
            base['mrc_ant_renov'] += b['mrc_ant_renov']
            base['mrc_act_renov'] += b['mrc_act_renov']
            base['mrc_cruz'] += b['mrc_cruz']
            base['rgu_renov'] += b['rgu_renov']
            base['rgu_cruz'] += b['rgu_cruz']
            for m in MESES_ORDEN:
                base['por_mes'][m] = base['por_mes'].get(m, 0) + b['por_mes'].get(m, 0)
                base['contactados_por_mes'][m] += b['contactados_por_mes'].get(m, 0)
                base['no_contesta_por_mes'][m] += b['no_contesta_por_mes'].get(m, 0)
                base['numeros_por_mes'][m] += b['numeros_por_mes'].get(m, 0)
                base['cruzada_por_mes'][m] += b['cruzada_por_mes'].get(m, 0)
                # duración: promedio ponderado por llamadas del mes
                llam_bb = b['por_mes'].get(m, 0)
                llam_bb_antes = base['por_mes'].get(m, 0) - llam_bb
                if llam_bb and (llam_bb_antes + llam_bb):
                    base['duracion_por_mes'][m] = round(
                        (base['duracion_por_mes'].get(m, 0) * llam_bb_antes + b['duracion_por_mes'].get(m, 0) * llam_bb)
                        / (llam_bb_antes + llam_bb), 0) if llam_bb_antes else b['duracion_por_mes'].get(m, 0)
                bm = b['ventas_por_mes'].get(m, {})
                for kk in ('movil', 'fijo', 'lineas', 'mrc', 'rgu', 'renov', 'cruz',
                           'mrc_ant_renov', 'mrc_act_renov', 'mrc_cruz', 'rgu_renov', 'rgu_cruz'):
                    base['ventas_por_mes'][m][kk] += bm.get(kk, 0)
                for kk in ('tramite', 'tipoventa'):
                    for k_, v_ in bm.get(kk, {}).items():
                        base['ventas_por_mes'][m][kk][k_] = base['ventas_por_mes'][m][kk].get(k_, 0) + v_
            for d, d2 in ((base['tramites'], b['tramites']), (base['tipoventa'], b['tipoventa']),
                          (base['status_fijo'], b['status_fijo'])):
                for k_, v_ in d2.items():
                    d[k_] = d.get(k_, 0) + v_
            merged.add(k2)
        # tasa de contacto recalculada
        base['tasa_contacto'] = round(base['contactados'] / base['llamadas'] * 100, 1) if base['llamadas'] else 0
        base['intentos_prom'] = round(base['llamadas'] / base['numeros'], 2) if base['numeros'] else 0
        # usar el nombre más completo del grupo
        base['agente'] = max([agentes[k]['agente'] for k in grupo], key=len)
        # BORRAR todas las claves del grupo y reinsertar con el nombre más completo
        for k in grupo:
            agentes.pop(k, None)
        agentes[(base['campana'], norm(base['agente']))] = base

# --- Metricas de eficiencia (despues de merges) ---
for a in agentes.values():
    a['ef_ventas_x_llamada'] = round(a['ventas_total'] / a['llamadas'], 3) if a['llamadas'] else 0.0
    a['ef_conv_pct'] = round(a['ventas_total'] / a['llamadas'] * 100, 2) if a['llamadas'] else 0.0
    a['ef_mrc_x_venta'] = round(a['mrc'] / a['ventas_total'], 2) if a['ventas_total'] else 0.0
    # MRC promedio (requisito de la reunion): MRC actual de renovacion / RGU renovacion
    a['mrc_prom'] = round(a['mrc_act_renov'] / a['rgu_renov'], 2) if a['rgu_renov'] else 0.0

# --- Roster completo POR CAMPAÑA (HC + overrides): los integrantes del roster
# sin actividad aparecen con ceros. Se arma para TODAS las campañas para que el
# selector de campaña del dashboard pueda mostrar cada equipo. ---
def armar_equipo(campana):
    """Devuelve (con_datos, sin_actividad) para una campaña, incluyendo los
    integrantes del roster sin actividad con ceros."""
    over = [o[2] for o in OVERRIDES if o[1] == campana]
    roster_camp = sorted(set(norm(m) for m in equipos[campana]['miembros']) | set(norm(o) for o in over))

    por_key = {k: a for k, a in agentes.items() if a['campana'] == campana}
    mapa = {}
    for k, a in por_key.items():
        mapa[norm(a['agente'])] = a
        mapa[norm(k[1])] = a

    for miembro_norm in roster_camp:
        if miembro_norm in mapa:
            mapa[miembro_norm]['en_roster'] = True
        else:
            k = (campana, miembro_norm)
            agentes[k] = {
                'agente': miembro_norm, 'campana': campana,
                'llamadas': 0, 'no_contesta': 0, 'contactados': 0, 'tasa_contacto': 0.0,
                'numeros': 0, 'intentos_prom': 0.0, 'duracion_prom_s': 0.0,
                'ventas_movil': 0, 'ventas_fijo': 0, 'lineas_fact': 0, 'ventas_total': 0,
                'rgu': 0.0, 'mrc': 0.0, 'venta_cruzada': 0, 'gestiones_cross': 0,
                'ventas_renov': 0, 'ventas_cruz': 0,
                'mrc_ant_renov': 0.0, 'mrc_act_renov': 0.0, 'mrc_cruz': 0.0, 'mrc_prom': 0.0,
                'rgu_renov': 0.0, 'rgu_cruz': 0.0,
                'tramites': {}, 'tipoventa': {}, 'status_fijo': {},
                'por_mes': {m: 0 for m in MESES_ORDEN},
                'contactados_por_mes': {m: 0 for m in MESES_ORDEN},
                'no_contesta_por_mes': {m: 0 for m in MESES_ORDEN},
                'numeros_por_mes': {m: 0 for m in MESES_ORDEN},
                'duracion_por_mes': {m: 0.0 for m in MESES_ORDEN},
                'ventas_por_mes': {m: {'movil':0,'fijo':0,'lineas':0,'mrc':0.0,'rgu':0.0,'tramite':{},'tipoventa':{},
                                        'renov':0,'cruz':0,'mrc_ant_renov':0.0,'mrc_act_renov':0.0,'mrc_cruz':0.0,
                                        'rgu_renov':0.0,'rgu_cruz':0.0} for m in MESES_ORDEN},
                'cruzada_por_mes': {m: 0 for m in MESES_ORDEN},
                'ef_ventas_x_llamada': 0.0, 'ef_conv_pct': 0.0, 'ef_mrc_x_venta': 0.0,
                'en_roster': True,
                'sin_actividad': True,
            }
            mapa[miembro_norm] = agentes[k]

    lista = sorted(mapa.values(),
                   key=lambda x: (not x.get('en_roster', False),
                                  -(x['ventas_total'] + x['mrc']/1000),
                                  -x['llamadas']))
    lista = [a for a in lista if a['llamadas'] > 0 or a['ventas_total'] > 0 or a.get('en_roster', False)]
    con_datos = [a for a in lista if a['llamadas'] > 0 or a['ventas_total'] > 0 or a['venta_cruzada'] > 0]
    sin_actividad = [a for a in lista if a['llamadas'] == 0 and a['ventas_total'] == 0 and a['venta_cruzada'] == 0]
    return con_datos, sin_actividad

campanas_data = {}
for c in equipos:
    campanas_data[c] = dict(zip(['con_datos', 'sin_actividad'], armar_equipo(c)))

renov_con_datos = campanas_data['RENOVACION']['con_datos']
renov_sin_actividad = campanas_data['RENOVACION']['sin_actividad']

sin_eq = sorted([a for k, a in agentes.items() if a['campana'] == 'SIN EQUIPO'],
                key=lambda x: x['llamadas'], reverse=True)

# Totales del equipo RENOVACION por mes (para KPIs y gráfico de evolución)
# Usa renov_con_datos (solo agentes con actividad) para no depender de la
# variable 'renov' que ahora incluye entradas del roster sin actividad.
renov_por_mes = {}
for m in MESES_ORDEN:
    renov_por_mes[m] = {
        'llamadas': sum(a['por_mes'].get(m, 0) for a in renov_con_datos),
        'contactados': sum(a['contactados_por_mes'].get(m, 0) for a in renov_con_datos),
        'no_contesta': sum(a['no_contesta_por_mes'].get(m, 0) for a in renov_con_datos),
        'numeros': sum(a['numeros_por_mes'].get(m, 0) for a in renov_con_datos),
        'ventas_movil': sum(a['ventas_por_mes'][m]['movil'] for a in renov_con_datos),
        'ventas_fijo': sum(a['ventas_por_mes'][m]['fijo'] for a in renov_con_datos),
        'lineas_fact': sum(a['ventas_por_mes'][m]['lineas'] for a in renov_con_datos),
        'mrc': round(sum(a['ventas_por_mes'][m]['mrc'] for a in renov_con_datos), 2),
        'rgu': round(sum(a['ventas_por_mes'][m]['rgu'] for a in renov_con_datos), 1),
        'cruzada': sum(a['cruzada_por_mes'].get(m, 0) for a in renov_con_datos),
    }
    renov_por_mes[m]['tasa_contacto'] = round(
        renov_por_mes[m]['contactados'] / renov_por_mes[m]['llamadas'] * 100, 1) if renov_por_mes[m]['llamadas'] else 0

# =============================================================================
# 6a. EQUIPO DE JULIO (campaña SOHO: Winback + Cross-selling)
# =============================================================================
if 'SOHO' in campanas_data:
    soho_con_datos = campanas_data['SOHO']['con_datos']
    soho_sin_actividad = campanas_data['SOHO']['sin_actividad']
else:
    soho_con_datos, soho_sin_actividad = [], []

soho_por_mes = {}
for m in MESES_ORDEN:
    soho_por_mes[m] = {
        'llamadas': sum(a['por_mes'].get(m, 0) for a in soho_con_datos),
        'contactados': sum(a['contactados_por_mes'].get(m, 0) for a in soho_con_datos),
        'no_contesta': sum(a['no_contesta_por_mes'].get(m, 0) for a in soho_con_datos),
        'numeros': sum(a['numeros_por_mes'].get(m, 0) for a in soho_con_datos),
        'ventas_movil': sum(a['ventas_por_mes'][m]['movil'] for a in soho_con_datos),
        'ventas_fijo': sum(a['ventas_por_mes'][m]['fijo'] for a in soho_con_datos),
        'lineas_fact': sum(a['ventas_por_mes'][m]['lineas'] for a in soho_con_datos),
        'mrc': round(sum(a['ventas_por_mes'][m]['mrc'] for a in soho_con_datos), 2),
        'rgu': round(sum(a['ventas_por_mes'][m]['rgu'] for a in soho_con_datos), 1),
        'cruzada': sum(a['cruzada_por_mes'].get(m, 0) for a in soho_con_datos),
    }
    soho_por_mes[m]['tasa_contacto'] = round(
        soho_por_mes[m]['contactados'] / soho_por_mes[m]['llamadas'] * 100, 1) if soho_por_mes[m]['llamadas'] else 0
print('Equipo SOHO (Julio):', len(soho_con_datos), 'agentes con datos |', len(soho_sin_actividad), 'sin actividad')

# =============================================================================
# 6ab. RESUMEN POR CAMPAÑA + RECORRIDO POR CAMPAÑA (vista 'Resumen de Campanas')
# =============================================================================
def resumen_equipo(campana, con_datos):
    llam = sum(a['llamadas'] for a in con_datos)
    cont = sum(a['contactados'] for a in con_datos)
    tot_rgu = sum(a['rgu'] for a in con_datos)
    return {
        'campana': campana,
        'llamadas': llam,
        'contactados': cont,
        'tasa_contacto': round(cont / llam * 100, 1) if llam else 0,
        'ventas_movil': sum(a['ventas_movil'] for a in con_datos),
        'ventas_fijo': sum(a['ventas_fijo'] for a in con_datos),
        'ventas_total': sum(a['ventas_total'] for a in con_datos),
        'mrc': round(sum(a['mrc'] for a in con_datos), 2),
        'rgu': round(tot_rgu, 1),
        'mrc_prom': round(sum(a['mrc'] for a in con_datos) / tot_rgu, 2) if tot_rgu else 0,
    }

resumen_campanas = {
    'RENOVACION': dict({
        'supervisor': 'MARINEL MORENO',
        'contactos_base': recorrido_base['total'],
        'marcados_base': recorrido_base['marcados'],
        'pct_recorrido': recorrido_base['pct_recorrido'],
        'facturadas': sum(a['ventas_renov'] for a in renov_con_datos),
        'mrc_anterior': round(sum(a['mrc_ant_renov'] for a in renov_con_datos), 2),
        'mrc_actual': round(sum(a['mrc_act_renov'] for a in renov_con_datos), 2),
        'mrc_cruzada': round(sum(a['mrc_cruz'] for a in renov_con_datos), 2),
        'ventas_cruzadas': sum(a['ventas_cruz'] for a in renov_con_datos),
        'rgu_renov': int(round(sum(a['rgu_renov'] for a in renov_con_datos))),
        'rgu_cruz': int(round(sum(a['rgu_cruz'] for a in renov_con_datos))),
    }, **resumen_equipo('RENOVACION', renov_con_datos)),
    'SOHO': dict({
        'supervisor': 'JULIO CESAR ARAUZ',
        'contactos_base': None, 'marcados_base': None, 'pct_recorrido': None,
        'facturadas': sum(a['ventas_total'] for a in soho_con_datos),
        'mrc_anterior': 0, 'mrc_actual': 0, 'mrc_cruzada': 0,
        'ventas_cruzadas': sum(a['ventas_cruz'] for a in soho_con_datos),
        'rgu_renov': 0, 'rgu_cruz': 0,
    }, **resumen_equipo('SOHO', soho_con_datos)),
}
resumen_campanas['WINBACK'] = {
    'campana': 'WINBACK', 'supervisor': 'JULIO CESAR ARAUZ',
    'clientes': winback['resumen'].get('clientes', 0),
    'rgu': winback['resumen'].get('rgu', 0),
    'mrc': winback['resumen'].get('mrc', 0),
    'razones': winback['resumen'].get('razones', 0),
    'principal': winback['por_razon'][0]['r'] if winback['por_razon'] else '',
}
resumen_campanas['CROSS_SELL'] = {
    'campana': 'CROSS_SELL', 'supervisor': 'JULIO CESAR ARAUZ',
    'gestiones': cross_data['resumen'].get('gestiones', 0),
    'con_tel': cross_data['resumen'].get('con_tel', 0),
    'sin_tel_base': cross_data['resumen'].get('sin_tel_base', 0),
    'ejecutivas': cross_data['resumen'].get('ejecutivas', 0),
}

# Recorrido de las bases Winback/Cross contra las llamadas SOHO (equipo de Julio)
soho_llam = [r for r in llamadas if r['campana'] == 'SOHO']
por_tel_soho = defaultdict(list)
for r in soho_llam:
    t = tel8(r['phone'])
    if t:
        por_tel_soho[t].append(r)

def recorrido_de_bases(lista_tels):
    tels = set(t for t in lista_tels if t)
    marc = set(t for t in tels if t in por_tel_soho)
    cont = set(t for t in marc if any(r['status'] and r['status'] not in STATUS_NO_CONTACTO
                                      for r in por_tel_soho[t]))
    return {
        'total': len(tels),
        'marcados': len(marc),
        'contactados': len(cont),
        'pct': round(len(marc) / len(tels) * 100, 1) if tels else 0,
    }

win_tels = [tel8(t) for d in winback['detalle']
            for t in (d['t1'], d['t2']) if t and t not in ('#N/A',)]
cross_tels = [tel8(d['tel']) for d in cross_data['detalle']
              if d['tel'] and d['tel'] != '#N/A']
recorrido_campanas = {
    'WINBACK': recorrido_de_bases(win_tels),
    'CROSS_SELL': recorrido_de_bases(cross_tels),
}
print('Recorrido Winback:', recorrido_campanas['WINBACK'])
print('Recorrido Cross-sell:', recorrido_campanas['CROSS_SELL'])

# =============================================================================
# 6ad. CAMPANAS WINBACK Y CROSS-SELLING COMPLETAS (como Renovacion pero con
#      sus condiciones): llamadas SOHO sobre la base, tipificaciones, recorrido
#      por agente/mes, y ventas del equipo de Julio
# =============================================================================
def campana_base(tels_set):
    """Estadisticas de las llamadas SOHO (equipo de Julio) sobre una base."""
    llam = [r for r in llamadas if r['campana'] == 'SOHO' and tel8(r['phone']) in tels_set]
    stats = stats_vici(llam)
    tipif = Counter(r['status'] or '(SIN TIPIFICACION)' for r in llam).most_common()
    por_mes = {}
    for m in MESES_ORDEN:
        lm = [r for r in llam if (r['mes'] or '') == m]
        mt = set(tel8(r['phone']) for r in lm)
        ct = set(t for t in mt if any(r['status'] and r['status'] not in STATUS_NO_CONTACTO
                                      for r in lm if tel8(r['phone']) == t))
        por_mes[m] = {'marcados': len(mt), 'contactados': len(ct)}
    por_ag = defaultdict(lambda: {'tels': set(), 'llamadas': 0, 'contactados': set()})
    for r in llam:
        a = por_ag[r['miembro'] or 'SIN AGENTE']
        a['tels'].add(tel8(r['phone']))
        a['llamadas'] += 1
        if r['status'] and r['status'] not in STATUS_NO_CONTACTO:
            a['contactados'].add(tel8(r['phone']))
    por_ag = [{'agente': k, 'tels': len(v['tels']), 'llamadas': v['llamadas'],
               'contactados': len(v['contactados'])}
              for k, v in por_ag.items()]
    por_ag.sort(key=lambda x: -x['tels'])
    return {'stats': stats, 'tipif': tipif, 'por_mes': por_mes, 'por_agente': por_ag}

winback_tels_set = set(t for t in win_tels if t)
cross_tels_set = set(t for t in cross_tels if t)
winback_camp = campana_base(winback_tels_set)
cross_camp = campana_base(cross_tels_set)

# Enriquecer el detalle de cada base con marcado/llamadas/agente
por_tel_wb = defaultdict(list)
por_tel_cross = defaultdict(list)
for r in llamadas:
    if r['campana'] != 'SOHO':
        continue
    t = tel8(r['phone'])
    if t in winback_tels_set:
        por_tel_wb[t].append(r)
    if t in cross_tels_set:
        por_tel_cross[t].append(r)
for d in winback['detalle']:
    ts = [tel8(x) for x in (d['t1'], d['t2']) if x and x != '#N/A']
    hit = [t for t in ts if t in por_tel_wb]
    d['m'] = 1 if hit else 0
    d['ll'] = sum(len(por_tel_wb[t]) for t in hit)
    d['ag'] = por_tel_wb[hit[-1]][-1]['miembro'] if hit else ''
for d in cross_data['detalle']:
    t = tel8(d['tel']) if d['tel'] and d['tel'] != '#N/A' else ''
    if t and t in por_tel_cross:
        d['m'] = 1
        d['ll'] = len(por_tel_cross[t])
        d['ag'] = por_tel_cross[t][-1]['miembro']
    else:
        d['m'] = 0
        d['ll'] = 0
        d['ag'] = ''

# Ventas del equipo de Julio (cross-selling = ventas de la campana SOHO)
soho_vmov = [s for s in ventas_moviles if s['campana'] == 'SOHO']
soho_vfij = [s for s in ventas_fijo if s['campana'] == 'SOHO']
soho_lin = [s for s in lineas_fact if s['campana'] == 'SOHO']
cross_camp['ventas'] = {
    'movil': len(soho_vmov),
    'fijo': len(soho_vfij),
    'lineas': len(soho_lin),
    'mrc': round(sum(s['mrc'] for s in soho_vmov) + sum(s['mrc'] for s in soho_vfij), 2),
    'rgu': round(sum(s['rgu'] for s in soho_vfij) + sum(s['cant'] for s in soho_lin), 1),
    'por_mes': {m: {
        'movil': sum(1 for s in soho_vmov if s['mes'] == m),
        'fijo': sum(1 for s in soho_vfij if s['mes'] == m),
        'mrc': round(sum(s['mrc'] for s in soho_vmov if s['mes'] == m)
                     + sum(s['mrc'] for s in soho_vfij if s['mes'] == m), 2),
    } for m in MESES_ORDEN},
}
winback_camp['resumen'] = winback['resumen']
winback_camp['por_ejecutiva'] = winback['por_ejecutiva']
winback_camp['por_razon'] = winback['por_razon']
winback_camp['detalle'] = winback['detalle']
cross_camp['resumen'] = cross_data['resumen']
cross_camp['por_ejecutiva'] = cross_data['por_ejecutiva']
cross_camp['detalle'] = cross_data['detalle']
print('Campana Winback: llamadas SOHO sobre base =', winback_camp['stats']['llamadas'])
print('Campana Cross: llamadas SOHO sobre base =', cross_camp['stats']['llamadas'],
      '| ventas SOHO:', cross_camp['ventas']['movil'], 'movil /',
      cross_camp['ventas']['fijo'], 'fijo')

# =============================================================================
# 6b. RESUMEN MRC / VENTA CRUZADA (visual de Amir): MRC anterior, MRC actual y
# MRC de venta cruzada por separado, + RGU de renovación y RGU cruzados
# =============================================================================
# MRC anterior/actual SOLO del trámite RENOVACION (hoja Ventas Moviles, 150 ventas).
# MRC de venta cruzada = LINEA NUEVA + PORTABILIDAD + PORTABILIDAD EXTERNA (aparte).
# RGU renovación = cantidad de líneas RENOVACION | RGU cruzados = líneas cruzadas.
mrc_resumen = {
    'clientes_base': recorrido_base['total'],
    'mrc_anterior': round(sum(a['mrc_ant_renov'] for a in renov_con_datos), 2),
    'mrc_actual': round(sum(a['mrc_act_renov'] for a in renov_con_datos), 2),
    'mrc_cruzada': round(sum(a['mrc_cruz'] for a in renov_con_datos), 2),
    'rgu_renovacion': int(round(sum(a['rgu_renov'] for a in renov_con_datos))),
    'rgu_cruzada': int(round(sum(a['rgu_cruz'] for a in renov_con_datos))),
    'ventas_renovacion': sum(a['ventas_renov'] for a in renov_con_datos),
    'ventas_cruzadas': sum(a['ventas_cruz'] for a in renov_con_datos),
    'mrc_promedio': round(sum(a['mrc_act_renov'] for a in renov_con_datos)
                          / sum(a['rgu_renov'] for a in renov_con_datos), 2)
        if sum(a['rgu_renov'] for a in renov_con_datos) else 0,
}
mrc_por_mes = {}
for m in MESES_ORDEN:
    mrc_por_mes[m] = {
        'mrc_anterior': round(sum(a['ventas_por_mes'][m]['mrc_ant_renov'] for a in renov_con_datos), 2),
        'mrc_actual': round(sum(a['ventas_por_mes'][m]['mrc_act_renov'] for a in renov_con_datos), 2),
        'mrc_cruzada': round(sum(a['ventas_por_mes'][m]['mrc_cruz'] for a in renov_con_datos), 2),
        'rgu_renovacion': int(round(sum(a['ventas_por_mes'][m]['rgu_renov'] for a in renov_con_datos))),
        'rgu_cruzada': int(round(sum(a['ventas_por_mes'][m]['rgu_cruz'] for a in renov_con_datos))),
        'ventas_renovacion': sum(a['ventas_por_mes'][m]['renov'] for a in renov_con_datos),
        'ventas_cruzadas': sum(a['ventas_por_mes'][m]['cruz'] for a in renov_con_datos),
    }
mrc_resumen['por_mes'] = mrc_por_mes
print('Resumen MRC: anterior $%s -> actual $%s | cruzada $%s | RGU renov %s / cruz %s' % (
    f"{mrc_resumen['mrc_anterior']:,.2f}", f"{mrc_resumen['mrc_actual']:,.2f}",
    f"{mrc_resumen['mrc_cruzada']:,.2f}", mrc_resumen['rgu_renovacion'], mrc_resumen['rgu_cruzada']))

# =============================================================================
# 7. DATOS PARA EL DASHBOARD
# =============================================================================
data = {
    'generado': datetime.date.today().isoformat(),
    'fuente': 'Reporte_soho_junio_25 (1).xlsb + extrainfo/HC + bases',
    'global_vici': global_vici,
    'global_por_mes': global_por_mes,
    'tipif_global': tipif_global.most_common(),
    'tipif_renovacion': tipif_renov.most_common(),
    'tipif_glob_por_mes': tipif_glob_por_mes,
    'tipif_renov_por_mes': tipif_renov_por_mes,
    'meses': MESES_ORDEN,
    'renov_por_mes': renov_por_mes,
    'equipo_renovacion': renov_con_datos,
    'renov_sin_actividad': renov_sin_actividad,
    'equipo_soho': soho_con_datos,
    'soho_sin_actividad': soho_sin_actividad,
    'soho_por_mes': soho_por_mes,
    'resumen_campanas': resumen_campanas,
    'recorrido_campanas': recorrido_campanas,
    'winback_camp': winback_camp,
    'cross_camp': cross_camp,
    'metas': {'RENOVACION': {'rgu': 0, 'mrc': 0}, 'SOHO': {'rgu': 0, 'mrc': 0}},
    'recorrido_base': recorrido_base,
    'winback': winback,
    'cross_sell': cross_data,
    'mrc_resumen': mrc_resumen,
    'totales_ventas': {
        'moviles': len(ventas_moviles),
        'fijo': len(ventas_fijo),
        'lineas': len(lineas_fact),
        'cross_sell': sum(cross_por_agente.values()),
    },
}

os.makedirs(os.path.join(BASE, 'data'), exist_ok=True)
json_path = os.path.join(BASE, 'data', 'renovacion_data.json')
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=1)
print('JSON generado:', json_path)

# =============================================================================
# 8. GENERAR HTML
# =============================================================================
embed = json.dumps(data, ensure_ascii=False)

html = """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Renovación SOHO — Equipo Marinel Moreno</title>
<script src="chart.umd.min.js"></script>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:'Segoe UI',Arial,sans-serif;background:#0f1220;color:#e8eaf6;padding:16px}
.header{background:linear-gradient(135deg,#7b1fa2,#1a1a2e 60%,#16213e);padding:22px 26px;border-radius:14px;margin-bottom:16px;display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:10px;border:1px solid rgba(255,255,255,.08)}
.header h1{font-size:22px;letter-spacing:.5px}
.header .sub{font-size:12px;opacity:.75;margin-top:4px}
.badge-camp{background:#ffd54f;color:#1a1a2e;font-weight:700;font-size:11px;padding:3px 10px;border-radius:20px;letter-spacing:1px}
.kpi-row{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:12px;margin-bottom:16px}
.kpi{background:rgba(255,255,255,.04);border:1px solid rgba(255,255,255,.08);border-radius:12px;padding:14px 16px;position:relative;overflow:hidden}
.kpi::before{content:'';position:absolute;left:0;top:0;bottom:0;width:4px}
.kpi .lbl{font-size:10px;text-transform:uppercase;letter-spacing:1px;color:#9fa8da}
.kpi .val{font-size:24px;font-weight:800;margin-top:4px}
.kpi .sub{font-size:11px;color:#7986cb;margin-top:2px}
.kpi.c-purple::before{background:#ab47bc}.kpi.c-blue::before{background:#42a5f5}
.kpi.c-green::before{background:#66bb6a}.kpi.c-red::before{background:#ef5350}
.kpi.c-orange::before{background:#ffa726}.kpi.c-teal::before{background:#26a69a}
.tabs{display:flex;gap:8px;margin-bottom:16px;flex-wrap:wrap}
.tab{padding:9px 22px;border-radius:8px;cursor:pointer;font-size:13px;font-weight:600;background:rgba(255,255,255,.06);color:#b0bec5;border:1px solid transparent;transition:.2s}
.tab:hover{background:rgba(255,255,255,.12)}
.tab.active{background:#7b1fa2;color:#fff;border-color:#9c27b0}
.camp-bar{display:flex;gap:8px;margin-bottom:14px;flex-wrap:wrap}
.camp-btn{padding:10px 22px;border-radius:8px;cursor:pointer;font-size:13px;font-weight:800;background:rgba(255,255,255,.06);color:#b0bec5;border:1px solid rgba(255,255,255,.1);transition:.2s}
.camp-btn:hover{background:rgba(255,255,255,.12);transform:translateY(-1px)}
.camp-btn.active{background:linear-gradient(135deg,#1a73e8,#7b1fa2);color:#fff;border-color:rgba(255,255,255,.25);box-shadow:0 2px 10px rgba(26,115,232,.3)}
.hidden{display:none!important}
.content{display:none}.content.active{display:block}
.grid-2{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:14px}
.grid-3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:14px;margin-bottom:14px}
.card{background:rgba(255,255,255,.04);border:1px solid rgba(255,255,255,.08);border-radius:12px;padding:16px}
.card h2{font-size:13px;margin-bottom:12px;color:#c5cae9;border-bottom:1px solid rgba(255,255,255,.08);padding-bottom:8px;letter-spacing:.5px}
.card.full{grid-column:1/-1}
.chart-box{position:relative;height:280px}
.chart-box-sm{position:relative;height:230px}
table{width:100%;border-collapse:collapse;font-size:12px}
th{text-align:left;padding:8px;background:rgba(255,255,255,.06);color:#9fa8da;font-weight:700;font-size:11px;text-transform:uppercase;letter-spacing:.5px;position:sticky;top:0}
td{padding:7px 8px;border-bottom:1px solid rgba(255,255,255,.06)}
tr:hover td{background:rgba(123,31,162,.12)}
.rt{text-align:right}.ct{text-align:center}
.pos{display:inline-flex;width:24px;height:24px;border-radius:50%;align-items:center;justify-content:center;font-weight:800;font-size:11px;background:rgba(255,255,255,.08)}
.pos-1{background:#ffd54f;color:#1a1a2e}.pos-2{background:#b0bec5;color:#1a1a2e}.pos-3{background:#ff8a65;color:#1a1a2e}
.bar{height:8px;border-radius:4px;background:linear-gradient(90deg,#ab47bc,#7b1fa2);min-width:2px}
@media(max-width:900px){.grid-2,.grid-3{grid-template-columns:1fr}}
.note{font-size:11px;color:#7986cb;margin-top:8px}
.filter-bar{display:flex;gap:8px;align-items:center;flex-wrap:wrap;margin-bottom:14px;background:rgba(255,255,255,.04);border:1px solid rgba(255,255,255,.08);padding:10px 14px;border-radius:12px}
.mes-btn{padding:6px 14px;border-radius:7px;cursor:pointer;font-size:12px;font-weight:700;background:rgba(255,255,255,.06);color:#b0bec5;border:1px solid rgba(255,255,255,.12);transition:.18s}
.mes-btn:hover{background:rgba(255,255,255,.14);transform:translateY(-1px)}
.mes-btn.active{background:#ffd54f;color:#1a1a2e;border-color:#ffd54f}
.mes-lbl{font-size:11px;font-weight:800;letter-spacing:1px;color:#9fa8da;margin-right:4px}
.period-chip{font-size:12px;font-weight:800;color:#ffd54f;background:rgba(255,213,79,.12);border:1px solid rgba(255,213,79,.3);padding:4px 12px;border-radius:20px;letter-spacing:.5px}
.st-badge{display:inline-flex;align-items:center;gap:4px;font-size:10px;font-weight:800;padding:2px 9px;border-radius:20px;letter-spacing:.4px}
.st-top{background:rgba(255,213,79,.15);color:#ffd54f;border:1px solid rgba(255,213,79,.4)}
.st-act{background:rgba(102,187,106,.12);color:#81c784;border:1px solid rgba(102,187,106,.35)}
.st-low{background:rgba(239,83,80,.12);color:#ef9a9a;border:1px solid rgba(239,83,80,.35)}
.st-warn{background:rgba(255,167,38,.12);color:#ffb74d;border:1px solid rgba(255,167,38,.35)}
.tr-low td{background:rgba(239,83,80,.05)!important;opacity:.55}
.tr-top td{background:rgba(255,213,79,.04)!important}
.ef-bar{display:inline-block;height:8px;border-radius:4px;background:linear-gradient(90deg,#66bb6a,#26a69a);vertical-align:middle}
</style>
</head>
<body>

<div class="header">
  <div>
    <span class="badge-camp">CANAL SOHO · TODAS LAS CAMPAÑAS</span>
    <h1>Renovación · Winback · Cross-selling</h1>
    <div class="sub"><span id="supAct" style="color:#ffd54f;font-weight:700">Supervisor: Marinel Moreno</span> · SOHO · Iniciativa Nextphone · Gerente: Amir Josue Rodriguez Chavarria · Datos: <span id="fechaGen"></span></div>
  </div>
  <div style="text-align:right;font-size:12px;opacity:.8;">
    <div id="totCall"></div>
  </div>
</div>

<div class="filter-bar">
  <span class="mes-lbl">PERIODO:</span>
  <button class="mes-btn active" data-mes="TODOS" onclick="setMes('TODOS')">TODOS</button>
  <button class="mes-btn" data-mes="ENERO" onclick="setMes('ENERO')">ENE</button>
  <button class="mes-btn" data-mes="FEBRERO" onclick="setMes('FEBRERO')">FEB</button>
  <button class="mes-btn" data-mes="MARZO" onclick="setMes('MARZO')">MAR</button>
  <button class="mes-btn" data-mes="ABRIL" onclick="setMes('ABRIL')">ABR</button>
  <button class="mes-btn" data-mes="MAYO" onclick="setMes('MAYO')">MAY</button>
  <button class="mes-btn" data-mes="JUNIO" onclick="setMes('JUNIO')">JUN</button>
  <span style="margin-left:auto;display:flex;gap:6px;align-items:center;">
    <button class="mes-btn" onclick="navMes(-1)" title="Mes anterior">◀</button>
    <span class="period-chip" id="mesActLbl">ENE — JUN 2026</span>
    <button class="mes-btn" onclick="navMes(1)" title="Mes siguiente">▶</button>
  </span>
</div>

<div class="kpi-row" id="kpiRow"></div>

<div class="camp-bar">
  <button class="camp-btn active" data-camp="RENOVACION" onclick="setCamp('RENOVACION',this)">🏆 Renovación · Marinel</button>
  <button class="camp-btn" data-camp="WINBACK" onclick="setCamp('WINBACK',this)">🔄 Winback · Julio</button>
  <button class="camp-btn" data-camp="CROSS" onclick="setCamp('CROSS',this)">🛒 Cross-selling · Julio</button>
  <button class="camp-btn" data-camp="GLOBAL" onclick="setCamp('GLOBAL',this)">🌐 Global (Resumen · Equipos · Metas)</button>
</div>

<div class="tabs subtab-bar" data-camp="RENOVACION">
  <button class="tab active" onclick="sw('renov',this)">🏆 Equipo Renovación</button>
  <button class="tab" onclick="sw('rend',this)">📊 Rendimiento del Equipo</button>
  <button class="tab" onclick="sw('marc',this)">📞 Marcaciones</button>
  <button class="tab" onclick="sw('tipif',this)">🏷️ Tipificación</button>
  <button class="tab" onclick="sw('ventas',this)">💰 Ventas & RGU</button>
  <button class="tab" onclick="sw('rec',this)">🗺️ Recorrido de la Base</button>
</div>
<div class="tabs subtab-bar hidden" data-camp="WINBACK">
  <button class="tab active" onclick="sw('winback',this)">🔄 Base Winback</button>
  <button class="tab" onclick="sw('winb-marc',this)">📞 Marcaciones</button>
  <button class="tab" onclick="sw('winb-tip',this)">🏷️ Tipificación</button>
  <button class="tab" onclick="sw('winb-rec',this)">🗺️ Recorrido</button>
</div>
<div class="tabs subtab-bar hidden" data-camp="CROSS">
  <button class="tab active" onclick="sw('cross',this)">🛒 Base Cross-selling</button>
  <button class="tab" onclick="sw('cross-marc',this)">📞 Marcaciones</button>
  <button class="tab" onclick="sw('cross-tip',this)">🏷️ Tipificación</button>
  <button class="tab" onclick="sw('cross-ventas',this)">💰 Ventas</button>
  <button class="tab" onclick="sw('cross-rec',this)">🗺️ Recorrido</button>
</div>
<div class="tabs subtab-bar hidden" data-camp="GLOBAL">
  <button class="tab active" onclick="sw('resumen',this)">📋 Resumen de Campañas</button>
  <button class="tab" onclick="sw('julio',this)">👥 Equipo Julio</button>
  <button class="tab" onclick="sw('metas',this)">🎯 Metas y Cumplimiento</button>
</div>

<div id="v-renov" class="content active">
  <div class="card full">
    <h2>🏆 Top Performance — Equipo Renovación <span class="period-chip" id="topPeriod" style="font-size:10px">TODOS LOS MESES</span> <span style="font-weight:400;color:#7986cb;font-size:11px">(clic en columna para ordenar)</span></h2>
    <div style="overflow-x:auto;max-height:520px;overflow-y:auto;">
    <table id="tabRenov"><thead><tr>
      <th>#</th><th>Agente</th><th class="rt">Llamadas</th><th class="rt">Contactados</th><th class="rt">% Contacto</th>
      <th class="rt">Ventas Renov</th><th class="rt">Ventas Cruz</th><th class="rt">RGU</th><th class="rt">MRC $</th><th class="rt">MRC Prom $</th><th>Progreso</th>
    </tr></thead><tbody></tbody></table>
    </div>
  </div>
  <div class="card full">
    <h2>💰 Cartera MRC — Solo RENOVACION <span class="period-chip" id="mrcPeriod" style="font-size:10px">TODOS LOS MESES</span></h2>
    <div class="kpi-row" id="mrcKpiRow"></div>
    <div class="grid-2">
      <div class="card"><h2>💰 MRC Anterior vs Actual vs Cruzada (por agente)</h2><div class="chart-box"><canvas id="chMrcAg"></canvas></div></div>
      <div class="card"><h2>📊 RGU: Renovación vs Cruzados (por agente)</h2><div class="chart-box"><canvas id="chRguAg"></canvas></div></div>
    </div>
    <div class="note" id="mrcNota"></div>
  </div>
  <div class="grid-2">
    <div class="card"><h2>📈 Marcaciones por Mes — EQUIPO RENOVACIÓN</h2><div class="chart-box-sm"><canvas id="chMesRenov"></canvas></div></div>
    <div class="card"><h2>🔢 Llamadas por Agente</h2><div class="chart-box-sm"><canvas id="chAgLlam"></canvas></div></div>
  </div>
</div>

<div id="v-marc" class="content">
  <div class="grid-3" style="margin-bottom:14px">
    <div class="card"><h2>🔥 Quién llama MÁS (marcaciones)</h2><div class="chart-box-sm"><canvas id="chLlamMas"></canvas></div></div>
    <div class="card"><h2>🌙 Quién llama MENOS</h2><div class="chart-box-sm"><canvas id="chLlamMenos"></canvas></div></div>
    <div class="card"><h2>💬 Contactados vs No Contesta</h2><div class="chart-box-sm"><canvas id="chContVsNc"></canvas></div></div>
  </div>
  <div class="card full">
    <h2>📞 Ranking de Marcaciones — de quien llama más a quien llama menos <span class="period-chip" id="marcPeriod" style="font-size:10px">TODOS LOS MESES</span></h2>
    <div style="overflow-x:auto;max-height:560px;overflow-y:auto;">
    <table id="tabMarc"><thead><tr>
      <th>#</th><th>Agente</th><th class="rt">Llamadas</th><th class="rt">Contactados</th>
      <th class="rt">% Contacto</th><th class="rt">No Contesta</th><th class="rt">Números únicos</th>
      <th class="rt">Intentos/núm</th><th class="rt">Duración prom (s)</th><th>% del total</th>
    </tr></thead><tbody></tbody></table>
    </div>
    <div class="note" id="marcNota"></div>
  </div>
</div>

<div id="v-rend" class="content">
  <div class="grid-3" style="margin-bottom:14px">
    <div class="card"><h2>🥇 Top Vendedores (por ventas + MRC)</h2><div class="chart-box-sm"><canvas id="chTopVen"></canvas></div></div>
    <div class="card"><h2>📉 Eficiencia — Ventas por Llamada</h2><div class="chart-box-sm"><canvas id="chEfVxL"></canvas></div></div>
    <div class="card"><h2>🎯 Conversión — Ventas / Llamadas %</h2><div class="chart-box-sm"><canvas id="chEfConv"></canvas></div></div>
  </div>
  <div class="card full">
    <h2>📊 Ranking Completo del Equipo (incluye integrantes sin actividad) <span class="period-chip" id="rendPeriod" style="font-size:10px">TODOS LOS MESES</span></h2>
    <div style="overflow-x:auto;max-height:560px;overflow-y:auto;">
    <table id="tabRend"><thead><tr>
      <th>#</th><th>Agente</th><th>Estado</th><th class="rt">Llamadas</th><th class="rt">Ventas</th>
      <th class="rt">MRC $</th><th class="rt">RGU</th><th class="rt">Ventas/Llamada</th>
      <th class="rt">% Conversión</th><th class="rt">MRC/Venta $</th>
    </tr></thead><tbody></tbody></table>
    </div>
    <div class="note" id="rendNota"></div>
  </div>
</div>

<div id="v-tipif" class="content">
  <div class="grid-3">
    <div class="card"><h2>Tipificación — Equipo EQUIPO RENOVACIÓN</h2><div class="chart-box-sm"><canvas id="chTipRenov"></canvas></div></div>
    <div class="card"><h2>Tipificación — Global</h2><div class="chart-box-sm"><canvas id="chTipGlob"></canvas></div></div>
    <div class="card"><h2>Métricas de Marcación</h2><div class="chart-box-sm"><canvas id="chMarc"></canvas></div></div>
  </div>
  <div class="card full">
    <h2>Detalle de Tipificaciones (Renovación)</h2>
    <table><thead><tr><th>#</th><th>Tipificación</th><th class="rt">Llamadas</th><th class="rt">% del Total</th><th>Interpretación</th></tr></thead>
    <tbody id="tabTip"></tbody></table>
  </div>
</div>

<div id="v-ventas" class="content">
  <div class="grid-3">
    <div class="card"><h2>Ventas por Agente (Móvil + Fijo + Líneas)</h2><div class="chart-box-sm"><canvas id="chVentas"></canvas></div></div>
    <div class="card"><h2>MRC Generado por Agente</h2><div class="chart-box-sm"><canvas id="chMrc"></canvas></div></div>
    <div class="card"><h2>RGU por Agente</h2><div class="chart-box-sm"><canvas id="chRgu"></canvas></div></div>
  </div>
  <div class="grid-2">
    <div class="card"><h2>Ventas Móviles por Trámite</h2><div class="chart-box-sm"><canvas id="chTramite"></canvas></div></div>
    <div class="card"><h2>Líneas Facturadas por Tipo de Venta</h2><div class="chart-box-sm"><canvas id="chTipoventa"></canvas></div></div>
  </div>
</div>

<div id="v-rec" class="content">
  <div class="kpi-row" id="recKpiRow"></div>
  <div class="grid-3" style="margin-bottom:14px">
    <div class="card"><h2>🚗 Marcados vs Sin Tocar</h2><div class="chart-box-sm"><canvas id="chRecDonut"></canvas></div></div>
    <div class="card"><h2>👤 Tel. de la base marcados por Agente</h2><div class="chart-box-sm"><canvas id="chRecAg"></canvas></div></div>
    <div class="card"><h2>📅 Marcados por Mes</h2><div class="chart-box-sm"><canvas id="chRecMes"></canvas></div></div>
  </div>
  <div class="card full">
    <h2>🗺️ Recorrido por Agente RENOVACIÓN <span class="period-chip" id="recPeriod" style="font-size:10px">TODOS LOS MESES</span></h2>
    <div style="overflow-x:auto;max-height:400px;overflow-y:auto;">
    <table id="tabRecAg"><thead><tr>
      <th>#</th><th>Agente</th><th class="rt">Tel. base marcados</th><th class="rt">% de la base</th>
      <th class="rt">Llamadas</th><th class="rt">Contactados</th>
    </tr></thead><tbody></tbody></table>
    </div>
    <div class="note" id="recNota"></div>
  </div>
  <div class="card full">
    <h2>🔍 Detalle de la Base (marcados primero)</h2>
    <div class="filter-bar" style="margin-bottom:8px">
      <label>Buscar teléfono / cuenta / nombre:</label>
      <input type="text" id="recQ" onkeyup="renderRecDetalle()" placeholder="Ej: 6887 o EDIL..." style="min-width:200px">
      <label>Base:</label>
      <select id="recFBase" onchange="renderRecDetalle()"><option value="">Todas</option><option>NEXTPHONE</option><option>MOVIL</option></select>
      <label>Estado:</label>
      <select id="recFEst" onchange="renderRecDetalle()"><option value="">Todos</option><option value="1">Marcados</option><option value="0">Sin tocar</option></select>
      <span style="flex:1"></span><span id="recCount" style="font-size:12px;color:#7986cb"></span>
    </div>
    <div style="overflow-x:auto;max-height:520px;overflow-y:auto;">
    <table id="tabRecDet"><thead><tr>
      <th>Base</th><th class="ct">Cuenta</th><th class="ct">Teléfono</th><th>Cliente</th><th>Ejecutiva (base)</th>
      <th class="ct">Marcado</th><th class="ct">Contactado</th><th class="rt">Llamadas</th><th>Agente que marcó</th><th>Meses</th>
    </tr></thead><tbody></tbody></table>
    </div>
  </div>
</div>

<div id="v-winback" class="content">
  <div class="kpi-row" id="winKpiRow"></div>
  <div class="grid-2">
    <div class="card"><h2>🎯 Clientes a recuperar por Razón de Salida</h2><div class="chart-box"><canvas id="chWinRazon"></canvas></div></div>
    <div class="card"><h2>👥 Clientes por Ejecutiva</h2><div class="chart-box"><canvas id="chWinEje"></canvas></div></div>
  </div>
  <div class="card full">
    <h2>📋 Base Winback — clientes a recuperar (campaña de Julio)</h2>
    <div class="filter-bar" style="margin-bottom:8px">
      <label>Buscar cliente / plan:</label>
      <input type="text" id="winQ" onkeyup="renderWinTabla()" placeholder="Ej: COIMSA..." style="min-width:200px">
      <label>Ejecutiva:</label>
      <select id="winFEje" onchange="renderWinTabla()"><option value="">Todas</option></select>
      <label>Razón:</label>
      <select id="winFRazon" onchange="renderWinTabla()"><option value="">Todas</option></select>
      <span style="flex:1"></span><span id="winCount" style="font-size:12px;color:#7986cb"></span>
    </div>
    <div style="overflow-x:auto;max-height:520px;overflow-y:auto;">
    <table id="tabWin"><thead><tr>
      <th>Ejecutiva</th><th>Cliente</th><th>Plan</th><th>Razón de salida</th>
      <th class="rt">RGU</th><th class="rt">MRC $</th><th>Contacto</th><th>Correo</th><th class="ct">Tel 1</th><th class="ct">Tel 2</th>
    </tr></thead><tbody></tbody></table>
    </div>
  </div>
</div>

<div id="v-cross" class="content">
  <div class="kpi-row" id="crossKpiRow"></div>
  <div class="grid-2">
    <div class="card"><h2>🛒 Gestiones por Ejecutiva</h2><div class="chart-box"><canvas id="chCrossEje"></canvas></div></div>
    <div class="card"><h2>📞 Con vs Sin Teléfono (gestiones)</h2><div class="chart-box"><canvas id="chCrossTel"></canvas></div></div>
  </div>
  <div class="card full">
    <h2>📋 Gestiones Cross-selling — clientes móvil (se les ofrece fijo/internet)</h2>
    <div class="filter-bar" style="margin-bottom:8px">
      <label>Buscar cliente / cuenta:</label>
      <input type="text" id="crossQ" onkeyup="renderCrossTabla()" placeholder="Ej: VIELKA..." style="min-width:200px">
      <label>Ejecutiva:</label>
      <select id="crossFEje" onchange="renderCrossTabla()"><option value="">Todas</option></select>
      <label>Con teléfono:</label>
      <select id="crossFTel" onchange="renderCrossTabla()"><option value="">Todos</option><option value="1">Sí</option><option value="0">No</option></select>
      <span style="flex:1"></span><span id="crossCount" style="font-size:12px;color:#7986cb"></span>
    </div>
    <div style="overflow-x:auto;max-height:520px;overflow-y:auto;">
    <table id="tabCross"><thead><tr>
      <th>Fecha</th><th>Cuenta</th><th>Cliente</th><th>Provincia</th><th class="ct">Teléfono</th>
      <th>Ejecutiva</th><th>Estatus / Comentario</th>
    </tr></thead><tbody></tbody></table>
    </div>
  </div>
</div>

<div id="v-resumen" class="content">
  <div class="kpi-row" id="resKpiRow"></div>
  <div class="card full">
    <h2>📊 Comparativa por Campaña (un vistazo)</h2>
    <div style="overflow-x:auto">
    <table id="tabResumen"><thead><tr>
      <th>Campaña</th><th>Supervisor</th><th class="rt">Contactos Base</th><th class="rt">Marcaciones</th>
      <th class="rt">% Recorrido</th><th class="rt">Ventas</th><th class="rt">Fijo</th><th class="rt">Móvil</th>
      <th class="rt">MRC $</th><th class="rt">RGU</th><th class="rt">MRC Prom $</th>
    </tr></thead><tbody></tbody></table>
    </div>
    <div class="note" id="resNota"></div>
  </div>
  <div class="card full">
    <h2>🗺️ Recorrido de las Bases Winback y Cross-selling (llamadas SOHO de Julio)</h2>
    <div style="overflow-x:auto">
    <table id="tabRecCamp"><thead><tr>
      <th>Base</th><th class="rt">Teléfonos únicos</th><th class="rt">Marcados</th><th class="rt">Contactados</th><th class="rt">% Recorrido</th>
    </tr></thead><tbody></tbody></table>
    </div>
    <div class="note" id="recCampNota"></div>
  </div>
</div>

<div id="v-julio" class="content">
  <div class="kpi-row" id="julioKpiRow"></div>
  <div class="grid-3" style="margin-bottom:14px">
    <div class="card"><h2>🥇 Top Vendedores Equipo Julio</h2><div class="chart-box-sm"><canvas id="chJulioTop"></canvas></div></div>
    <div class="card"><h2>📞 Marcaciones por Agente</h2><div class="chart-box-sm"><canvas id="chJulioLlam"></canvas></div></div>
    <div class="card"><h2>💰 MRC por Agente</h2><div class="chart-box-sm"><canvas id="chJulioMrc"></canvas></div></div>
  </div>
  <div class="card full">
    <h2>📊 Rendimiento del Equipo de Julio (SOHO: Winback + Cross-selling)</h2>
    <div style="overflow-x:auto;max-height:560px;overflow-y:auto;">
    <table id="tabJulio"><thead><tr>
      <th>#</th><th>Agente</th><th class="rt">Llamadas</th><th class="rt">Contactados</th><th class="rt">% Contacto</th>
      <th class="rt">Ventas Móvil</th><th class="rt">Ventas Fijo</th><th class="rt">Ventas Tot</th>
      <th class="rt">RGU</th><th class="rt">MRC $</th><th class="rt">MRC Prom $</th>
    </tr></thead><tbody></tbody></table>
    </div>
    <div class="note" id="julioNota"></div>
  </div>
</div>

<div id="v-metas" class="content">
  <div class="card full">
    <h2>🎯 Metas y % de Cumplimiento por Campaña</h2>
    <div class="note" style="margin-bottom:10px">Escribe la meta de RGU y MRC de cada campaña y presiona Guardar (se guardan en este navegador). El % se calcula con las ventas actuales del reporte. Cuando Julio comparta las metas oficiales, solo hay que escribirlas aquí.</div>
    <div style="overflow-x:auto">
    <table id="tabMetas"><thead><tr>
      <th>Campaña</th><th class="rt">RGU Actual</th><th class="rt">Meta RGU</th><th class="rt">% RGU</th>
      <th class="rt">MRC Actual $</th><th class="rt">Meta MRC $</th><th class="rt">% MRC</th>
    </tr></thead><tbody></tbody></table>
    </div>
    <div style="margin-top:12px">
      <button class="mes-btn" onclick="guardarMetas()">💾 Guardar Metas</button>
      <button class="mes-btn" onclick="limpiarMetas()">🗑️ Limpiar</button>
      <span id="metasMsg" style="margin-left:10px;font-size:12px;color:#81c784"></span>
    </div>
    <div class="note" id="metasNota"></div>
  </div>
</div>

<div id="v-winb-marc" class="content">
  <div class="kpi-row" id="winbKpiRow"></div>
  <div class="grid-2">
    <div class="card"><h2>👤 Quién marcó los teléfonos de la base Winback</h2><div class="chart-box-sm"><canvas id="winbAgChart"></canvas></div></div>
    <div class="card"><h2>📅 Marcados por Mes</h2><div class="chart-box-sm"><canvas id="winbMes"></canvas></div></div>
  </div>
  <div class="card full">
    <h2>📞 Agentes (equipo de Julio) que llamaron a la base Winback</h2>
    <div style="overflow-x:auto;max-height:420px;overflow-y:auto;">
    <table id="winbAgTab"><thead><tr>
      <th>Agente</th><th class="rt">Tel. de la base marcados</th><th class="rt">Llamadas</th><th class="rt">Contactados</th>
    </tr></thead><tbody></tbody></table>
    </div>
    <div class="note" id="winbMarcNota"></div>
  </div>
</div>

<div id="v-winb-tip" class="content">
  <div class="grid-2">
    <div class="card"><h2>🏷️ Tipificación de las llamadas a la base Winback</h2><div class="chart-box"><canvas id="winbTip"></canvas></div></div>
    <div class="card"><h2>Detalle (barras)</h2><div class="chart-box"><canvas id="winbTipBar"></canvas></div></div>
  </div>
  <div class="card full">
    <h2>Detalle de Tipificaciones (base Winback)</h2>
    <table><thead><tr><th>#</th><th>Tipificación</th><th class="rt">Llamadas</th><th class="rt">% del Total</th></tr></thead>
    <tbody id="winbTipTab"></tbody></table>
  </div>
</div>

<div id="v-winb-rec" class="content">
  <div class="kpi-row" id="winbRecKpi"></div>
  <div class="card full">
    <h2>🗺️ Detalle de la Base Winback (marcados primero)</h2>
    <div class="filter-bar" style="margin-bottom:8px">
      <label>Buscar cliente / plan:</label>
      <input type="text" id="winbDetQ" onkeyup="renderWinbDet()" placeholder="Ej: COIMSA..." style="min-width:200px">
      <label>Estado:</label>
      <select id="winbDetF" onchange="renderWinbDet()"><option value="">Todos</option><option value="1">Marcados</option><option value="0">Sin tocar</option></select>
      <span style="flex:1"></span><span id="winbDetCount" style="font-size:12px;color:#7986cb"></span>
    </div>
    <div style="overflow-x:auto;max-height:520px;overflow-y:auto;">
    <table id="winbDetTab"><thead><tr>
      <th>Ejecutiva</th><th>Cliente</th><th>Plan</th><th>Razón</th><th class="rt">RGU</th><th class="rt">MRC $</th>
      <th class="ct">Marcado</th><th class="rt">Llamadas</th><th>Agente que llamó</th>
    </tr></thead><tbody></tbody></table>
    </div>
  </div>
</div>

<div id="v-cross-marc" class="content">
  <div class="kpi-row" id="crossKpiRow"></div>
  <div class="grid-2">
    <div class="card"><h2>👤 Quién marcó los teléfonos de la base Cross-selling</h2><div class="chart-box-sm"><canvas id="crossAgChart"></canvas></div></div>
    <div class="card"><h2>📅 Marcados por Mes</h2><div class="chart-box-sm"><canvas id="crossMes"></canvas></div></div>
  </div>
  <div class="card full">
    <h2>📞 Agentes (equipo de Julio) que llamaron a la base Cross-selling</h2>
    <div style="overflow-x:auto;max-height:420px;overflow-y:auto;">
    <table id="crossAgTab"><thead><tr>
      <th>Agente</th><th class="rt">Tel. de la base marcados</th><th class="rt">Llamadas</th><th class="rt">Contactados</th>
    </tr></thead><tbody></tbody></table>
    </div>
    <div class="note" id="crossMarcNota"></div>
  </div>
</div>

<div id="v-cross-tip" class="content">
  <div class="grid-2">
    <div class="card"><h2>🏷️ Tipificación de las llamadas a la base Cross-selling</h2><div class="chart-box"><canvas id="crossTip"></canvas></div></div>
    <div class="card"><h2>Detalle (barras)</h2><div class="chart-box"><canvas id="crossTipBar"></canvas></div></div>
  </div>
  <div class="card full">
    <h2>Detalle de Tipificaciones (base Cross-selling)</h2>
    <table><thead><tr><th>#</th><th>Tipificación</th><th class="rt">Llamadas</th><th class="rt">% del Total</th></tr></thead>
    <tbody id="crossTipTab"></tbody></table>
  </div>
</div>

<div id="v-cross-ventas" class="content">
  <div class="kpi-row" id="crossVentKpi"></div>
  <div class="grid-2">
    <div class="card"><h2>📊 Ventas Móvil vs Fijo por Mes (equipo de Julio)</h2><div class="chart-box"><canvas id="crossVentMes"></canvas></div></div>
    <div class="card"><h2>💰 MRC por Mes</h2><div class="chart-box"><canvas id="crossVentMrc"></canvas></div></div>
  </div>
  <div class="card full">
    <h2>Ventas por Vendedor (equipo de Julio)</h2>
    <div style="overflow-x:auto;max-height:460px;overflow-y:auto;">
    <table id="crossVentTab"><thead><tr>
      <th>Agente</th><th class="rt">Ventas Móvil</th><th class="rt">Ventas Fijo</th><th class="rt">MRC $</th><th class="rt">RGU</th>
    </tr></thead><tbody></tbody></table>
    </div>
    <div class="note" id="crossVentNota"></div>
  </div>
</div>

<div id="v-cross-rec" class="content">
  <div class="kpi-row" id="crossRecKpi"></div>
  <div class="card full">
    <h2>🗺️ Detalle de la Base Cross-selling (marcados primero)</h2>
    <div class="filter-bar" style="margin-bottom:8px">
      <label>Buscar cliente / cuenta:</label>
      <input type="text" id="crossDetQ" onkeyup="renderCrossDet()" placeholder="Ej: VIELKA..." style="min-width:200px">
      <label>Estado:</label>
      <select id="crossDetF" onchange="renderCrossDet()"><option value="">Todos</option><option value="1">Marcados</option><option value="0">Sin tocar</option></select>
      <span style="flex:1"></span><span id="crossDetCount" style="font-size:12px;color:#7986cb"></span>
    </div>
    <div style="overflow-x:auto;max-height:520px;overflow-y:auto;">
    <table id="crossDetTab"><thead><tr>
      <th>Fecha</th><th>Cuenta</th><th>Cliente</th><th>Provincia</th><th class="ct">Teléfono</th>
      <th>Ejecutiva</th><th class="ct">Marcado</th><th class="rt">Llamadas</th><th>Agente que llamó</th>
    </tr></thead><tbody></tbody></table>
    </div>
  </div>
</div>

<script>
const D = __DATA__;
const fmt$ = v => '$' + Number(v).toLocaleString('en-US', {minimumFractionDigits:2, maximumFractionDigits:2});
const COLORS = ['#ab47bc','#42a5f5','#66bb6a','#ffa726','#ef5350','#26a69a','#ec407a','#5c6bc0','#ffca28','#8d6e63'];
const MESES = D.meses;
const charts = {};

document.getElementById('fechaGen').textContent = D.generado;

function fmtCtx(c){
  const el = document.getElementById(c);
  return el ? el.getContext('2d') : null;
}

function nuevoChart(id, cfg){
  const c = fmtCtx(id); if(!c) return;
  if(charts[id]) charts[id].destroy();
  charts[id] = new Chart(c, cfg);
}

let mesAct = 'TODOS';
function campData(){
  return {con: D.equipo_renovacion||[], sin: D.renov_sin_actividad||[]};
}

// --- Filtrado por mes ---
function agFiltrado(a){
  if(mesAct === 'TODOS'){
    return Object.assign({}, a);
  }
  const m = mesAct;
  const pm = a.por_mes[m]||0, cm = a.contactados_por_mes[m]||0;
  const nm = a.no_contesta_por_mes[m]||0, num = a.numeros_por_mes[m]||0;
  const vp = a.ventas_por_mes[m]||{};
  const mov = vp.movil||0, fij = vp.fijo||0, lin = vp.lineas||0;
  const vt = mov+fij+lin, mrc = vp.mrc||0, rgu = vp.rgu||0;
  return {
    agente: a.agente, campana: a.campana,
    llamadas: pm, contactados: cm, no_contesta: nm, numeros: num,
    tasa_contacto: pm ? +(cm/pm*100).toFixed(1) : 0,
    ventas_movil: mov, ventas_fijo: fij, lineas_fact: lin,
    ventas_total: vt, mrc: mrc, rgu: rgu,
    venta_cruzada: vp.cruz||0,
    ventas_renov: vp.renov||0,
    ventas_cruz: vp.cruz||0,
    mrc_ant_renov: vp.mrc_ant_renov||0,
    mrc_act_renov: vp.mrc_act_renov||0,
    mrc_cruz: vp.mrc_cruz||0,
    mrc_prom: (vp.rgu_renov||0) ? +((vp.mrc_act_renov||0)/(vp.rgu_renov)).toFixed(2) : 0,
    rgu_renov: vp.rgu_renov||0,
    rgu_cruz: vp.rgu_cruz||0,
    tramites: vp.tramite||{}, tipoventa: vp.tipoventa||{},
    ef_ventas_x_llamada: pm ? +(vt/pm).toFixed(3) : 0,
    ef_conv_pct: pm ? +(vt/pm*100).toFixed(2) : 0,
    ef_mrc_x_venta: vt ? +(mrc/vt).toFixed(2) : 0,
    intentos_prom: num ? +(pm/num).toFixed(2) : 0,
    duracion_prom_s: (a.duracion_por_mes&&a.duracion_por_mes[m]) ? a.duracion_por_mes[m] : 0,
    en_roster: !!a.en_roster, sin_actividad: !!a.sin_actividad
  };
}
function agentesActuales(){
  return campData().con.map(agFiltrado);
}
// Roster completo: con datos + los del HC sin actividad (ceros)
function rosterActual(){
  const conDatos = agentesActuales();
  const sinAct = campData().sin.map(agFiltrado);
  return conDatos.concat(sinAct);
}
function tipifRenovActual(){
  return mesAct==='TODOS' ? (D.tipif_renovacion||[]) : ((D.tipif_renov_por_mes||{})[mesAct]||[]);
}
function tipifGlobActual(){
  return mesAct==='TODOS' ? D.tipif_global : (D.tipif_glob_por_mes[mesAct]||[]);
}
function mesLabel(){
  return mesAct==='TODOS' ? 'ENE — JUN 2026' : mesAct;
}

function setMes(m){
  mesAct = m;
  document.querySelectorAll('.mes-btn').forEach(b=>b.classList.toggle('active', b.getAttribute('data-mes')===m));
  renderAll();
}
function navMes(d){
  if(mesAct === 'TODOS'){
    setMes(d>0 ? MESES[0] : MESES[MESES.length-1]);
    return;
  }
  const i = MESES.indexOf(mesAct);
  const n = (i + d + MESES.length) % MESES.length;
  setMes(MESES[n]);
}

function kpis(){
  const g = mesAct==='TODOS' ? D.global_vici : D.global_por_mes[mesAct];
  const r = agentesActuales();
  const sum = f => r.reduce((a,x)=>a+f(x),0);
  const llamR = sum(x=>x.llamadas), contR = sum(x=>x.contactados);
  const ventasR = sum(x=>x.ventas_total), mrcR = sum(x=>x.mrc);
  const rguR = sum(x=>x.rgu), xcR = sum(x=>x.venta_cruzada);
  const lbl = 'RENOVACIÓN';
  const vsub = mesAct==='TODOS'
    ? D.totales_ventas.moviles+' móv. · '+D.totales_ventas.fijo+' fijo · '+D.totales_ventas.lineas+' líneas'
    : sum(x=>x.ventas_movil)+' móv. · '+sum(x=>x.ventas_fijo)+' fijo · '+sum(x=>x.lineas_fact)+' líneas';
  const arr = [
    ['Llamadas '+lbl, llamR.toLocaleString(), g.llamadas? (llamR/g.llamadas*100).toFixed(1)+'% del global ('+g.llamadas.toLocaleString()+')':'', 'purple'],
    ['Contactados', contR.toLocaleString(), llamR? (contR/llamR*100).toFixed(1)+'% tasa contacto':'-', 'green'],
    ['Ventas Subidas', ventasR, vsub, 'blue'],
    ['MRC Generado', fmt$(mrcR), 'móvil + fijo + líneas', 'orange'],
    ['RGU Total', rguR.toLocaleString(), 'servicios vendidos', 'teal'],
    ['Venta Cruzada', xcR, 'línea nueva + portabilidad + port. externa', 'red'],
  ];
  document.getElementById('kpiRow').innerHTML = arr.map(a=>
    `<div class="kpi c-${a[3]}"><div class="lbl">${a[0]}</div><div class="val">${a[1]}</div><div class="sub">${a[2]}</div></div>`
  ).join('');
  document.getElementById('totCall').innerHTML = `<strong>${g.llamadas.toLocaleString()}</strong> llamadas ${lbl.toLowerCase()} · ${g.numeros_unicos.toLocaleString()} números únicos`;
  document.getElementById('mesActLbl').textContent = mesLabel();
  document.getElementById('topPeriod').textContent = mesAct==='TODOS' ? 'TODOS LOS MESES' : mesAct;
}

function sw(id, btn){
  document.querySelectorAll('.tab').forEach(t=>t.classList.remove('active'));
  document.querySelectorAll('.content').forEach(c=>c.classList.remove('active'));
  btn.classList.add('active');
  const el = document.getElementById('v-'+id);
  if(el) el.classList.add('active');
  // Re-render al cambiar de pestana: los charts creados con el div oculto
  // quedan en 0x0; al re-renderizar se dibujan con el tamano correcto.
  renderAll();
}

// ---------- RENOVACION ----------
let sortKey = 'llamadas', sortDir = -1;
function renderRenov(){
  const rows = agentesActuales();
  rows.sort((a,b)=> (sortKey==='agente' ? String(a[sortKey]).localeCompare(String(b[sortKey])) : a[sortKey]-b[sortKey])*sortDir);
  const maxL = Math.max(...rows.map(r=>r.llamadas),1);
  document.getElementById('tabRenov').querySelector('tbody').innerHTML = rows.map((r,i)=>{
    const pct = (r.llamadas/maxL*100).toFixed(0);
    return `<tr>
      <td><span class="pos pos-${i+1}">${i+1}</span></td>
      <td><strong>${r.agente}</strong></td>
      <td class="rt">${r.llamadas.toLocaleString()}</td>
      <td class="rt">${r.contactados.toLocaleString()}</td>
      <td class="rt">${r.tasa_contacto}%</td>
      <td class="rt"><strong>${r.ventas_renov}</strong></td>
      <td class="rt">${r.ventas_cruz}</td>
      <td class="rt">${r.rgu.toLocaleString()}</td>
      <td class="rt">${fmt$(r.mrc)}</td>
      <td class="rt">${r.mrc_prom?fmt$(r.mrc_prom):'—'}</td>
      <td style="min-width:140px"><div class="bar" style="width:${pct}%"></div><span style="font-size:10px;color:#7986cb">${pct}% de máx</span></td>
    </tr>`;
  }).join('');
}
document.querySelector('#tabRenov thead').addEventListener('click', e=>{
  const th = e.target.closest('th'); if(!th) return;
  const map = {0:'agente',1:'agente',2:'llamadas',3:'contactados',4:'tasa_contacto',5:'ventas_renov',6:'ventas_cruz',7:'rgu',8:'mrc',9:'mrc_prom'};
  const idx = Array.from(th.parentNode.children).indexOf(th);
  const k = map[idx]; if(!k) return;
  if(sortKey===k) sortDir*=-1; else {sortKey=k; sortDir = (k==='agente')?1:-1;}
  renderRenov();
});

function renderMesRenov(){
  const data = MESES.map(m=> (D.renov_por_mes[m]||{}).llamadas||0);
  const bg = MESES.map(m=> (mesAct==='TODOS'||mesAct===m) ? '#ab47bc' : 'rgba(171,71,188,.25)');
  nuevoChart('chMesRenov',{type:'bar',data:{labels:MESES,datasets:[{label:'Llamadas',data,backgroundColor:bg,borderRadius:4}]},
    options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}},
      scales:{y:{beginAtZero:true,grid:{color:'rgba(255,255,255,.06)'},ticks:{color:'#9fa8da'}},x:{ticks:{color:'#9fa8da'}}}}});
}
function renderAgLlam(){
  const top = agentesActuales().sort((a,b)=>b.llamadas-a.llamadas).slice(0,10);
  nuevoChart('chAgLlam',{type:'bar',data:{labels:top.map(r=>r.agente),datasets:[{label:'Llamadas',data:top.map(r=>r.llamadas),backgroundColor:COLORS,borderRadius:4}]},
    options:{indexAxis:'y',responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}},
      scales:{x:{beginAtZero:true,grid:{color:'rgba(255,255,255,.06)'},ticks:{color:'#9fa8da'}},y:{ticks:{color:'#9fa8da',font:{size:10}}}}}});
}

// ---------- TIPIFICACION ----------
function renderTipif(){
  const tRen = tipifRenovActual();
  const c1 = fmtCtx('chTipRenov'); if(c1){
    const t = tRen.slice(0,10);
    const otros = tRen.slice(10).reduce((s,x)=>s+x[1],0);
    const labs = t.map(x=>x[0]), vals = t.map(x=>x[1]);
    if(otros>0){labs.push('Otros');vals.push(otros);}
    nuevoChart('chTipRenov',{type:'doughnut',data:{labels:labs,datasets:[{data:vals,backgroundColor:COLORS,borderWidth:0}]},
      options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{position:'bottom',labels:{color:'#9fa8da',boxWidth:10,font:{size:10}}}}}});
  }
  const c2 = fmtCtx('chTipGlob'); if(c2){
    const tGlob = tipifGlobActual();
    const t = tGlob.slice(0,10);
    const otros = tGlob.slice(10).reduce((s,x)=>s+x[1],0);
    const labs = t.map(x=>x[0]), vals = t.map(x=>x[1]);
    if(otros>0){labs.push('Otros');vals.push(otros);}
    nuevoChart('chTipGlob',{type:'bar',data:{labels:labs.map(l=>l.length>18?l.slice(0,18)+'…':l),datasets:[{label:'Llamadas',data:vals,backgroundColor:'#42a5f5',borderRadius:4}]},
      options:{indexAxis:'y',responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}},
        scales:{x:{beginAtZero:true,grid:{color:'rgba(255,255,255,.06)'},ticks:{color:'#9fa8da'}},y:{ticks:{color:'#9fa8da',font:{size:9}}}}}});
  }
  const c3 = fmtCtx('chMarc'); if(c3){
    const r = agentesActuales();
    nuevoChart('chMarc',{type:'bar',data:{labels:['Llamadas','Contactados','No Contesta','Números Únicos'],
      datasets:[{label:'',data:[r.reduce((a,x)=>a+x.llamadas,0),r.reduce((a,x)=>a+x.contactados,0),r.reduce((a,x)=>a+x.no_contesta,0),r.reduce((a,x)=>a+x.numeros,0)],backgroundColor:['#ab47bc','#66bb6a','#ef5350','#42a5f5'],borderRadius:4}]},
      options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}},
        scales:{y:{beginAtZero:true,grid:{color:'rgba(255,255,255,.06)'},ticks:{color:'#9fa8da'}},x:{ticks:{color:'#9fa8da'}}}}});
  }
  const total = tRen.reduce((s,x)=>s+x[1],0) || 1;
  const interp = s=>{
    if(/NO CONTESTA|LLAMADA NO CONTESTA/.test(s)) return 'Cliente no respondió la llamada';
    if(/BUZON|CONTESTADOR/.test(s)) return 'Contestador / buzón de voz';
    if(/SEGUIMIENTO|VOLVER A LLAMAR|CALL BACK/.test(s)) return 'Requiere seguimiento / rellamada';
    if(/PROPUESTA/.test(s)) return 'Se envió propuesta al cliente';
    if(/NO ACEPTA|RECHAZA|NO INTERESADO/.test(s)) return 'Cliente rechazó la oferta';
    if(/VENTA|RENOVO|APROBACION/.test(s)) return 'Resultado positivo / venta';
    if(/TITULAR|TOMADOR/.test(s)) return 'No es la persona responsable';
    if(/CUELGA/.test(s)) return 'Cliente colgó';
    if(/DEUDA/.test(s)) return 'Cliente con deuda pendiente';
    if(/OCUPADO|BUSY|MUDA|LEAD|OUTBOUND/.test(s)) return 'Sin contacto efectivo';
    return 'Gestión completada';
  };
  document.getElementById('tabTip').innerHTML = tRen.map((x,i)=>{
    const p = (x[1]/total*100).toFixed(1);
    return `<tr><td>${i+1}</td><td><strong>${x[0]}</strong></td><td class="rt">${x[1].toLocaleString()}</td><td class="rt">${p}%</td><td style="font-size:11px;color:#9fa8da">${interp(x[0])}</td></tr>`;
  }).join('');
}

// ---------- VENTAS ----------
function renderVentas(){
  const r = agentesActuales().sort((a,b)=>b.ventas_total-a.ventas_total).slice(0,10);
  const c1 = fmtCtx('chVentas'); if(c1){
    nuevoChart('chVentas',{
      type:'bar',
      data:{
        labels: r.map(x=>x.agente),
        datasets:[
          {label:'Móvil', data:r.map(x=>x.ventas_movil), backgroundColor:'#42a5f5', borderRadius:3},
          {label:'Fijo', data:r.map(x=>x.ventas_fijo), backgroundColor:'#66bb6a', borderRadius:3},
          {label:'Líneas', data:r.map(x=>x.lineas_fact), backgroundColor:'#ffa726', borderRadius:3}
        ]
      },
      options:{
        responsive:true,
        maintainAspectRatio:false,
        plugins:{legend:{position:'top', labels:{color:'#9fa8da', boxWidth:10}}},
        scales:{
          y:{beginAtZero:true, stacked:true, grid:{color:'rgba(255,255,255,.06)'}, ticks:{color:'#9fa8da'}},
          x:{stacked:true, ticks:{color:'#9fa8da', font:{size:9}}}
        }
      }
    });
  }
  const c2 = fmtCtx('chMrc'); if(c2){
    nuevoChart('chMrc',{type:'bar',data:{labels:r.map(x=>x.agente),datasets:[{label:'MRC $',data:r.map(x=>x.mrc),backgroundColor:'#ab47bc',borderRadius:4}]},
      options:{indexAxis:'y',responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}},
        scales:{x:{beginAtZero:true,grid:{color:'rgba(255,255,255,.06)'},ticks:{color:'#9fa8da'}},y:{ticks:{color:'#9fa8da',font:{size:9}}}}}});
  }
  const c3 = fmtCtx('chRgu'); if(c3){
    nuevoChart('chRgu',{type:'bar',data:{labels:r.map(x=>x.agente),datasets:[{label:'RGU',data:r.map(x=>x.rgu),backgroundColor:'#26a69a',borderRadius:4}]},
      options:{indexAxis:'y',responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}},
        scales:{x:{beginAtZero:true,grid:{color:'rgba(255,255,255,.06)'},ticks:{color:'#9fa8da'}},y:{ticks:{color:'#9fa8da',font:{size:9}}}}}});
  }
  const c4 = fmtCtx('chTramite'); if(c4){
    const agg = {};
    agentesActuales().forEach(x=>{Object.entries(x.tramites||{}).forEach(([k,v])=>agg[k]=(agg[k]||0)+v);});
    const e = Object.entries(agg).sort((a,b)=>b[1]-a[1]);
    nuevoChart('chTramite',{type:'doughnut',data:{labels:e.map(x=>x[0]),datasets:[{data:e.map(x=>x[1]),backgroundColor:COLORS,borderWidth:0}]},
      options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{position:'bottom',labels:{color:'#9fa8da',boxWidth:10,font:{size:10}}}}}});
  }
  const c5 = fmtCtx('chTipoventa'); if(c5){
    const agg = {};
    agentesActuales().forEach(x=>{Object.entries(x.tipoventa||{}).forEach(([k,v])=>agg[k]=(agg[k]||0)+v);});
    const e = Object.entries(agg).sort((a,b)=>b[1]-a[1]);
    nuevoChart('chTipoventa',{type:'doughnut',data:{labels:e.map(x=>x[0]),datasets:[{data:e.map(x=>x[1]),backgroundColor:COLORS,borderWidth:0}]},
      options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{position:'bottom',labels:{color:'#9fa8da',boxWidth:10,font:{size:10}}}}}});
  }
}

// ---------- RENDIMIENTO ----------
function renderRend(){
  const roster = rosterActual();
  // Ranking: por ventas+MRC, luego llamadas
  const ranked = roster.slice().sort((a,b)=> (b.ventas_total + b.mrc/1000) - (a.ventas_total + a.mrc/1000) || b.llamadas - a.llamadas);
  // Top 3 vendedores
  const top3 = ranked.filter(x=>x.ventas_total>0).slice(0,3).map(x=>x.agente);
  // Grafico 1: Top vendedores
  const c1 = fmtCtx('chTopVen'); if(c1){
    const t = ranked.filter(x=>x.ventas_total>0).slice(0,8);
    if(!t.length){ nuevoChart('chTopVen',{type:'bar',data:{labels:['Sin ventas en este período'],datasets:[{data:[0],backgroundColor:'#ef5350'}]},options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}},scales:{y:{beginAtZero:true,grid:{color:'rgba(255,255,255,.06)'},ticks:{color:'#9fa8da'}},x:{ticks:{color:'#9fa8da'}}}}}); }
    else
    nuevoChart('chTopVen',{type:'bar',data:{labels:t.map(x=>x.agente.length>16?x.agente.slice(0,16)+'…':x.agente),
      datasets:[{label:'Ventas',data:t.map(x=>x.ventas_total),backgroundColor:COLORS,borderRadius:4},
                {label:'MRC $',data:t.map(x=>x.mrc),backgroundColor:'#42a5f5',borderRadius:4}]},
      options:{indexAxis:'y',responsive:true,maintainAspectRatio:false,plugins:{legend:{position:'top',labels:{color:'#9fa8da',boxWidth:10,font:{size:9}}}},
        scales:{x:{beginAtZero:true,grid:{color:'rgba(255,255,255,.06)'},ticks:{color:'#9fa8da'}},y:{ticks:{color:'#9fa8da',font:{size:9}}}}}});
  }
  // Grafico 2: eficiencia ventas/llamada
  const c2 = fmtCtx('chEfVxL'); if(c2){
    const t = ranked.filter(x=>x.llamadas>0).slice(0,8);
    if(!t.length){ nuevoChart('chEfVxL',{type:'bar',data:{labels:['Sin llamadas en este período'],datasets:[{data:[0],backgroundColor:'#ef5350'}]},options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}},scales:{y:{beginAtZero:true,grid:{color:'rgba(255,255,255,.06)'},ticks:{color:'#9fa8da'}},x:{ticks:{color:'#9fa8da'}}}}}); }
    else
    nuevoChart('chEfVxL',{type:'bar',data:{labels:t.map(x=>x.agente.length>16?x.agente.slice(0,16)+'…':x.agente),
      datasets:[{label:'Ventas por llamada',data:t.map(x=>x.ef_ventas_x_llamada),backgroundColor:'#26a69a',borderRadius:4}]},
      options:{indexAxis:'y',responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}},
        scales:{x:{beginAtZero:true,grid:{color:'rgba(255,255,255,.06)'},ticks:{color:'#9fa8da'}},y:{ticks:{color:'#9fa8da',font:{size:9}}}}}});
  }
  // Grafico 3: conversion %
  const c3 = fmtCtx('chEfConv'); if(c3){
    const t = ranked.filter(x=>x.llamadas>0).slice(0,8);
    if(!t.length){ nuevoChart('chEfConv',{type:'bar',data:{labels:['Sin llamadas en este período'],datasets:[{data:[0],backgroundColor:'#ef5350'}]},options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}},scales:{y:{beginAtZero:true,grid:{color:'rgba(255,255,255,.06)'},ticks:{color:'#9fa8da'}},x:{ticks:{color:'#9fa8da'}}}}}); }
    else
    nuevoChart('chEfConv',{type:'bar',data:{labels:t.map(x=>x.agente.length>16?x.agente.slice(0,16)+'…':x.agente),
      datasets:[{label:'% Conversión',data:t.map(x=>x.ef_conv_pct),backgroundColor:'#66bb6a',borderRadius:4}]},
      options:{indexAxis:'y',responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}},
        scales:{x:{beginAtZero:true,grid:{color:'rgba(255,255,255,.06)'},ticks:{color:'#9fa8da'}},y:{ticks:{color:'#9fa8da',font:{size:9}}}}}});
  }
  // Tabla ranking completo
  const maxV = Math.max(...ranked.map(x=>x.ventas_total),1);
  document.getElementById('tabRend').querySelector('tbody').innerHTML = ranked.map((r,i)=>{
    let est, cls = '', badge;
    if(r.ventas_total === 0 && r.llamadas === 0){
      est = 'Sin actividad'; cls = 'tr-low'; badge = '<span class="st-badge st-low">⚠ SIN ACTIVIDAD</span>';
    } else if(top3.includes(r.agente)){
      est = 'Top vendedor'; cls = 'tr-top'; badge = '<span class="st-badge st-top">🥇 TOP '+(top3.indexOf(r.agente)+1)+'</span>';
    } else if(r.ef_conv_pct >= 2){
      est = 'Activo'; badge = '<span class="st-badge st-act">✓ ACTIVO</span>';
    } else {
      est = 'Bajo rendimiento'; badge = '<span class="st-badge st-warn">⚠ BAJO</span>';
    }
    const pct = (r.ventas_total/maxV*100).toFixed(0);
    return `<tr class="${cls}">
      <td><span class="pos">${i+1}</span></td>
      <td><strong>${r.agente}</strong></td>
      <td>${badge}</td>
      <td class="rt">${r.llamadas.toLocaleString()}</td>
      <td class="rt"><strong>${r.ventas_total}</strong></td>
      <td class="rt">${fmt$(r.mrc)}</td>
      <td class="rt">${r.rgu.toLocaleString()}</td>
      <td class="rt">${r.ef_ventas_x_llamada? r.ef_ventas_x_llamada.toFixed(3) : '—'}</td>
      <td class="rt">${r.ef_conv_pct? r.ef_conv_pct.toFixed(2)+'%' : '—'}</td>
      <td class="rt">${r.ef_mrc_x_venta? fmt$(r.ef_mrc_x_venta) : '—'}</td>
    </tr>`;
  }).join('');
  const activos = ranked.filter(x=>x.ventas_total>0||x.llamadas>0).length;
  const sinAct = ranked.length - activos;
  document.getElementById('rendNota').textContent =
    `${ranked.length} integrantes en total · ${activos} con actividad · ${sinAct} sin actividad · Los TOP son los 3 con más ventas+MRC · 'Sin actividad' = integrantes del HC sin llamadas ni ventas en el período.`;
  document.getElementById('rendPeriod').textContent = mesAct==='TODOS' ? 'TODOS LOS MESES' : mesAct;
}

// ---------- MARCACIONES (quien llama mas / menos) ----------
function renderMarc(){
  const roster = rosterActual();
  const activos = roster.filter(x=>x.llamadas>0).sort((a,b)=>b.llamadas-a.llamadas);
  const totalLlam = activos.reduce((s,x)=>s+x.llamadas,0) || 1;
  // Grafico 1: quien llama mas (top 8)
  const c1 = fmtCtx('chLlamMas'); if(c1){
    const t = activos.slice(0,8);
    if(!t.length){ nuevoChart('chLlamMas',{type:'bar',data:{labels:['Sin llamadas'],datasets:[{data:[0],backgroundColor:'#ef5350'}]},options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}},scales:{y:{beginAtZero:true,grid:{color:'rgba(255,255,255,.06)'},ticks:{color:'#9fa8da'}},x:{ticks:{color:'#9fa8da'}}}}}); }
    else
    nuevoChart('chLlamMas',{type:'bar',data:{labels:t.map(x=>x.agente.length>15?x.agente.slice(0,15)+'…':x.agente),
      datasets:[{label:'Llamadas',data:t.map(x=>x.llamadas),backgroundColor:'#42a5f5',borderRadius:4}]},
      options:{indexAxis:'y',responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}},
        scales:{x:{beginAtZero:true,grid:{color:'rgba(255,255,255,.06)'},ticks:{color:'#9fa8da'}},y:{ticks:{color:'#9fa8da',font:{size:9}}}}}});
  }
  // Grafico 2: quien llama menos (bottom 8 con actividad)
  const c2 = fmtCtx('chLlamMenos'); if(c2){
    const t = activos.slice(-8).reverse();
    if(!t.length){ nuevoChart('chLlamMenos',{type:'bar',data:{labels:['Sin llamadas'],datasets:[{data:[0],backgroundColor:'#ef5350'}]},options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}},scales:{y:{beginAtZero:true,grid:{color:'rgba(255,255,255,.06)'},ticks:{color:'#9fa8da'}},x:{ticks:{color:'#9fa8da'}}}}}); }
    else
    nuevoChart('chLlamMenos',{type:'bar',data:{labels:t.map(x=>x.agente.length>15?x.agente.slice(0,15)+'…':x.agente),
      datasets:[{label:'Llamadas',data:t.map(x=>x.llamadas),backgroundColor:'#ffa726',borderRadius:4}]},
      options:{indexAxis:'y',responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}},
        scales:{x:{beginAtZero:true,grid:{color:'rgba(255,255,255,.06)'},ticks:{color:'#9fa8da'}},y:{ticks:{color:'#9fa8da',font:{size:9}}}}}});
  }
  // Grafico 3: contactados vs no contesta
  const c3 = fmtCtx('chContVsNc'); if(c3){
    const t = activos.slice(0,8);
    if(!t.length){ nuevoChart('chContVsNc',{type:'bar',data:{labels:['Sin datos'],datasets:[{data:[0]},{data:[0]}]},options:{responsive:true,maintainAspectRatio:false}}); }
    else
    nuevoChart('chContVsNc',{type:'bar',data:{labels:t.map(x=>x.agente.length>15?x.agente.slice(0,15)+'…':x.agente),
      datasets:[
        {label:'Contactados',data:t.map(x=>x.contactados),backgroundColor:'#66bb6a',borderRadius:3},
        {label:'No contesta',data:t.map(x=>x.no_contesta),backgroundColor:'#ef5350',borderRadius:3}
      ]},
      options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{position:'top',labels:{color:'#9fa8da',boxWidth:10,font:{size:9}}}},
        scales:{y:{beginAtZero:true,stacked:false,grid:{color:'rgba(255,255,255,.06)'},ticks:{color:'#9fa8da'}},x:{ticks:{color:'#9fa8da',font:{size:9}}}}}});
  }
  // Tabla ranking completo (incluye sin actividad al final)
  const todos = roster.slice().sort((a,b)=>b.llamadas-a.llamadas);
  document.getElementById('tabMarc').querySelector('tbody').innerHTML = todos.map((r,i)=>{
    const pctT = (r.llamadas/totalLlam*100).toFixed(1);
    const sinAct = r.llamadas===0 && r.ventas_total===0;
    const cls = sinAct ? 'tr-low' : '';
    const badge = sinAct ? '<span class="st-badge st-low">⚠ SIN ACTIVIDAD</span>' : '';
    return `<tr class="${cls}">
      <td><span class="pos">${i+1}</span></td>
      <td><strong>${r.agente}</strong> ${badge}</td>
      <td class="rt"><strong>${r.llamadas.toLocaleString()}</strong></td>
      <td class="rt">${r.contactados.toLocaleString()}</td>
      <td class="rt">${r.tasa_contacto}%</td>
      <td class="rt">${r.no_contesta.toLocaleString()}</td>
      <td class="rt">${r.numeros.toLocaleString()}</td>
      <td class="rt">${r.intentos_prom? r.intentos_prom.toFixed(2) : '—'}</td>
      <td class="rt">${r.duracion_prom_s? Math.round(r.duracion_prom_s)+' s' : '—'}</td>
      <td style="min-width:120px"><div class="bar" style="width:${Math.min(pctT*3,100)}%"></div><span style="font-size:10px;color:#7986cb">${pctT}%</span></td>
    </tr>`;
  }).join('');
  const sinActivos = todos.filter(x=>x.llamadas===0).length;
  document.getElementById('marcNota').textContent =
    `${activos.length} agentes con marcaciones · ${totalLlam.toLocaleString()} llamadas en total · ${sinActivos} integrantes sin ninguna llamada · Ordenado de quien más llama a quien menos.`;
  document.getElementById('marcPeriod').textContent = mesAct==='TODOS' ? 'TODOS LOS MESES' : mesAct;
}

function renderRecorrido(){
  const R = D.recorrido_base;
  if(!R || !R.total){ return; }
  const esTodos = mesAct==='TODOS';
  const marc = esTodos ? R.marcados : ((R.por_mes||{})[mesAct]||{}).marcados||0;
  const cont = esTodos ? R.contactados : ((R.por_mes||{})[mesAct]||{}).contactados||0;
  const sinT = R.total - marc;
  const pct = R.total ? (marc/R.total*100).toFixed(1) : 0;
  document.getElementById('recKpiRow').innerHTML = [
    ['Base Total', R.total.toLocaleString(), 'teléfonos únicos (Nextphone + Móvil sin fijo)', 'purple'],
    ['Marcados', marc.toLocaleString(), pct+'% de la base · NP '+((R.por_base.NEXTPHONE||{}).marcados||0)+' · MOV '+((R.por_base.MOVIL||{}).marcados||0), 'blue'],
    ['Sin Tocar', sinT.toLocaleString(), esTodos? 'en la base sin ninguna llamada' : 'sin llamada en este período', 'red'],
    ['Contactados', cont.toLocaleString(), marc? (cont/marc*100).toFixed(1)+'% de marcados':'—', 'green'],
    ['Intentos prom', esTodos? (R.intentos_prom||0).toFixed(2) : '—', 'llamadas por teléfono marcado', 'orange'],
  ].map(a=>`<div class="kpi c-${a[3]}"><div class="lbl">${a[0]}</div><div class="val">${a[1]}</div><div class="sub">${a[2]}</div></div>`).join('');
  nuevoChart('chRecDonut',{type:'doughnut',data:{labels:['Marcados','Sin tocar'],datasets:[{data:[marc,sinT],backgroundColor:['#66bb6a','#ef5350'],borderWidth:0}]},
    options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{position:'bottom',labels:{color:'#9fa8da',boxWidth:10,font:{size:10}}}}}});
  const ags = (R.por_agente||[]).slice().sort((a,b)=>b.telefonos-a.telefonos);
  const top = ags.slice(0,10);
  nuevoChart('chRecAg',{type:'bar',data:{labels:top.map(x=>x.agente),datasets:[{label:'Tel. base marcados',data:top.map(x=>x.telefonos),backgroundColor:COLORS,borderRadius:4}]},
    options:{indexAxis:'y',responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}},
      scales:{x:{beginAtZero:true,grid:{color:'rgba(255,255,255,.06)'},ticks:{color:'#9fa8da'}},y:{ticks:{color:'#9fa8da',font:{size:10}}}}}});
  nuevoChart('chRecMes',{type:'bar',data:{labels:MESES,datasets:[{label:'Marcados',data:MESES.map(mm=>(R.por_mes[mm]||{}).marcados||0),backgroundColor:MESES.map(mm=>(esTodos||mesAct===mm)?'#42a5f5':'rgba(66,165,245,.25)'),borderRadius:4}]},
    options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}},
      scales:{y:{beginAtZero:true,grid:{color:'rgba(255,255,255,.06)'},ticks:{color:'#9fa8da'}},x:{ticks:{color:'#9fa8da'}}}}});
  document.getElementById('tabRecAg').querySelector('tbody').innerHTML = ags.map((x,i)=>{
    const pctB = R.total ? (x.telefonos/R.total*100).toFixed(1) : 0;
    return `<tr><td><span class="pos pos-${i+1}">${i+1}</span></td><td><strong>${x.agente}</strong></td>
      <td class="rt">${x.telefonos}</td><td class="rt">${pctB}%</td>
      <td class="rt">${x.llamadas.toLocaleString()}</td><td class="rt">${x.contactados}</td></tr>`;
  }).join('') || '<tr><td colspan="6" class="ct">Sin datos</td></tr>';
  document.getElementById('recPeriod').textContent = esTodos ? 'TODOS LOS MESES' : mesAct;
  document.getElementById('recNota').textContent =
    'Llave de cruce: '+(R.llave||'telefono')+'. De los '+R.total.toLocaleString()+' teléfonos de la base, '+
    R.marcados.toLocaleString()+' ('+pct+'%) fueron marcados por el equipo RENOVACIÓN y '+
    R.contactados.toLocaleString()+' contactados.';
  renderRecDetalle();
}

function renderRecDetalle(){
  const R = D.recorrido_base; if(!R) return;
  const q = (document.getElementById('recQ').value||'').toUpperCase().trim();
  const fb = document.getElementById('recFBase').value;
  const fe = document.getElementById('recFEst').value;
  let rows = R.detalle||[];
  if(fb) rows = rows.filter(x=>x.b===fb);
  if(fe!=='') rows = rows.filter(x=>String(x.m)===fe);
  if(q) rows = rows.filter(x=>x.t.includes(q)||x.c.includes(q)||(x.n||'').toUpperCase().includes(q)||(x.e||'').toUpperCase().includes(q)||(x.ag||'').toUpperCase().includes(q));
  document.getElementById('recCount').textContent = rows.length.toLocaleString()+' de '+(R.detalle||[]).length.toLocaleString()+' registros';
  document.getElementById('tabRecDet').querySelector('tbody').innerHTML = rows.slice(0,500).map(x=>{
    const mesL = (x.ms||[]).join(', ') || (x.m?'?':'—');
    return `<tr class="${x.m?'':'tr-low'}">
      <td>${x.b}</td><td class="ct">${x.c||'—'}</td><td class="ct"><strong>${x.t}</strong></td>
      <td>${x.n||'—'}</td><td>${x.e||'—'}</td>
      <td class="ct">${x.m?'<span class="st-badge st-act">SÍ</span>':'<span class="st-badge st-low">NO</span>'}</td>
      <td class="ct">${x.co?'<span class="st-badge st-top">SÍ</span>':'—'}</td>
      <td class="rt">${x.ll||0}</td><td>${x.ag||'—'}</td><td>${mesL}</td>
    </tr>`;
  }).join('') || '<tr><td colspan="10" class="ct">Sin resultados</td></tr>';
}

function renderMrc(){
  const M = D.mrc_resumen;
  if(!M){ return; }
  const esTodos = mesAct==='TODOS';
  const pm = esTodos ? null : ((M.por_mes||{})[mesAct]||{});
  const val = k => pm ? (pm[k]||0) : (M[k]||0);
  const cli = esTodos ? M.clientes_base : (D.recorrido_base||{}).total;
  const kpisM = [
    ['Clientes Base', (cli||0).toLocaleString(), 'teléfonos únicos en la base', 'purple'],
    ['MRC Anterior', fmt$(val('mrc_anterior')), 'cartera antes de renovar (solo RENOVACION)', 'blue'],
    ['MRC Actual', fmt$(val('mrc_actual')), 'cartera con renovación hecha (solo RENOVACION)', 'green'],
    ['MRC Venta Cruzada', fmt$(val('mrc_cruzada')), 'línea nueva + portabilidad + port. externa', 'orange'],
    ['RGU Renovación', (val('rgu_renovacion')||0).toLocaleString(), 'cantidad de líneas RENOVACION', 'teal'],
    ['RGU Cruzados', (val('rgu_cruzada')||0).toLocaleString(), 'líneas de venta cruzada', 'red'],
    ['MRC Promedio', (val('rgu_renovacion')||0) ? fmt$((val('mrc_actual')) / (val('rgu_renovacion'))) : '—', 'MRC actual ÷ RGU renovación', 'teal'],
  ];
  const elM = document.getElementById('mrcKpiRow');
  if(elM) elM.innerHTML = kpisM.map(a=>
    `<div class="kpi c-${a[3]}"><div class="lbl">${a[0]}</div><div class="val">${a[1]}</div><div class="sub">${a[2]}</div></div>`
  ).join('');
  const elP = document.getElementById('mrcPeriod');
  if(elP) elP.textContent = esTodos ? 'TODOS LOS MESES' : mesAct;
  const elN = document.getElementById('mrcNota');
  if(elN) elN.textContent = 'MRC Anterior/Actual = cartera del trámite RENOVACION únicamente (sin venta cruzada). ' +
    'MRC Venta Cruzada = líneas nuevas, portabilidades y portabilidades externas (va aparte). ' +
    'La portabilidad interna no cuenta en ninguno de los dos grupos.';
  const rows = agentesActuales().slice().sort((a,b)=> (b.mrc_act_renov+b.mrc_cruz) - (a.mrc_act_renov+a.mrc_cruz)).slice(0,12);
  nuevoChart('chMrcAg',{type:'bar',
    data:{labels:rows.map(r=>r.agente.length>22?r.agente.slice(0,22)+'…':r.agente),
      datasets:[
        {label:'MRC Anterior',data:rows.map(r=>r.mrc_ant_renov),backgroundColor:'#42a5f5',borderRadius:3},
        {label:'MRC Actual (renov)',data:rows.map(r=>r.mrc_act_renov),backgroundColor:'#66bb6a',borderRadius:3},
        {label:'MRC Cruzada',data:rows.map(r=>r.mrc_cruz),backgroundColor:'#ffa726',borderRadius:3},
      ]},
    options:{indexAxis:'y',responsive:true,maintainAspectRatio:false,
      plugins:{legend:{position:'bottom',labels:{color:'#9fa8da',boxWidth:10,font:{size:9}}}},
      scales:{x:{beginAtZero:true,grid:{color:'rgba(255,255,255,.06)'},ticks:{color:'#9fa8da',font:{size:9}}},y:{ticks:{color:'#9fa8da',font:{size:9}}}}}});
  nuevoChart('chRguAg',{type:'bar',
    data:{labels:rows.map(r=>r.agente.length>22?r.agente.slice(0,22)+'…':r.agente),
      datasets:[
        {label:'RGU Renovación',data:rows.map(r=>r.rgu_renov),backgroundColor:'#ab47bc',borderRadius:3},
        {label:'RGU Cruzados',data:rows.map(r=>r.rgu_cruz),backgroundColor:'#ffa726',borderRadius:3},
      ]},
    options:{indexAxis:'y',responsive:true,maintainAspectRatio:false,
      plugins:{legend:{position:'bottom',labels:{color:'#9fa8da',boxWidth:10,font:{size:9}}}},
      scales:{x:{beginAtZero:true,grid:{color:'rgba(255,255,255,.06)'},ticks:{color:'#9fa8da',font:{size:9}}},y:{ticks:{color:'#9fa8da',font:{size:9}}}}}});
}

// ---------- WINBACK ----------
function renderWin(){
  const W = D.winback; if(!W || !W.resumen) return;
  const r = W.resumen;
  const kpisW = [
    ['Clientes a Recuperar', (r.clientes||0).toLocaleString(), 'base Winback (julio 2026)', 'purple'],
    ['RGU Potencial', (r.rgu||0).toLocaleString(), 'líneas que se podrían recuperar', 'blue'],
    ['MRC Acumulado', fmt$(r.mrc||0), 'cartera mensual que se recuperaría', 'green'],
    ['Razones de Salida', (r.razones||0), 'motivos distintos de baja', 'orange'],
  ];
  document.getElementById('winKpiRow').innerHTML = kpisW.map(a=>
    `<div class="kpi c-${a[3]}"><div class="lbl">${a[0]}</div><div class="val">${a[1]}</div><div class="sub">${a[2]}</div></div>`).join('');
  const porE = W.por_ejecutiva||[], porR = W.por_razon||[];
  const se = document.getElementById('winFEje');
  if(se && se.options.length<=1) se.innerHTML = '<option value="">Todas</option>'+porE.map(x=>`<option>${x.e}</option>`).join('');
  const sr = document.getElementById('winFRazon');
  if(sr && sr.options.length<=1) sr.innerHTML = '<option value="">Todas</option>'+porR.map(x=>`<option>${x.r}</option>`).join('');
  const raz = porR.slice(0,10);
  nuevoChart('chWinRazon',{type:'bar',data:{labels:raz.map(x=>x.r.length>24?x.r.slice(0,24)+'…':x.r),datasets:[{label:'Clientes',data:raz.map(x=>x.clientes),backgroundColor:COLORS,borderRadius:4}]},
    options:{indexAxis:'y',responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}},
      scales:{x:{beginAtZero:true,grid:{color:'rgba(255,255,255,.06)'},ticks:{color:'#9fa8da'}},y:{ticks:{color:'#9fa8da',font:{size:9}}}}}});
  const eje = porE.slice(0,12);
  nuevoChart('chWinEje',{type:'bar',data:{labels:eje.map(x=>x.e.length>22?x.e.slice(0,22)+'…':x.e),datasets:[{label:'Clientes',data:eje.map(x=>x.clientes),backgroundColor:'#ffa726',borderRadius:4}]},
    options:{indexAxis:'y',responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}},
      scales:{x:{beginAtZero:true,grid:{color:'rgba(255,255,255,.06)'},ticks:{color:'#9fa8da'}},y:{ticks:{color:'#9fa8da',font:{size:9}}}}}});
  renderWinTabla();
}
function renderWinTabla(){
  const W = D.winback; if(!W) return;
  const q = (document.getElementById('winQ').value||'').toUpperCase().trim();
  const fe = document.getElementById('winFEje').value;
  const fr = document.getElementById('winFRazon').value;
  const rows = W.detalle.filter(x=>
    (!q || (x.cli+' '+x.plan).toUpperCase().includes(q)) &&
    (!fe || x.e===fe) && (!fr || x.razon===fr));
  document.getElementById('winCount').textContent = rows.length.toLocaleString()+' de '+(W.detalle||[]).length.toLocaleString()+' registros';
  document.getElementById('tabWin').querySelector('tbody').innerHTML = rows.slice(0,400).map(x=>`<tr>
      <td>${x.e}</td><td><strong>${x.cli||'—'}</strong></td><td>${x.plan||'—'}</td><td>${x.razon}</td>
      <td class="rt">${x.rgu||0}</td><td class="rt">${fmt$(x.mrc||0)}</td>
      <td>${x.cont||'—'}</td><td style="max-width:180px;overflow:hidden;text-overflow:ellipsis">${x.correo||'—'}</td>
      <td class="ct">${x.t1||'—'}</td><td class="ct">${x.t2||'—'}</td>
    </tr>`).join('') || '<tr><td colspan="10" class="ct">Sin resultados</td></tr>';
}

// ---------- CROSS-SELL ----------
function renderCross(){
  const C = D.cross_sell; if(!C || !C.resumen) return;
  const r = C.resumen;
  const tot = r.gestiones||0, conT = r.con_tel||0;
  const kpisC = [
    ['Gestiones Totales', (tot).toLocaleString(), 'gestiones registradas en la base', 'purple'],
    ['Con Teléfono', (conT).toLocaleString(), tot? (conT/tot*100).toFixed(1)+'% del total':'', 'green'],
    ['Sin Teléfono (base)', (r.sin_tel_base||0).toLocaleString(), 'clientes de la base sin contacto', 'red'],
    ['Ejecutivas Activas', (r.ejecutivas||0), 'que registraron gestiones', 'blue'],
  ];
  document.getElementById('crossKpiRow').innerHTML = kpisC.map(a=>
    `<div class="kpi c-${a[3]}"><div class="lbl">${a[0]}</div><div class="val">${a[1]}</div><div class="sub">${a[2]}</div></div>`).join('');
  const porE = C.por_ejecutiva||[];
  const se = document.getElementById('crossFEje');
  if(se && se.options.length<=1) se.innerHTML = '<option value="">Todas</option>'+porE.map(x=>`<option>${x.e}</option>`).join('');
  const eje = porE.slice(0,12);
  nuevoChart('chCrossEje',{type:'bar',data:{labels:eje.map(x=>x.e.length>22?x.e.slice(0,22)+'…':x.e),datasets:[{label:'Gestiones',data:eje.map(x=>x.gestiones),backgroundColor:'#42a5f5',borderRadius:4}]},
    options:{indexAxis:'y',responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}},
      scales:{x:{beginAtZero:true,grid:{color:'rgba(255,255,255,.06)'},ticks:{color:'#9fa8da'}},y:{ticks:{color:'#9fa8da',font:{size:9}}}}}});
  nuevoChart('chCrossTel',{type:'doughnut',data:{labels:['Con teléfono','Sin teléfono'],datasets:[{data:[conT, Math.max(tot-conT,0)],backgroundColor:['#66bb6a','#ef5350'],borderWidth:0}]},
    options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{position:'bottom',labels:{color:'#9fa8da',boxWidth:10,font:{size:10}}}}}});
  renderCrossTabla();
}
function renderCrossTabla(){
  const C = D.cross_sell; if(!C) return;
  const q = (document.getElementById('crossQ').value||'').toUpperCase().trim();
  const fe = document.getElementById('crossFEje').value;
  const ft = document.getElementById('crossFTel').value;
  const rows = C.detalle.filter(x=>{
    const conT = x.tel && x.tel !== '#N/A';
    const okT = ft==='' || (ft==='1' ? conT : !conT);
    return (!q || (x.cli+' '+x.cuenta).toUpperCase().includes(q)) && (!fe || x.e===fe) && okT;
  });
  document.getElementById('crossCount').textContent = rows.length.toLocaleString()+' de '+(C.detalle||[]).length.toLocaleString()+' registros';
  document.getElementById('tabCross').querySelector('tbody').innerHTML = rows.slice(0,500).map(x=>`<tr>
      <td>${x.fecha||'—'}</td><td class="ct">${x.cuenta||'—'}</td><td><strong>${x.cli||'—'}</strong></td>
      <td>${x.prov||'—'}</td><td class="ct">${x.tel&&x.tel!=='#N/A'?x.tel:'—'}</td>
      <td>${x.e}</td><td style="max-width:300px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">${x.com||x.est||'—'}</td>
    </tr>`).join('') || '<tr><td colspan="7" class="ct">Sin resultados</td></tr>';
}

// ---------- RESUMEN DE CAMPANAS ----------
function renderResumen(){
  const R = D.resumen_campanas; if(!R) return;
  const ren = R.RENOVACION||{}, so = R.SOHO||{}, wb = R.WINBACK||{}, cs = R.CROSS_SELL||{};
  const cards = [
    ['Renovación (Marinel)', (ren.llamadas||0).toLocaleString()+' llamadas', (ren.ventas_total||0)+' ventas · '+fmt$(ren.mrc||0)+' MRC', 'purple'],
    ['Equipo Julio (SOHO)', (so.llamadas||0).toLocaleString()+' llamadas', (so.ventas_total||0)+' ventas · '+fmt$(so.mrc||0)+' MRC', 'blue'],
    ['Winback', (wb.clientes||0)+' clientes a recuperar', (wb.rgu||0)+' RGU · '+fmt$(wb.mrc||0)+' MRC acum.', 'orange'],
    ['Cross-selling', (cs.gestiones||0).toLocaleString()+' gestiones', (cs.con_tel||0)+' con teléfono · '+(cs.ejecutivas||0)+' ejecutivas', 'green'],
  ];
  document.getElementById('resKpiRow').innerHTML = cards.map(a=>
    `<div class="kpi c-${a[3]}"><div class="lbl">${a[0]}</div><div class="val">${a[1]}</div><div class="sub">${a[2]}</div></div>`).join('');
  const filas = [
    {n:'RENOVACIÓN', s:'Marinel Moreno', b:ren.contactos_base!=null?ren.contactos_base.toLocaleString():'—', ll:(ren.llamadas||0).toLocaleString(), p:ren.pct_recorrido!=null?ren.pct_recorrido+'%':'—', v:ren.ventas_total||0, f:ren.ventas_fijo||0, m:ren.ventas_movil||0, mrc:fmt$(ren.mrc||0), rgu:(ren.rgu||0).toLocaleString(), mp:fmt$(ren.mrc_prom||0)},
    {n:'SOHO (Julio)', s:'Julio César Arauz', b:so.contactos_base!=null?so.contactos_base.toLocaleString():'—', ll:(so.llamadas||0).toLocaleString(), p:so.pct_recorrido!=null?so.pct_recorrido+'%':'—', v:so.ventas_total||0, f:so.ventas_fijo||0, m:so.ventas_movil||0, mrc:fmt$(so.mrc||0), rgu:(so.rgu||0).toLocaleString(), mp:fmt$(so.mrc_prom||0)},
    {n:'WINBACK', s:'Julio César Arauz', b:(wb.clientes||0).toLocaleString()+' clientes', ll:'—', p:'—', v:'—', f:'—', m:'—', mrc:fmt$(wb.mrc||0), rgu:(wb.rgu||0).toLocaleString(), mp:'—'},
    {n:'CROSS-SELL', s:'Julio César Arauz', b:(cs.gestiones||0).toLocaleString()+' gestiones', ll:'—', p:'—', v:'—', f:'—', m:'—', mrc:'—', rgu:'—', mp:'—'},
  ];
  document.getElementById('tabResumen').querySelector('tbody').innerHTML = filas.map(x=>`<tr>
      <td><strong>${x.n}</strong></td><td>${x.s}</td><td class="rt">${x.b}</td><td class="rt">${x.ll}</td>
      <td class="rt">${x.p}</td><td class="rt">${x.v}</td><td class="rt">${x.f}</td><td class="rt">${x.m}</td>
      <td class="rt">${x.mrc}</td><td class="rt">${x.rgu}</td><td class="rt">${x.mp}</td>
    </tr>`).join('');
  const elN = document.getElementById('resNota');
  if(elN) elN.textContent = 'Renovación: contactos/marcaciones del recorrido de base (reporte ene-jun). Winback/Cross: bases de julio 2026. MRC Prom aquí = MRC total ÷ RGU total; el MRC Promedio de la pestana Cartera MRC usa solo renovación (MRC actual ÷ RGU renovación).';
  // recorrido por campana
  const RC = D.recorrido_campanas||{};
  const filasRC = [['WINBACK', RC.WINBACK], ['CROSS_SELL', RC.CROSS_SELL]].map(([k,v])=>{
    v = v||{};
    return `<tr><td><strong>${k}</strong></td><td class="rt">${(v.total||0).toLocaleString()}</td><td class="rt">${(v.marcados||0).toLocaleString()}</td><td class="rt">${(v.contactados||0).toLocaleString()}</td><td class="rt">${(v.pct||0)}%</td></tr>`;
  }).join('');
  document.getElementById('tabRecCamp').querySelector('tbody').innerHTML = filasRC || '<tr><td colspan="5" class="ct">Sin datos</td></tr>';
  const n2 = document.getElementById('recCampNota');
  if(n2) n2.textContent = 'Llave de cruce: teléfono (últimos 8 dígitos) contra las marcaciones SOHO del reporte (ene-jun). Las bases son de julio 2026; con bases y reporte del mismo periodo el % subirá.';
}

// ---------- EQUIPO JULIO ----------
function renderJulio(){
  if(!D.equipo_soho) return;
  const J = D.equipo_soho, Js = D.soho_sin_actividad||[];
  const sum = f => J.reduce((a,x)=>a+f(x),0);
  const totL = sum(x=>x.llamadas), totC = sum(x=>x.contactados);
  const totMrc = sum(x=>x.mrc), totRgu = sum(x=>x.rgu);
  const kpisJ = [
    ['Llamadas', totL.toLocaleString(), 'equipo SOHO (Julio)', 'purple'],
    ['Contactados', totC.toLocaleString(), totL? (totC/totL*100).toFixed(1)+'% tasa contacto':'', 'green'],
    ['Ventas Totales', sum(x=>x.ventas_total), 'móvil + fijo + líneas', 'blue'],
    ['MRC Generado', fmt$(totMrc), 'móvil + fijo + líneas', 'orange'],
    ['RGU Total', totRgu.toLocaleString(), 'servicios vendidos', 'teal'],
    ['MRC Promedio', totRgu? fmt$(totMrc/totRgu) : '—', 'MRC ÷ RGU del equipo', 'red'],
  ];
  const elJ = document.getElementById('julioKpiRow');
  if(elJ) elJ.innerHTML = kpisJ.map(a=>
    `<div class="kpi c-${a[3]}"><div class="lbl">${a[0]}</div><div class="val">${a[1]}</div><div class="sub">${a[2]}</div></div>`).join('');
  const rows = J.concat(Js).sort((a,b)=>(b.ventas_total+b.mrc/1000)-(a.ventas_total+a.mrc/1000));
  document.getElementById('tabJulio').querySelector('tbody').innerHTML = rows.map((r,i)=>`<tr class="${r.llamadas===0&&r.ventas_total===0?'tr-low':''}">
      <td><span class="pos pos-${i+1}">${i+1}</span></td>
      <td><strong>${r.agente}</strong></td>
      <td class="rt">${r.llamadas.toLocaleString()}</td><td class="rt">${r.contactados.toLocaleString()}</td>
      <td class="rt">${r.tasa_contacto}%</td>
      <td class="rt">${r.ventas_movil}</td><td class="rt">${r.ventas_fijo}</td><td class="rt"><strong>${r.ventas_total}</strong></td>
      <td class="rt">${r.rgu.toLocaleString()}</td><td class="rt">${fmt$(r.mrc)}</td><td class="rt">${r.mrc_prom?fmt$(r.mrc_prom):'—'}</td>
    </tr>`).join('') || '<tr><td colspan="11" class="ct">Sin resultados</td></tr>';
  const nJ = document.getElementById('julioNota');
  if(nJ) nJ.textContent = 'Ventas del reporte ene-jun para el equipo de Julio (campaña SOHO: Winback y Cross-selling, móvil y fijo). Los integrantes sin actividad aparecen en rojo.';
  const top = J.slice().sort((a,b)=>b.ventas_total-a.ventas_total).slice(0,10);
  nuevoChart('chJulioTop',{type:'bar',data:{labels:top.map(r=>r.agente.length>22?r.agente.slice(0,22)+'…':r.agente),datasets:[{label:'Ventas',data:top.map(r=>r.ventas_total),backgroundColor:COLORS,borderRadius:4}]},
    options:{indexAxis:'y',responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}},
      scales:{x:{beginAtZero:true,grid:{color:'rgba(255,255,255,.06)'},ticks:{color:'#9fa8da'}},y:{ticks:{color:'#9fa8da',font:{size:9}}}}}});
  const lm = J.slice().sort((a,b)=>b.llamadas-a.llamadas).slice(0,10);
  nuevoChart('chJulioLlam',{type:'bar',data:{labels:lm.map(r=>r.agente.length>22?r.agente.slice(0,22)+'…':r.agente),datasets:[{label:'Llamadas',data:lm.map(r=>r.llamadas),backgroundColor:'#42a5f5',borderRadius:4}]},
    options:{indexAxis:'y',responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}},
      scales:{x:{beginAtZero:true,grid:{color:'rgba(255,255,255,.06)'},ticks:{color:'#9fa8da'}},y:{ticks:{color:'#9fa8da',font:{size:9}}}}}});
  const mm = J.slice().sort((a,b)=>b.mrc-a.mrc).slice(0,10);
  nuevoChart('chJulioMrc',{type:'bar',data:{labels:mm.map(r=>r.agente.length>22?r.agente.slice(0,22)+'…':r.agente),datasets:[{label:'MRC $',data:mm.map(r=>r.mrc),backgroundColor:'#ffa726',borderRadius:4}]},
    options:{indexAxis:'y',responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}},
      scales:{x:{beginAtZero:true,grid:{color:'rgba(255,255,255,.06)'},ticks:{color:'#9fa8da'}},y:{ticks:{color:'#9fa8da',font:{size:9}}}}}});
}

// ---------- METAS Y CUMPLIMIENTO ----------
function metasCargadas(){
  try { return JSON.parse(localStorage.getItem('soho_metas')||'{}'); } catch(e){ return {}; }
}
function renderMetas(){
  const elV = document.getElementById('v-metas');
  if(elV && !elV.classList.contains('active')) return; // solo al estar visible
  const M = metasCargadas();
  const R = D.resumen_campanas||{};
  const def = [['RENOVACION','Renovación (Marinel Moreno)','MARINEL MORENO', R.RENOVACION||{}],
               ['SOHO','Equipo Julio (Winback + Cross-selling)','JULIO CESAR ARAUZ', R.SOHO||{}]];
  document.getElementById('tabMetas').querySelector('tbody').innerHTML = def.map(([key,nom,sup,d])=>{
    const rguA = d.rgu||0, mrcA = d.mrc||0;
    const meta = M[key]||{};
    const rguM = meta.rgu||'', mrcM = meta.mrc||'';
    const pR = rguM ? (rguA/rguM*100).toFixed(0)+'%' : '—';
    const pM = mrcM ? (mrcA/mrcM*100).toFixed(0)+'%' : '—';
    return `<tr>
      <td><strong>${nom}</strong><div style="font-size:10px;color:#7986cb">${sup}</div></td>
      <td class="rt">${rguA.toLocaleString()}</td>
      <td class="rt"><input type="number" id="meta_rgu_${key}" value="${rguM}" style="width:90px;background:rgba(255,255,255,.06);color:#e8eaf6;border:1px solid rgba(255,255,255,.15);border-radius:6px;padding:5px;text-align:right"></td>
      <td class="rt">${pR}</td>
      <td class="rt">${fmt$(mrcA)}</td>
      <td class="rt"><input type="number" id="meta_mrc_${key}" value="${mrcM}" style="width:110px;background:rgba(255,255,255,.06);color:#e8eaf6;border:1px solid rgba(255,255,255,.15);border-radius:6px;padding:5px;text-align:right"></td>
      <td class="rt">${pM}</td>
    </tr>`;
  }).join('');
  const nM = document.getElementById('metasNota');
  if(nM) nM.textContent = 'Los valores actuales (RGU y MRC) salen del reporte ene-jun. Las metas las define la gerencia por campaña; una vez guardadas, el % de cumplimiento se calcula automáticamente.';
}
function guardarMetas(){
  const M = metasCargadas();
  ['RENOVACION','SOHO'].forEach(key=>{
    const r = document.getElementById('meta_rgu_'+key), m = document.getElementById('meta_mrc_'+key);
    if(!r || !m) return;
    M[key] = {rgu: Number(r.value)||0, mrc: Number(m.value)||0};
  });
  localStorage.setItem('soho_metas', JSON.stringify(M));
  const msg = document.getElementById('metasMsg');
  if(msg){ msg.textContent = '✅ Metas guardadas'; setTimeout(()=>msg.textContent='',2500); }
  renderMetas();
}
function limpiarMetas(){
  localStorage.removeItem('soho_metas');
  renderMetas();
}

// ---------- SELECTOR DE CAMPANA (cada campana con sus pestanas y supervisor) ----------
const SUPS = {'RENOVACION':'Marinel Moreno','WINBACK':'Julio César Arauz','CROSS':'Julio César Arauz','GLOBAL':'todas las campañas'};
const FIRST_TAB = {'RENOVACION':'renov','WINBACK':'winback','CROSS':'cross','GLOBAL':'resumen'};
function setCamp(c, btn){
  document.querySelectorAll('.camp-btn').forEach(b=>b.classList.toggle('active', b===btn));
  document.querySelectorAll('.subtab-bar').forEach(b=>b.classList.toggle('hidden', b.getAttribute('data-camp')!==c));
  const sup = document.getElementById('supAct');
  if(sup) sup.textContent = 'Supervisor: '+(SUPS[c]||'');
  const bar = document.querySelector('.subtab-bar[data-camp="'+c+'"]');
  const btns = bar ? bar.querySelectorAll('.tab') : [];
  if(btns.length) sw(FIRST_TAB[c], btns[0]);
}

// ---------- VISTAS DE CAMPANA (Winback / Cross) ----------
function renderBaseMarc(key, pref, cfg){
  const C = D[key]; if(!C) return;
  const s = C.stats||{};
  const el = document.getElementById(pref+'KpiRow');
  if(el) el.innerHTML = [
    ['Llamadas a la base', (s.llamadas||0).toLocaleString(), 'del equipo de Julio (SOHO)', 'purple'],
    ['Contactados', (s.contactados||0).toLocaleString(), (s.tasa_contacto||0)+'% tasa contacto', 'green'],
    ['Tel. únicos marcados', (s.numeros_unicos||0).toLocaleString(), 'de la base '+cfg.base, 'blue'],
    ['Intentos / tel', (s.intentos_prom||0), 'promedio', 'orange'],
  ].map(a=>`<div class="kpi c-${a[3]}"><div class="lbl">${a[0]}</div><div class="val">${a[1]}</div><div class="sub">${a[2]}</div></div>`).join('');
  const ag = C.por_agente||[];
  const t = document.getElementById(pref+'AgTab');
  if(t) t.querySelector('tbody').innerHTML = ag.map(x=>`<tr>
      <td><strong>${x.agente}</strong></td><td class="rt">${x.tels}</td><td class="rt">${x.llamadas}</td><td class="rt">${x.contactados}</td>
    </tr>`).join('') || '<tr><td colspan="4" class="ct">Sin resultados</td></tr>';
  const lm = ag.slice(0,10);
  nuevoChart(pref+'AgChart',{type:'bar',data:{labels:lm.map(x=>x.agente.length>22?x.agente.slice(0,22)+'…':x.agente),datasets:[{label:'Tel. base marcados',data:lm.map(x=>x.tels),backgroundColor:'#42a5f5',borderRadius:4}]},
    options:{indexAxis:'y',responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}},
      scales:{x:{beginAtZero:true,grid:{color:'rgba(255,255,255,.06)'},ticks:{color:'#9fa8da'}},y:{ticks:{color:'#9fa8da',font:{size:9}}}}}});
  const pm = C.por_mes||{};
  nuevoChart(pref+'Mes',{type:'bar',data:{labels:MESES,datasets:[
    {label:'Marcados',data:MESES.map(m=>(pm[m]||{}).marcados||0),backgroundColor:'#ab47bc',borderRadius:3},
    {label:'Contactados',data:MESES.map(m=>(pm[m]||{}).contactados||0),backgroundColor:'#66bb6a',borderRadius:3}]},
    options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{position:'bottom',labels:{color:'#9fa8da',boxWidth:10,font:{size:9}}}},
      scales:{y:{beginAtZero:true,grid:{color:'rgba(255,255,255,.06)'},ticks:{color:'#9fa8da'}},x:{ticks:{color:'#9fa8da'}}}}});
  const n = document.getElementById(pref+'MarcNota');
  if(n) n.textContent = cfg.nota;
}
function renderBaseTip(key, pref){
  const C = D[key]; if(!C) return;
  const tip = C.tipif||[];
  const t = tip.slice(0,10);
  const otros = tip.slice(10).reduce((s,x)=>s+x[1],0);
  const labs = t.map(x=>x[0]), vals = t.map(x=>x[1]);
  if(otros>0){labs.push('Otros');vals.push(otros);}
  nuevoChart(pref+'Tip',{type:'doughnut',data:{labels:labs,datasets:[{data:vals,backgroundColor:COLORS,borderWidth:0}]},
    options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{position:'bottom',labels:{color:'#9fa8da',boxWidth:10,font:{size:10}}}}}});
  const tb = tip.slice(0,10);
  nuevoChart(pref+'TipBar',{type:'bar',data:{labels:tb.map(x=>x[0].length>24?x[0].slice(0,24)+'…':x[0]),datasets:[{label:'Llamadas',data:tb.map(x=>x[1]),backgroundColor:'#ffa726',borderRadius:4}]},
    options:{indexAxis:'y',responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}},
      scales:{x:{beginAtZero:true,grid:{color:'rgba(255,255,255,.06)'},ticks:{color:'#9fa8da'}},y:{ticks:{color:'#9fa8da',font:{size:9}}}}}});
  const el = document.getElementById(pref+'TipTab');
  if(el){ const total = tip.reduce((s,x)=>s+x[1],0)||1;
    el.innerHTML = tip.slice(0,25).map((x,i)=>`<tr>
      <td>${i+1}</td><td>${x[0]}</td><td class="rt">${x[1]}</td><td class="rt">${(x[1]/total*100).toFixed(1)}%</td>
    </tr>`).join(''); }
}
function renderBaseRec(key, pref){
  const C = D[key]; if(!C) return;
  const tot = C.total||0, mar = C.marcados||0, cont = C.contactados||0;
  const el = document.getElementById(pref+'RecKpi');
  if(el) el.innerHTML = [
    ['Base Total', tot.toLocaleString(), 'tel. únicos en la base', 'purple'],
    ['Marcados', mar.toLocaleString(), (C.pct||0)+'% recorrido', 'blue'],
    ['Contactados', cont.toLocaleString(), mar? (cont/mar*100).toFixed(1)+'% de los marcados':'', 'green'],
    ['Sin Tocar', (tot-mar).toLocaleString(), 'pendientes de llamar', 'red'],
  ].map(a=>`<div class="kpi c-${a[3]}"><div class="lbl">${a[0]}</div><div class="val">${a[1]}</div><div class="sub">${a[2]}</div></div>`).join('');
  if(pref==='winb') renderWinbDet(); else renderCrossDet();
}

function renderWinbMarc(){ renderBaseMarc('winback_camp','winb',{base:'Winback',nota:'Llamadas SOHO (equipo de Julio) que tocaron teléfonos de la base Winback (julio 2026) vs reporte ene-jun. Con bases del mismo periodo el cruce subirá.'}); }
function renderWinbTip(){ renderBaseTip('winback_camp','winb'); }
function renderWinbRec(){ renderBaseRec('winback_camp','winb'); }
function renderWinbDet(){
  const C = D.winback_camp; if(!C) return;
  const q = (document.getElementById('winbDetQ').value||'').toUpperCase().trim();
  const f = document.getElementById('winbDetF').value;
  const rows = (C.detalle||[]).filter(x=>{
    const okM = f==='' || String(x.m)===f;
    return okM && (!q || (x.cli+' '+(x.plan||'')).toUpperCase().includes(q));
  });
  document.getElementById('winbDetCount').textContent = rows.length.toLocaleString()+' de '+(C.detalle||[]).length.toLocaleString()+' registros';
  document.getElementById('winbDetTab').querySelector('tbody').innerHTML = rows.slice(0,400).map(x=>`<tr class="${x.m?'':'tr-low'}">
      <td>${x.e}</td><td><strong>${x.cli||'—'}</strong></td><td>${x.plan||'—'}</td><td>${x.razon||'—'}</td>
      <td class="rt">${x.rgu||0}</td><td class="rt">${fmt$(x.mrc||0)}</td>
      <td class="ct">${x.m?'<span class="st-badge st-act">SÍ</span>':'<span class="st-badge st-low">NO</span>'}</td>
      <td class="rt">${x.ll||0}</td><td>${x.ag||'—'}</td>
    </tr>`).join('') || '<tr><td colspan="9" class="ct">Sin resultados</td></tr>';
}

function renderCrossMarc(){ renderBaseMarc('cross_camp','cross',{base:'Cross-selling',nota:'Llamadas SOHO (equipo de Julio) que tocaron teléfonos de la base Cross-selling (julio 2026) vs reporte ene-jun.'}); }
function renderCrossTip(){ renderBaseTip('cross_camp','cross'); }
function renderCrossRec(){ renderBaseRec('cross_camp','cross'); }
function renderCrossDet(){
  const C = D.cross_camp; if(!C) return;
  const q = (document.getElementById('crossDetQ').value||'').toUpperCase().trim();
  const f = document.getElementById('crossDetF').value;
  const rows = (C.detalle||[]).filter(x=>{
    const okM = f==='' || String(x.m)===f;
    return okM && (!q || (x.cli+' '+x.cuenta).toUpperCase().includes(q));
  });
  document.getElementById('crossDetCount').textContent = rows.length.toLocaleString()+' de '+(C.detalle||[]).length.toLocaleString()+' registros';
  document.getElementById('crossDetTab').querySelector('tbody').innerHTML = rows.slice(0,500).map(x=>`<tr class="${x.m?'':'tr-low'}">
      <td>${x.fecha||'—'}</td><td class="ct">${x.cuenta||'—'}</td><td><strong>${x.cli||'—'}</strong></td>
      <td>${x.prov||'—'}</td><td class="ct">${x.tel&&x.tel!=='#N/A'?x.tel:'—'}</td>
      <td>${x.e}</td>
      <td class="ct">${x.m?'<span class="st-badge st-act">SÍ</span>':'<span class="st-badge st-low">NO</span>'}</td>
      <td class="rt">${x.ll||0}</td><td>${x.ag||'—'}</td>
    </tr>`).join('') || '<tr><td colspan="9" class="ct">Sin resultados</td></tr>';
}
function renderCrossVentas(){
  const C = D.cross_camp; if(!C || !C.ventas) return;
  const v = C.ventas;
  const el = document.getElementById('crossVentKpi');
  if(el) el.innerHTML = [
    ['Ventas Móvil', (v.movil||0), 'del equipo de Julio', 'purple'],
    ['Ventas Fijo', (v.fijo||0), 'cross-selling (fijo/internet)', 'blue'],
    ['MRC Generado', fmt$(v.mrc||0), 'móvil + fijo', 'green'],
    ['RGU', (v.rgu||0), 'servicios', 'orange'],
  ].map(a=>`<div class="kpi c-${a[3]}"><div class="lbl">${a[0]}</div><div class="val">${a[1]}</div><div class="sub">${a[2]}</div></div>`).join('');
  const pm = v.por_mes||{};
  nuevoChart('crossVentMes',{type:'bar',data:{labels:MESES,datasets:[
    {label:'Móvil',data:MESES.map(m=>(pm[m]||{}).movil||0),backgroundColor:'#42a5f5',borderRadius:3},
    {label:'Fijo',data:MESES.map(m=>(pm[m]||{}).fijo||0),backgroundColor:'#ffa726',borderRadius:3}]},
    options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{position:'bottom',labels:{color:'#9fa8da',boxWidth:10,font:{size:9}}}},
      scales:{y:{beginAtZero:true,grid:{color:'rgba(255,255,255,.06)'},ticks:{color:'#9fa8da'}},x:{ticks:{color:'#9fa8da'}}}}});
  nuevoChart('crossVentMrc',{type:'bar',data:{labels:MESES,datasets:[{label:'MRC $',data:MESES.map(m=>(pm[m]||{}).mrc||0),backgroundColor:'#66bb6a',borderRadius:4}]},
    options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}},
      scales:{y:{beginAtZero:true,grid:{color:'rgba(255,255,255,.06)'},ticks:{color:'#9fa8da'}},x:{ticks:{color:'#9fa8da'}}}}});
  const J = D.equipo_soho||[];
  const rows = J.filter(a=>a.ventas_total>0).sort((a,b)=>b.ventas_total-a.ventas_total);
  const t = document.getElementById('crossVentTab');
  if(t) t.querySelector('tbody').innerHTML = rows.map(x=>`<tr>
      <td><strong>${x.agente}</strong></td><td class="rt">${x.ventas_movil}</td><td class="rt">${x.ventas_fijo}</td>
      <td class="rt">${fmt$(x.mrc)}</td><td class="rt">${x.rgu.toLocaleString()}</td>
    </tr>`).join('') || '<tr><td colspan="5" class="ct">Sin ventas</td></tr>';
  const n = document.getElementById('crossVentNota');
  if(n) n.textContent = 'Ventas del reporte ene-jun del equipo de Julio (campaña SOHO = Winback y Cross-selling, móvil y fijo). El detalle completo por vendedor está en la vista global "Equipo Julio".';
}

function renderAll(){
  const safe = fn => { try { fn(); } catch(e){ console.error('Error en '+fn.name+':', e); } };
  safe(kpis);
  safe(renderRenov); safe(renderMesRenov); safe(renderAgLlam);
  safe(renderMrc);
  safe(renderRend);
  safe(renderMarc);
  safe(renderTipif);
  safe(renderVentas);
  safe(renderRecorrido);
  safe(renderWin);
  safe(renderCross);
  safe(renderResumen);
  safe(renderJulio);
  safe(renderMetas);
  safe(renderWinbMarc);
  safe(renderWinbTip);
  safe(renderWinbRec);
  safe(renderCrossMarc);
  safe(renderCrossTip);
  safe(renderCrossVentas);
  safe(renderCrossRec);
}

renderAll();
</script>
</body>
</html>
"""

html = html.replace('__DATA__', embed)
with open(os.path.join(BASE, 'index.html'), 'w', encoding='utf-8') as f:
    f.write(html)
print('Dashboard generado:', os.path.join(BASE, 'index.html'))
