#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Inspecciona hojas, columnas y muestras de las bases Excel SOHO (robusto)."""
import sys
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from openpyxl import load_workbook

BASES = [
    "Amir/soho_prod/bases/Winback SOHO 2026 Julio-Iniciativa Nextphone 17 julio 2026.xlsx",
    "Amir/soho_prod/bases/Base para campaña cross-sell movil Nextphone.xlsx",
    "Amir/soho_prod/bases/Base de datos Nextphone 14072026 Compartida Hoy 14 de julio 2026.xlsx",
    "Amir/soho_prod/bases/Base Móvil sin fijo 02072026 ren.xlsx",
]


def normalizar(v):
    if v is None:
        return ""
    s = str(v).strip()
    return s[:32]


for ruta in BASES:
    p = Path(ruta)
    print("=" * 90)
    print(f"ARCHIVO: {p.name}")
    if not p.exists():
        print("  !! NO EXISTE")
        continue
    try:
        wb = load_workbook(str(p), read_only=True, data_only=True)
    except Exception as e:
        print(f"  !! ERROR al abrir: {e}")
        continue
    for hoja in wb.sheetnames:
        print(f"\n  --- HOJA: '{hoja}' ---")
        ws = wb[hoja]
        try:
            filas = ws.max_row
            cols = ws.max_column
        except Exception as e:
            filas, cols = None, None
            print(f"    (dimensiones no disponibles: {e})")
        print(f"    filas={filas}, columnas={cols}")
        # primera fila (headers)
        try:
            headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            print(f"    COLUMNAS ({len(headers)}):")
            for i, h in enumerate(headers):
                print(f"      [{i}] {normalizar(h)}")
        except Exception as e:
            print(f"    (sin headers: {e})")
        # muestras de datos
        try:
            for fila in ws.iter_rows(min_row=2, max_row=4, values_only=True):
                vals = [normalizar(v) for v in fila]
                print(f"    MUESTRA: {vals}")
        except Exception as e:
            print(f"    (sin muestras: {e})")
    wb.close()
print("\nFIN")
