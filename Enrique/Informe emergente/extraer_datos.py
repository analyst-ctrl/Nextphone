# -*- coding: utf-8 -*-
"""Extrae los datos de Sharep_ago_17_2026.xlsx y genera ventas_data.js
con registros agregados por (fecha, asesor, supervisor, tipo de venta, categoria).
"""
import json
import re
import openpyxl

SRC = "Sharep_ago_17_2026.xlsx"
OUT = "ventas_data.js"

def normalizar(s):
    if s is None:
        return ""
    s = str(s).replace("\xa0", " ").strip()
    return s

def categoria(status):
    s = status.upper()
    if any(k in s for k in ("EXITOSA", "ACTIVADA", "ACTIVADO", "REGALIA ACTIVADA", "COMPLETADOS", "SINERGIA")):
        return "exitosa"
    if "CALIBRADOR" in s:
        return "calibrador"
    if any(k in s for k in ("DESISTE", "NO CONTACTADO", "REPROGRAMA", "NO TIENE EL ABONO")):
        return "desistida"
    if "SOLICITUD CANCELADA" in s:
        return "cancelada"
    if "RECHAZADO" in s:
        return "rechazada"
    if any(k in s for k in ("BO (", "MENSAJERIA", "CONFIRMAR NIP", "PENDIENTE", "APLICAR DESCUENTO", "SOC POST ENTREGA", "ASEP")):
        return "en_proceso"
    return "otros"

def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb.active

    agregado = {}
    precio_por_clave = {}
    total = 0
    for r in range(2, ws.max_row + 1):
        fecha = ws.cell(row=r, column=36).value  # Created
        if not fecha:
            continue
        fecha = fecha.date().isoformat()
        asesor = normalizar(ws.cell(row=r, column=33).value)
        supervisor = normalizar(ws.cell(row=r, column=34).value)
        tipo = normalizar(ws.cell(row=r, column=10).value)
        status = normalizar(ws.cell(row=r, column=31).value)
        cat = categoria(status)
        precio = ws.cell(row=r, column=14).value
        try:
            precio = float(precio) if precio is not None else 0.0
        except (TypeError, ValueError):
            precio = 0.0

        clave = (fecha, asesor, supervisor, tipo, cat)
        agregado[clave] = agregado.get(clave, 0) + 1
        precio_por_clave[clave] = precio_por_clave.get(clave, 0.0) + precio
        total += 1

    filas = [
        {
            "f": f, "a": a, "s": s, "t": t, "c": c,
            "n": n, "p": round(precio_por_clave[(f, a, s, t, c)], 2),
        }
        for (f, a, s, t, c), n in sorted(agregado.items())
    ]

    data = {
        "meta": {
            "fuente": SRC,
            "registros": total,
            "generado": "2026-08-17",
        },
        "rows": filas,
    }

    with open(OUT, "w", encoding="utf-8") as fh:
        fh.write("// Datos generados por extraer_datos.py - no editar a mano\n")
        fh.write("const VENTAS_DATA = ")
        json.dump(data, fh, ensure_ascii=False, separators=(",", ":"))
        fh.write(";\n")

    print(f"OK: {total} registros -> {len(filas)} filas agregadas -> {OUT}")

if __name__ == "__main__":
    main()
