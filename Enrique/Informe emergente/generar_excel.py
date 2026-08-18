# -*- coding: utf-8 -*-
"""Genera Reporte_ventas_dashboard.xlsx a partir de Sharep_ago_17_2026.xlsx,
con las mismas tablas y gráficas del dashboard (dashboard_ventas.html)."""
from collections import Counter
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, DoughnutChart, Reference

from extraer_datos import normalizar, categoria

SRC = "Sharep_ago_17_2026.xlsx"
OUT = "Reporte_ventas_dashboard.xlsx"

# ---------- estilos ----------
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
SUB_FONT = Font(size=10, color="555555")
BOLD = Font(bold=True)
THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right")

# ---------- paleta del dashboard ----------
COL_EXITOSA = "22C55E"
COL_NO = "EF4444"
COL_PROCESO = "3B82F6"
COL_TOTAL = "94A3B8"
COL_PCT = "FACC15"
CAT_COLOR = {
    "exitosa": COL_EXITOSA, "desistida": "EF4444", "cancelada": "F59E0B",
    "calibrador": "FACC15", "rechazada": "EC4899", "en_proceso": COL_PROCESO, "otros": COL_TOTAL,
}
TIPO_COLORES = ["3B82F6", "8B5CF6", "06B6D4", "F59E0B", "EC4899", "22C55E", "94A3B8"]

CAT_LABEL = {
    "exitosa": "Exitosas", "desistida": "Cliente desiste", "cancelada": "Canceladas",
    "calibrador": "Calibrador rechaza", "rechazada": "Rechazadas",
    "en_proceso": "En proceso", "otros": "Otros",
}
CAT_ORDER = ["exitosa", "desistida", "cancelada", "calibrador", "rechazada", "en_proceso", "otros"]
NO_EXITOSA = ["desistida", "cancelada", "calibrador", "rechazada"]


def estilo_tabla(ws, fila_inicio, num_cols, fila_fin=None):
    fila_fin = fila_fin or ws.max_row
    for c in range(1, num_cols + 1):
        celda = ws.cell(row=fila_inicio, column=c)
        celda.fill = HEADER_FILL
        celda.font = HEADER_FONT
        celda.alignment = CENTER
        ws.column_dimensions[get_column_letter(c)].width = 22
    for r in range(fila_inicio, fila_fin + 1):
        for c in range(1, num_cols + 1):
            celda = ws.cell(row=r, column=c)
            celda.border = BORDER
            if c > 1 and isinstance(celda.value, (int, float)):
                celda.alignment = RIGHT
    ws.freeze_panes = ws.cell(row=fila_inicio + 1, column=1)


def escribir_titulo(ws, titulo, sub, fila=1):
    ws.cell(row=fila, column=1, value=titulo).font = TITLE_FONT
    ws.cell(row=fila + 1, column=1, value=sub).font = SUB_FONT


def kpis(rows):
    por = Counter(r["c"] for r in rows)
    total = len(rows)
    exitosas = por["exitosa"]
    no = sum(por[c] for c in NO_EXITOSA)
    proc = por["en_proceso"] + por["otros"]
    valor = sum(r["p"] for r in rows if r["c"] == "exitosa")
    pct = (exitosas / total * 100) if total else 0
    return total, exitosas, pct, no, proc, valor


def fila_kpi(ws, fila, label, total, exitosas, pct, no, proc, valor):
    ws.cell(row=fila, column=1, value=label).font = BOLD
    ws.cell(row=fila, column=2, value=total).number_format = "#,##0"
    ws.cell(row=fila, column=3, value=exitosas).number_format = "#,##0"
    ws.cell(row=fila, column=4, value=round(pct, 1)).number_format = '0.0"%"'
    ws.cell(row=fila, column=5, value=no).number_format = "#,##0"
    ws.cell(row=fila, column=6, value=proc).number_format = "#,##0"
    ws.cell(row=fila, column=7, value=round(valor, 2)).number_format = '#,##0.00'


def encabezado_kpi(ws, fila):
    for c, h in enumerate(["Período", "Total", "Exitosas", "% Éxito", "No exitosas", "En proceso", "Valor planes (exitosas)"], 1):
        ws.cell(row=fila, column=c, value=h)


def encabezado_cat(ws, fila):
    for c, h in enumerate(["Categoría", "Cantidad", "% del total"], 1):
        ws.cell(row=fila, column=c, value=h)


def color_series(chart, colores):
    for serie, color in zip(chart.series, colores):
        serie.graphicalProperties.solidFill = color
        serie.graphicalProperties.line.solidFill = color
        serie.graphicalProperties.line.width = 20000  # 2pt en EMU


def estilo_chart(chart, titulo, ancho=16, alto=9):
    chart.title = titulo
    chart.width = ancho
    chart.height = alto
    chart.legend.position = "b"
    return chart


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb.active

    rows = []
    for r in range(2, ws.max_row + 1):
        fecha = ws.cell(row=r, column=36).value
        if not fecha:
            continue
        precio = ws.cell(row=r, column=14).value
        try:
            precio = float(precio) if precio is not None else 0.0
        except (TypeError, ValueError):
            precio = 0.0
        rows.append({
            "f": fecha.date().isoformat(),
            "a": normalizar(ws.cell(row=r, column=33).value),
            "s": normalizar(ws.cell(row=r, column=34).value),
            "t": normalizar(ws.cell(row=r, column=10).value),
            "c": categoria(normalizar(ws.cell(row=r, column=31).value)),
            "p": precio,
        })

    todos = rows
    ago = [r for r in rows if r["f"] >= "2026-08-01" and r["f"] <= "2026-08-31"]
    MESES = ["ene", "feb", "mar", "abr", "may", "jun", "jul", "ago", "sep", "oct", "nov", "dic"]

    out = openpyxl.Workbook()

    # =====================================================================
    # 1. Resumen (KPIs + doughnuts por resultado y por tipo)
    # =====================================================================
    s = out.active
    s.title = "Resumen"
    escribir_titulo(s, "Reporte de Ventas — Sharep",
                    f"Fuente: {SRC} · {len(rows):,} registros · Generado automáticamente")
    fila = 4
    encabezado_kpi(s, fila)
    fila_kpi(s, fila + 1, "Todo el historial (ene 2025 → ago 2026)", *kpis(todos))
    fila_kpi(s, fila + 2, "Agosto 2026 (mes actual)", *kpis(ago))
    estilo_tabla(s, fila, 7)

    # tabla: por resultado
    fila = fila + 4
    s.cell(row=fila, column=1, value="Ventas por resultado (todo el historial)").font = BOLD
    fila += 1
    encabezado_cat(s, fila)
    por = Counter(r["c"] for r in todos)
    rr = fila + 1
    for c in CAT_ORDER:
        s.cell(row=rr, column=1, value=CAT_LABEL[c])
        s.cell(row=rr, column=2, value=por[c]).number_format = "#,##0"
        s.cell(row=rr, column=3, value=round(por[c] / len(todos) * 100, 1)).number_format = '0.0"%"'
        rr += 1
    s.cell(row=rr, column=1, value="TOTAL").font = BOLD
    s.cell(row=rr, column=2, value=len(todos)).font = BOLD
    s.cell(row=rr, column=2).number_format = "#,##0"
    estilo_tabla(s, fila, 3, rr)

    # tabla: por tipo de venta
    fila_tipo = rr + 2
    s.cell(row=fila_tipo, column=1, value="Ventas por tipo de venta (todo el historial)").font = BOLD
    fila_tipo += 1
    for c, h in enumerate(["Tipo de venta", "Cantidad", "% del total"], 1):
        s.cell(row=fila_tipo, column=c, value=h)
    por_tipo = Counter(r["t"] for r in todos)
    rr_tipo = fila_tipo + 1
    for tipo, cant in por_tipo.most_common():
        s.cell(row=rr_tipo, column=1, value=tipo)
        s.cell(row=rr_tipo, column=2, value=cant).number_format = "#,##0"
        s.cell(row=rr_tipo, column=3, value=round(cant / len(todos) * 100, 1)).number_format = '0.0"%"'
        rr_tipo += 1
    estilo_tabla(s, fila_tipo, 3, rr_tipo - 1)

    # doughnut por resultado
    d1 = estilo_chart(DoughnutChart(), "Ventas por resultado")
    d1.holeSize = 55
    d1.add_data(Reference(s, min_col=2, min_row=fila + 1, max_row=rr - 1), titles_from_data=False)
    d1.set_categories(Reference(s, min_col=1, min_row=fila + 1, max_row=rr - 1))
    color_series(d1, [CAT_COLOR[c] for c in CAT_ORDER])
    s.add_chart(d1, "F5")

    # doughnut por tipo
    d2 = estilo_chart(DoughnutChart(), "Ventas por tipo de venta")
    d2.holeSize = 55
    d2.add_data(Reference(s, min_col=2, min_row=fila_tipo + 1, max_row=rr_tipo - 1), titles_from_data=False)
    d2.set_categories(Reference(s, min_col=1, min_row=fila_tipo + 1, max_row=rr_tipo - 1))
    color_series(d2, TIPO_COLORES[: rr_tipo - fila_tipo - 1])
    s.add_chart(d2, "F22")

    # =====================================================================
    # 2. Evolución diaria (tabla + línea total vs exitosas)
    # =====================================================================
    s = out.create_sheet("Evolución diaria")
    escribir_titulo(s, "Evolución diaria de ventas", "Total vs exitosas por día")
    por_dia = {}
    for r in todos:
        por_dia.setdefault(r["f"], Counter())[r["c"]] += 1
    fila = 4
    for c, h in enumerate(["Fecha", "Total", "Exitosas", "No exitosas", "En proceso"], 1):
        s.cell(row=fila, column=c, value=h)
    rr = fila + 1
    for dia in sorted(por_dia):
        pc = por_dia[dia]
        total = sum(pc.values())
        s.cell(row=rr, column=1, value=dia)
        s.cell(row=rr, column=2, value=total).number_format = "#,##0"
        s.cell(row=rr, column=3, value=pc["exitosa"]).number_format = "#,##0"
        s.cell(row=rr, column=4, value=sum(pc[c] for c in NO_EXITOSA)).number_format = "#,##0"
        s.cell(row=rr, column=5, value=pc["en_proceso"] + pc["otros"]).number_format = "#,##0"
        rr += 1
    estilo_tabla(s, fila, 5, rr - 1)
    s.column_dimensions["A"].width = 12

    linea = estilo_chart(LineChart(), "Evolución diaria — total vs exitosas", ancho=24, alto=10)
    linea.add_data(Reference(s, min_col=2, min_row=fila, max_col=3, max_row=rr - 1), titles_from_data=True)
    linea.set_categories(Reference(s, min_col=1, min_row=fila + 1, max_row=rr - 1))
    color_series(linea, [COL_TOTAL, COL_EXITOSA])
    linea.y_axis.majorGridlines = None
    s.add_chart(linea, "G4")

    # =====================================================================
    # 3. Por mes (tabla + barras apiladas)
    # =====================================================================
    s = out.create_sheet("Por mes")
    escribir_titulo(s, "Ventas por mes", "Desglose mensual por resultado")
    por_mes = {}
    for r in todos:
        por_mes.setdefault(r["f"][:7], Counter())[r["c"]] += 1
    fila = 4
    for c, h in enumerate(["Mes", "Total", "Exitosas", "No exitosas", "En proceso", "% Éxito"], 1):
        s.cell(row=fila, column=c, value=h)
    rr = fila + 1
    for mes in sorted(por_mes):
        pc = por_mes[mes]
        total = sum(pc.values())
        ex = pc["exitosa"]
        no = sum(pc[c] for c in NO_EXITOSA)
        pr = pc["en_proceso"] + pc["otros"]
        y, m = mes.split("-")
        s.cell(row=rr, column=1, value=f"{MESES[int(m)-1]} {y}")
        s.cell(row=rr, column=2, value=total).number_format = "#,##0"
        s.cell(row=rr, column=3, value=ex).number_format = "#,##0"
        s.cell(row=rr, column=4, value=no).number_format = "#,##0"
        s.cell(row=rr, column=5, value=pr).number_format = "#,##0"
        s.cell(row=rr, column=6, value=round(ex / total * 100, 1)).number_format = '0.0"%"'
        rr += 1
    estilo_tabla(s, fila, 6, rr - 1)

    bar = estilo_chart(BarChart(), "Ventas por mes (por resultado)", ancho=24, alto=10)
    bar.type = "col"
    bar.grouping = "stacked"
    bar.overlap = 100
    bar.add_data(Reference(s, min_col=3, min_row=fila, max_col=5, max_row=rr - 1), titles_from_data=True)
    bar.set_categories(Reference(s, min_col=1, min_row=fila + 1, max_row=rr - 1))
    color_series(bar, [COL_EXITOSA, COL_NO, COL_PROCESO])
    s.add_chart(bar, "H4")

    # =====================================================================
    # 4. Por supervisor (tablas + combo barras/línea % éxito)
    # =====================================================================
    def tabla_supervisor(ws_, periodo_rows, fila_):
        por_sup = {}
        for r in periodo_rows:
            por_sup.setdefault(r["s"], Counter())[r["c"]] += 1
        for c, h in enumerate(["Supervisor", "Total", "Exitosas", "No exitosas", "En proceso", "% Éxito"], 1):
            ws_.cell(row=fila_, column=c, value=h)
        rr_ = fila_ + 1
        for sup in sorted(por_sup, key=lambda x: -sum(por_sup[x].values())):
            pc = por_sup[sup]
            total = sum(pc.values())
            ex = pc["exitosa"]
            ws_.cell(row=rr_, column=1, value=sup)
            ws_.cell(row=rr_, column=2, value=total).number_format = "#,##0"
            ws_.cell(row=rr_, column=3, value=ex).number_format = "#,##0"
            ws_.cell(row=rr_, column=4, value=sum(pc[c] for c in NO_EXITOSA)).number_format = "#,##0"
            ws_.cell(row=rr_, column=5, value=pc["en_proceso"] + pc["otros"]).number_format = "#,##0"
            ws_.cell(row=rr_, column=6, value=round(ex / total * 100, 1)).number_format = '0.0"%"'
            rr_ += 1
        estilo_tabla(ws_, fila_, 6, rr_ - 1)
        return fila_, rr_

    def combo_supervisor(ws_, fila_, rr_):
        """Barras Total/Exitosas + línea % éxito en eje secundario."""
        bar = BarChart()
        bar.type = "col"
        bar.title = "Ventas por supervisor"
        bar.width = 24
        bar.height = 11
        bar.add_data(Reference(ws_, min_col=2, min_row=fila_, max_col=3, max_row=rr_ - 1), titles_from_data=True)
        bar.set_categories(Reference(ws_, min_col=1, min_row=fila_ + 1, max_row=rr_ - 1))
        color_series(bar, [COL_TOTAL, COL_EXITOSA])
        bar.y_axis.axId = 100
        bar.y_axis.title = "Ventas"
        bar.y_axis.majorGridlines = None

        linea = LineChart()
        linea.add_data(Reference(ws_, min_col=6, min_row=fila_, max_col=6, max_row=rr_ - 1), titles_from_data=True)
        linea.y_axis.axId = 200
        linea.y_axis.title = "% Éxito"
        linea.y_axis.majorGridlines = None
        linea.y_axis.crosses = "max"
        color_series(linea, [COL_PCT])
        bar += linea
        bar.legend.position = "b"
        return bar

    s = out.create_sheet("Por supervisor")
    escribir_titulo(s, "Ventas por supervisor", "Todo el historial y agosto 2026")
    fila = 4
    s.cell(row=fila, column=1, value="TODO EL HISTORIAL").font = BOLD
    fh, rr = tabla_supervisor(s, todos, fila + 1)
    s.add_chart(combo_supervisor(s, fh, rr), "H5")
    fila2 = rr + 2
    s.cell(row=fila2, column=1, value="AGOSTO 2026").font = BOLD
    fh2, rr2 = tabla_supervisor(s, ago, fila2 + 1)
    s.add_chart(combo_supervisor(s, fh2, rr2), "H" + str(fila2 + 3))

    # =====================================================================
    # 5/6. Rankings de asesores (tablas + barra horizontal top 15)
    # =====================================================================
    def tabla_ranking(ws_, periodo_rows, fila_):
        por_a = {}
        for r in periodo_rows:
            por_a.setdefault((r["a"], r["s"]), Counter())[r["c"]] += 1
        for c, h in enumerate(["#", "Asesor", "Supervisor", "Total", "Exitosas", "% Éxito", "No exitosas", "En proceso", "Valor planes"], 1):
            ws_.cell(row=fila_, column=c, value=h)
        rr_ = fila_ + 1
        orden = sorted(por_a.items(), key=lambda kv: (-kv[1]["exitosa"], -sum(kv[1].values())))
        for i, ((a, sup), pc) in enumerate(orden, 1):
            total = sum(pc.values())
            ex = pc["exitosa"]
            ws_.cell(row=rr_, column=1, value=i)
            ws_.cell(row=rr_, column=2, value=a)
            ws_.cell(row=rr_, column=3, value=sup)
            ws_.cell(row=rr_, column=4, value=total).number_format = "#,##0"
            ws_.cell(row=rr_, column=5, value=ex).number_format = "#,##0"
            ws_.cell(row=rr_, column=6, value=round(ex / total * 100, 1)).number_format = '0.0"%"'
            ws_.cell(row=rr_, column=7, value=sum(pc[c] for c in NO_EXITOSA)).number_format = "#,##0"
            ws_.cell(row=rr_, column=8, value=pc["en_proceso"] + pc["otros"]).number_format = "#,##0"
            ws_.cell(row=rr_, column=9, value=round(
                sum(r["p"] for r in periodo_rows if r["a"] == a and r["s"] == sup and r["c"] == "exitosa"), 2)
            ).number_format = '#,##0.00'
            rr_ += 1
        estilo_tabla(ws_, fila_, 9, rr_ - 1)
        ws_.auto_filter.ref = f"A{fila_}:I{rr_ - 1}"
        return fila_, rr_

    def barra_top(ws_, fila_, rr_, titulo, limite=15):
        bar = BarChart()
        bar.type = "bar"
        bar.title = titulo
        bar.width = 24
        bar.height = 11
        fin = min(fila_ + limite, rr_ - 1)
        bar.add_data(Reference(ws_, min_col=4, min_row=fila_, max_col=5, max_row=fin), titles_from_data=True)
        bar.set_categories(Reference(ws_, min_col=2, min_row=fila_ + 1, max_row=fin))
        color_series(bar, [COL_TOTAL, COL_EXITOSA])
        bar.legend.position = "b"
        return bar

    s = out.create_sheet("Ranking asesores")
    escribir_titulo(s, "Ranking de asesores — TODO EL HISTORIAL", f"{len(todos):,} ventas")
    fh, rr = tabla_ranking(s, todos, 4)
    s.add_chart(barra_top(s, fh, rr, "Top 15 asesores (exitosas)"), "K5")

    s = out.create_sheet("Ranking asesores ago 2026")
    escribir_titulo(s, "Ranking de asesores — AGOSTO 2026",
                    f"{len(ago):,} ventas (al {max(r['f'] for r in ago)})")
    fh, rr = tabla_ranking(s, ago, 4)
    s.add_chart(barra_top(s, fh, rr, "Top 15 asesores agosto (exitosas)"), "K5")

    out.save(OUT)
    print(f"OK: {len(rows)} registros -> {OUT}")


if __name__ == "__main__":
    main()
