import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from collections import defaultdict, Counter
from datetime import datetime
import sys, io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

PROMO_DAYS = {0: 'Lunes', 2: 'Miércoles', 4: 'Viernes'}
PROMO_DAY_NAMES = ['Lunes', 'Miércoles', 'Viernes']

print("Leyendo Sharep.xlsx...")
wb = openpyxl.load_workbook('../Lixi/3diasporsemana/Sharep.xlsx', read_only=True, data_only=True)
ws = wb['owssvr (6)']

headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
col_idx = {h: i for i, h in enumerate(headers)}

cols = {k: col_idx[v] for k, v in {
    'modelo': 'Modelo del equipo',
    'promo': 'Lleva promoción',
    'supervisor': 'Supervisor',
    'created': 'Created',
    'cliente': 'Nombre del cliente',
    'asesor': 'Asesor',
    'plan': 'Plan Vendido',
    'precio': 'Precio del Plan',
    'status': 'Status del Cliente',
}.items()}

dia_sup = defaultdict(lambda: defaultdict(int))
sup_total = Counter()
dia_sup_detalle = defaultdict(list)
records = []
promo_days_set = set()

print("Procesando registros...")
count = 0
iphone_count = 0
julio_count = 0
promo_count = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    count += 1

    modelo = row[cols['modelo']]
    if not modelo or 'IPHONE' not in str(modelo).upper():
        continue
    iphone_count += 1

    promo = row[cols['promo']] if cols['promo'] < len(row) else None
    if not promo or str(promo).strip().upper() != 'SI':
        continue

    created = row[cols['created']] if cols['created'] < len(row) else None
    date_obj = None
    if created:
        try:
            if isinstance(created, datetime):
                date_obj = created
            elif isinstance(created, str):
                for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y %H:%M:%S', '%d/%m/%Y']:
                    try:
                        date_obj = datetime.strptime(created, fmt)
                        break
                    except:
                        pass
        except:
            pass

    if not date_obj or date_obj.year != 2026 or date_obj.month != 7:
        continue
    julio_count += 1

    wd = date_obj.weekday()
    if wd not in PROMO_DAYS:
        continue

    promo_count += 1
    dia_str = date_obj.strftime('%Y-%m-%d')
    dia_nombre = PROMO_DAYS[wd]
    supervisor = row[cols['supervisor']] if cols['supervisor'] < len(row) else 'Sin supervisor'
    if not supervisor:
        supervisor = 'Sin supervisor'

    dia_sup[dia_str][supervisor] += 1
    sup_total[supervisor] += 1
    dia_sup_detalle[dia_str].append({
        'supervisor': supervisor,
        'cliente': row[cols['cliente']],
        'asesor': row[cols['asesor']],
        'plan': row[cols['plan']],
        'precio': row[cols['precio']],
    })
    promo_days_set.add(dia_str)

    records.append({
        'cliente': row[cols['cliente']],
        'modelo': modelo,
        'supervisor': supervisor,
        'asesor': row[cols['asesor']],
        'dia': dia_str,
        'dia_nombre': dia_nombre,
        'plan': row[cols['plan']],
        'precio': row[cols['precio']],
        'status': row[cols['status']],
    })

wb.close()

print(f"\nTotal registros: {count}")
print(f"iPhone vendidos: {iphone_count}")
print(f"iPhone + Promo Si + Julio 2026: {julio_count}")
print(f"iPhone + Promo Si + Julio + Lun/Mie/Vie: {promo_count}")
print(f"Días promo únicos: {len(promo_days_set)} ({sorted(promo_days_set)})")
print(f"Supervisores con ventas: {len(sup_total)}")

# ========== GENERAR DASHBOARD ==========
print("\nGenerando dashboard...")
out_wb = openpyxl.Workbook()

hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
hdr_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
title_font = Font(name='Arial', bold=True, size=14, color='1F4E79')
sub_font = Font(name='Arial', bold=True, size=11, color='333333')
data_font = Font(name='Arial', size=10)
gold = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
silver = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
bronze = PatternFill(start_color='CD7F32', end_color='CD7F32', fill_type='solid')
alt = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')
promo_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0')
)
center = Alignment(horizontal='center', vertical='center')
left_c = Alignment(horizontal='left', vertical='center')

def set_hdr(ws, r, c_max):
    for c in range(1, c_max+1):
        cell = ws.cell(row=r, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = border

def set_border(ws, r, c_max):
    for c in range(1, c_max+1):
        ws.cell(row=r, column=c).border = border

def style_rank(r, i, c_max, ws):
    fill = None
    if i == 1: fill = gold
    elif i == 2: fill = silver
    elif i == 3: fill = bronze
    elif i % 2 == 0: fill = alt
    if fill:
        for c in range(1, c_max+1):
            ws.cell(row=r, column=c).fill = fill

# ===== SHEET 1: RANKING GENERAL (dias promo) =====
ws1 = out_wb.active
ws1.title = "Ranking 3 Dias"

ws1.cell(row=1, column=1, value="RANKING IPHONE - PROMO 3 DIAS POR SEMANA (LUN/MIE/VIE)").font = title_font
ws1.merge_cells('A1:F1')
ws1.cell(row=2, column=1, value=f"Julio 2026 | Ventas en días promo: {promo_count} | Días promo: {len(promo_days_set)}").font = sub_font
ws1.merge_cells('A2:F2')

h1 = ['Posicion', 'Supervisor', 'Ventas Promo', '% del Total', 'Dias con Venta', 'Promedio x Dia']
for c, h in enumerate(h1, 1):
    ws1.cell(row=4, column=c, value=h)
set_hdr(ws1, 4, len(h1))

sorted_sups = sorted(sup_total.items(), key=lambda x: -x[1])
total_promo = sum(sup_total.values())

for i, (sup, cnt) in enumerate(sorted_sups, 1):
    r = 4 + i
    sup_days = sum(1 for d, sups in dia_sup.items() if sup in sups)
    ws1.cell(row=r, column=1, value=i)
    ws1.cell(row=r, column=2, value=sup)
    ws1.cell(row=r, column=3, value=cnt)
    ws1.cell(row=r, column=4, value=round(cnt / total_promo * 100, 1))
    ws1.cell(row=r, column=5, value=sup_days)
    ws1.cell(row=r, column=6, value=round(cnt / sup_days, 1) if sup_days else 0)
    set_border(ws1, r, len(h1))
    style_rank(r, i, len(h1), ws1)
    for c in range(1, len(h1)+1):
        ws1.cell(row=r, column=c).font = data_font
        ws1.cell(row=r, column=c).alignment = center if c != 2 else left_c

ws1.column_dimensions['A'].width = 10
ws1.column_dimensions['B'].width = 25
ws1.column_dimensions['C'].width = 15
ws1.column_dimensions['D'].width = 14
ws1.column_dimensions['E'].width = 18
ws1.column_dimensions['F'].width = 18

# Chart
chart1 = BarChart()
chart1.type = "col"
chart1.title = "Top Supervisores - Ventas en Dias Promo"
chart1.y_axis.title = "Ventas iPhone"
chart1.x_axis.title = "Supervisor"
chart1.style = 10
data_ref = Reference(ws1, min_col=3, min_row=4, max_row=4 + min(10, len(sorted_sups)))
cats_ref = Reference(ws1, min_col=2, min_row=5, max_row=4 + min(10, len(sorted_sups)))
chart1.add_data(data_ref, titles_from_data=True)
chart1.set_categories(cats_ref)
chart1.shape = 4
ws1.add_chart(chart1, "H4")

# ===== SHEET 2: DIA A DIA =====
ws2 = out_wb.create_sheet("Dia a Dia")

ws2.cell(row=1, column=1, value="DESGLOSE POR DIA PROMO (LUN/MIE/VIE)").font = title_font
ws2.merge_cells('A1:E1')

r = 3
for dia in sorted(promo_days_set):
    sups = sorted(dia_sup[dia].items(), key=lambda x: -x[1])
    total_dia = sum(dia_sup[dia].values())
    date_obj = datetime.strptime(dia, '%Y-%m-%d')
    dia_nombre = PROMO_DAYS[date_obj.weekday()]

    ws2.cell(row=r, column=1, value=f"{dia} - {dia_nombre}").font = Font(name='Arial', bold=True, size=12, color='1F4E79')
    ws2.cell(row=r, column=3, value=f"Total: {total_dia} iPhones")
    ws2.cell(row=r, column=3).font = Font(name='Arial', bold=True, size=11, color='333333')
    for c in range(1, 4):
        ws2.cell(row=r, column=c).fill = promo_fill
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1

    h2 = ['#', 'Supervisor', 'Ventas', '% del Dia']
    for c, h in enumerate(h2, 1):
        ws2.cell(row=r, column=c, value=h)
    set_hdr(ws2, r, len(h2))
    r += 1

    for i, (sup, cnt) in enumerate(sups, 1):
        ws2.cell(row=r, column=1, value=i)
        ws2.cell(row=r, column=2, value=sup)
        ws2.cell(row=r, column=3, value=cnt)
        ws2.cell(row=r, column=4, value=round(cnt / total_dia * 100, 1))
        set_border(ws2, r, len(h2))
        for c in range(1, len(h2)+1):
            ws2.cell(row=r, column=c).font = data_font
            ws2.cell(row=r, column=c).alignment = center if c != 2 else left_c
        r += 1
    r += 1

ws2.column_dimensions['A'].width = 22
ws2.column_dimensions['B'].width = 25
ws2.column_dimensions['C'].width = 10
ws2.column_dimensions['D'].width = 12

# ===== SHEET 3: COMPARATIVA =====
ws3 = out_wb.create_sheet("Comparativa")

ws3.cell(row=1, column=1, value="COMPARATIVA: DIAS PROMO VS TOTAL JULIO").font = title_font
ws3.merge_cells('A1:E1')

from datetime import date
today = date.today()
july_days = [d for d in range(1, 32) if date(2026, 7, d) <= today]
total_july_days = len(july_days)

lmj_days = []
other_days = []
for d in july_days:
    wd = date(2026, 7, d).weekday()
    if wd in (0, 2, 4):
        lmj_days.append(d)
    else:
        other_days.append(d)

# Re-count totals for comparison (all days julio, promo=Si, iPhone)
print("Recontando comparativas...")
wb2 = openpyxl.load_workbook('../Lixi/3diasporsemana/Sharep.xlsx', read_only=True, data_only=True)
ws2 = wb2['owssvr (6)']
_ = [cell.value for cell in next(ws2.iter_rows(min_row=1, max_row=1))]

julio_total_all = 0
julio_total_lmj = 0
julio_total_other = 0

for row in ws2.iter_rows(min_row=2, values_only=True):
    modelo = row[cols['modelo']] if cols['modelo'] < len(row) else None
    if not modelo or 'IPHONE' not in str(modelo).upper():
        continue
    promo = row[cols['promo']] if cols['promo'] < len(row) else None
    if not promo or str(promo).strip().upper() != 'SI':
        continue
    created = row[cols['created']] if cols['created'] < len(row) else None
    date_obj = None
    if created:
        try:
            if isinstance(created, datetime):
                date_obj = created
            elif isinstance(created, str):
                for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y %H:%M:%S', '%d/%m/%Y']:
                    try:
                        date_obj = datetime.strptime(created, fmt)
                        break
                    except:
                        pass
        except:
            pass
    if not date_obj or date_obj.year != 2026 or date_obj.month != 7:
        continue
    julio_total_all += 1
    if date_obj.weekday() in (0, 2, 4):
        julio_total_lmj += 1
    else:
        julio_total_other += 1
wb2.close()

ws3.cell(row=3, column=1, value="METRICA").font = Font(name='Arial', bold=True, size=11)
ws3.cell(row=3, column=2, value="VALOR").font = Font(name='Arial', bold=True, size=11)
ws3.cell(row=3, column=3, value="%").font = Font(name='Arial', bold=True, size=11)
set_hdr(ws3, 3, 3)

metrics = [
    ('Días transcurridos Julio', total_july_days, '100%'),
    ('Días Promo (L/M/J) transcurridos', len(lmj_days), f'{round(len(lmj_days)/total_july_days*100,1)}%'),
    ('Días No Promo transcurridos', len(other_days), f'{round(len(other_days)/total_july_days*100,1)}%'),
    ('', None, ''),
    ('Ventas iPhone + Promo Si (Total Julio)', julio_total_all, '100%'),
    ('Ventas en Días Promo (L/M/J)', julio_total_lmj, f'{round(julio_total_lmj/julio_total_all*100,1) if julio_total_all else 0}%'),
    ('Ventas en Días No Promo', julio_total_other, f'{round(julio_total_other/julio_total_all*100,1) if julio_total_all else 0}%'),
    ('', None, ''),
    ('Promedio x día (general)', round(julio_total_all/total_july_days, 1) if total_july_days else 0, ''),
    ('Promedio x día promo', round(julio_total_lmj/len(lmj_days), 1) if lmj_days else 0, ''),
]

for i, (label, val, pct) in enumerate(metrics, 4):
    ws3.cell(row=i, column=1, value=label)
    if val is not None:
        ws3.cell(row=i, column=2, value=val)
    ws3.cell(row=i, column=3, value=pct)
    ws3.cell(row=i, column=1).font = data_font
    ws3.cell(row=i, column=2).font = data_font
    ws3.cell(row=i, column=3).font = data_font

ws3.column_dimensions['A'].width = 40
ws3.column_dimensions['B'].width = 15
ws3.column_dimensions['C'].width = 10

# Chart comparativo
chart3 = BarChart()
chart3.type = "col"
chart3.title = "Ventas: Dias Promo vs No Promo"
chart3.y_axis.title = "Ventas"
chart3.style = 10

# Small data table for chart
ws3.cell(row=16, column=1, value="Tipo Dia")
ws3.cell(row=16, column=2, value="Ventas")
ws3.cell(row=17, column=1, value="Dias Promo (L/M/J)")
ws3.cell(row=17, column=2, value=julio_total_lmj)
ws3.cell(row=18, column=1, value="Dias No Promo")
ws3.cell(row=18, column=2, value=julio_total_other)

data_ref3 = Reference(ws3, min_col=2, min_row=16, max_row=18)
cats_ref3 = Reference(ws3, min_col=1, min_row=17, max_row=18)
chart3.add_data(data_ref3, titles_from_data=True)
chart3.set_categories(cats_ref3)
ws3.add_chart(chart3, "E3")

# ===== SHEET 4: DETALLE VENTAS =====
ws4 = out_wb.create_sheet("Detalle Ventas")

ws4.cell(row=1, column=1, value="DETALLE VENTAS IPHONE - PROMO 3 DIAS").font = title_font
ws4.merge_cells('A1:I1')
ws4.cell(row=2, column=1, value=f"Registros: {len(records)}").font = sub_font
ws4.merge_cells('A2:I2')

h4 = ['#', 'Cliente', 'Modelo', 'Supervisor', 'Asesor', 'Dia', 'Dia Semana', 'Plan', 'Status']
for c, h in enumerate(h4, 1):
    ws4.cell(row=4, column=c, value=h)
set_hdr(ws4, 4, len(h4))

for i, rec in enumerate(records, 1):
    r = 4 + i
    ws4.cell(row=r, column=1, value=i)
    ws4.cell(row=r, column=2, value=rec['cliente'])
    ws4.cell(row=r, column=3, value=rec['modelo'])
    ws4.cell(row=r, column=4, value=rec['supervisor'])
    ws4.cell(row=r, column=5, value=rec['asesor'])
    ws4.cell(row=r, column=6, value=rec['dia'])
    ws4.cell(row=r, column=7, value=rec['dia_nombre'])
    ws4.cell(row=r, column=8, value=rec['plan'])
    ws4.cell(row=r, column=9, value=rec['status'])
    set_border(ws4, r, len(h4))
    for c in range(1, len(h4)+1):
        ws4.cell(row=r, column=c).font = data_font
        ws4.cell(row=r, column=c).alignment = center if c not in (2, 3) else left_c
        if i % 2 == 0:
            ws4.cell(row=r, column=c).fill = alt

ws4.column_dimensions['A'].width = 6
ws4.column_dimensions['B'].width = 30
ws4.column_dimensions['C'].width = 30
ws4.column_dimensions['D'].width = 22
ws4.column_dimensions['E'].width = 22
ws4.column_dimensions['F'].width = 14
ws4.column_dimensions['G'].width = 14
ws4.column_dimensions['H'].width = 12
ws4.column_dimensions['I'].width = 30

outfile = '../Lixi/3diasporsemana/Ranking 3 Dias por Semana - Julio 2026.xlsx'
out_wb.save(outfile)
print(f"\nDashboard generado: {outfile}")
print("\nHojas del dashboard:")
print("  1. Ranking 3 Dias - Ranking general solo L/M/J")
print("  2. Dia a Dia - Desglose por cada día promo")
print("  3. Comparativa - Días promo vs días no promo")
print("  4. Detalle Ventas - Todos los registros")
