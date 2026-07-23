import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.utils import get_column_letter
from collections import defaultdict, Counter
from datetime import datetime, date
import sys, io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

PROMO_DAYS = {0: 'Lunes', 2: 'Miércoles', 4: 'Viernes'}
TODAY = date.today()
EXCLUDE = {'', 'N/A', 'NINGUNO', 'NO APLICA', 'NA', 'S/N', 'NO LLEVA EQUIPO', 'SIN EQUIPO', 'EQUIPO PROPIO'}

def week_of_july(d):
    return (d.day - 1) // 7 + 1

print("Procesando datos...")
wb = openpyxl.load_workbook('../Lixi/3diasporsemana/Sharep.xlsx', read_only=True, data_only=True)
ws = wb['owssvr (6)']
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
col_idx = {h: i for i, h in enumerate(headers)}
cols = {k: col_idx[v] for k, v in {
    'modelo': 'Modelo del equipo', 'promo': 'Lleva promoción',
    'supervisor': 'Supervisor', 'created': 'Created',
    'cliente': 'Nombre del cliente', 'asesor': 'Asesor',
    'plan': 'Plan Vendido', 'status': 'Status del Cliente',
}.items()}

all_rows = []
dia_sup = defaultdict(lambda: defaultdict(int))
sup_total = defaultdict(int)
sup_total_iphone = defaultdict(int)
sup_dias = defaultdict(set)
dia_info = {}
modelos_set = set()

for row in ws.iter_rows(min_row=2, values_only=True):
    modelo_raw = row[cols['modelo']]
    modelo = str(modelo_raw or '').strip()
    if not modelo or modelo.upper() in EXCLUDE:
        continue
    modelo_up = modelo.upper()
    modelos_set.add(modelo)
    promo = row[cols['promo']] if cols['promo'] < len(row) else None
    if not promo or str(promo).strip().upper() != 'SI':
        continue
    created = row[cols['created']] if cols['created'] < len(row) else None
    date_obj = None
    if created:
        try:
            if isinstance(created, datetime): date_obj = created
            elif isinstance(created, str):
                for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y %H:%M:%S', '%d/%m/%Y']:
                    try: date_obj = datetime.strptime(created, fmt); break
                    except: pass
        except: pass
    if not date_obj or date_obj.year != 2026 or date_obj.month != 7:
        continue
    wd = date_obj.weekday()
    if wd not in PROMO_DAYS:
        continue
    dia_str = date_obj.strftime('%Y-%m-%d')
    supervisor = str(row[cols['supervisor']] or 'Sin supervisor')
    es_iphone = 'IPHONE' in modelo_up
    dia_sup[dia_str][supervisor] += 1
    sup_total[supervisor] += 1
    if es_iphone:
        sup_total_iphone[supervisor] += 1
    sup_dias[supervisor].add(dia_str)
    if dia_str not in dia_info:
        dia_info[dia_str] = {'semana': week_of_july(date_obj), 'nombre': PROMO_DAYS[wd], 'weekday': wd}
    all_rows.append({
        'cliente': str(row[cols['cliente']] or ''),
        'modelo': modelo,
        'supervisor': supervisor,
        'asesor': str(row[cols['asesor']] or ''),
        'dia': dia_str, 'dia_nombre': PROMO_DAYS[wd],
        'semana': week_of_july(date_obj),
        'es_iphone': es_iphone,
        'plan': str(row[cols['plan']] or ''),
        'status': str(row[cols['status']] or ''),
    })
wb.close()
print(f"Datos: {len(all_rows)} registros, {len(sup_total)} supervisores, {len(dia_info)} dias, {len(modelos_set)} modelos")

# Comparison data (full julio)
wb2 = openpyxl.load_workbook('../Lixi/3diasporsemana/Sharep.xlsx', read_only=True, data_only=True)
ws2 = wb2['owssvr (6)']
_ = [cell.value for cell in next(ws2.iter_rows(min_row=1, max_row=1))]
julio_all_total = julio_all_lmj = 0
julio_iphone_total = julio_iphone_lmj = 0
for row in ws2.iter_rows(min_row=2, values_only=True):
    modelo_raw = row[cols['modelo']] if cols['modelo'] < len(row) else None
    m = str(modelo_raw or '').strip().upper()
    if not m or m in EXCLUDE: continue
    promo = row[cols['promo']] if cols['promo'] < len(row) else None
    if not promo or str(promo).strip().upper() != 'SI': continue
    created = row[cols['created']] if cols['created'] < len(row) else None
    date_obj = None
    if created:
        try:
            if isinstance(created, datetime): date_obj = created
            elif isinstance(created, str):
                for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y %H:%M:%S', '%d/%m/%Y']:
                    try: date_obj = datetime.strptime(created, fmt); break
                    except: pass
        except: pass
    if not date_obj or date_obj.year != 2026 or date_obj.month != 7: continue
    es_iphone = 'IPHONE' in m
    julio_all_total += 1
    if es_iphone: julio_iphone_total += 1
    if date_obj.weekday() in (0, 2, 4):
        julio_all_lmj += 1
        if es_iphone: julio_iphone_lmj += 1
wb2.close()

july_days_total = [d for d in range(1, 32) if date(2026, 7, d) <= TODAY]
total_july_days = len(july_days_total)
lmj_count = sum(1 for d in july_days_total if date(2026, 7, d).weekday() in (0, 2, 4))
other_count = total_july_days - lmj_count
dias_sorted = sorted(dia_info.keys())
sup_sorted = sorted(sup_total.keys(), key=lambda x: -sup_total[x])

# Styles
hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
hdr_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
title_font = Font(name='Arial', bold=True, size=14, color='1F4E79')
sub_font = Font(name='Arial', bold=True, size=11, color='333333')
data_font = Font(name='Arial', size=10)
bold_font = Font(name='Arial', bold=True, size=10)
gold = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
silver = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
bronze = PatternFill(start_color='CD7F32', end_color='CD7F32', fill_type='solid')
alt = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')
green_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
iphone_fill = PatternFill(start_color='F3E8FF', end_color='F3E8FF', fill_type='solid')
border = Border(
    left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0')
)
center = Alignment(horizontal='center', vertical='center')
left_a = Alignment(horizontal='left', vertical='center')
wrap_a = Alignment(horizontal='center', vertical='center', wrap_text=True)

def set_hdr(ws, r, c_max):
    for c in range(1, c_max+1):
        cell = ws.cell(row=r, column=c)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = wrap_a; cell.border = border

def style_cell(ws, r, c, align=None, font=None):
    cell = ws.cell(row=r, column=c)
    cell.font = font or data_font; cell.border = border
    cell.alignment = align or center

def fmt_border(ws, r, c_max):
    for c in range(1, c_max+1):
        ws.cell(row=r, column=c).border = border

print("Generando Excel...")
out = openpyxl.Workbook()

# ===== SHEET 1: TABLA DINAMICA (Todos) =====
ws1 = out.active
ws1.title = "Tabla Dinamica"

ws1.cell(row=1, column=1, value="TABLA DINAMICA - TODOS LOS EQUIPOS (Promo L/M/J)").font = title_font
ws1.merge_cells('A1:H1')
ws1.cell(row=2, column=1, value=f"Julio 2026 | Corte: {TODAY} | {len(modelos_set)} modelos diferentes").font = sub_font
ws1.merge_cells('A2:H2')

hd = ['Supervisor', 'Total'] + dias_sorted + ['% Total', 'iPhones']
for c, h in enumerate(hd, 1):
    ws1.cell(row=4, column=c, value=h)
set_hdr(ws1, 4, len(hd))

for i, sup in enumerate(sup_sorted):
    r = 5 + i
    ws1.cell(row=r, column=1, value=sup)
    style_cell(ws1, r, 1, left_a)
    ws1.cell(row=r, column=2, value=sup_total[sup])
    style_cell(ws1, r, 2, font=bold_font)
    for j, dia in enumerate(dias_sorted):
        val = dia_sup[dia].get(sup, 0)
        ws1.cell(row=r, column=3+j, value=val if val else '')
        style_cell(ws1, r, 3+j, center, Font(name='Arial', size=10, color='333333' if val else 'CCCCCC'))
    pct = round(sup_total[sup] / sum(sup_total.values()) * 100, 1)
    ws1.cell(row=r, column=len(hd)-1, value=pct)
    style_cell(ws1, r, len(hd)-1)
    ws1.cell(row=r, column=len(hd), value=sup_total_iphone.get(sup, 0))
    style_cell(ws1, r, len(hd))
    if i % 2 == 1:
        for c in range(1, len(hd)+1):
            ws1.cell(row=r, column=c).fill = alt

# Totals
rt = 5 + len(sup_sorted)
ws1.cell(row=rt, column=1, value="TOTAL")
style_cell(ws1, rt, 1, left_a, bold_font)
ws1.cell(row=rt, column=2, value=sum(sup_total.values()))
style_cell(ws1, rt, 2, font=bold_font)
for j, dia in enumerate(dias_sorted):
    val = sum(dia_sup[dia].values())
    ws1.cell(row=rt, column=3+j, value=val)
    style_cell(ws1, rt, 3+j, font=bold_font)
ws1.cell(row=rt, column=len(hd)-1, value=100)
style_cell(ws1, rt, len(hd)-1, font=bold_font)
ws1.cell(row=rt, column=len(hd), value=sum(sup_total_iphone.values()))
style_cell(ws1, rt, len(hd), font=bold_font)

ws1.column_dimensions['A'].width = 25; ws1.column_dimensions['B'].width = 10
for j in range(len(dias_sorted)):
    ws1.column_dimensions[get_column_letter(3+j)].width = 14
ws1.column_dimensions[get_column_letter(len(hd)-1)].width = 10
ws1.column_dimensions[get_column_letter(len(hd))].width = 10

# Stacked chart
csr = rt + 3
ws1.cell(row=csr, column=1, value="Supervisor")
ws1.cell(row=csr, column=2, value="Total")
for j, d in enumerate(dias_sorted):
    ws1.cell(row=csr, column=3+j, value=d)
for i, sup in enumerate(sup_sorted):
    ws1.cell(row=csr+1+i, column=1, value=sup)
    ws1.cell(row=csr+1+i, column=2, value=sup_total[sup])
    for j, d in enumerate(dias_sorted):
        ws1.cell(row=csr+1+i, column=3+j, value=dia_sup[d].get(sup, 0))

chart1 = BarChart()
chart1.type = "col"; chart1.title = "Ventas x Supervisor y Dia"; chart1.style = 10; chart1.grouping = "stacked"
data1 = Reference(ws1, min_col=2, min_row=csr, max_row=csr+len(sup_sorted), max_col=1+len(dias_sorted))
cats1 = Reference(ws1, min_col=1, min_row=csr+1, max_row=csr+len(sup_sorted))
chart1.add_data(data1, titles_from_data=True, from_rows=True)
chart1.set_categories(cats1)
ws1.add_chart(chart1, f"A{csr + len(sup_sorted) + 3}")

# ===== SHEET 2: RANKING (Todos) =====
ws2 = out.create_sheet("Ranking General")
ws2.cell(row=1, column=1, value="RANKING SUPERVISORES - TODOS LOS EQUIPOS (Promo L/M/J)").font = title_font
ws2.merge_cells('A1:G1')
ws2.cell(row=2, column=1, value=f"Total: {sum(sup_total.values())} equipos | iPhones: {sum(sup_total_iphone.values())} | {len(dia_info)} dias | Corte: {TODAY}").font = sub_font
ws2.merge_cells('A2:G2')

h2 = ['#', 'Supervisor', 'Total', 'iPhones', '% del Total', 'Dias', 'Prom/Dia']
for c, h in enumerate(h2, 1):
    ws2.cell(row=4, column=c, value=h)
set_hdr(ws2, 4, len(h2))

sorted_sups = sorted(sup_total.items(), key=lambda x: -x[1])
total_gral = sum(sup_total.values())
for i, (sup, cnt) in enumerate(sorted_sups, 1):
    r = 4 + i
    d_cnt = len(sup_dias[sup])
    ip_cnt = sup_total_iphone.get(sup, 0)
    ws2.cell(row=r, column=1, value=i)
    ws2.cell(row=r, column=2, value=sup)
    ws2.cell(row=r, column=3, value=cnt)
    ws2.cell(row=r, column=4, value=ip_cnt)
    ws2.cell(row=r, column=5, value=round(cnt / total_gral * 100, 1))
    ws2.cell(row=r, column=6, value=d_cnt)
    ws2.cell(row=r, column=7, value=round(cnt / d_cnt, 1) if d_cnt else 0)
    fmt_border(ws2, r, len(h2))
    fill = gold if i==1 else silver if i==2 else bronze if i==3 else (alt if i%2==0 else None)
    if fill:
        for c in range(1, len(h2)+1):
            ws2.cell(row=r, column=c).fill = fill
    for c in range(1, len(h2)+1):
        style_cell(ws2, r, c, left_a if c==2 else None)

ws2.column_dimensions['A'].width = 8; ws2.column_dimensions['B'].width = 25
ws2.column_dimensions['C'].width = 10; ws2.column_dimensions['D'].width = 10
ws2.column_dimensions['E'].width = 12; ws2.column_dimensions['F'].width = 10
ws2.column_dimensions['G'].width = 12

chart2 = BarChart()
chart2.type = "col"; chart2.title = "Top Supervisores - Todos los Equipos"
chart2.y_axis.title = "Ventas"; chart2.style = 10
d2 = Reference(ws2, min_col=3, min_row=4, max_row=4+len(sorted_sups))
c2 = Reference(ws2, min_col=2, min_row=5, max_row=4+len(sorted_sups))
chart2.add_data(d2, titles_from_data=True); chart2.set_categories(c2)
ws2.add_chart(chart2, "I4")

# ===== SHEET 3: SOLO IPHONE =====
ws3 = out.create_sheet("Ranking iPhone")

iphone_rows = [r for r in all_rows if r['es_iphone']]
iphone_sup = Counter()
iphone_dia_sup = defaultdict(lambda: defaultdict(int))
iphone_dias = set()
for r in iphone_rows:
    iphone_sup[r['supervisor']] += 1
    iphone_dia_sup[r['dia']][r['supervisor']] += 1
    iphone_dias.add(r['dia'])

ws3.cell(row=1, column=1, value="RANKING IPHONE - PROMO 3 DIAS").font = title_font
ws3.merge_cells('A1:F1')
ws3.cell(row=2, column=1, value=f"Total: {len(iphone_rows)} iPhones | {len(iphone_sup)} supervisores | {len(iphone_dias)} dias").font = sub_font
ws3.merge_cells('A2:F2')

h3 = ['#', 'Supervisor', 'iPhones', '%', 'Dias', 'Prom/Dia']
for c, h in enumerate(h3, 1):
    ws3.cell(row=4, column=c, value=h)
set_hdr(ws3, 4, len(h3))

sorted_iphone = sorted(iphone_sup.items(), key=lambda x: -x[1])
total_iph = len(iphone_rows)
for i, (sup, cnt) in enumerate(sorted_iphone, 1):
    r = 4 + i
    d_cnt = sum(1 for d, sups in iphone_dia_sup.items() if sup in sups)
    ws3.cell(row=r, column=1, value=i)
    ws3.cell(row=r, column=2, value=sup)
    ws3.cell(row=r, column=3, value=cnt)
    ws3.cell(row=r, column=4, value=round(cnt / total_iph * 100, 1) if total_iph else 0)
    ws3.cell(row=r, column=5, value=d_cnt)
    ws3.cell(row=r, column=6, value=round(cnt / d_cnt, 1) if d_cnt else 0)
    fmt_border(ws3, r, len(h3))
    fill = gold if i==1 else silver if i==2 else bronze if i==3 else (alt if i%2==0 else None)
    if fill:
        for c in range(1, len(h3)+1):
            ws3.cell(row=r, column=c).fill = fill
    for c in range(1, len(h3)+1):
        style_cell(ws3, r, c, left_a if c==2 else None)

ws3.column_dimensions['A'].width = 8; ws3.column_dimensions['B'].width = 25
ws3.column_dimensions['C'].width = 10; ws3.column_dimensions['D'].width = 8
ws3.column_dimensions['E'].width = 8; ws3.column_dimensions['F'].width = 10

chart3 = BarChart()
chart3.type = "col"; chart3.title = "Ranking iPhone - Promo 3 Dias"
chart3.style = 10
ws3.cell(row=20, column=1, value="Sup"); ws3.cell(row=20, column=2, value="Ventas")
for i, (s, c) in enumerate(sorted_iphone):
    ws3.cell(row=21+i, column=1, value=s); ws3.cell(row=21+i, column=2, value=c)
d3 = Reference(ws3, min_col=2, min_row=20, max_row=20+len(sorted_iphone))
c3 = Reference(ws3, min_col=1, min_row=21, max_row=20+len(sorted_iphone))
chart3.add_data(d3, titles_from_data=True); chart3.set_categories(c3)
ws3.add_chart(chart3, "H4")

# ===== SHEET 4: DESGLOSE DIARIO =====
ws4 = out.create_sheet("Desglose Diario")
ws4.cell(row=1, column=1, value="DESGLOSE POR DIA PROMO (LUN/MIE/VIE)").font = title_font
ws4.merge_cells('A1:D1')

r = 3
for dia in dias_sorted:
    info = dia_info[dia]
    sups_data = sorted(dia_sup[dia].items(), key=lambda x: -x[1])
    total_dia = sum(dia_sup[dia].values())
    ws4.cell(row=r, column=1, value=f"{dia} - {info['nombre']} (Semana {info['semana']})")
    ws4.cell(row=r, column=1).font = Font(name='Arial', bold=True, size=12, color='1F4E79')
    ws4.cell(row=r, column=3, value=f"Total: {total_dia}")
    ws4.cell(row=r, column=3).font = Font(name='Arial', bold=True, size=11, color='333333')
    for c in range(1, 4): ws4.cell(row=r, column=c).fill = green_fill
    ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1
    h4 = ['#', 'Supervisor', 'Ventas', '% del Dia']
    for c, h in enumerate(h4, 1): ws4.cell(row=r, column=c, value=h)
    set_hdr(ws4, r, len(h4)); r += 1
    for i, (sup, cnt) in enumerate(sups_data, 1):
        ws4.cell(row=r, column=1, value=i); ws4.cell(row=r, column=2, value=sup)
        ws4.cell(row=r, column=3, value=cnt); ws4.cell(row=r, column=4, value=round(cnt/total_dia*100, 1))
        fmt_border(ws4, r, len(h4))
        for c in range(1, len(h4)+1): style_cell(ws4, r, c, left_a if c==2 else None)
        r += 1
    r += 1

ws4.column_dimensions['A'].width = 30; ws4.column_dimensions['B'].width = 25
ws4.column_dimensions['C'].width = 10; ws4.column_dimensions['D'].width = 12

# ===== SHEET 5: COMPARATIVA =====
ws5 = out.create_sheet("Comparativa")
ws5.cell(row=1, column=1, value="COMPARATIVA: DIAS PROMO VS TOTAL JULIO").font = title_font
ws5.merge_cells('A1:D1')

labels_5 = [
    ('DIAS', '', '', ''),
    ('Dias transcurridos Julio', total_july_days, '100%', ''),
    ('Dias Promo (L/M/J)', lmj_count, f'{round(lmj_count/total_july_days*100,1)}%' if total_july_days else '0%', ''),
    ('Dias No Promo', other_count, f'{round(other_count/total_july_days*100,1)}%' if total_july_days else '0%', ''),
    ('', '', '', ''),
    ('VENTAS - TODOS LOS EQUIPOS', '', '', ''),
    ('Total Julio', julio_all_total, '100%', ''),
    ('En Dias Promo (L/M/J)', julio_all_lmj, f'{round(julio_all_lmj/julio_all_total*100,1)}%' if julio_all_total else '0%', ''),
    ('En Dias No Promo', julio_all_total - julio_all_lmj, f'{round((julio_all_total-julio_all_lmj)/julio_all_total*100,1)}%' if julio_all_total else '0%', ''),
    ('', '', '', ''),
    ('VENTAS - SOLO IPHONE', '', '', ''),
    ('Total Julio', julio_iphone_total, '100%', ''),
    ('En Dias Promo (L/M/J)', julio_iphone_lmj, f'{round(julio_iphone_lmj/julio_iphone_total*100,1)}%' if julio_iphone_total else '0%', ''),
    ('En Dias No Promo', julio_iphone_total - julio_iphone_lmj, f'{round((julio_iphone_total-julio_iphone_lmj)/julio_iphone_total*100,1)}%' if julio_iphone_total else '0%', ''),
    ('', '', '', ''),
    ('PROMEDIOS (Todos)', '', '', ''),
    ('Promedio x dia general', round(julio_all_total/total_july_days, 1) if total_july_days else 0, '', ''),
    ('Promedio x dia promo', round(julio_all_lmj/lmj_count, 1) if lmj_count else 0, '', ''),
    ('', '', '', ''),
    ('PROMEDIOS (iPhone)', '', '', ''),
    ('Promedio x dia general', round(julio_iphone_total/total_july_days, 1) if total_july_days else 0, '', ''),
    ('Promedio x dia promo', round(julio_iphone_lmj/lmj_count, 1) if lmj_count else 0, '', ''),
]

for i, (l, v, p, _) in enumerate(labels_5, 3):
    ws5.cell(row=i, column=1, value=l)
    ws5.cell(row=i, column=2, value=v if v != '' else '')
    ws5.cell(row=i, column=3, value=p)
    if l.startswith('DIA') or l.startswith('VENTA') or l.startswith('PROM'):
        ws5.cell(row=i, column=1).font = Font(name='Arial', bold=True, size=11, color='1F4E79')
    else:
        style_cell(ws5, i, 1, left_a)
        style_cell(ws5, i, 2); style_cell(ws5, i, 3)

ws5.column_dimensions['A'].width = 35; ws5.column_dimensions['B'].width = 15
ws5.column_dimensions['C'].width = 12; ws5.column_dimensions['D'].width = 12

# Comparison charts
chart5a = BarChart(); chart5a.type = "col"; chart5a.title = "Todos los Equipos"; chart5a.style = 10
ws5.cell(row=30, column=1, value="Tipo"); ws5.cell(row=30, column=2, value="Ventas")
ws5.cell(row=31, column=1, value="Dias Promo"); ws5.cell(row=31, column=2, value=julio_all_lmj)
ws5.cell(row=32, column=1, value="Dias No Promo"); ws5.cell(row=32, column=2, value=julio_all_total-julio_all_lmj)
d5a = Reference(ws5, min_col=2, min_row=30, max_row=32)
c5a = Reference(ws5, min_col=1, min_row=31, max_row=32)
chart5a.add_data(d5a, titles_from_data=True); chart5a.set_categories(c5a)
ws5.add_chart(chart5a, "E3")

chart5b = BarChart(); chart5b.type = "col"; chart5b.title = "Solo iPhone"; chart5b.style = 10
ws5.cell(row=30, column=4, value="Tipo"); ws5.cell(row=30, column=5, value="Ventas")
ws5.cell(row=31, column=4, value="Dias Promo"); ws5.cell(row=31, column=5, value=julio_iphone_lmj)
ws5.cell(row=32, column=4, value="Dias No Promo"); ws5.cell(row=32, column=5, value=julio_iphone_total-julio_iphone_lmj)
d5b = Reference(ws5, min_col=5, min_row=30, max_row=32)
c5b = Reference(ws5, min_col=4, min_row=31, max_row=32)
chart5b.add_data(d5b, titles_from_data=True); chart5b.set_categories(c5b)
ws5.add_chart(chart5b, "E18")

# ===== SHEET 6: DETALLE =====
ws6 = out.create_sheet("Detalle Ventas")
ws6.cell(row=1, column=1, value="DETALLE DE VENTAS - PROMO 3 DIAS").font = title_font
ws6.merge_cells('A1:J1')
ws6.cell(row=2, column=1, value=f"Total: {len(all_rows)} registros | {sum(1 for r in all_rows if r['es_iphone'])} iPhones").font = sub_font
ws6.merge_cells('A2:J2')

h6 = ['#', 'Cliente', 'Modelo', 'Tipo', 'Supervisor', 'Asesor', 'Dia', 'Dia Semana', 'Semana', 'Status']
for c, h in enumerate(h6, 1):
    ws6.cell(row=4, column=c, value=h)
set_hdr(ws6, 4, len(h6))

for i, rec in enumerate(all_rows, 1):
    r = 4 + i
    ws6.cell(row=r, column=1, value=i)
    ws6.cell(row=r, column=2, value=rec['cliente'])
    ws6.cell(row=r, column=3, value=rec['modelo'])
    ws6.cell(row=r, column=4, value='iPhone' if rec['es_iphone'] else 'Otro')
    ws6.cell(row=r, column=5, value=rec['supervisor'])
    ws6.cell(row=r, column=6, value=rec['asesor'])
    ws6.cell(row=r, column=7, value=rec['dia'])
    ws6.cell(row=r, column=8, value=rec['dia_nombre'])
    ws6.cell(row=r, column=9, value=f"Semana {rec['semana']}")
    ws6.cell(row=r, column=10, value=rec['status'])
    fmt_border(ws6, r, len(h6))
    fill = iphone_fill if rec['es_iphone'] else (alt if i%2==0 else None)
    for c in range(1, len(h6)+1):
        style_cell(ws6, r, c, left_a if c in (2,3,5,6) else None)
        if fill: ws6.cell(row=r, column=c).fill = fill

ws6.column_dimensions['A'].width = 6; ws6.column_dimensions['B'].width = 28
ws6.column_dimensions['C'].width = 28; ws6.column_dimensions['D'].width = 8
ws6.column_dimensions['E'].width = 22; ws6.column_dimensions['F'].width = 22
ws6.column_dimensions['G'].width = 14; ws6.column_dimensions['H'].width = 14
ws6.column_dimensions['I'].width = 10; ws6.column_dimensions['J'].width = 28

outfile = '../Lixi/3diasporsemana/Ranking_3_Dias_x_Semana_Julio_2026.xlsx'
out.save(outfile)
print(f"Excel generado: {outfile}")
print(f"6 hojas: Tabla Dinamica, Ranking General, Ranking iPhone, Desglose Diario, Comparativa, Detalle Ventas")
