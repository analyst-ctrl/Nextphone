import openpyxl
from datetime import datetime
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

XLSX = r'C:\Users\Near\Documents\Chamba Panama\excels\Iphone (1).xlsx'
wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
ws = wb['Sharep']
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
col_idx = {h: i for i, h in enumerate(headers)}

def cv(v): return str(v).strip() if v is not None else ''
def parse_date(v):
    if not v: return None
    if isinstance(v, datetime): return v
    s = str(v).strip()
    for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d']:
        try: return datetime.strptime(s, fmt)
        except: pass
    return None

tabla_models = [
    'IPHONE 17 256GB (Black)', 'iPhone 17 256GB Black',
    'IPHONE 17E 256GB (Black)', 'iPhone 17e 256GB Black',
    'IPHONE 17 256GB (Blue)', 'Iphone 17e'
]

records = []
for row in ws.iter_rows(min_row=2, values_only=True):
    created_raw = row[col_idx.get('Created', 38)]
    d = parse_date(created_raw)
    if not d or d.year != 2026 or d.month != 7: continue
    modelo = cv(row[col_idx.get('Modelo del equipo', 21)])
    if 'IPHONE' not in modelo.upper(): continue
    sup = cv(row[col_idx.get('Supervisor', 37)])
    if sup not in ['Manuel García', 'Jonathan Ortega']: continue
    dia_crea = cv(row[col_idx.get('Dia Crea', 1)])
    records.append({'supervisor': sup, 'modelo': modelo, 'dia_crea': dia_crea})

wb.close()

# Target: MG=18, JO=13
# Try removing one day at a time
from itertools import combinations

all_dias = sorted(set(r['dia_crea'] for r in records))

print("Dias disponibles:", all_dias)
print(f"Total registros: {len(records)}")
print(f"MG: {sum(1 for r in records if r['supervisor']=='Manuel García')}")
print(f"JO: {sum(1 for r in records if r['supervisor']=='Jonathan Ortega')}")
print()

# Try excluding combinations of days to hit MG=18, JO=13
for n in range(1, 6):
    found = False
    for combo in combinations(all_dias, n):
        excluded = set(combo)
        mg = sum(1 for r in records if r['supervisor']=='Manuel García' and r['dia_crea'] not in excluded)
        jo = sum(1 for r in records if r['supervisor']=='Jonathan Ortega' and r['dia_crea'] not in excluded)
        if mg == 18 and jo == 13:
            print(f"Encontrado! Excluir {n} dia(s): {combo}")
            print(f"  MG={mg}, JO={jo}")
            found = True
    if found: break

if not found:
    print("No se encontro combinacion exacta excluyendo <=5 dias")
    print("\nProbando exclusion por rango de fechas...")
    # Try excluding days before a certain date
    for i, d in enumerate(all_dias):
        excluded = set(all_dias[:i])
        mg = sum(1 for r in records if r['supervisor']=='Manuel García' and r['dia_crea'] not in excluded)
        jo = sum(1 for r in records if r['supervisor']=='Jonathan Ortega' and r['dia_crea'] not in excluded)
        if mg == 18 and jo == 13:
            print(f"Excluir dias antes de {d}: MG={mg}, JO={jo}")
            break
        excluded2 = set(all_dias[i+1:])
        mg2 = sum(1 for r in records if r['supervisor']=='Manuel García' and r['dia_crea'] not in excluded2)
        jo2 = sum(1 for r in records if r['supervisor']=='Jonathan Ortega' and r['dia_crea'] not in excluded2)
        if mg2 == 18 and jo2 == 13:
            print(f"Excluir dias despues de {d}: MG={mg2}, JO={jo2}")
            break
