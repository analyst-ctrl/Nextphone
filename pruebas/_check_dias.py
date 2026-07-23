import openpyxl
from datetime import datetime
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

XLSX = r'C:\Users\Near\Documents\Chamba Panama\excels\Iphone (1).xlsx'

wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
ws = wb['Sharep']

headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
col_idx = {h: i for i, h in enumerate(headers)}

def cv(v):
    return str(v).strip() if v is not None else ''

def parse_date(v):
    if not v: return None
    if isinstance(v, datetime): return v
    s = str(v).strip()
    for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d']:
        try: return datetime.strptime(s, fmt)
        except: pass
    return None

# Collect ALL iPhone July data with dia_crea
all_records = []
for row in ws.iter_rows(min_row=2, values_only=True):
    created_raw = row[col_idx.get('Created', 38)]
    d = parse_date(created_raw)
    if not d or d.year != 2026 or d.month != 7:
        continue
    modelo = cv(row[col_idx.get('Modelo del equipo', 21)])
    if 'IPHONE' not in modelo.upper():
        continue
    
    sup = cv(row[col_idx.get('Supervisor', 37)])
    ase = cv(row[col_idx.get('Asesor', 36)])
    st = cv(row[col_idx.get('Status del Cliente', 34)])
    dia_crea = cv(row[col_idx.get('Dia Crea', 1)])
    
    all_records.append({
        'supervisor': sup, 'asesor': ase, 'modelo': modelo,
        'status': st, 'dia_crea': dia_crea, 'fecha': d.strftime('%Y-%m-%d')
    })

wb.close()

# Tabla expected counts (from manual reading)
# Manuel Garcia CREACION: 18
# Jonathan Ortega CREACION: 13
# Total: 31

# Our dashboard "Todos" shows 55
# Difference: 55 - 31 = 24... that's not 2

# Wait, let me re-check. The user said difference of 2.
# Let me count by supervisor and model matching the Tabla's models

# Tabla models (Manuel Garcia): IPHONE 17 256GB (Black), iPhone 17 256GB Black, IPHONE 17E 256GB (Black), iPhone 17e 256GB Black
# Tabla models (Jonathan Ortega): IPHONE 17 256GB (Black), IPHONE 17 256GB (Blue), Iphone 17e, IPHONE 17E 256GB (Black)

tabla_models_mg = ['IPHONE 17 256GB (Black)', 'iPhone 17 256GB Black', 'IPHONE 17E 256GB (Black)', 'iPhone 17e 256GB Black']
tabla_models_jo = ['IPHONE 17 256GB (Black)', 'IPHONE 17 256GB (Blue)', 'Iphone 17e', 'IPHONE 17E 256GB (Black)']

print("=== Registros iPhone 17 en julio (todos los dias) ===")
mg_all = [r for r in all_records if r['supervisor'] == 'Manuel García' and r['modelo'] in tabla_models_mg]
jo_all = [r for r in all_records if r['supervisor'] == 'Jonathan Ortega' and r['modelo'] in tabla_models_jo]

print(f"\nManuel Garcia - iPhone 17 (todos los dias): {len(mg_all)}")
print(f"Jonathan Ortega - iPhone 17 (todos los dias): {len(jo_all)}")
print(f"Total: {len(mg_all) + len(jo_all)}")

# By dia_crea
print("\n=== Manuel Garcia - por Dia Crea ===")
mg_dias = {}
for r in mg_all:
    d = r['dia_crea']
    mg_dias[d] = mg_dias.get(d, 0) + 1
for d in sorted(mg_dias.keys()):
    print(f"  {d}: {mg_dias[d]}")

print("\n=== Jonathan Ortega - por Dia Crea ===")
jo_dias = {}
for r in jo_all:
    d = r['dia_crea']
    jo_dias[d] = jo_dias.get(d, 0) + 1
for d in sorted(jo_dias.keys()):
    print(f"  {d}: {jo_dias[d]}")

# Tabla expected: MG=18, JO=13, Total=31
# Our total with all days: MG=?, JO=?

# Let me also check what dates appear
print("\n=== Todas las fechas de creacion (iPhone 17, julio) ===")
fechas = {}
for r in all_records:
    f = r['fecha']
    fechas[f] = fechas.get(f, 0) + 1
for f in sorted(fchas.keys()):
    print(f"  {f}: {fechas[f]}")
