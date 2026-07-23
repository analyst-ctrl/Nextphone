import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict, Counter
from datetime import datetime
import sys
import io

# Fix encoding for Windows
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

print("Leyendo Sharep.xlsx...")
wb = openpyxl.load_workbook('../Lixi/3diasporsemana/Sharep.xlsx', read_only=True, data_only=True)
ws = wb['owssvr (6)']

headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

# Map column indices
col_idx = {h: i for i, h in enumerate(headers)}

MODELO_IDX = col_idx['Modelo del equipo']
PROMO_IDX = col_idx.get('Lleva promoci\u00f3n')
SUPERVISOR_IDX = col_idx['Supervisor']
CREATED_IDX = col_idx['Created']
CLIENTE_IDX = col_idx['Nombre del cliente']
PLAN_IDX = col_idx['Plan Vendido']
STATUS_IDX = col_idx['Status del Cliente']
TIPO_VENTA_IDX = col_idx['Tipo de venta']
ASESOR_IDX = col_idx['Asesor']
PRECIO_IDX = col_idx['Precio del Plan']

# Data structures
month_day_sup = defaultdict(lambda: defaultdict(int))
sup_total = Counter()
month_sup = defaultdict(lambda: defaultdict(int))
iphone_records = []

print("Procesando 38,094 registros...")
count = 0
iphone_count = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    count += 1
    if count % 5000 == 0:
        print(f"  Procesados {count} registros...")
    
    modelo = row[MODELO_IDX]
    if modelo and 'IPHONE' in str(modelo).upper():
        iphone_count += 1
        
        created = row[CREATED_IDX] if CREATED_IDX < len(row) else None
        date_obj = None
        month_year = "Sin fecha"
        day_str = "Sin fecha"
        
        if created:
            try:
                if isinstance(created, datetime):
                    date_obj = created
                elif isinstance(created, str):
                    for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y %H:%M:%S', '%d/%m/%Y']:
                        try:
                            date_obj = datetime.strptime(created, fmt)
                            break
                        except:
                            pass
                
                if date_obj:
                    month_year = date_obj.strftime('%Y-%m')
                    day_str = date_obj.strftime('%Y-%m-%d')
            except:
                month_year = str(created)[:7] if created else "Sin fecha"
                day_str = str(created)[:10] if created else "Sin fecha"
        
        supervisor = row[SUPERVISOR_IDX] if SUPERVISOR_IDX < len(row) else "Sin supervisor"
        if not supervisor:
            supervisor = "Sin supervisor"
        
        promo = row[PROMO_IDX] if PROMO_IDX is not None and PROMO_IDX < len(row) else None
        cliente = row[CLIENTE_IDX] if CLIENTE_IDX < len(row) else ""
        plan = row[PLAN_IDX] if PLAN_IDX < len(row) else ""
        status = row[STATUS_IDX] if STATUS_IDX < len(row) else ""
        tipo_venta = row[TIPO_VENTA_IDX] if TIPO_VENTA_IDX < len(row) else ""
        asesor = row[ASESOR_IDX] if ASESOR_IDX < len(row) else ""
        
        month_day_sup[(month_year, day_str)][supervisor] += 1
        sup_total[supervisor] += 1
        month_sup[month_year][supervisor] += 1
        
        iphone_records.append({
            'cliente': cliente,
            'modelo': modelo,
            'supervisor': supervisor,
            'asesor': asesor,
            'mes': month_year,
            'dia': day_str,
            'promo': promo,
            'plan': plan,
            'status': status,
            'tipo_venta': tipo_venta
        })

print(f"\nProcesados {count} registros totales")
print(f"Encontrados {iphone_count} registros de iPhone")

wb.close()

# Create output workbook
print("\nGenerando archivo Ranking iPhone.xlsx...")

out_wb = openpyxl.Workbook()

# Styles
header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
title_font = Font(name='Arial', bold=True, size=14, color='1F4E79')
subtitle_font = Font(name='Arial', bold=True, size=11, color='333333')
data_font = Font(name='Arial', size=10)
rank1_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
rank2_fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
rank3_fill = PatternFill(start_color='CD7F32', end_color='CD7F32', fill_type='solid')
alt_fill = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')
thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0')
)

def style_header(ws, row_num, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def apply_border(ws, row_num, max_col):
    for col in range(1, max_col + 1):
        ws.cell(row=row_num, column=col).border = thin_border

# ============================================================
# Sheet 1: Ranking General de Supervisores
# ============================================================
ws1 = out_wb.active
ws1.title = "Ranking General"

ws1.cell(row=1, column=1, value="RANKING DE INCENTIVOS IPHONE").font = title_font
ws1.merge_cells('A1:E1')
ws1.cell(row=2, column=1, value=f"Total de iPhones vendidos: {iphone_count}").font = subtitle_font
ws1.merge_cells('A2:E2')

headers1 = ['Posicion', 'Supervisor', 'Total iPhones', '% del Total', 'Promedio por Dia']
for col, h in enumerate(headers1, 1):
    ws1.cell(row=4, column=col, value=h)
style_header(ws1, 4, len(headers1))

sorted_sups = sorted(sup_total.items(), key=lambda x: -x[1])
total = sum(sup_total.values())

all_days = set()
for (month, day) in month_day_sup.keys():
    all_days.add(day)
num_days = len(all_days) if all_days else 1

for i, (sup, count) in enumerate(sorted_sups, 1):
    row_num = 4 + i
    ws1.cell(row=row_num, column=1, value=i)
    ws1.cell(row=row_num, column=2, value=sup)
    ws1.cell(row=row_num, column=3, value=count)
    ws1.cell(row=row_num, column=4, value=round(count / total * 100, 1))
    ws1.cell(row=row_num, column=5, value=round(count / num_days, 1))
    
    apply_border(ws1, row_num, len(headers1))
    
    if i == 1:
        for col in range(1, len(headers1) + 1):
            ws1.cell(row=row_num, column=col).fill = rank1_fill
    elif i == 2:
        for col in range(1, len(headers1) + 1):
            ws1.cell(row=row_num, column=col).fill = rank2_fill
    elif i == 3:
        for col in range(1, len(headers1) + 1):
            ws1.cell(row=row_num, column=col).fill = rank3_fill
    elif i % 2 == 0:
        for col in range(1, len(headers1) + 1):
            ws1.cell(row=row_num, column=col).fill = alt_fill
    
    for col in range(1, len(headers1) + 1):
        ws1.cell(row=row_num, column=col).font = data_font
        ws1.cell(row=row_num, column=col).alignment = Alignment(horizontal='center' if col != 2 else 'left', vertical='center')

ws1.column_dimensions['A'].width = 10
ws1.column_dimensions['B'].width = 25
ws1.column_dimensions['C'].width = 15
ws1.column_dimensions['D'].width = 15
ws1.column_dimensions['E'].width = 18

# ============================================================
# Sheet 2: Por Mes
# ============================================================
ws2 = out_wb.create_sheet("Por Mes")

ws2.cell(row=1, column=1, value="RANKING IPHONE POR MES").font = title_font
ws2.merge_cells('A1:F1')

current_row = 3
sorted_months = sorted(month_sup.keys())

for month in sorted_months:
    ws2.cell(row=current_row, column=1, value=f"Mes: {month}").font = Font(name='Arial', bold=True, size=12, color='1F4E79')
    ws2.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    current_row += 1
    
    headers_m = ['Posicion', 'Supervisor', 'Total iPhones', '% del Mes']
    for col, h in enumerate(headers_m, 1):
        ws2.cell(row=current_row, column=col, value=h)
    style_header(ws2, current_row, len(headers_m))
    current_row += 1
    
    month_sups = sorted(month_sup[month].items(), key=lambda x: -x[1])
    month_total = sum(month_sup[month].values())
    
    for i, (sup, cnt) in enumerate(month_sups, 1):
        ws2.cell(row=current_row, column=1, value=i)
        ws2.cell(row=current_row, column=2, value=sup)
        ws2.cell(row=current_row, column=3, value=cnt)
        ws2.cell(row=current_row, column=4, value=round(cnt / month_total * 100, 1))
        
        apply_border(ws2, current_row, len(headers_m))
        
        for col in range(1, len(headers_m) + 1):
            ws2.cell(row=current_row, column=col).font = data_font
            ws2.cell(row=current_row, column=col).alignment = Alignment(horizontal='center' if col != 2 else 'left', vertical='center')
            if i % 2 == 0:
                ws2.cell(row=current_row, column=col).fill = alt_fill
        
        current_row += 1
    
    current_row += 1

ws2.column_dimensions['A'].width = 10
ws2.column_dimensions['B'].width = 25
ws2.column_dimensions['C'].width = 15
ws2.column_dimensions['D'].width = 15

# ============================================================
# Sheet 3: Por Dia
# ============================================================
ws3 = out_wb.create_sheet("Por Dia")

ws3.cell(row=1, column=1, value="RANKING IPHONE POR DIA").font = title_font
ws3.merge_cells('A1:F1')

current_row = 3
sorted_days = sorted(month_day_sup.keys(), key=lambda x: x[1])

for (month, day) in sorted_days:
    sup_data = month_day_sup[(month, day)]
    
    ws3.cell(row=current_row, column=1, value=f"Fecha: {day}").font = Font(name='Arial', bold=True, size=11, color='1F4E79')
    ws3.cell(row=current_row, column=3, value=f"Total: {sum(sup_data.values())}")
    ws3.cell(row=current_row, column=3).font = Font(name='Arial', bold=True, size=11, color='333333')
    ws3.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    current_row += 1
    
    headers_d = ['Posicion', 'Supervisor', 'Ventas iPhone', 'Acumulado %']
    for col, h in enumerate(headers_d, 1):
        ws3.cell(row=current_row, column=col, value=h)
    style_header(ws3, current_row, len(headers_d))
    current_row += 1
    
    day_sups = sorted(sup_data.items(), key=lambda x: -x[1])
    day_total = sum(sup_data.values())
    acum = 0
    
    for i, (sup, cnt) in enumerate(day_sups, 1):
        acum += cnt
        ws3.cell(row=current_row, column=1, value=i)
        ws3.cell(row=current_row, column=2, value=sup)
        ws3.cell(row=current_row, column=3, value=cnt)
        ws3.cell(row=current_row, column=4, value=round(acum / day_total * 100, 1))
        
        apply_border(ws3, current_row, len(headers_d))
        
        for col in range(1, len(headers_d) + 1):
            ws3.cell(row=current_row, column=col).font = data_font
            ws3.cell(row=current_row, column=col).alignment = Alignment(horizontal='center' if col != 2 else 'left', vertical='center')
            if i % 2 == 0:
                ws3.cell(row=current_row, column=col).fill = alt_fill
        
        current_row += 1
    
    current_row += 1

ws3.column_dimensions['A'].width = 10
ws3.column_dimensions['B'].width = 25
ws3.column_dimensions['C'].width = 15
ws3.column_dimensions['D'].width = 15

# ============================================================
# Sheet 4: Detalle de Ventas iPhone
# ============================================================
ws4 = out_wb.create_sheet("Detalle Ventas")

ws4.cell(row=1, column=1, value="DETALLE DE VENTAS IPHONE").font = title_font
ws4.merge_cells('A1:H1')
ws4.cell(row=2, column=1, value=f"Total: {len(iphone_records)} registros").font = subtitle_font
ws4.merge_cells('A2:H2')

headers_det = ['#', 'Cliente', 'Modelo iPhone', 'Supervisor', 'Asesor', 'Mes', 'Dia', 'Status']
for col, h in enumerate(headers_det, 1):
    ws4.cell(row=4, column=col, value=h)
style_header(ws4, 4, len(headers_det))

for i, rec in enumerate(iphone_records, 1):
    row_num = 4 + i
    ws4.cell(row=row_num, column=1, value=i)
    ws4.cell(row=row_num, column=2, value=rec['cliente'])
    ws4.cell(row=row_num, column=3, value=rec['modelo'])
    ws4.cell(row=row_num, column=4, value=rec['supervisor'])
    ws4.cell(row=row_num, column=5, value=rec['asesor'])
    ws4.cell(row=row_num, column=6, value=rec['mes'])
    ws4.cell(row=row_num, column=7, value=rec['dia'])
    ws4.cell(row=row_num, column=8, value=rec['status'])
    
    apply_border(ws4, row_num, len(headers_det))
    
    for col in range(1, len(headers_det) + 1):
        ws4.cell(row=row_num, column=col).font = data_font
        ws4.cell(row=row_num, column=col).alignment = Alignment(horizontal='center' if col != 2 and col != 3 else 'left', vertical='center')
        if i % 2 == 0:
            ws4.cell(row=row_num, column=col).fill = alt_fill

ws4.column_dimensions['A'].width = 6
ws4.column_dimensions['B'].width = 30
ws4.column_dimensions['C'].width = 30
ws4.column_dimensions['D'].width = 22
ws4.column_dimensions['E'].width = 22
ws4.column_dimensions['F'].width = 12
ws4.column_dimensions['G'].width = 14
ws4.column_dimensions['H'].width = 30

# ============================================================
# Sheet 5: Ranking por Supervisor (vista completa)
# ============================================================
ws5 = out_wb.create_sheet("Ranking Supervisor")

ws5.cell(row=1, column=1, value="VENTAS IPHONE POR SUPERVISOR (Detalle Mensual)").font = title_font
ws5.merge_cells('A1:H1')

current_row = 3

headers_sup = ['Supervisor', 'Total iPhones', 'Ranking'] + [m for m in sorted_months]
for col, h in enumerate(headers_sup, 1):
    ws5.cell(row=current_row, column=col, value=h)
style_header(ws5, current_row, len(headers_sup))
current_row += 1

for i, (sup, total) in enumerate(sorted_sups, 1):
    ws5.cell(row=current_row, column=1, value=sup)
    ws5.cell(row=current_row, column=2, value=total)
    ws5.cell(row=current_row, column=3, value=i)
    
    for j, m in enumerate(sorted_months, 4):
        val = month_sup[m].get(sup, 0)
        ws5.cell(row=current_row, column=j, value=val if val > 0 else 0)
    
    apply_border(ws5, current_row, len(headers_sup))
    
    for col in range(1, len(headers_sup) + 1):
        ws5.cell(row=current_row, column=col).font = data_font
        ws5.cell(row=current_row, column=col).alignment = Alignment(horizontal='center' if col != 1 else 'left', vertical='center')
        if i % 2 == 0:
            ws5.cell(row=current_row, column=col).fill = alt_fill
    
    current_row += 1

ws5.column_dimensions['A'].width = 25
ws5.column_dimensions['B'].width = 14
ws5.column_dimensions['C'].width = 10
for col_idx in range(4, len(headers_sup) + 1):
    ws5.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'AA'].width = 12

# Save
output_file = '../Lixi/3diasporsemana/Ranking iPhone.xlsx'
out_wb.save(output_file)
print(f"\nArchivo generado: {output_file}")
print("\nResumen:")
print("  - Sheet 1: Ranking General")
print("  - Sheet 2: Por Mes")
print("  - Sheet 3: Por Dia")
print("  - Sheet 4: Detalle Ventas")
print("  - Sheet 5: Ranking Supervisor")
