import openpyxl
from collections import defaultdict
from datetime import datetime
import sqlite3
import sys, io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

print("Paso 1: Extrayendo datos de Iphone_07_22_2026_sharep.xlsx...")
wb = openpyxl.load_workbook('../Lixi/3diasporsemana/modelos/Iphone_07_22_2026_sharep.xlsx', read_only=True, data_only=True)
ws = wb['Sharep']
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
col_idx = {h: i for i, h in enumerate(headers)}

def cv(v): return str(v).strip() if v is not None else ''

def parse_date(v):
    if not v: return None
    if isinstance(v, datetime): return v
    for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y %H:%M:%S', '%d/%m/%Y']:
        try: return datetime.strptime(str(v).strip(), fmt)
        except: pass
    return None

# Temporary sets for building dimension tables
supervisores = {}
asesores = {}
sup_asesor_pairs = set()
modelos = {}
planes = {}
statuses = {}
tipo_ventas = {}
provincias = {}
distritos = {}
corregimientos = {}
origenes = {}
metodos_pago = {}
abonos = {}
seguros = {}
equipos = {}
clientes = {}

ventas_data = []
id_counter = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    id_counter += 1
    # Supervisores
    sup_nom = cv(row[col_idx['Supervisor']])
    if sup_nom and sup_nom not in supervisores:
        supervisores[sup_nom] = len(supervisores) + 1

    # Asesores
    ase_nom = cv(row[col_idx['Asesor']])
    if ase_nom and ase_nom not in asesores:
        asesores[ase_nom] = len(asesores) + 1

    if sup_nom and ase_nom:
        sup_asesor_pairs.add((supervisores[sup_nom], asesores[ase_nom]))

    # Modelos
    mod = cv(row[col_idx['Modelo del equipo']])
    if mod and mod not in modelos:
        modelos[mod] = len(modelos) + 1

    # Planes
    plan = cv(row[col_idx['Plan Vendido']])
    if plan and plan not in planes:
        planes[plan] = len(planes) + 1

    # Status
    st = cv(row[col_idx['Status del Cliente']])
    if st and st not in statuses:
        statuses[st] = len(statuses) + 1

    # Tipo venta
    tv = cv(row[col_idx['Tipo de venta']])
    if tv and tv not in tipo_ventas:
        tipo_ventas[tv] = len(tipo_ventas) + 1

    # Origen
    ori = cv(row[col_idx['Origen del Cliente']])
    if ori and ori not in origenes:
        origenes[ori] = len(origenes) + 1

    # Metodo pago
    mp = cv(row[col_idx['Método de pago']])
    if mp and mp not in metodos_pago:
        metodos_pago[mp] = len(metodos_pago) + 1

    # Abono
    ab = cv(row[col_idx['Tipo de Abono']])
    if ab and ab not in abonos:
        abonos[ab] = len(abonos) + 1

    # Seguro
    sg = cv(row[col_idx['Lleva Seguro Movil']])
    if sg and sg not in seguros:
        seguros[sg] = len(seguros) + 1

    # Equipo lleva
    eq = cv(row[col_idx['Lleva Equipo']])
    if eq and eq not in equipos:
        equipos[eq] = len(equipos) + 1

    # Provincias / Distritos / Corregimientos
    prov = cv(row[col_idx['Provincia']])
    if prov and prov not in provincias:
        provincias[prov] = len(provincias) + 1
    dist = cv(row[col_idx['Distrito']])
    if dist and dist not in distritos:
        distritos[dist] = len(distritos) + 1
    corr = cv(row[col_idx['Corregimiento']])
    if corr and corr not in corregimientos:
        corregimientos[corr] = len(corregimientos) + 1

    # Clientes
    cli_nom = cv(row[col_idx['Nombre del cliente']])
    cli_id = cv(row[col_idx['Identificación (cedula o pasaporte)']])
    cli_key = (cli_nom, cli_id)
    if cli_nom and cli_key not in clientes:
        clientes[cli_key] = len(clientes) + 1

    # Venta
    fecha_creacion = parse_date(row[col_idx['Created']])
    fecha_entrega = parse_date(row[col_idx['Día y hora de la entrega']])
    precio = row[col_idx['Precio del Plan']]
    try: precio_val = float(precio) if precio else None
    except: precio_val = None

    promo = cv(row[col_idx['Lleva promoción']])
    como_lleva = cv(row[col_idx['Como lleva el equipo']])
    calle = cv(row[col_idx['Calle']])
    casa = cv(row[col_idx['Casa']])
    direccion_entrega = cv(row[col_idx['Dirección completa de la entrega']])
    direccion_cliente = cv(row[col_idx['Dirección completa del cliente']])
    imei = cv(row[col_idx['IMEI del equipo']])
    sim = cv(row[col_idx['Número de SIM CARD']])
    soc = cv(row[col_idx['Numero de SOC \\ VB2C']])
    contacto = cv(row[col_idx['# Contacto por WhatsApp']])
    email = cv(row[col_idx['Correo Electrónico']])
    num_cuenta = cv(row[col_idx['Número de Cuenta']])
    num_portar = cv(row[col_idx.get('Número a portar o renovar', col_idx.get('Número', -1))]) if col_idx.get('Número a portar o renovar', col_idx.get('Número')) else ''
    num_ref = cv(row[col_idx['Número de referencia']])
    barrio = cv(row[col_idx['Barrio o barriada']])

    ventas_data.append((
        id_counter,
        supervisores.get(sup_nom),
        asesores.get(ase_nom),
        modelos.get(mod),
        planes.get(plan),
        statuses.get(st),
        tipo_ventas.get(tv),
        origenes.get(ori),
        metodos_pago.get(mp),
        abonos.get(ab),
        seguros.get(sg),
        equipos.get(eq),
        provincias.get(prov),
        distritos.get(dist),
        corregimientos.get(corr),
        clientes.get(cli_key),
        precio_val,
        promo,
        como_lleva,
        calle, casa, barrio,
        direccion_entrega, direccion_cliente,
        imei, sim, soc, contacto, email,
        num_cuenta, num_portar, num_ref,
        fecha_creacion, fecha_entrega,
    ))

wb.close()
print(f"  {len(ventas_data)} ventas, {len(supervisores)} supervisores, {len(asesores)} asesores, {len(modelos)} modelos, {len(clientes)} clientes")

print("Paso 2: Creando base de datos SQLite...")
conn = sqlite3.connect('../Lixi/3diasporsemana/sharep.db')
c = conn.cursor()
c.execute('PRAGMA foreign_keys = ON')

c.execute('''
CREATE TABLE supervisores (
    id INTEGER PRIMARY KEY, nombre TEXT NOT NULL UNIQUE
)''')
c.execute('''
CREATE TABLE asesores (
    id INTEGER PRIMARY KEY, nombre TEXT NOT NULL UNIQUE
)''')
c.execute('''
CREATE TABLE supervisor_asesor (
    supervisor_id INTEGER NOT NULL,
    asesor_id INTEGER NOT NULL,
    PRIMARY KEY (supervisor_id, asesor_id),
    FOREIGN KEY (supervisor_id) REFERENCES supervisores(id),
    FOREIGN KEY (asesor_id) REFERENCES asesores(id)
)''')
c.execute('''
CREATE TABLE modelos (
    id INTEGER PRIMARY KEY, nombre TEXT NOT NULL UNIQUE
)''')
c.execute('''
CREATE TABLE planes (
    id INTEGER PRIMARY KEY, nombre TEXT NOT NULL UNIQUE
)''')
c.execute('''
CREATE TABLE status_cliente (
    id INTEGER PRIMARY KEY, nombre TEXT NOT NULL UNIQUE
)''')
c.execute('''
CREATE TABLE tipo_venta (
    id INTEGER PRIMARY KEY, nombre TEXT NOT NULL UNIQUE
)''')
c.execute('''
CREATE TABLE origen_cliente (
    id INTEGER PRIMARY KEY, nombre TEXT NOT NULL UNIQUE
)''')
c.execute('''
CREATE TABLE metodo_pago (
    id INTEGER PRIMARY KEY, nombre TEXT NOT NULL UNIQUE
)''')
c.execute('''
CREATE TABLE tipo_abono (
    id INTEGER PRIMARY KEY, nombre TEXT NOT NULL UNIQUE
)''')
c.execute('''
CREATE TABLE seguro_movil (
    id INTEGER PRIMARY KEY, nombre TEXT NOT NULL UNIQUE
)''')
c.execute('''
CREATE TABLE lleva_equipo (
    id INTEGER PRIMARY KEY, nombre TEXT NOT NULL UNIQUE
)''')
c.execute('''
CREATE TABLE provincias (
    id INTEGER PRIMARY KEY, nombre TEXT NOT NULL UNIQUE
)''')
c.execute('''
CREATE TABLE distritos (
    id INTEGER PRIMARY KEY, nombre TEXT NOT NULL UNIQUE
)''')
c.execute('''
CREATE TABLE corregimientos (
    id INTEGER PRIMARY KEY, nombre TEXT NOT NULL UNIQUE
)''')
c.execute('''
CREATE TABLE clientes (
    id INTEGER PRIMARY KEY,
    nombre TEXT NOT NULL,
    identificacion TEXT,
    UNIQUE(nombre, identificacion)
)''')
c.execute('''
CREATE TABLE ventas (
    id INTEGER PRIMARY KEY,
    supervisor_id INTEGER,
    asesor_id INTEGER,
    modelo_id INTEGER,
    plan_id INTEGER,
    status_id INTEGER,
    tipo_venta_id INTEGER,
    origen_id INTEGER,
    metodo_pago_id INTEGER,
    abono_id INTEGER,
    seguro_id INTEGER,
    equipo_id INTEGER,
    provincia_id INTEGER,
    distrito_id INTEGER,
    corregimiento_id INTEGER,
    cliente_id INTEGER,
    precio_plan REAL,
    lleva_promocion TEXT,
    como_lleva_equipo TEXT,
    calle TEXT, casa TEXT, barrio TEXT,
    direccion_entrega TEXT,
    direccion_cliente TEXT,
    imei TEXT, sim_card TEXT, soc TEXT,
    contacto TEXT, email TEXT,
    num_cuenta TEXT, num_portar TEXT, num_referencia TEXT,
    fecha_creacion TIMESTAMP,
    fecha_entrega TIMESTAMP,
    FOREIGN KEY (supervisor_id) REFERENCES supervisores(id),
    FOREIGN KEY (asesor_id) REFERENCES asesores(id),
    FOREIGN KEY (modelo_id) REFERENCES modelos(id),
    FOREIGN KEY (plan_id) REFERENCES planes(id),
    FOREIGN KEY (status_id) REFERENCES status_cliente(id),
    FOREIGN KEY (tipo_venta_id) REFERENCES tipo_venta(id),
    FOREIGN KEY (origen_id) REFERENCES origen_cliente(id),
    FOREIGN KEY (metodo_pago_id) REFERENCES metodo_pago(id),
    FOREIGN KEY (abono_id) REFERENCES tipo_abono(id),
    FOREIGN KEY (seguro_id) REFERENCES seguro_movil(id),
    FOREIGN KEY (equipo_id) REFERENCES lleva_equipo(id),
    FOREIGN KEY (provincia_id) REFERENCES provincias(id),
    FOREIGN KEY (distrito_id) REFERENCES distritos(id),
    FOREIGN KEY (corregimiento_id) REFERENCES corregimientos(id),
    FOREIGN KEY (cliente_id) REFERENCES clientes(id)
)''')

# Also create a view for easy querying
c.execute('''
CREATE VIEW v_ventas AS
SELECT
    v.id,
    cl.nombre AS cliente,
    cl.identificacion,
    s.nombre AS supervisor,
    a.nombre AS asesor,
    m.nombre AS modelo,
    p.nombre AS plan,
    v.precio_plan,
    st.nombre AS status,
    tv.nombre AS tipo_venta,
    oc.nombre AS origen,
    mp.nombre AS metodo_pago,
    ta.nombre AS tipo_abono,
    sm.nombre AS seguro,
    le.nombre AS lleva_equipo,
    pr.nombre AS provincia,
    d.nombre AS distrito,
    co.nombre AS corregimiento,
    v.lleva_promocion,
    v.fecha_creacion,
    v.fecha_entrega
FROM ventas v
LEFT JOIN clientes cl ON v.cliente_id = cl.id
LEFT JOIN supervisores s ON v.supervisor_id = s.id
LEFT JOIN asesores a ON v.asesor_id = a.id
LEFT JOIN modelos m ON v.modelo_id = m.id
LEFT JOIN planes p ON v.plan_id = p.id
LEFT JOIN status_cliente st ON v.status_id = st.id
LEFT JOIN tipo_venta tv ON v.tipo_venta_id = tv.id
LEFT JOIN origen_cliente oc ON v.origen_id = oc.id
LEFT JOIN metodo_pago mp ON v.metodo_pago_id = mp.id
LEFT JOIN tipo_abono ta ON v.abono_id = ta.id
LEFT JOIN seguro_movil sm ON v.seguro_id = sm.id
LEFT JOIN lleva_equipo le ON v.equipo_id = le.id
LEFT JOIN provincias pr ON v.provincia_id = pr.id
LEFT JOIN distritos d ON v.distrito_id = d.id
LEFT JOIN corregimientos co ON v.corregimiento_id = co.id
''')

print("Paso 3: Insertando datos...")

# Dimension tables
tables = [
    ('supervisores', supervisores),
    ('asesores', asesores),
    ('modelos', modelos),
    ('planes', planes),
    ('status_cliente', statuses),
    ('tipo_venta', tipo_ventas),
    ('origen_cliente', origenes),
    ('metodo_pago', metodos_pago),
    ('tipo_abono', abonos),
    ('seguro_movil', seguros),
    ('lleva_equipo', equipos),
    ('provincias', provincias),
    ('distritos', distritos),
    ('corregimientos', corregimientos),
]
for table, data in tables:
    c.executemany(f'INSERT OR IGNORE INTO {table} (id, nombre) VALUES (?, ?)',
                  [(v, k) for k, v in data.items()])

# Supervisor-Asesor relationships
c.executemany('INSERT OR IGNORE INTO supervisor_asesor (supervisor_id, asesor_id) VALUES (?, ?)',
              list(sup_asesor_pairs))

# Clientes
c.executemany('INSERT OR IGNORE INTO clientes (id, nombre, identificacion) VALUES (?, ?, ?)',
              [(v, k[0], k[1]) for k, v in clientes.items()])

# Ventas (in batches)
batch_size = 500
for i in range(0, len(ventas_data), batch_size):
    batch = ventas_data[i:i+batch_size]
    c.executemany('''
        INSERT INTO ventas (
            id, supervisor_id, asesor_id, modelo_id, plan_id,
            status_id, tipo_venta_id, origen_id, metodo_pago_id,
            abono_id, seguro_id, equipo_id, provincia_id,
            distrito_id, corregimiento_id, cliente_id,
            precio_plan, lleva_promocion, como_lleva_equipo,
            calle, casa, barrio, direccion_entrega, direccion_cliente,
            imei, sim_card, soc, contacto, email,
            num_cuenta, num_portar, num_referencia,
            fecha_creacion, fecha_entrega
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    ''', batch)

conn.commit()

# Verify
c.execute('SELECT COUNT(*) FROM ventas')
ventas_count = c.fetchone()[0]
c.execute('SELECT COUNT(*) FROM supervisores')
sup_count = c.fetchone()[0]
c.execute('SELECT COUNT(*) FROM asesores')
ase_count = c.fetchone()[0]
c.execute('SELECT COUNT(*) FROM supervisor_asesor')
rel_count = c.fetchone()[0]
c.execute('SELECT COUNT(*) FROM modelos')
mod_count = c.fetchone()[0]
c.execute('SELECT COUNT(*) FROM clientes')
cli_count = c.fetchone()[0]

conn.close()

print(f"\nBase de datos creada: sharep.db")
print(f"\nResumen:")
print(f"  supervisores:          {sup_count}")
print(f"  asesores:              {ase_count}")
print(f"  supervisor_asesor:     {rel_count} relaciones")
print(f"  modelos:               {mod_count}")
print(f"  clientes:              {cli_count}")
print(f"  ventas:                {ventas_count}")
print(f"  + 9 tablas dimensionales (planes, status, tipo_venta, etc)")
print(f"\nVista creada: v_ventas (join completo para consultas)")
