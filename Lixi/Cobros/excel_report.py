import sqlite3, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

db = r'C:\xampp\htdocs\Chamba Panama\Lixi\Cobros\cobros.db'
OUT = r'C:\xampp\htdocs\Chamba Panama\Lixi\Cobros\Reporte_Cobros.xlsx'
conn = sqlite3.connect(db)

wb = Workbook()

# ---- COLOR PALETTE ----
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
KPI_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
KPI_FONT = Font(name="Calibri", bold=True, size=14, color="1F4E79")
KPI_VAL_FONT = Font(name="Calibri", bold=True, size=20, color="1F4E79")
TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="1F4E79")
SECTION_FONT = Font(name="Calibri", bold=True, size=12, color="1F4E79")
THIN_BORDER = Border(
    left=Side(style='thin', color='B0B0B0'),
    right=Side(style='thin', color='B0B0B0'),
    top=Side(style='thin', color='B0B0B0'),
    bottom=Side(style='thin', color='B0B0B0'),
)
ALT_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

def style_header(ws, row, max_col):
    for c in range(1, max_col+1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

def style_data(ws, start_row, end_row, max_col):
    for r in range(start_row, end_row+1):
        for c in range(1, max_col+1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if (r - start_row) % 2 == 1:
                cell.fill = ALT_FILL

# ============================================================
# SHEET 1: DASHBOARD
# ============================================================
ws = wb.active
ws.title = "Dashboard"
ws.sheet_properties.tabColor = "1F4E79"

# Title
ws.merge_cells('A1:F1')
ws['A1'] = "REPORTE DE COBROS"
ws['A1'].font = Font(name="Calibri", bold=True, size=22, color="1F4E79")
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 40

# ---- KPI ROW ----
total_cuentas = conn.execute("SELECT COUNT(*) FROM list_51124").fetchone()[0]
al_dia = conn.execute("SELECT COUNT(*) FROM list_51124 WHERE estatus='Al dia'").fetchone()[0]
en_mora = conn.execute("SELECT COUNT(*) FROM list_51124 WHERE estatus IN ('Mora','Mora Real','Cero Pago')").fetchone()[0]
total_ventas = conn.execute("SELECT ROUND(SUM(CAST(REPLACE(precio,',','') AS REAL)),2) FROM list_51124 WHERE CAST(REPLACE(precio,',','') AS REAL) > 0").fetchone()[0]
total_llamadas = conn.execute("SELECT COUNT(*) FROM call_report").fetchone()[0]
clientes_gestion = conn.execute("SELECT COUNT(DISTINCT full_name) FROM call_report").fetchone()[0]
prom_dias_mora = conn.execute("SELECT ROUND(AVG(CAST(dias_atraso AS INTEGER)),0) FROM list_51124 WHERE CAST(dias_atraso AS INTEGER) > 0 AND estatus IN ('Mora','Mora Real','Cero Pago')").fetchone()[0]

kpis = [
    ("TOTAL CUENTAS", f"{total_cuentas:,}", "Clientes registrados"),
    ("AL DIA", f"{al_dia:,}", f"{al_dia/total_cuentas*100:.1f}% del total"),
    ("EN MORA", f"{en_mora:,}", f"{en_mora/total_cuentas*100:.1f}% del total"),
    ("VENTAS TOTALES", f"${total_ventas:,.0f}" if total_ventas else "$0", "Suma total financiada"),
    ("LLAMADAS", f"{total_llamadas:,}", f"{clientes_gestion} clientes gestionados"),
    ("DIAS PROM MORA", f"{prom_dias_mora:,.0f}", "Promedio de atraso"),
]

for i, (label, val, sub) in enumerate(kpis):
    col = 1 + i * 2
    if i > 2:
        col = 1 + (i-3) * 2
        row = 5
    else:
        row = 3

    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
    ws.cell(row=row, column=col, value=label).font = KPI_FONT
    ws.cell(row=row, column=col).fill = KPI_FILL
    ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=row, column=col).border = THIN_BORDER

    ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
    ws.cell(row=row+1, column=col, value=val).font = KPI_VAL_FONT
    ws.cell(row=row+1, column=col).fill = KPI_FILL
    ws.cell(row=row+1, column=col).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=row+1, column=col).border = THIN_BORDER

    ws.merge_cells(start_row=row+2, start_column=col, end_row=row+2, end_column=col+1)
    ws.cell(row=row+2, column=col, value=sub).font = Font(name="Calibri", italic=True, size=9, color="666666")
    ws.cell(row=row+2, column=col).fill = KPI_FILL
    ws.cell(row=row+2, column=col).alignment = Alignment(horizontal='center', vertical='center')

for r in range(3, 8):
    ws.row_dimensions[r].height = 28

# ---- CHART: Status Cuentas ----
status_data = conn.execute("SELECT estatus, COUNT(*) FROM list_51124 WHERE estatus != '' GROUP BY estatus ORDER BY 2 DESC").fetchall()
ws.cell(row=10, column=1, value="DISTRIBUCION DE ESTADO DE CUENTAS").font = SECTION_FONT
ws.cell(row=10, column=1).alignment = Alignment(horizontal='left')
ws.merge_cells('A10:B10')

ws.cell(row=11, column=1, value="Estado").font = Font(bold=True)
ws.cell(row=11, column=2, value="Cantidad").font = Font(bold=True)
style_header(ws, 11, 2)
for i, (s, c) in enumerate(status_data):
    ws.cell(row=12+i, column=1, value=s)
    ws.cell(row=12+i, column=2, value=c)
style_data(ws, 12, 12+len(status_data)-1, 2)

pie = PieChart()
pie.title = "Estado de Cuentas"
pie.style = 10
pie.width = 16
pie.height = 12
data = Reference(ws, min_col=2, min_row=11, max_row=11+len(status_data))
cats = Reference(ws, min_col=1, min_row=12, max_row=11+len(status_data))
pie.add_data(data, titles_from_data=True)
pie.set_categories(cats)
pie.dataLabels = DataLabelList()
pie.dataLabels.showPercent = True
pie.dataLabels.showCatName = True
ws.add_chart(pie, "D1")

# ---- CHART: Resultado Llamadas ----
call_data = conn.execute("SELECT status_name, COUNT(*) FROM call_report GROUP BY status_name ORDER BY 2 DESC LIMIT 8").fetchall()
row_start = 12 + len(status_data) + 2
ws.cell(row=row_start, column=1, value="TOP RESULTADOS DE LLAMADAS").font = SECTION_FONT
ws.cell(row=row_start, column=1).alignment = Alignment(horizontal='left')
ws.merge_cells(f'A{row_start}:B{row_start}')

ws.cell(row=row_start+1, column=1, value="Resultado").font = Font(bold=True)
ws.cell(row=row_start+1, column=2, value="Cantidad").font = Font(bold=True)
style_header(ws, row_start+1, 2)
for i, (s, c) in enumerate(call_data):
    ws.cell(row=row_start+2+i, column=1, value=s)
    ws.cell(row=row_start+2+i, column=2, value=c)
style_data(ws, row_start+2, row_start+2+len(call_data)-1, 2)

bar = BarChart()
bar.type = "col"
bar.title = "Resultado de Llamadas (Top 8)"
bar.style = 10
bar.width = 16
bar.height = 12
data = Reference(ws, min_col=2, min_row=row_start+1, max_row=row_start+1+len(call_data))
cats = Reference(ws, min_col=1, min_row=row_start+2, max_row=row_start+1+len(call_data))
bar.add_data(data, titles_from_data=True)
bar.set_categories(cats)
bar.series[0].graphicalProperties.solidFill = "1F4E79"
ws.add_chart(bar, f"D{row_start}")

# ---- Col widths ----
for c in range(1, 14):
    ws.column_dimensions[get_column_letter(c)].width = 18

# ============================================================
# SHEET 2: DETALLE CUENTAS
# ============================================================
ws2 = wb.create_sheet("Cuentas")
ws2.sheet_properties.tabColor = "2E75B6"

headers2 = ["Nombre", "Cedula", "Modelo", "Marca", "Precio", "Cuotas", "Pagadas", "Estado", "Dias Atraso", "Tienda", "Aliado"]
ws2.cell(row=1, column=1, value="DETALLE DE CUENTAS").font = TITLE_FONT
ws2.merge_cells('A1:K1')
ws2.row_dimensions[1].height = 30

for i, h in enumerate(headers2, 1):
    ws2.cell(row=2, column=i, value=h)
style_header(ws2, 2, len(headers2))

data2 = conn.execute("""
    SELECT first_name || ' ' || last_name, numero_identificacion, modelo_telefono, marca,
           CAST(precio AS REAL), CAST(monto_cuotas AS REAL), CAST(cuotas_pagadas AS INTEGER),
           estatus, CAST(dias_atraso AS INTEGER), tienda, aliado
    FROM list_51124
    WHERE estatus != '' AND CAST(dias_atraso AS INTEGER) >= 0
    ORDER BY CAST(dias_atraso AS INTEGER) DESC
    LIMIT 2000
""").fetchall()

for i, row_data in enumerate(data2):
    for j, val in enumerate(row_data):
        ws2.cell(row=3+i, column=j+1, value=val)

style_data(ws2, 3, 3+len(data2)-1, len(headers2))

for c in range(1, len(headers2)+1):
    ws2.column_dimensions[get_column_letter(c)].width = 22

# Auto filter
ws2.auto_filter.ref = f"A2:K{2+len(data2)}"

# ============================================================
# SHEET 3: TOP DEUDORES
# ============================================================
ws3 = wb.create_sheet("Top Deudores")
ws3.sheet_properties.tabColor = "C00000"

h3 = ["#", "Nombre", "Identificacion", "Modelo", "Marca", "Precio", "Cuotas x Pagar", "Dias Atraso", "Aliado"]
ws3.cell(row=1, column=1, value="TOP 50 DEUDORES - MAYOR ATRASO").font = TITLE_FONT
ws3.merge_cells('A1:I1')
ws3.row_dimensions[1].height = 30

for i, h in enumerate(h3, 1):
    ws3.cell(row=2, column=i, value=h)
style_header(ws3, 2, len(h3))

top50 = conn.execute("""
    SELECT first_name || ' ' || last_name, numero_identificacion, modelo_telefono, marca,
           CAST(precio AS REAL), CAST(monto_cuotas AS REAL) - CAST(cuotas_pagadas AS INTEGER),
           CAST(dias_atraso AS INTEGER), aliado
    FROM list_51124
    WHERE CAST(dias_atraso AS INTEGER) > 0 AND estatus IN ('Mora','Mora Real','Cero Pago')
    ORDER BY CAST(dias_atraso AS INTEGER) DESC
    LIMIT 50
""").fetchall()

for i, row_data in enumerate(top50):
    ws3.cell(row=3+i, column=1, value=i+1)
    for j, val in enumerate(row_data):
        ws3.cell(row=3+i, column=j+2, value=val)

style_data(ws3, 3, 3+len(top50)-1, len(h3))

for c in range(1, len(h3)+1):
    ws3.column_dimensions[get_column_letter(c)].width = 22

# Highlight red for top debtors
RED_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
for r in range(3, 3+10):
    for c in range(1, len(h3)+1):
        ws3.cell(row=r, column=c).fill = RED_FILL

# ============================================================
# SAVE
# ============================================================
wb.save(OUT)
conn.close()
print(f"Excel creado: {OUT}")
print(f"Tamano: {os.path.getsize(OUT)/1024:.1f} KB")
