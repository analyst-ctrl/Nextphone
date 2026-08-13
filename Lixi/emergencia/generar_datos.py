import openpyxl, sys, datetime, json

def parse_date(v):
    if v is None: return None
    if isinstance(v, datetime.datetime): return v.date()
    if isinstance(v, datetime.date): return v
    s = str(v).strip()
    if s.lower() in ('activo', ''): return None
    for fmt in ('%m/%d/%Y', '%d/%m/%Y', '%Y-%m-%d'):
        try: return datetime.datetime.strptime(s, fmt).date()
        except Exception: pass
    return None

def load(path, cols):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['HC']
    out = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for name, c in cols.items():
            row[name] = ws.cell(r, c).value
        if row.get('Nombre') is None: continue
        row['FIngreso'] = parse_date(row['Ingreso'])
        out.append(row)
    return out

AGO_PATH = r'HC Actualizado Agosto.xlsx'
JUL_PATH = r'..\3diasporsemana\Headcount 1\HC Actualizado Julio 2026.xlsx'

ago = load(AGO_PATH, {'Nombre': 6, 'Cargo': 10, 'Departamento': 11, 'Campana': 12,
                      'Supervisor': 14, 'Estado': 15, 'Ingreso': 16})
jul = load(JUL_PATH, {'Nombre': 6, 'Cargo': 9, 'Departamento': 10, 'Campana': 11,
                      'Supervisor': 13, 'Estado': 14, 'Ingreso': 15})

def slim(rows):
    out = []
    for r in rows:
        out.append({
            'n': str(r['Nombre']),
            'c': str(r['Campana']) if r['Campana'] is not None else 'Sin campa\u00f1a',
            'r': str(r['Cargo']) if r['Cargo'] is not None else 'Sin cargo',
            'e': str(r['Estado']) if r['Estado'] is not None else 'Activo',
            's': str(r['Supervisor']) if r['Supervisor'] is not None else 'Sin supervisor',
            'f': r['FIngreso'].isoformat() if r['FIngreso'] else None,
        })
    return out

raw = {
    'Julio 2026': slim(jul),
    'Agosto 2026': slim(ago),
}

with open('datos.json', 'w', encoding='utf-8') as f:
    json.dump(raw, f, ensure_ascii=False)

print('Julio:', len(raw['Julio 2026']), 'Agosto:', len(raw['Agosto 2026']))
