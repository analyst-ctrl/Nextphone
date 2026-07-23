import sqlite3, json, os
from collections import defaultdict
from datetime import date, datetime
import sys, io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

TODAY = date.today()
OUT_DIR = 'output'
os.makedirs(OUT_DIR, exist_ok=True)

PROMO_DAYS = {0: 'Todos', 1: 'Lunes', 3: 'Miércoles', 5: 'Viernes'}
WEEKDAYS = {1: 'Lunes', 3: 'Miércoles', 5: 'Viernes'}

print("Conectando a sharep.db...")
conn = sqlite3.connect('sharep.db')
conn.row_factory = sqlite3.Row
c = conn.cursor()

# Get brand for each model
def get_brand(name):
    u = name.upper()
    if 'IPHONE' in u: return 'iPhone'
    if 'HONOR' in u: return 'Honor'
    if 'SAMSUNG' in u: return 'Samsung'
    if 'XIAOMI' in u or 'REDMI' in u: return 'Xiaomi'
    if 'MOTOROLA' in u: return 'Motorola'
    if 'OPPO' in u: return 'Oppo'
    if 'HUAWEI' in u: return 'Huawei'
    if 'LG' in u: return 'LG'
    if 'NOKIA' in u: return 'Nokia'
    if 'TECNO' in u: return 'Tecno'
    return 'Otra'

# Load all ventas with promo in Julio LMJ
print("Cargando datos...")
c.execute('''
    SELECT v.id, v.fecha_creacion, v.lleva_promocion, v.precio_plan,
           s.nombre as supervisor, a.nombre as asesor,
           m.nombre as modelo, cl.nombre as cliente,
           p.nombre as plan, st.nombre as status
    FROM ventas v
    JOIN supervisores s ON v.supervisor_id = s.id
    LEFT JOIN asesores a ON v.asesor_id = a.id
    JOIN modelos m ON v.modelo_id = m.id
    LEFT JOIN planes p ON v.plan_id = p.id
    LEFT JOIN status_cliente st ON v.status_id = st.id
    LEFT JOIN clientes cl ON v.cliente_id = cl.id
    WHERE strftime('%Y-%m', v.fecha_creacion) = '2026-07'
      AND v.lleva_promocion = 'Si'
      AND strftime('%w', v.fecha_creacion) IN ('1','3','5')
    ORDER BY v.fecha_creacion
''')

registros = []
marcas_set = set()
supervisores_set = set()
for row in c.fetchall():
    fecha = row['fecha_creacion']
    if isinstance(fecha, str):
        fecha = datetime.strptime(fecha[:19], '%Y-%m-%d %H:%M:%S')
    wd = fecha.weekday()
    if wd == 6: wd = 0  # SQLite Sunday=0 -> Monday=1 -> Python Monday=0
    else: wd += 1
    dia_semana = WEEKDAYS.get(wd, '')
    semana = (fecha.day - 1) // 7 + 1
    marca = get_brand(row['modelo'] or '')
    marcas_set.add(marca)
    supervisores_set.add(row['supervisor'])
    registros.append({
        'id': row['id'],
        'cliente': row['cliente'] or '',
        'modelo': row['modelo'] or '',
        'marca': marca,
        'supervisor': row['supervisor'],
        'asesor': row['asesor'] or '',
        'dia': fecha.strftime('%Y-%m-%d'),
        'dia_nombre': dia_semana,
        'weekday': wd,
        'semana': semana,
        'plan': row['plan'] or '',
        'precio': row['precio_plan'],
        'status': row['status'] or '',
    })

conn.close()

marcas_ordenadas = sorted(marcas_set, key=lambda m: -sum(1 for r in registros if r['marca'] == m))

# Comparison data (all July, all phones)
conn2 = sqlite3.connect('sharep.db')
c2 = conn2.cursor()
c2.execute('''
    SELECT
        SUM(CASE WHEN strftime('%w', fecha_creacion) IN ('1','3','5') THEN 1 ELSE 0 END) as lmj,
        COUNT(*) as total
    FROM ventas
    WHERE strftime('%Y-%m', fecha_creacion) = '2026-07'
      AND lleva_promocion = 'Si'
''')
r = c2.fetchone()
julio_total = r[1]
julio_lmj = r[0]
julio_other = julio_total - julio_lmj
conn2.close()

july_days = [d for d in range(1, 32) if date(2026, 7, d) <= TODAY]
total_dias = len(july_days)
lmj_dias = sum(1 for d in july_days if date(2026, 7, d).weekday() in (0, 2, 4))
other_dias = total_dias - lmj_dias

semanas_list = sorted(set(r['semana'] for r in registros))
supervisores_list = sorted(supervisores_set)
marcas_list = marcas_ordenadas

print(f"  {len(registros)} registros, {len(marcas_list)} marcas, {len(supervisores_list)} supervisores")

# ===== GERAR HTML =====
print("Generando dashboard HTML...")
data_json = json.dumps({
    'registros': registros,
    'marcas': marcas_list,
    'semanas': semanas_list,
    'supervisores': supervisores_list,
    'julioTotal': julio_total,
    'julioLmj': julio_lmj,
    'julioOther': julio_other,
    'totalDias': total_dias,
    'lmjDias': lmj_dias,
    'otherDias': other_dias,
    'today': TODAY.strftime('%Y-%m-%d'),
}, ensure_ascii=False)

# Brand colors
brand_colors = {
    'iPhone': '#555555',
    'Samsung': '#1428A0',
    'Honor': '#D82030',
    'Xiaomi': '#FF6900',
    'Motorola': '#5B5B5B',
    'Oppo': '#1A6D32',
    'Huawei': '#CF0A2C',
    'LG': '#A50034',
    'Nokia': '#004680',
    'Tecno': '#00A651',
    'Otra': '#999999',
}
brand_colors_json = json.dumps(brand_colors)

html = '''<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Dashboard Promo - 3 Dias x Semana</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4"></script>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:'Segoe UI',Arial,sans-serif;background:#f0f2f5;color:#333;padding:20px}
.container{max-width:1400px;margin:0 auto}
h1{color:#1F4E79;font-size:22px;margin-bottom:3px}
.subtitle{color:#666;font-size:13px;margin-bottom:20px}
.card{background:#fff;border-radius:10px;box-shadow:0 2px 8px rgba(0,0,0,0.08);padding:20px;margin-bottom:20px}
.card h2{color:#1F4E79;font-size:17px;margin-bottom:12px;padding-bottom:6px;border-bottom:2px solid #1F4E79}
table{width:100%;border-collapse:collapse;font-size:13px}
th{background:#1F4E79;color:#fff;padding:8px 10px;text-align:center;font-weight:600;font-size:12px}
td{padding:6px 10px;border-bottom:1px solid #e0e0e0;text-align:center;font-size:13px}
td.left{text-align:left}
tr:nth-child(even){background:#f8f9fa}
.gold{background:#FFD700!important;font-weight:bold}
.silver{background:#C0C0C0!important}
.bronze{background:#CD7F32!important;color:#fff}
.grid2{display:grid;grid-template-columns:1fr 1fr;gap:20px}
.grid3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:20px}
.metric{text-align:center;padding:15px}
.metric .num{font-size:28px;font-weight:bold;color:#1F4E79}
.metric .label{font-size:12px;color:#666;margin-top:4px}
.metric .sub{font-size:11px;color:#999}
.day-section{margin-bottom:15px}
.day-header{background:#D4EDDA;padding:8px 15px;border-radius:8px;font-weight:bold;color:#155724;margin-bottom:8px;display:flex;justify-content:space-between;font-size:14px}
.chart-container{height:260px;margin-top:8px}
@media(max-width:768px){.grid2,.grid3{grid-template-columns:1fr}}
.detail-table{font-size:11px}
.detail-table th{font-size:11px;padding:5px 6px}
.detail-table td{padding:3px 6px}
.filters{display:flex;gap:12px;flex-wrap:wrap;margin-bottom:15px;align-items:center}
.filters label{font-weight:600;font-size:13px;color:#333}
.filters select{padding:5px 10px;border:2px solid #1F4E79;border-radius:6px;font-size:13px;background:#fff;cursor:pointer}
.filters select:focus{outline:none;box-shadow:0 0 0 2px rgba(31,78,121,0.3)}
.filters .radio-group{display:flex;gap:2px;align-items:center;background:#fff;border:2px solid #1F4E79;border-radius:6px;padding:2px}
.filters .radio-group label{padding:3px 10px;border-radius:4px;cursor:pointer;font-weight:400;font-size:12px;margin:0}
.filters .radio-group input{display:none}
.filters .radio-group input:checked+label{background:#1F4E79;color:#fff;font-weight:600}
.badge{padding:2px 7px;border-radius:10px;font-size:10px;font-weight:bold}
.badge-lun{background:#cce5ff;color:#004085}
.badge-mie{background:#d4edda;color:#155724}
.badge-vie{background:#fff3cd;color:#856404}
.marca-tag{display:inline-block;padding:1px 7px;border-radius:4px;font-size:10px;font-weight:bold;color:#fff;margin-right:3px}
</style>
</head>
<body>
<div class="container">
<h1>Dashboard Promo - 3 Dias por Semana (Lun/Mie/Vie)</h1>
<p class="subtitle" id="subtitleText">Julio 2026 | Corte: $TODAY$</p>

<div class="filters">
  <div>
    <label>Equipos:</label>
    <div class="radio-group">
      <input type="radio" name="filtroTipo" id="tipoTodos" value="todos" checked><label for="tipoTodos">Todos</label>
      <input type="radio" name="filtroTipo" id="tipoCel" value="celulares"><label for="tipoCel">Celulares</label>
    </div>
  </div>
  <div>
    <label for="filterMarca">Marca:</label>
    <select id="filterMarca"><option value="todas">Todas</option></select>
  </div>
  <div>
    <label for="filterSemana">Semana:</label>
    <select id="filterSemana"><option value="todas">Todas</option></select>
  </div>
  <div>
    <label for="filterSup">Supervisor:</label>
    <select id="filterSup"><option value="todos">Todos</option></select>
  </div>
  <div id="filterStats" style="font-size:13px;color:#666;margin-left:auto"></div>
</div>

<div class="grid3" id="metricsCards"></div>

<div class="grid2">
  <div class="card"><h2>Ranking Supervisores</h2><div id="rankingContainer"></div></div>
  <div class="card"><h2>Grafico Ventas</h2><div class="chart-container"><canvas id="chartRanking"></canvas></div></div>
</div>

<div class="grid2">
  <div class="card"><h2>Ventas por Marca</h2><div class="chart-container"><canvas id="chartMarca"></canvas></div></div>
  <div class="card"><h2>Distribucion Temporal</h2><div class="chart-container"><canvas id="chartTemporal"></canvas></div></div>
</div>

<div class="card"><h2>Desglose por Dia Promo</h2><div id="dailyContainer"></div></div>

<div class="card"><h2>Comparativa: Dias Promo vs Total Julio</h2>
  <div class="grid2">
    <div id="comparativaTable"></div>
    <div><div class="chart-container"><canvas id="chartComparativo"></canvas></div></div>
  </div>
</div>

<div class="card">
  <h2>Detalle de Ventas</h2>
  <div style="max-height:350px;overflow-y:auto"><table class="detail-table" id="detailTable"></table></div>
</div>

</div>

<script>
const DATA = $DATA$;
const BRAND_COLORS = $BRAND_COLORS$;
const DAY_NAMES = {1:'Lunes',3:'Miercoles',5:'Viernes'};
const WEEKDAYS = {'Lunes':1,'Miercoles':3,'Viernes':5};

DATA.marcas.forEach(m => {
  document.getElementById('filterMarca').innerHTML += '<option value="'+m+'">'+m+'</option>';
});
DATA.semanas.forEach(s => {
  document.getElementById('filterSemana').innerHTML += '<option value="'+s+'">Semana '+s+'</option>';
});
DATA.supervisores.forEach(s => {
  document.getElementById('filterSup').innerHTML += '<option value="'+s+'">'+s+'</option>';
});

function brandColor(m) { return BRAND_COLORS[m] || '#999'; }

const EXCLUDE_MODELOS = new Set(['N/A','NINGUNO','NO APLICA','NA','S/N','NO LLEVA EQUIPO','SIN EQUIPO','EQUIPO PROPIO']);

function filterData() {
  const soloCel = document.getElementById('tipoCel').checked;
  const marca = document.getElementById('filterMarca').value;
  const sem = document.getElementById('filterSemana').value;
  const sup = document.getElementById('filterSup').value;
  return DATA.registros.filter(r => {
    if (soloCel && EXCLUDE_MODELOS.has(r.modelo.toUpperCase())) return false;
    if (marca !== 'todas' && r.marca !== marca) return false;
    if (sem !== 'todas' && r.semana !== parseInt(sem)) return false;
    if (sup !== 'todos' && r.supervisor !== sup) return false;
    return true;
  });
}

function render() {
  const filtered = filterData();
  const cnt = filtered.length;
  const sups = [...new Set(filtered.map(r=>r.supervisor))].length;
  const dias = [...new Set(filtered.map(r=>r.dia))].length;

  document.getElementById('metricsCards').innerHTML =
    '<div class="card metric"><div class="num">'+cnt+'</div><div class="label">Equipos Vendidos</div><div class="sub">En '+dias+' dias promo</div></div>' +
    '<div class="card metric"><div class="num">'+sups+'</div><div class="label">Supervisores</div><div class="sub">Con ventas</div></div>' +
    '<div class="card metric"><div class="num">'+(dias?(cnt/dias).toFixed(1):'0')+'</div><div class="label">Promedio x Dia</div><div class="sub">'+dias+' dias</div></div>';
  document.getElementById('filterStats').textContent = 'Mostrando '+cnt+' registros';

  // Ranking
  const sc = {}; const sd = {};
  filtered.forEach(r => {
    sc[r.supervisor] = (sc[r.supervisor]||0)+1;
    sd[r.supervisor] = sd[r.supervisor]||new Set(); sd[r.supervisor].add(r.dia);
  });
  const sorted = Object.entries(sc).sort((a,b)=>b[1]-a[1]);
  const total = cnt||1;
  let rh = '<table><tr><th>#</th><th class="left">Supervisor</th><th>Ventas</th><th>%</th><th>Dias</th><th>Prom/Dia</th></tr>';
  sorted.forEach(([s,c],i)=>{
    const cl = i===0?'gold':i===1?'silver':i===2?'bronze':'';
    const d = sd[s].size;
    rh += '<tr class="'+cl+'"><td>'+(i+1)+'</td><td class="left">'+s+'</td><td>'+c+'</td><td>'+(c/total*100).toFixed(1)+'%</td><td>'+d+'</td><td>'+(c/d).toFixed(1)+'</td></tr>';
  });
  rh += '</table>';
  document.getElementById('rankingContainer').innerHTML = rh;

  // Chart ranking
  if (window.chRanking) window.chRanking.destroy();
  window.chRanking = new Chart(document.getElementById('chartRanking'), {
    type: 'bar',
    data: { labels: sorted.slice(0,10).map(s=>s[0]), datasets: [{ label: 'Ventas', data: sorted.slice(0,10).map(s=>s[1]), backgroundColor: ['#FFD700','#C0C0C0','#CD7F32','#1F4E79','#2E75B6','#5B9BD5','#9DC3E6','#E8F0FE','#D4EDDA','#FFF3CD'], borderRadius: 4 }] },
    options: { responsive: true, maintainAspectRatio: false, plugins: { legend: { display: false } }, scales: { y: { beginAtZero: true, ticks: { stepSize: 1 } } } }
  });

  // Brand chart
  const brandCnt = {};
  filtered.forEach(r => { brandCnt[r.marca] = (brandCnt[r.marca]||0) + 1; });
  const brandSorted = Object.entries(brandCnt).sort((a,b)=>b[1]-a[1]);
  if (window.chMarca) window.chMarca.destroy();
  window.chMarca = new Chart(document.getElementById('chartMarca'), {
    type: 'doughnut',
    data: {
      labels: brandSorted.map(s=>s[0]),
      datasets: [{ data: brandSorted.map(s=>s[1]), backgroundColor: brandSorted.map(s=>brandColor(s[0])), borderWidth: 2 }]
    },
    options: { responsive: true, maintainAspectRatio: false, plugins: { legend: { position: 'right', labels: { font: {size:11} } } } }
  });

  // Temporal chart (by day)
  const dayCnt = {};
  filtered.forEach(r => { dayCnt[r.dia] = (dayCnt[r.dia]||0) + 1; });
  const daysSorted = Object.keys(dayCnt).sort();
  if (window.chTemp) window.chTemp.destroy();
  window.chTemp = new Chart(document.getElementById('chartTemporal'), {
    type: 'line',
    data: {
      labels: daysSorted,
      datasets: [{ label: 'Ventas', data: daysSorted.map(d=>dayCnt[d]), borderColor: '#1F4E79', backgroundColor: 'rgba(31,78,121,0.1)', fill: true, tension: 0.3, pointRadius: 5, pointBackgroundColor: '#1F4E79' }]
    },
    options: { responsive: true, maintainAspectRatio: false, plugins: { legend: { display: false } }, scales: { y: { beginAtZero: true, ticks: { stepSize: 1 } } } }
  });

  // Daily breakdown
  const dg = {};
  filtered.forEach(r => {
    if (!dg[r.dia]) dg[r.dia] = {fecha:r.dia, nombre:r.dia_nombre, weekday:r.weekday, sups:{}};
    dg[r.dia].sups[r.supervisor] = (dg[r.dia].sups[r.supervisor]||0)+1;
  });
  let dh = '';
  Object.keys(dg).sort().forEach(dia=>{
    const g = dg[dia];
    const arr = Object.entries(g.sups).sort((a,b)=>b[1]-a[1]);
    const td = arr.reduce((s,x)=>s+x[1],0);
    dh += '<div class="day-section"><div class="day-header"><span>'+g.fecha+' - '+g.nombre+'</span><span>Total: '+td+'</span></div>' +
      '<table><tr><th>#</th><th class="left">Supervisor</th><th>Ventas</th><th>%</th></tr>';
    arr.forEach(([s,c],i)=>{ dh += '<tr><td>'+(i+1)+'</td><td class="left">'+s+'</td><td>'+c+'</td><td>'+(c/td*100).toFixed(1)+'%</td></tr>'; });
    dh += '</table></div>';
  });
  document.getElementById('dailyContainer').innerHTML = dh || '<p style="color:#999;padding:10px">Sin datos.</p>';

  // Comparativa
  document.getElementById('comparativaTable').innerHTML =
    '<table><tr><th class="left">Metrica</th><th>Valor</th><th>%</th></tr>' +
    '<tr><td class="left">Dias transcurridos Julio</td><td>'+DATA.totalDias+'</td><td>100%</td></tr>' +
    '<tr><td class="left">Dias Promo (L/M/J)</td><td>'+DATA.lmjDias+'</td><td>'+(DATA.lmjDias/DATA.totalDias*100).toFixed(1)+'%</td></tr>' +
    '<tr><td class="left">Dias No Promo</td><td>'+DATA.otherDias+'</td><td>'+(DATA.otherDias/DATA.totalDias*100).toFixed(1)+'%</td></tr>' +
    '<tr><td colspan="3" style="padding:2px"></td></tr>' +
    '<tr><td class="left">Ventas (Total Julio)</td><td>'+DATA.julioTotal+'</td><td>100%</td></tr>' +
    '<tr><td class="left">Ventas en Dias Promo</td><td>'+DATA.julioLmj+'</td><td>'+(DATA.julioTotal?(DATA.julioLmj/DATA.julioTotal*100).toFixed(1):'0')+'%</td></tr>' +
    '<tr><td class="left">Ventas en Dias No Promo</td><td>'+DATA.julioOther+'</td><td>'+(DATA.julioTotal?(DATA.julioOther/DATA.julioTotal*100).toFixed(1):'0')+'%</td></tr>' +
    '<tr><td colspan="3" style="padding:2px"></td></tr>' +
    '<tr><td class="left">Promedio x dia general</td><td>'+(DATA.totalDias?(DATA.julioTotal/DATA.totalDias).toFixed(1):'0')+'</td><td></td></tr>' +
    '<tr><td class="left">Promedio x dia promo</td><td>'+(DATA.lmjDias?(DATA.julioLmj/DATA.lmjDias).toFixed(1):'0')+'</td><td></td></tr></table>';

  if (window.chComp) window.chComp.destroy();
  window.chComp = new Chart(document.getElementById('chartComparativo'), {
    type: 'bar',
    data: { labels: ['Dias Promo', 'Dias No Promo'], datasets: [{ label: 'Ventas', data: [DATA.julioLmj, DATA.julioOther], backgroundColor: ['#28a745','#dc3545'], borderRadius: 4 }] },
    options: { responsive: true, maintainAspectRatio: false, plugins: { legend: { display: false } }, scales: { y: { beginAtZero: true } } }
  });

  // Detail
  let dth = '<tr><th>#</th><th class="left">Cliente</th><th class="left">Modelo</th><th>Marca</th><th class="left">Supervisor</th><th class="left">Asesor</th><th>Dia</th></tr>';
  filtered.forEach((r,i)=>{
    dth += '<tr><td>'+(i+1)+'</td><td class="left">'+r.cliente+'</td><td class="left">'+r.modelo+'</td><td><span class="marca-tag" style="background:'+brandColor(r.marca)+'">'+r.marca+'</span></td><td class="left">'+r.supervisor+'</td><td class="left">'+r.asesor+'</td><td>'+r.dia+'<br><small>'+r.dia_nombre+'</small></td></tr>';
  });
  document.getElementById('detailTable').innerHTML = dth;
  document.getElementById('subtitleText').textContent = 'Julio 2026 | Corte: '+DATA.today+' | '+DATA.marcas.length+' marcas';
}

document.querySelectorAll('input[name="filtroTipo"]').forEach(el => el.addEventListener('change', render));
['filterMarca','filterSemana','filterSup'].forEach(id => document.getElementById(id).addEventListener('change', render));
render();
</script>
</body>
</html>'''

html = html.replace('$DATA$', data_json).replace('$BRAND_COLORS$', brand_colors_json).replace('$TODAY$', TODAY.strftime('%Y-%m-%d'))

html_path = os.path.join(OUT_DIR, 'dashboard_3diasporsemana.html')
with open(html_path, 'w', encoding='utf-8') as f:
    f.write(html)
print(f"  HTML: {html_path}")

# ===== GENERATE XLSX =====
print("Generando Excel en output/...")
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

data_xlsx = registros

wb = openpyxl.Workbook()

hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
hdr_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
title_font = Font(name='Arial', bold=True, size=14, color='1F4E79')
sub_font = Font(name='Arial', bold=True, size=11, color='333333')
data_font = Font(name='Arial', size=10)
bold_f = Font(name='Arial', bold=True, size=10)
gold = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
silver = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
bronze = PatternFill(start_color='CD7F32', end_color='CD7F32', fill_type='solid')
alt = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')
border = Border(left=Side(style='thin',color='D0D0D0'),right=Side(style='thin',color='D0D0D0'),top=Side(style='thin',color='D0D0D0'),bottom=Side(style='thin',color='D0D0D0'))
center = Alignment(horizontal='center',vertical='center')
left_a = Alignment(horizontal='left',vertical='center')
wrap_a = Alignment(horizontal='center',vertical='center',wrap_text=True)

def sh(ws, r, cm):
    for c in range(1, cm+1):
        cl = ws.cell(row=r, column=c); cl.font = hdr_font; cl.fill = hdr_fill; cl.alignment = wrap_a; cl.border = border

def sb(ws, r, cm):
    for c in range(1, cm+1): ws.cell(row=r, column=c).border = border

def sc(ws, r, c, a=None, fnt=None):
    cl = ws.cell(row=r, column=c); cl.font = fnt or data_font; cl.border = border; cl.alignment = a or center

# Sheet 1: Ranking
ws1 = wb.active
ws1.title = 'Ranking'
ws1.cell(row=1, column=1, value='RANKING SUPERVISORES - PROMO 3 DIAS').font = title_font
ws1.merge_cells('A1:G1')
ws1.cell(row=2, column=1, value=f'{len(data_xlsx)} ventas | {TODAY}').font = sub_font
ws1.merge_cells('A2:G2')

sup_tot = defaultdict(int)
sup_dias = defaultdict(set)
for r in data_xlsx:
    sup_tot[r['supervisor']] += 1
    sup_dias[r['supervisor']].add(r['dia'])

h1 = ['#','Supervisor','Ventas','%','Dias','Prom/Dia','Marca Principal']
for c, h in enumerate(h1, 1): ws1.cell(row=4, column=c, value=h)
sh(ws1, 4, len(h1))

tot_v = len(data_xlsx)
sorted_s = sorted(sup_tot.items(), key=lambda x: -x[1])
for i, (s, c) in enumerate(sorted_s, 1):
    r = 4 + i
    d = len(sup_dias[s])
    # Main brand
    brands_sup = defaultdict(int)
    for rec in data_xlsx:
        if rec['supervisor'] == s: brands_sup[rec['marca']] += 1
    main_brand = max(brands_sup, key=brands_sup.get) if brands_sup else ''
    ws1.cell(row=r, column=1, value=i)
    ws1.cell(row=r, column=2, value=s); sc(ws1, r, 2, left_a)
    ws1.cell(row=r, column=3, value=c); sc(ws1, r, 3, fnt=bold_f)
    ws1.cell(row=r, column=4, value=round(c/tot_v*100,1) if tot_v else 0); sc(ws1, r, 4)
    ws1.cell(row=r, column=5, value=d); sc(ws1, r, 5)
    ws1.cell(row=r, column=6, value=round(c/d,1) if d else 0); sc(ws1, r, 6)
    ws1.cell(row=r, column=7, value=main_brand); sc(ws1, r, 7, left_a)
    sb(ws1, r, len(h1))
    fill = gold if i==1 else silver if i==2 else bronze if i==3 else (alt if i%2==0 else None)
    if fill:
        for c in range(1, len(h1)+1): ws1.cell(row=r, column=c).fill = fill

ws1.column_dimensions['A'].width = 6; ws1.column_dimensions['B'].width = 22
ws1.column_dimensions['C'].width = 8; ws1.column_dimensions['D'].width = 8
ws1.column_dimensions['E'].width = 8; ws1.column_dimensions['F'].width = 10
ws1.column_dimensions['G'].width = 18

# Sheet 2: Por Marca
ws2 = wb.create_sheet('Por Marca')
ws2.cell(row=1, column=1, value='VENTAS POR MARCA').font = title_font
ws2.merge_cells('A1:D1')

marca_tot = defaultdict(int)
for r in data_xlsx: marca_tot[r['marca']] += 1
h2 = ['Marca','Ventas','%','Modelos Distintos']
for c, h in enumerate(h2, 1): ws2.cell(row=3, column=c, value=h)
sh(ws2, 3, len(h2))
sorted_m = sorted(marca_tot.items(), key=lambda x: -x[1])
for i, (m, c) in enumerate(sorted_m, 1):
    r = 3 + i
    mods = len(set(rec['modelo'] for rec in data_xlsx if rec['marca'] == m))
    ws2.cell(row=r, column=1, value=m); sc(ws2, r, 1, left_a)
    ws2.cell(row=r, column=2, value=c); sc(ws2, r, 2)
    ws2.cell(row=r, column=3, value=round(c/tot_v*100,1)); sc(ws2, r, 3)
    ws2.cell(row=r, column=4, value=mods); sc(ws2, r, 4)
    sb(ws2, r, len(h2))
ws2.column_dimensions['A'].width = 20; ws2.column_dimensions['B'].width = 10
ws2.column_dimensions['C'].width = 8; ws2.column_dimensions['D'].width = 18

# Sheet 3: Detalle
ws3 = wb.create_sheet('Detalle')
ws3.cell(row=1, column=1, value=f'DETALLE DE VENTAS ({len(data_xlsx)} registros)').font = title_font
ws3.merge_cells('A1:H1')
h3 = ['#','Cliente','Modelo','Marca','Supervisor','Asesor','Dia','Status']
for c, h in enumerate(h3, 1): ws3.cell(row=3, column=c, value=h)
sh(ws3, 3, len(h3))
for i, r in enumerate(data_xlsx, 1):
    row = 3 + i
    ws3.cell(row=row, column=1, value=i); sc(ws3, row, 1)
    ws3.cell(row=row, column=2, value=r['cliente']); sc(ws3, row, 2, left_a)
    ws3.cell(row=row, column=3, value=r['modelo']); sc(ws3, row, 3, left_a)
    ws3.cell(row=row, column=4, value=r['marca']); sc(ws3, row, 4)
    ws3.cell(row=row, column=5, value=r['supervisor']); sc(ws3, row, 5, left_a)
    ws3.cell(row=row, column=6, value=r['asesor']); sc(ws3, row, 6, left_a)
    ws3.cell(row=row, column=7, value=r['dia']); sc(ws3, row, 7)
    ws3.cell(row=row, column=8, value=r['status']); sc(ws3, row, 8, left_a)
    sb(ws3, row, len(h3))
    if i % 2 == 0:
        for c in range(1, len(h3)+1): ws3.cell(row=row, column=c).fill = alt
ws3.column_dimensions['A'].width = 6; ws3.column_dimensions['B'].width = 25; ws3.column_dimensions['C'].width = 30
ws3.column_dimensions['D'].width = 12; ws3.column_dimensions['E'].width = 20; ws3.column_dimensions['F'].width = 20
ws3.column_dimensions['G'].width = 14; ws3.column_dimensions['H'].width = 25

xlsx_path = os.path.join(OUT_DIR, 'Ranking_3_Dias_x_Semana_Julio_2026.xlsx')
wb.save(xlsx_path)
print(f"  XLSX: {xlsx_path}")

print("\nListo! Archivos en output/:")
print(f"  dashboard_3diasporsemana.html")
print(f"  Ranking_3_Dias_x_Semana_Julio_2026.xlsx")
