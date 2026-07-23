import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict, Counter
from datetime import datetime
import sys, io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

print("Leyendo Sharep.xlsx...")
wb = openpyxl.load_workbook('Sharep.xlsx', read_only=True, data_only=True)
ws = wb['owssvr (6)']

headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
col_idx = {h: i for i, h in enumerate(headers)}

cols = {
    'modelo': col_idx['Modelo del equipo'],
    'supervisor': col_idx['Supervisor'],
    'created': col_idx['Created'],
    'cliente': col_idx['Nombre del cliente'],
    'asesor': col_idx['Asesor'],
    'status': col_idx['Status del Cliente']
}

# Data: { (mes, dia): { supervisor: count } }
dia_sup = defaultdict(lambda: defaultdict(int))
sup_total = Counter()
records = []

print("Procesando 38,094 registros...")
count = 0
iphone_count = 0
julio_count = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    count += 1
    if count % 5000 == 0:
        print(f"  Procesados {count} registros...")

    modelo = row[cols['modelo']]
    if not modelo or 'IPHONE' not in str(modelo).upper():
        continue

    iphone_count += 1
    created = row[cols['created']] if cols['created'] < len(row) else None

    # Parse date
    date_obj = None
    if created:
        try:
            if isinstance(created, datetime):
                date_obj = created
            elif isinstance(created, str):
                for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y %H:%M:%S', '%d/%m/%Y']:
                    try:
                        date_obj = datetime.strptime(created, fmt)
                        break
                    except:
                        pass
        except:
            pass

    # Only July 2026
    if not date_obj or date_obj.year != 2026 or date_obj.month != 7:
        continue

    julio_count += 1
    mes = date_obj.strftime('%Y-%m')
    dia = date_obj.strftime('%Y-%m-%d')
    supervisor = row[cols['supervisor']] if cols['supervisor'] < len(row) else "Sin supervisor"
    if not supervisor:
        supervisor = "Sin supervisor"

    dia_sup[(mes, dia)][supervisor] += 1
    sup_total[supervisor] += 1

    records.append({
        'cliente': row[cols['cliente']],
        'modelo': modelo,
        'supervisor': supervisor,
        'asesor': row[cols['asesor']],
        'dia': dia,
        'status': row[cols['status']]
    })

wb.close()

print(f"\nTotal registros: {count}")
print(f"iPhone total: {iphone_count}")
print(f"iPhone Julio 2026: {julio_count}")

# ========== CREATE XLSX ==========
print("\nGenerando Ranking iPhone - Julio 2026.xlsx...")
out_wb = openpyxl.Workbook()

# Styles
hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
hdr_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
title_font = Font(name='Arial', bold=True, size=14, color='1F4E79')
sub_font = Font(name='Arial', bold=True, size=11, color='333333')
data_font = Font(name='Arial', size=10)
gold = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
silver = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
bronze = PatternFill(start_color='CD7F32', end_color='CD7F32', fill_type='solid')
alt = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')
today = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0')
)

def set_hdr(ws, r, c_max):
    for c in range(1, c_max+1):
        cell = ws.cell(row=r, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

def set_border(ws, r, c_max):
    for c in range(1, c_max+1):
        ws.cell(row=r, column=c).border = border

# ===== SHEET 1: Ranking General Julio =====
ws1 = out_wb.active
ws1.title = "Ranking Julio"

ws1.cell(row=1, column=1, value="RANKING IPHONE - JULIO 2026").font = title_font
ws1.merge_cells('A1:E1')
ws1.cell(row=2, column=1, value=f"Total iPhones vendidos en Julio: {julio_count}").font = sub_font
ws1.merge_cells('A2:E2')

h1 = ['Posicion', 'Supervisor', 'Total Julio', '% del Mes', 'Dias con ventas']
for c, h in enumerate(h1, 1):
    ws1.cell(row=4, column=c, value=h)
set_hdr(ws1, 4, len(h1))

sorted_sups = sorted(sup_total.items(), key=lambda x: -x[1])
total_julio = sum(sup_total.values())
unique_days = len(set(d for (m, d) in dia_sup.keys()))

for i, (sup, cnt) in enumerate(sorted_sups, 1):
    r = 4 + i
    # Count days this supervisor sold
    sup_days = sum(1 for (m, d), sups in dia_sup.items() if sup in sups)
    ws1.cell(row=r, column=1, value=i)
    ws1.cell(row=r, column=2, value=sup)
    ws1.cell(row=r, column=3, value=cnt)
    ws1.cell(row=r, column=4, value=round(cnt / total_julio * 100, 1))
    ws1.cell(row=r, column=5, value=sup_days)
    set_border(ws1, r, len(h1))

    fill = None
    if i == 1: fill = gold
    elif i == 2: fill = silver
    elif i == 3: fill = bronze
    elif i % 2 == 0: fill = alt
    if fill:
        for c in range(1, len(h1)+1):
            ws1.cell(row=r, column=c).fill = fill

    for c in range(1, len(h1)+1):
        ws1.cell(row=r, column=c).font = data_font
        ws1.cell(row=r, column=c).alignment = Alignment(horizontal='center' if c != 2 else 'left', vertical='center')

ws1.column_dimensions['A'].width = 10
ws1.column_dimensions['B'].width = 25
ws1.column_dimensions['C'].width = 14
ws1.column_dimensions['D'].width = 14
ws1.column_dimensions['E'].width = 18

# ===== SHEET 2: Dia a Dia =====
ws2 = out_wb.create_sheet("Dia a Dia")

ws2.cell(row=1, column=1, value="VENTAS IPHONE JULIO 2026 - DIA POR DIA").font = title_font
ws2.merge_cells('A1:E1')

r = 3
for (mes, dia) in sorted(dia_sup.keys(), key=lambda x: x[1]):
    sups = sorted(dia_sup[(mes, dia)].items(), key=lambda x: -x[1])
    total_dia = sum(dia_sup[(mes, dia)].values())

    # Check if today
    is_today = dia == '2026-07-22'
    day_fill_today = today if is_today else None

    ws2.cell(row=r, column=1, value=f"{dia}").font = Font(name='Arial', bold=True, size=12, color='1F4E79')
    ws2.cell(row=r, column=3, value=f"Total: {total_dia} iPhones")
    ws2.cell(row=r, column=3).font = Font(name='Arial', bold=True, size=11, color='333333')
    if day_fill_today:
        for c in range(1, 4):
            ws2.cell(row=r, column=c).fill = day_fill_today
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1

    h2 = ['#', 'Supervisor', 'Ventas', '% del Dia']
    for c, h in enumerate(h2, 1):
        ws2.cell(row=r, column=c, value=h)
    set_hdr(ws2, r, len(h2))
    r += 1

    for i, (sup, cnt) in enumerate(sups, 1):
        ws2.cell(row=r, column=1, value=i)
        ws2.cell(row=r, column=2, value=sup)
        ws2.cell(row=r, column=3, value=cnt)
        ws2.cell(row=r, column=4, value=round(cnt / total_dia * 100, 1))
        set_border(ws2, r, len(h2))

        if day_fill_today:
            for c in range(1, len(h2)+1):
                ws2.cell(row=r, column=c).fill = day_fill_today

        for c in range(1, len(h2)+1):
            ws2.cell(row=r, column=c).font = data_font
            ws2.cell(row=r, column=c).alignment = Alignment(horizontal='center' if c != 2 else 'left', vertical='center')
        r += 1
    r += 1

ws2.column_dimensions['A'].width = 8
ws2.column_dimensions['B'].width = 25
ws2.column_dimensions['C'].width = 10
ws2.column_dimensions['D'].width = 12

# ===== SHEET 3: Detalle Ventas Julio =====
ws3 = out_wb.create_sheet("Detalle Ventas")

ws3.cell(row=1, column=1, value="DETALLE VENTAS IPHONE - JULIO 2026").font = title_font
ws3.merge_cells('A1:G1')
ws3.cell(row=2, column=1, value=f"Total: {len(records)} registros").font = sub_font
ws3.merge_cells('A2:G2')

h3 = ['#', 'Cliente', 'Modelo iPhone', 'Supervisor', 'Asesor', 'Dia', 'Status']
for c, h in enumerate(h3, 1):
    ws3.cell(row=4, column=c, value=h)
set_hdr(ws3, 4, len(h3))

for i, rec in enumerate(records, 1):
    r = 4 + i
    ws3.cell(row=r, column=1, value=i)
    ws3.cell(row=r, column=2, value=rec['cliente'])
    ws3.cell(row=r, column=3, value=rec['modelo'])
    ws3.cell(row=r, column=4, value=rec['supervisor'])
    ws3.cell(row=r, column=5, value=rec['asesor'])
    ws3.cell(row=r, column=6, value=rec['dia'])
    ws3.cell(row=r, column=7, value=rec['status'])
    set_border(ws3, r, len(h3))
    for c in range(1, len(h3)+1):
        ws3.cell(row=r, column=c).font = data_font
        ws3.cell(row=r, column=c).alignment = Alignment(horizontal='center' if c != 2 and c != 3 else 'left', vertical='center')
        if i % 2 == 0:
            ws3.cell(row=r, column=c).fill = alt

ws3.column_dimensions['A'].width = 6
ws3.column_dimensions['B'].width = 30
ws3.column_dimensions['C'].width = 30
ws3.column_dimensions['D'].width = 22
ws3.column_dimensions['E'].width = 22
ws3.column_dimensions['F'].width = 14
ws3.column_dimensions['G'].width = 30

outfile = 'Ranking iPhone - Julio 2026.xlsx'
out_wb.save(outfile)
print(f"\nArchivo generado: {outfile}")
print("\nHojas:")
print("  1. Ranking Julio - Ranking general del mes")
print("  2. Dia a Dia - Desglose diario con posiciones")
print("  3. Detalle Ventas - Todos los registros individuales")
