import openpyxl, shutil, zipfile, os, re, io, sys
from datetime import datetime
from copy import deepcopy

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

SRC = r'excels\Iphone (1).xlsx'
TGT = r'excels\hoy.xlsx'
TMP = r'excels\hoy_tmp.xlsx'

# ===== STEP 1: Add 4 columns to hoy.xlsx =====
print('=== PASO 1: Agregando columnas a Sharep ===')
wb = openpyxl.load_workbook(TGT)
ws = wb['Sharep']

MES_MAP = {1:'ene',2:'feb',3:'mar',4:'abr',5:'may',6:'jun',7:'jul',8:'ago',9:'sep',10:'oct',11:'nov',12:'dic'}

def format_mes(dt):
    if dt is None:
        return None
    if isinstance(dt, str):
        try:
            dt = datetime.strptime(dt[:19], '%Y-%m-%d %H:%M:%S')
        except:
            return None
    return f'{MES_MAP[dt.month]}/{str(dt.year)[-2:]}'

def format_dia(dt):
    if dt is None:
        return None
    if isinstance(dt, str):
        try:
            dt = datetime.strptime(dt[:19], '%Y-%m-%d %H:%M:%S')
        except:
            return None
    return f'{dt.day:02d}/{MES_MAP[dt.month]}/{str(dt.year)[-2:]}'

# Get column indices for Created and Modified
# Current headers: col 36=Created, col 38=Modified (0-based: 35, 37)
# After insert: need to shift everything right by 4
max_row = ws.max_row
max_col = ws.max_column
print(f'  Hoja actual: {max_row} filas x {max_col} cols')

# Read all data into memory first
print('  Leyendo datos...')
all_data = []
for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
    all_data.append(list(row))

# Create new data with 4 extra columns at the beginning
# Original: Issue ID(0), Nombre(1), ..., Created(35), Created By(36), Modified(37), ...
# New: Mes Crea, Dia Crea, Dia Mod, Mes Mod, Issue ID, Nombre, ..., Created, Created By, Modified, ...
CREATED_IDX = 35  # 0-based index of Created column in original
MODIFIED_IDX = 37  # 0-based index of Modified column in original

print('  Calculando Mes Crea, Dia Crea, Dia Mod, Mes Mod...')
new_data = []
for r_idx, row in enumerate(all_data):
    if r_idx == 0:
        # Header row: add 4 new headers at the beginning
        new_row = ['Mes Crea', 'Dia Crea', 'Dia Mod', 'Mes Mod'] + row
        new_data.append(new_row)
    else:
        created = row[CREATED_IDX]
        modified = row[MODIFIED_IDX]
        new_row = [format_mes(created), format_dia(created), format_dia(modified), format_mes(modified)] + row
        new_data.append(new_row)

print(f'  Nuevos datos: {len(new_data)} filas x {len(new_data[0])} cols')

# Clear sheet and write new data
print('  Limpiando hoja Sharep...')
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
    for cell in row:
        cell.value = None

print('  Escribiendo datos...')
for r_idx, row_data in enumerate(new_data, 1):
    for c_idx, value in enumerate(row_data, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)
    if r_idx % 5000 == 0:
        print(f'    {r_idx}/{len(new_data)}...')

print(f'  Guardando...')
wb.save(TGT)
wb.close()
print('  OK! Sharep ahora tiene 45 cols')

# ===== STEP 2: XML manipulation =====
print('\n=== PASO 2: Copiando XMLs de pivot tables ===')

# Read source XMLs
with zipfile.ZipFile(SRC, 'r') as z_src:
    src_pivots = {}
    for i in range(1, 5):
        src_pivots[f'xl/pivotTables/pivotTable{i}.xml'] = z_src.read(f'xl/pivotTables/pivotTable{i}.xml')
    src_cache_def = z_src.read('xl/pivotCache/pivotCacheDefinition1.xml')
    src_table = z_src.read('xl/tables/table1.xml')
    src_sheet2 = z_src.read('xl/worksheets/sheet2.xml')
    src_sheet2_rels = z_src.read('xl/worksheets/_rels/sheet2.xml.rels')
    src_cache_rels = z_src.read('xl/pivotCache/_rels/pivotCacheDefinition1.xml.rels')

# Fix table definition: update ref to match new data (A1:AT{max_row}) 
# Original Iphone (1).xlsx: A1:AS38770 (45 cols), hoy: A1:AT{rows} (45 cols after adding 4)
table_xml = src_table.decode('utf-8')
new_ref = f'A1:AT{len(new_data)}'
table_xml = re.sub(r'ref="[^"]*"', f'ref="{new_ref}"', table_xml)
print(f'  Table ref updated to: {new_ref}')

# Fix cache definition: update recordCount, remove refreshOnLoad if present
cache_xml = src_cache_def.decode('utf-8')
cache_xml = re.sub(r'recordCount="\d+"', f'recordCount="{len(new_data)-1}"', cache_xml)
# Add refreshOnLoad
cache_xml = cache_xml.replace('<cacheSource', '<cacheSource', 1)
if 'refreshOnLoad' not in cache_xml:
    cache_xml = cache_xml.replace('minRefreshableVersion', 'refreshOnLoad="1" minRefreshableVersion')
print(f'  Cache recordCount: {len(new_data)-1}')

# Write the new xlsx
shutil.copy2(TGT, TMP)

with zipfile.ZipFile(TMP, 'r') as zin:
    with zipfile.ZipFile(TGT, 'w', zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            
            if item.filename == 'xl/worksheets/sheet2.xml':
                data = src_sheet2
            elif item.filename == 'xl/worksheets/_rels/sheet2.xml.rels':
                data = src_sheet2_rels
            elif item.filename.startswith('xl/pivotTables/pivotTable') and item.filename.endswith('.xml'):
                data = src_pivots[item.filename]
            elif item.filename == 'xl/pivotCache/pivotCacheDefinition1.xml':
                data = cache_xml.encode('utf-8')
            elif item.filename == 'xl/pivotCache/_rels/pivotCacheDefinition1.xml.rels':
                data = src_cache_rels
            elif item.filename == 'xl/tables/table1.xml':
                data = table_xml.encode('utf-8')
            elif item.filename == 'xl/tables/_rels/table1.xml.rels':
                # Add query table rel if not exists
                data = data  # keep as-is or add
            elif item.filename == '[Content_Types].xml':
                content = data.decode('utf-8')
                # Ensure pivotCache content type exists
                if 'pivotCache' not in content:
                    content = content.replace('</Types>', 
                        '<Override PartName="/xl/pivotCache/pivotCacheDefinition1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.pivotCacheDefinition+xml"/>'
                        '<Override PartName="/xl/pivotCache/pivotCacheRecords1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.pivotCacheRecords+xml"/>'
                        '</Types>')
                # Ensure pivotTable content types exist
                for i in range(1, 5):
                    part = f'/xl/pivotTables/pivotTable{i}.xml'
                    if part not in content:
                        content = content.replace('</Types>',
                            f'<Override PartName="{part}" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.pivotTable+xml"/></Types>')
                data = content.encode('utf-8')
            elif item.filename == 'xl/workbook.xml':
                content = data.decode('utf-8')
                # Add pivotCaches if not present
                if 'pivotCaches' not in content:
                    content = content.replace('</workbook>',
                        '<pivotCaches><pivotCache cacheId="1" r:id="rId3"/></pivotCaches></workbook>')
                data = content.encode('utf-8')
            elif item.filename == 'xl/_rels/workbook.xml.rels':
                content = data.decode('utf-8')
                # Ensure pivotCacheDefinition rel exists
                if 'pivotCacheDefinition' not in content:
                    content = content.replace('</Relationships>',
                        '<Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotCacheDefinition" Target="pivotCache/pivotCacheDefinition1.xml"/></Relationships>')
                data = content.encode('utf-8')
            
            zout.writestr(item, data)

os.remove(TMP)
print('  XMLs copiados!')

# ===== STEP 3: Verify =====
print('\n=== PASO 3: Verificacion ===')
with zipfile.ZipFile(TGT, 'r') as z:
    sheets = z.read('xl/workbook.xml').decode('utf-8')
    print(f'  Sheets: {re.findall(r"name=\"([^\"]+)\"", sheets)}')
    
    cache = z.read('xl/pivotCache/pivotCacheDefinition1.xml').decode('utf-8')
    wss = re.search(r'<worksheetSource[^>]*/>', cache)
    rc = re.search(r'recordCount="\d+"', cache)
    ro = re.search(r'refreshOnLoad="[^"]*"', cache)
    print(f'  Cache source: {wss.group(0) if wss else "N/A"}')
    print(f'  RecordCount: {rc.group(0) if rc else "N/A"}')
    print(f'  RefreshOnLoad: {ro.group(0) if ro else "N/A"}')
    
    table = z.read('xl/tables/table1.xml').decode('utf-8')
    tref = re.search(r'ref="([^"]+)"', table)
    tname = re.search(r'name="([^"]+)"', table)
    print(f'  Table name: {tname.group(1) if tname else "N/A"}')
    print(f'  Table ref: {tref.group(1) if tref else "N/A"}')
    
    for i in range(1, 5):
        pt = z.read(f'xl/pivotTables/pivotTable{i}.xml').decode('utf-8')
        name = re.search(r'name="([^"]+)"', pt)
        print(f'  Pivot {i}: {name.group(1) if name else "N/A"}')

print('\nListo! Abrir hoy.xlsx en Excel y refrescar (Ctrl+Alt+F5)')
