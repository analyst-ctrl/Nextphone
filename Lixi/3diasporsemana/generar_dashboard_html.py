import openpyxl
from collections import defaultdict, Counter
from datetime import datetime, date
import sys, io, json

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

PROMO_DAYS = {0: 'Lunes', 2: 'Miércoles', 4: 'Viernes'}
PROMO_DAYS_SHORT = {0: 'Lun', 2: 'Mie', 4: 'Vie'}
TODAY = date.today()
EXCLUDE = {'', 'N/A', 'NINGUNO', 'NO APLICA', 'NA', 'S/N', 'NO LLEVA EQUIPO', 'SIN EQUIPO', 'EQUIPO PROPIO', 'N/A ', 'NINGUNO ', 'NO APLICA '}

def week_of_july(d):
    return (d.day - 1) // 7 + 1

print("Leyendo Sharep.xlsx...")
wb = openpyxl.load_workbook('Sharep.xlsx', read_only=True, data_only=True)
ws = wb['owssvr (6)']
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
col_idx = {h: i for i, h in enumerate(headers)}

cols = {k: col_idx[v] for k, v in {
    'modelo': 'Modelo del equipo', 'promo': 'Lleva promoción',
    'supervisor': 'Supervisor', 'created': 'Created',
    'cliente': 'Nombre del cliente', 'asesor': 'Asesor',
    'plan': 'Plan Vendido', 'status': 'Status del Cliente',
}.items()}

all_data = []
sup_total_all = Counter()
sup_total_iphone = Counter()
dia_sup_all = defaultdict(lambda: defaultdict(int))
dia_sup_iphone = defaultdict(lambda: defaultdict(int))
promo_days_set = set()
modelos_set = set()

for row in ws.iter_rows(min_row=2, values_only=True):
    modelo_raw = row[cols['modelo']]
    modelo = str(modelo_raw or '').strip()
    modelo_up = modelo.upper()
    if not modelo or modelo_up in EXCLUDE:
        continue
    promo = row[cols['promo']] if cols['promo'] < len(row) else None
    if not promo or str(promo).strip().upper() != 'SI':
        continue
    created = row[cols['created']] if cols['created'] < len(row) else None
    date_obj = None
    if created:
        try:
            if isinstance(created, datetime):
                date_obj = created
            elif isinstance(created, str):
                for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y %H:%M:%S', '%d/%m/%Y']:
                    try:
                        date_obj = datetime.strptime(created, fmt)
                        break
                    except:
                        pass
        except:
            pass
    if not date_obj or date_obj.year != 2026 or date_obj.month != 7:
        continue
    wd = date_obj.weekday()
    if wd not in PROMO_DAYS:
        continue

    dia_str = date_obj.strftime('%Y-%m-%d')
    dia_nombre = PROMO_DAYS[wd]
    supervisor = str(row[cols['supervisor']] or 'Sin supervisor')
    es_iphone = 'IPHONE' in modelo_up
    modelos_set.add(modelo)

    dia_sup_all[dia_str][supervisor] += 1
    sup_total_all[supervisor] += 1
    if es_iphone:
        dia_sup_iphone[dia_str][supervisor] += 1
        sup_total_iphone[supervisor] += 1
    promo_days_set.add(dia_str)

    all_data.append({
        'cliente': str(row[cols['cliente']] or ''),
        'modelo': modelo,
        'supervisor': supervisor,
        'asesor': str(row[cols['asesor']] or ''),
        'dia': dia_str,
        'dia_nombre': dia_nombre,
        'dia_short': PROMO_DAYS_SHORT[wd],
        'weekday': wd,
        'semana': week_of_july(date_obj),
        'es_iphone': es_iphone,
        'plan': str(row[cols['plan']] or ''),
        'status': str(row[cols['status']] or ''),
    })
wb.close()

# Comparison data (all julio, promo=Si, any phone)
wb2 = openpyxl.load_workbook('Sharep.xlsx', read_only=True, data_only=True)
ws2 = wb2['owssvr (6)']
_ = [cell.value for cell in next(ws2.iter_rows(min_row=1, max_row=1))]
julio_all_total = julio_all_lmj = 0
julio_iphone_total = julio_iphone_lmj = 0
for row in ws2.iter_rows(min_row=2, values_only=True):
    modelo_raw = row[cols['modelo']] if cols['modelo'] < len(row) else None
    modelo_str = str(modelo_raw or '').strip().upper()
    if not modelo_str or modelo_str in EXCLUDE: continue
    promo = row[cols['promo']] if cols['promo'] < len(row) else None
    if not promo or str(promo).strip().upper() != 'SI': continue
    created = row[cols['created']] if cols['created'] < len(row) else None
    date_obj = None
    if created:
        try:
            if isinstance(created, datetime): date_obj = created
            elif isinstance(created, str):
                for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y %H:%M:%S', '%d/%m/%Y']:
                    try: date_obj = datetime.strptime(created, fmt); break
                    except: pass
        except: pass
    if not date_obj or date_obj.year != 2026 or date_obj.month != 7: continue
    es_iphone = 'IPHONE' in modelo_str
    julio_all_total += 1
    if es_iphone: julio_iphone_total += 1
    if date_obj.weekday() in (0, 2, 4):
        julio_all_lmj += 1
        if es_iphone: julio_iphone_lmj += 1
wb2.close()
julio_all_other = julio_all_total - julio_all_lmj
julio_iphone_other = julio_iphone_total - julio_iphone_lmj

july_days = [d for d in range(1, 32) if date(2026, 7, d) <= TODAY]
total_july_days = len(july_days)
lmj_days_count = sum(1 for d in july_days if date(2026, 7, d).weekday() in (0, 2, 4))
other_days_count = total_july_days - lmj_days_count

semanas_con_datos = sorted(set(d['semana'] for d in all_data))
supervisores_list = sorted(set(d['supervisor'] for d in all_data))

data_json = json.dumps({
    'registros': all_data,
    'totalPromo': sum(sup_total_all.values()),
    'totalIphone': sum(sup_total_iphone.values()),
    'totalSups': len(sup_total_all),
    'totalDias': len(promo_days_set),
    'julioAllTotal': julio_all_total,
    'julioAllLmj': julio_all_lmj,
    'julioAllOther': julio_all_other,
    'julioIphoneTotal': julio_iphone_total,
    'julioIphoneLmj': julio_iphone_lmj,
    'julioIphoneOther': julio_iphone_other,
    'totalJulyDays': total_july_days,
    'lmjDaysCount': lmj_days_count,
    'otherDaysCount': other_days_count,
    'semanas': semanas_con_datos,
    'supervisores': supervisores_list,
    'modelosCount': len(modelos_set),
    'today': TODAY.strftime('%Y-%m-%d'),
}, ensure_ascii=False)

html = '''<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Dashboard Promo - 3 Dias por Semana</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4"></script>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:'Segoe UI',Arial,sans-serif;background:#f0f2f5;color:#333;padding:20px}
.container{max-width:1400px;margin:0 auto}
h1{color:#1F4E79;font-size:22px;margin-bottom:3px}
.subtitle{color:#666;font-size:13px;margin-bottom:20px}
.card{background:#fff;border-radius:10px;box-shadow:0 2px 8px rgba(0,0,0,0.08);padding:20px;margin-bottom:20px}
.card h2{color:#1F4E79;font-size:17px;margin-bottom:12px;padding-bottom:6px;border-bottom:2px solid #1F4E79}
table{width:100%;border-collapse:collapse;font-size:13px}
th{background:#1F4E79;color:#fff;padding:8px 10px;text-align:center;font-weight:600;font-size:12px}
td{padding:6px 10px;border-bottom:1px solid #e0e0e0;text-align:center;font-size:13px}
td.left{text-align:left}
tr:nth-child(even){background:#f8f9fa}
.gold{background:#FFD700!important;font-weight:bold}
.silver{background:#C0C0C0!important}
.bronze{background:#CD7F32!important;color:#fff}
.grid2{display:grid;grid-template-columns:1fr 1fr;gap:20px}
.grid3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:20px}
.metric{text-align:center;padding:15px}
.metric .num{font-size:30px;font-weight:bold;color:#1F4E79}
.metric .label{font-size:12px;color:#666;margin-top:4px}
.metric .sub{font-size:11px;color:#999}
.day-section{margin-bottom:15px}
.day-header{background:#D4EDDA;padding:8px 15px;border-radius:8px;font-weight:bold;color:#155724;margin-bottom:8px;display:flex;justify-content:space-between;font-size:14px}
.chart-container{height:260px;margin-top:8px}
@media(max-width:768px){.grid2,.grid3{grid-template-columns:1fr}}
.detail-table{font-size:11px}
.detail-table th{font-size:11px;padding:5px 6px}
.detail-table td{padding:3px 6px}
.filters{display:flex;gap:15px;flex-wrap:wrap;margin-bottom:15px;align-items:center}
.filters label{font-weight:600;font-size:13px;color:#333}
.filters select,.filters input[type=radio]{padding:6px 12px;border:2px solid #1F4E79;border-radius:6px;font-size:13px;background:#fff;cursor:pointer;min-width:120px}
.filters select:focus{outline:none;box-shadow:0 0 0 2px rgba(31,78,121,0.3)}
.filters .radio-group{display:flex;gap:5px;align-items:center;background:#fff;border:2px solid #1F4E79;border-radius:6px;padding:3px}
.filters .radio-group label{padding:4px 12px;border-radius:4px;cursor:pointer;font-weight:400;font-size:12px;margin:0}
.filters .radio-group input{display:none}
.filters .radio-group input:checked+label{background:#1F4E79;color:#fff;font-weight:600}
.badge{padding:2px 8px;border-radius:10px;font-size:11px;font-weight:bold}
.badge-lun{background:#cce5ff;color:#004085}
.badge-mie{background:#d4edda;color:#155724}
.badge-vie{background:#fff3cd;color:#856404}
.badge-iphone{background:#e8d5f5;color:#6a1b9a}
.iphone-row{background:#f3e8ff!important}
</style>
</head>
<body>
<div class="container">
<h1>Dashboard Promo - 3 Dias por Semana (Lun/Mie/Vie)</h1>
<p class="subtitle" id="subtitleText">Julio 2026 | Corte: $TODAY_STR$</p>

<div class="filters">
  <div>
    <label>Equipos:</label>
    <div class="radio-group">
      <input type="radio" name="filtroTipo" id="tipoTodos" value="todos" checked><label for="tipoTodos">Todos</label>
      <input type="radio" name="filtroTipo" id="tipoIphone" value="iphone"><label for="tipoIphone">Solo iPhone</label>
    </div>
  </div>
  <div>
    <label for="filterSemana">Semana:</label>
    <select id="filterSemana"><option value="todas">Todas</option></select>
  </div>
  <div>
    <label for="filterDia">Dia:</label>
    <select id="filterDia"><option value="todos">Todos (L/M/J)</option><option value="0">Lunes</option><option value="2">Miercoles</option><option value="4">Viernes</option></select>
  </div>
  <div>
    <label for="filterSup">Supervisor:</label>
    <select id="filterSup"><option value="todos">Todos</option></select>
  </div>
  <div id="filterStats" style="font-size:13px;color:#666;margin-left:auto"></div>
</div>

<div class="grid3" id="metricsCards"></div>

<div class="grid2">
  <div class="card"><h2>Ranking Supervisores</h2><div id="rankingContainer"></div></div>
  <div class="card"><h2>Grafico Ventas</h2><div class="chart-container"><canvas id="chartRanking"></canvas></div></div>
</div>

<div class="card"><h2>Desglose por Dia Promo</h2><div id="dailyContainer"></div></div>

<div class="card"><h2>Comparativa: Dias Promo vs Total Julio</h2>
  <div class="grid2">
    <div id="comparativaTable"></div>
    <div><div class="chart-container"><canvas id="chartComparativo"></canvas></div></div>
  </div>
</div>

<div class="card">
  <h2>Detalle de Ventas</h2>
  <div style="max-height:350px;overflow-y:auto"><table class="detail-table" id="detailTable"></table></div>
</div>

</div>

<script>
const DATA = $DATA_JSON$;
const DAY_NAMES = {0:'Lunes',2:'Miercoles',4:'Viernes'};
const DAY_SHORT = {0:'Lun',2:'Mie',4:'Vie'};
const DAY_CLASS = {0:'badge-lun',2:'badge-mie',4:'badge-vie'};

// Populate filters
[...new Set(DATA.registros.map(r=>r.semana))].sort().forEach(s => {
  document.getElementById('filterSemana').innerHTML += '<option value="'+s+'">Semana '+s+'</option>';
});
[...new Set(DATA.registros.map(r=>r.supervisor))].sort().forEach(s => {
  document.getElementById('filterSup').innerHTML += '<option value="'+s+'">'+s+'</option>';
});

function filterData() {
  const soloIphone = document.getElementById('tipoIphone').checked;
  const sem = document.getElementById('filterSemana').value;
  const dia = document.getElementById('filterDia').value;
  const sup = document.getElementById('filterSup').value;
  return DATA.registros.filter(r => {
    if (soloIphone && !r.es_iphone) return false;
    if (sem !== 'todas' && r.semana !== parseInt(sem)) return false;
    if (dia !== 'todos' && r.weekday !== parseInt(dia)) return false;
    if (sup !== 'todos' && r.supervisor !== sup) return false;
    return true;
  });
}

function render() {
  const filtered = filterData();
  const filteredCount = filtered.length;
  const filteredSups = [...new Set(filtered.map(r=>r.supervisor))].length;
  const filteredDays = [...new Set(filtered.map(r=>r.dia))].length;
  const iphoneCount = filtered.filter(r=>r.es_iphone).length;

  // Metrics
  const isIphone = document.getElementById('tipoIphone').checked;
  const totalRef = isIphone ? DATA.julioIphoneTotal : DATA.julioAllTotal;
  document.getElementById('metricsCards').innerHTML =
    '<div class="card metric"><div class="num">'+filteredCount+'</div><div class="label">Equipos Vendidos (Promo L/M/J)</div><div class="sub">En '+filteredDays+' dias promo'+(isIphone?' | '+iphoneCount+' iPhone':'')+'</div></div>' +
    '<div class="card metric"><div class="num">'+filteredSups+'</div><div class="label">Supervisores</div><div class="sub">Con ventas en promo</div></div>' +
    '<div class="card metric"><div class="num">'+(filteredDays ? (filteredCount/filteredDays).toFixed(1) : '0')+'</div><div class="label">Promedio x Dia</div><div class="sub">'+filteredDays+' dias filtrados</div></div>';

  document.getElementById('filterStats').textContent = 'Mostrando '+filteredCount+' de '+(isIphone ? DATA.julioIphoneLmj : DATA.julioAllLmj)+' registros en dias promo';

  // Ranking
  const supCount = {}; const supDays = {}; const supIphone = {};
  filtered.forEach(r => {
    supCount[r.supervisor] = (supCount[r.supervisor] || 0) + 1;
    supDays[r.supervisor] = supDays[r.supervisor] || new Set();
    supDays[r.supervisor].add(r.dia);
    if (r.es_iphone) supIphone[r.supervisor] = (supIphone[r.supervisor] || 0) + 1;
  });
  const sorted = Object.entries(supCount).sort((a,b) => b[1] - a[1]);
  const total = filteredCount || 1;
  let html = '<table><tr><th>#</th><th class="left">Supervisor</th><th>Ventas</th><th>%</th><th>Dias</th><th>Prom/Dia</th>'+(isIphone?'':'<th>iPhone</th>')+'</tr>';
  sorted.forEach(([sup, cnt], i) => {
    const cls = i===0?'gold':i===1?'silver':i===2?'bronze':'';
    const d = supDays[sup].size;
    const ip = supIphone[sup] || 0;
    html += '<tr class="'+cls+'"><td>'+(i+1)+'</td><td class="left">'+sup+'</td><td>'+cnt+'</td><td>'+(cnt/total*100).toFixed(1)+'%</td><td>'+d+'</td><td>'+(cnt/d).toFixed(1)+'</td>'+(isIphone?'':'<td>'+ip+'</td>')+'</tr>';
  });
  html += '</table>';
  document.getElementById('rankingContainer').innerHTML = html;

  // Chart
  const ctx = document.getElementById('chartRanking');
  if (window.rankingChart) window.rankingChart.destroy();
  window.rankingChart = new Chart(ctx, {
    type: 'bar',
    data: {
      labels: sorted.slice(0,10).map(s=>s[0]),
      datasets: [{
        label: 'Ventas', data: sorted.slice(0,10).map(s=>s[1]),
        backgroundColor: ['#FFD700','#C0C0C0','#CD7F32','#1F4E79','#2E75B6','#5B9BD5','#9DC3E6','#E8F0FE','#D4EDDA','#FFF3CD'],
        borderRadius: 4,
      }]
    },
    options: { responsive: true, maintainAspectRatio: false, plugins: { legend: { display: false } }, scales: { y: { beginAtZero: true, ticks: { stepSize: 1 } } } }
  });

  // Daily
  const dayGroups = {};
  filtered.forEach(r => {
    if (!dayGroups[r.dia]) dayGroups[r.dia] = {fecha:r.dia, nombre:r.dia_nombre, weekday:r.weekday, sups:{}};
    dayGroups[r.dia].sups[r.supervisor] = (dayGroups[r.dia].sups[r.supervisor] || 0) + 1;
  });
  let dailyHtml = '';
  Object.keys(dayGroups).sort().forEach(dia => {
    const dg = dayGroups[dia];
    const supsArr = Object.entries(dg.sups).sort((a,b) => b[1] - a[1]);
    const totalDia = supsArr.reduce((s, x) => s + x[1], 0);
    const badgeClass = DAY_CLASS[dg.weekday];
    dailyHtml += '<div class="day-section">' +
      '<div class="day-header"><span>'+dg.fecha+' - '+dg.nombre+' <span class="badge '+badgeClass+'">'+DAY_SHORT[dg.weekday]+'</span></span><span>Total: '+totalDia+' equipos</span></div>' +
      '<table><tr><th>#</th><th class="left">Supervisor</th><th>Ventas</th><th>% del Dia</th></tr>';
    supsArr.forEach(([sup, cnt], i) => {
      dailyHtml += '<tr><td>'+(i+1)+'</td><td class="left">'+sup+'</td><td>'+cnt+'</td><td>'+(cnt/totalDia*100).toFixed(1)+'%</td></tr>';
    });
    dailyHtml += '</table></div>';
  });
  document.getElementById('dailyContainer').innerHTML = dailyHtml || '<p style="color:#999;padding:10px">Sin datos con estos filtros.</p>';

  // Comparativa
  const cmp = isIphone ? {
    total: DATA.julioIphoneTotal, lmj: DATA.julioIphoneLmj, other: DATA.julioIphoneOther, label: 'iPhone'
  } : {
    total: DATA.julioAllTotal, lmj: DATA.julioAllLmj, other: DATA.julioAllOther, label: 'Equipos'
  };
  document.getElementById('comparativaTable').innerHTML =
    '<table><tr><th class="left">Metrica</th><th>Valor</th><th>%</th></tr>' +
    '<tr><td class="left">Dias transcurridos Julio</td><td>'+DATA.totalJulyDays+'</td><td>100%</td></tr>' +
    '<tr><td class="left">Dias Promo (L/M/J)</td><td>'+DATA.lmjDaysCount+'</td><td>'+(DATA.lmjDaysCount/DATA.totalJulyDays*100).toFixed(1)+'%</td></tr>' +
    '<tr><td class="left">Dias No Promo</td><td>'+DATA.otherDaysCount+'</td><td>'+(DATA.otherDaysCount/DATA.totalJulyDays*100).toFixed(1)+'%</td></tr>' +
    '<tr><td colspan="3" style="padding:2px"></td></tr>' +
    '<tr><td class="left">'+cmp.label+' vendidos (Total Julio)</td><td>'+cmp.total+'</td><td>100%</td></tr>' +
    '<tr><td class="left">En Dias Promo (L/M/J)</td><td>'+cmp.lmj+'</td><td>'+(cmp.total?(cmp.lmj/cmp.total*100).toFixed(1):'0')+'%</td></tr>' +
    '<tr><td class="left">En Dias No Promo</td><td>'+cmp.other+'</td><td>'+(cmp.total?(cmp.other/cmp.total*100).toFixed(1):'0')+'%</td></tr>' +
    '<tr><td colspan="3" style="padding:2px"></td></tr>' +
    '<tr><td class="left">Promedio x dia (general)</td><td>'+(DATA.totalJulyDays?(cmp.total/DATA.totalJulyDays).toFixed(1):'0')+'</td><td></td></tr>' +
    '<tr><td class="left">Promedio x dia promo</td><td>'+(DATA.lmjDaysCount?(cmp.lmj/DATA.lmjDaysCount).toFixed(1):'0')+'</td><td></td></tr>' +
    '</table>';

  const ctx2 = document.getElementById('chartComparativo');
  if (window.compChart) window.compChart.destroy();
  window.compChart = new Chart(ctx2, {
    type: 'bar',
    data: {
      labels: ['Dias Promo (L/M/J)', 'Dias No Promo'],
      datasets: [{
        label: 'Ventas '+cmp.label, data: [cmp.lmj, cmp.other],
        backgroundColor: ['#28a745', '#dc3545'], borderRadius: 4,
      }]
    },
    options: { responsive: true, maintainAspectRatio: false, plugins: { legend: { display: false } }, scales: { y: { beginAtZero: true, ticks: { stepSize: 1 } } } }
  });

  // Detail
  let dh = '<tr><th>#</th><th class="left">Cliente</th><th class="left">Modelo</th><th class="left">Supervisor</th><th class="left">Asesor</th><th>Dia</th><th>Status</th></tr>';
  filtered.forEach((r, i) => {
    const cls = r.es_iphone ? ' class="iphone-row"' : '';
    dh += '<tr'+cls+'><td>'+(i+1)+'</td><td class="left">'+r.cliente+'</td><td class="left">'+r.modelo+(r.es_iphone?' <span class="badge badge-iphone">iPhone</span>':'')+'</td><td class="left">'+r.supervisor+'</td><td class="left">'+r.asesor+'</td><td>'+r.dia+'<br><small>'+r.dia_nombre+'</small></td><td>'+r.status+'</td></tr>';
  });
  document.getElementById('detailTable').innerHTML = dh;
  document.getElementById('subtitleText').textContent = 'Julio 2026 | Corte: '+DATA.today+' | '+DATA.modelosCount+' modelos diferentes';
}

document.querySelectorAll('input[name="filtroTipo"]').forEach(el => el.addEventListener('change', render));
['filterSemana','filterDia','filterSup'].forEach(id => {
  document.getElementById(id).addEventListener('change', render);
});

render();
</script>
</body>
</html>'''

html = html.replace('$DATA_JSON$', data_json).replace('$TODAY_STR$', TODAY.strftime('%Y-%m-%d'))

with open('dashboard_3diasporsemana.html', 'w', encoding='utf-8') as f:
    f.write(html)

print(f"Dashboard generado: dashboard_3diasporsemana.html")
print(f"Registros: {len(all_data)} | Semanas: {semanas_con_datos} | Supervisores: {len(supervisores_list)}")
print(f"Modelos unicos: {len(modelos_set)} | iPhones: {sum(1 for d in all_data if d['es_iphone'])}")
