import openpyxl
from collections import defaultdict, Counter
from datetime import datetime, date
import sqlite3
import json
import sys, io, os

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

BASE = os.path.dirname(os.path.abspath(__file__))
PROJ = os.path.join(BASE, '..')
XLSX_PATH = os.path.join(PROJ, 'excels', 'Iphone_7_24_2026.xlsx')
OUT_DIR = os.path.join(PROJ, 'Lixi', '3diasporsemana')
DB_PATH = os.path.join(OUT_DIR, 'iphone_julio_7_24.db')
OUT_HTML = os.path.join(OUT_DIR, 'dashboard_iphone_julio_7_24.html')

TODAY = date.today()

MES_MAP = {1:'ene',2:'feb',3:'mar',4:'abr',5:'may',6:'jun',7:'jul',8:'ago',9:'sep',10:'oct',11:'nov',12:'dic'}

print("=" * 60)
print("  PASO 1: Leyendo Iphone_7_24_2026.xlsx ...")
print("=" * 60)

wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
print(f"  Hoja: {ws.title} | {ws.max_row} filas x {ws.max_column} cols")

headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
col_idx = {h: i for i, h in enumerate(headers)}
print(f"  Columnas: {headers}")

def cv(v):
    return str(v).strip() if v is not None else ''

def parse_date(v):
    if not v:
        return None
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y %H:%M:%S', '%d/%m/%Y']:
        try:
            return datetime.strptime(s, fmt)
        except:
            pass
    return None

def format_mes(dt):
    if dt is None: return None
    return f'{MES_MAP[dt.month]}/{str(dt.year)[-2:]}'

def format_dia(dt):
    if dt is None: return None
    return f'{dt.day:02d}/{MES_MAP[dt.month]}/{str(dt.year)[-2:]}'

print("\n  Leyendo registros iPhone julio 2026...")

dim_supervisores = {}
dim_asesores = {}
dim_modelos = {}
dim_statuses = {}
dim_clientes = {}

ventas_raw = []
count = 0
iphone_count = 0
julio_count = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    count += 1

    created_raw = row[col_idx.get('Created', 35)] if 'Created' in col_idx else None
    d = parse_date(created_raw)
    if not d or d.year != 2026 or d.month != 7:
        continue

    modelo = cv(row[col_idx.get('Modelo del equipo', 18)])
    if not modelo or 'IPHONE' not in modelo.upper():
        continue

    julio_count += 1

    modified_raw = row[col_idx.get('Modified', 38)] if 'Modified' in col_idx else None
    dm = parse_date(modified_raw)

    sup = cv(row[col_idx.get('Supervisor', 33)]) or 'Sin Supervisor'
    ase = cv(row[col_idx.get('Asesor', 32)]) or 'Sin Asesor'
    st = cv(row[col_idx.get('Status del Cliente', 30)])
    cli_nom = cv(row[col_idx.get('Nombre del cliente', 1)])
    cli_id = cv(row[col_idx.get('Identificación (cedula o pasaporte)', 2)])

    dia_crea = format_dia(d)
    mes_crea = format_mes(d)
    dia_mod = format_dia(dm) if dm else None
    mes_mod = format_mes(dm) if dm else None

    if sup not in dim_supervisores:
        dim_supervisores[sup] = len(dim_supervisores) + 1
    if ase not in dim_asesores:
        dim_asesores[ase] = len(dim_asesores) + 1
    if modelo not in dim_modelos:
        dim_modelos[modelo] = len(dim_modelos) + 1
    if st and st not in dim_statuses:
        dim_statuses[st] = len(dim_statuses) + 1
    cli_key = (cli_nom, cli_id)
    if cli_nom and cli_key not in dim_clientes:
        dim_clientes[cli_key] = len(dim_clientes) + 1

    ventas_raw.append({
        'sup_id': dim_supervisores[sup],
        'ase_id': dim_asesores[ase],
        'mod_id': dim_modelos[modelo],
        'st_id': dim_statuses.get(st),
        'cli_id': dim_clientes.get(cli_key),
        'modelo': modelo, 'supervisor': sup, 'asesor': ase, 'status': st,
        'fecha_created': d, 'fecha_entrega': parse_date(row[col_idx.get('Día y hora de la entrega', 19)]),
        'dia_crea': dia_crea, 'mes_crea': mes_crea,
        'dia_mod': dia_mod, 'mes_mod': mes_mod,
        'cliente_nombre': cli_nom, 'cliente_id_str': cli_id,
    })

wb.close()
print(f"  Total registros leidos: {count}")
print(f"  iPhone julio 2026: {julio_count}")
print(f"  Supervisores: {len(dim_supervisores)}")
print(f"  Asesores: {len(dim_asesores)}")
print(f"  Modelos: {len(dim_modelos)}")

# ========================================
# PASO 2: Crear SQLite
# ========================================
print("\n" + "=" * 60)
print("  PASO 2: Creando base de datos SQLite ...")
print("=" * 60)

os.makedirs(OUT_DIR, exist_ok=True)

if os.path.exists(DB_PATH):
    os.remove(DB_PATH)

conn = sqlite3.connect(DB_PATH)
c = conn.cursor()
c.execute('PRAGMA journal_mode=WAL')

c.execute('CREATE TABLE supervisores (id INTEGER PRIMARY KEY, nombre TEXT NOT NULL UNIQUE)')
c.execute('CREATE TABLE asesores (id INTEGER PRIMARY KEY, nombre TEXT NOT NULL UNIQUE)')
c.execute('CREATE TABLE modelos (id INTEGER PRIMARY KEY, nombre TEXT NOT NULL UNIQUE)')
c.execute('CREATE TABLE status_cliente (id INTEGER PRIMARY KEY, nombre TEXT NOT NULL UNIQUE)')
c.execute('CREATE TABLE clientes (id INTEGER PRIMARY KEY, nombre TEXT NOT NULL, identificacion TEXT, UNIQUE(nombre, identificacion))')

c.execute('''CREATE TABLE ventas (
    id INTEGER PRIMARY KEY,
    supervisor_id INTEGER, asesor_id INTEGER, modelo_id INTEGER,
    status_id INTEGER, cliente_id INTEGER,
    fecha_creacion TEXT, fecha_entrega TEXT,
    dia_crea TEXT, mes_crea TEXT, dia_mod TEXT, mes_mod TEXT,
    FOREIGN KEY (supervisor_id) REFERENCES supervisores(id),
    FOREIGN KEY (asesor_id) REFERENCES asesores(id),
    FOREIGN KEY (modelo_id) REFERENCES modelos(id),
    FOREIGN KEY (status_id) REFERENCES status_cliente(id),
    FOREIGN KEY (cliente_id) REFERENCES clientes(id)
)''')

for table, data in [('supervisores', dim_supervisores), ('asesores', dim_asesores),
                     ('modelos', dim_modelos), ('status_cliente', dim_statuses)]:
    c.executemany(f'INSERT INTO {table} (id, nombre) VALUES (?, ?)',
                  [(v, k) for k, v in data.items()])

c.executemany('INSERT INTO clientes (id, nombre, identificacion) VALUES (?, ?, ?)',
              [(v, k[0], k[1]) for k, v in dim_clientes.items()])

batch = []
for i, v in enumerate(ventas_raw):
    batch.append((
        i + 1, v['sup_id'], v['ase_id'], v['mod_id'],
        v['st_id'], v['cli_id'],
        str(v['fecha_created']) if v['fecha_created'] else None,
        str(v['fecha_entrega']) if v['fecha_entrega'] else None,
        v['dia_crea'], v['mes_crea'], v['dia_mod'], v['mes_mod'],
    ))

c.executemany('''INSERT INTO ventas (
    id, supervisor_id, asesor_id, modelo_id,
    status_id, cliente_id,
    fecha_creacion, fecha_entrega,
    dia_crea, mes_crea, dia_mod, mes_mod
) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)''', batch)

conn.commit()
db_size = os.path.getsize(DB_PATH)
print(f"  DB: {DB_PATH}")
print(f"  Tamano: {db_size / 1024:.1f} KB")

# ========================================
# PASO 3: Datos para dashboard
# ========================================
print("\n" + "=" * 60)
print("  PASO 3: Obteniendo datos para dashboard ...")
print("=" * 60)

c.execute('SELECT DISTINCT nombre FROM supervisores ORDER BY nombre')
all_sups = [r[0] for r in c.fetchall()]

c.execute('SELECT DISTINCT nombre FROM modelos ORDER BY nombre')
all_models = [r[0] for r in c.fetchall()]

c.execute('SELECT DISTINCT dia_crea FROM ventas WHERE dia_crea IS NOT NULL ORDER BY dia_crea')
all_dia_crea = [r[0] for r in c.fetchall()]

c.execute('SELECT DISTINCT dia_mod FROM ventas WHERE dia_mod IS NOT NULL ORDER BY dia_mod')
all_dia_mod = [r[0] for r in c.fetchall()]

c.execute('SELECT DISTINCT nombre FROM status_cliente ORDER BY nombre')
all_statuses = [r[0] for r in c.fetchall()]

c.execute('''SELECT
    v.id,
    s.nombre AS supervisor,
    a.nombre AS asesor,
    m.nombre AS modelo,
    st.nombre AS status,
    v.fecha_creacion,
    v.fecha_entrega,
    v.dia_crea, v.mes_crea,
    v.dia_mod, v.mes_mod,
    cl.nombre AS cliente
FROM ventas v
LEFT JOIN supervisores s ON v.supervisor_id = s.id
LEFT JOIN asesores a ON v.asesor_id = a.id
LEFT JOIN modelos m ON v.modelo_id = m.id
LEFT JOIN status_cliente st ON v.status_id = st.id
LEFT JOIN clientes cl ON v.cliente_id = cl.id
ORDER BY v.id
''')

all_ventas = []
for row in c.fetchall():
    all_ventas.append({
        'id': row[0], 'supervisor': row[1], 'asesor': row[2],
        'modelo': row[3], 'status': row[4],
        'fecha_creacion': row[5], 'fecha_entrega': row[6],
        'dia_crea': row[7], 'mes_crea': row[8],
        'dia_mod': row[9], 'mes_mod': row[10],
        'cliente': row[11],
    })

conn.close()

dashboard_data = {
    'ventas': all_ventas,
    'supervisores': all_sups,
    'modelos': all_models,
    'dia_crea': all_dia_crea,
    'dia_mod': all_dia_mod,
    'statuses': all_statuses,
    'total': len(all_ventas),
    'db_size': f"{db_size / 1024:.1f} KB",
    'today': TODAY.strftime('%Y-%m-%d'),
}

data_json = json.dumps(dashboard_data, ensure_ascii=False)

# ========================================
# PASO 4: Dashboard HTML
# ========================================
print("\n" + "=" * 60)
print("  PASO 4: Generando dashboard ...")
print("=" * 60)

html = '''<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Tabla Dinamica iPhone Julio 2026</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:'Segoe UI',Arial,sans-serif;background:#f0f2f5;color:#333;padding:20px}
.container{max-width:1400px;margin:0 auto}
h1{color:#1F4E79;font-size:22px;margin-bottom:3px}
.subtitle{color:#666;font-size:13px;margin-bottom:20px}
.card{background:#fff;border-radius:10px;box-shadow:0 2px 8px rgba(0,0,0,0.08);padding:20px;margin-bottom:20px}
.card h2{color:#1F4E79;font-size:17px;margin-bottom:12px;padding-bottom:6px;border-bottom:2px solid #1F4E79}
table{width:100%;border-collapse:collapse;font-size:13px}
th{background:#1F4E79;color:#fff;padding:8px 10px;text-align:center;font-weight:600;font-size:12px}
td{padding:6px 10px;border-bottom:1px solid #e0e0e0;text-align:center;font-size:13px}
td.left{text-align:left}
tr:nth-child(even){background:#f8f9fa}
.gold{background:#FFD700!important;font-weight:bold}
.total-row{background:#D4EDDA!important;font-weight:bold}
.total-col{background:#FFF3CD!important;font-weight:bold}
.filters{display:flex;gap:15px;flex-wrap:wrap;margin-bottom:20px;align-items:end}
.filter-group{display:flex;flex-direction:column;gap:4px}
.filter-group label{font-weight:600;font-size:12px;color:#555}
.filter-group select{padding:8px 12px;border:2px solid #1F4E79;border-radius:6px;font-size:13px;background:#fff;cursor:pointer;min-width:180px}
.filter-group select:focus{outline:none;box-shadow:0 0 0 2px rgba(31,78,121,0.3)}
.metric{display:inline-block;background:#fff;border-radius:10px;padding:15px 25px;margin-right:15px;box-shadow:0 2px 8px rgba(0,0,0,0.08);text-align:center}
.metric .num{font-size:28px;font-weight:bold;color:#1F4E79}
.metric .label{font-size:11px;color:#666;margin-top:4px}
.grid2{display:grid;grid-template-columns:1fr 1fr;gap:20px}
.mode-btn{padding:8px 20px;border:2px solid #1F4E79;border-radius:6px;font-size:13px;cursor:pointer;font-weight:600;background:#fff;color:#1F4E79}
.mode-btn.active{background:#1F4E79;color:#fff}
.mode-btn:hover{opacity:0.9}
.status-facturada{color:#28a745;font-weight:bold}
.status-cancelada{color:#dc3545}
.badge{padding:2px 8px;border-radius:10px;font-size:11px;font-weight:bold}
.badge-creacion{background:#cce5ff;color:#004085}
.badge-facturada{background:#d4edda;color:#155724}
@media(max-width:768px){.grid2{grid-template-columns:1fr}.filters{flex-direction:column}}
</style>
</head>
<body>
<div class="container">
<h1>Tabla Dinamica iPhone - Julio 2026</h1>
<p class="subtitle">SQLite: ''' + f"{db_size / 1024:.1f} KB" + ''' | Corte: ''' + TODAY.strftime('%Y-%m-%d') + ''' | Fuente: Iphone_7_24_2026.xlsx</p>

<div class="filters">
  <div class="filter-group">
    <label>Supervisor</label>
    <select id="filterSup"><option value="todos">Todos</option></select>
  </div>
  <div class="filter-group">
    <label>Dia Creacion</label>
    <select id="filterDiaCrea"><option value="todos">Todos</option></select>
  </div>
  <div class="filter-group">
    <label>Dia Modificacion</label>
    <select id="filterDiaMod"><option value="todos">Todos</option></select>
  </div>
  <div class="filter-group">
    <label>Modelo</label>
    <select id="filterModelo"><option value="todos">Todos</option></select>
  </div>
  <div class="filter-group">
    <label>Modo Vista</label>
    <div style="display:flex;gap:5px">
      <button class="mode-btn active" onclick="setMode('creacion')">CREACION</button>
      <button class="mode-btn" onclick="setMode('facturadas')">FACTURADAS</button>
      <button class="mode-btn" onclick="setMode('ambas')">AMBAS</button>
    </div>
  </div>
</div>

<div id="metricsContainer" style="margin-bottom:20px"></div>

<div id="pivotContainer" class="card">
  <h2>Pivot: Asesor x Modelo</h2>
  <div style="overflow-x:auto" id="pivotTable"></div>
</div>

<div class="grid2">
  <div class="card">
    <h2>Resumen por Supervisor</h2>
    <div id="resumenSup"></div>
  </div>
  <div class="card">
    <h2>Resumen por Modelo</h2>
    <div id="resumenMod"></div>
  </div>
</div>

<div class="card">
  <h2>Detalle de Registros</h2>
  <div style="max-height:400px;overflow-y:auto" id="detalleTable"></div>
</div>

</div>

<script>
const DATA = ''' + data_json + ''';
let currentMode = 'creacion';

function populateSelect(id, values) {
  const sel = document.getElementById(id);
  values.forEach(v => {
    const opt = document.createElement('option');
    opt.value = v; opt.textContent = v; sel.appendChild(opt);
  });
}
populateSelect('filterSup', DATA.supervisores);
populateSelect('filterDiaCrea', DATA.dia_crea);
populateSelect('filterDiaMod', DATA.dia_mod);
populateSelect('filterModelo', DATA.modelos);

function setMode(mode) {
  currentMode = mode;
  document.querySelectorAll('.mode-btn').forEach(b => b.classList.remove('active'));
  event.target.classList.add('active');
  render();
}

function getFiltered() {
  const sup = document.getElementById('filterSup').value;
  const diaCrea = document.getElementById('filterDiaCrea').value;
  const diaMod = document.getElementById('filterDiaMod').value;
  const modelo = document.getElementById('filterModelo').value;
  return DATA.ventas.filter(v => {
    if (sup !== 'todos' && v.supervisor !== sup) return false;
    if (diaCrea !== 'todos' && v.dia_crea !== diaCrea) return false;
    if (diaMod !== 'todos' && v.dia_mod !== diaMod) return false;
    if (modelo !== 'todos' && v.modelo !== modelo) return false;
    return true;
  });
}

function isFacturada(status) {
  return status && (status.toUpperCase().includes('RENOVACION EXITOSA') || status.toUpperCase().includes('VENTAS') && status.toUpperCase().includes('EXITOSA'));
}

function render() {
  const filtered = getFiltered();
  let creacion = [], facturadas = [];
  if (currentMode === 'creacion') { creacion = filtered; }
  else if (currentMode === 'facturadas') { facturadas = filtered.filter(v => isFacturada(v.status)); }
  else { creacion = filtered; facturadas = filtered.filter(v => isFacturada(v.status)); }

  const totalC = creacion.length;
  const totalF = facturadas.length;
  document.getElementById('metricsContainer').innerHTML =
    '<div class="metric"><div class="num">' + totalC + '</div><div class="label">CREACION (todos status)</div></div>' +
    '<div class="metric"><div class="num" style="color:#28a745">' + totalF + '</div><div class="label">FACTURADAS</div></div>' +
    '<div class="metric"><div class="num">' + filtered.length + '</div><div class="label">Total Filtrado</div></div>';

  const data = currentMode === 'facturadas' ? facturadas : creacion;
  const pivotData = {};
  const allAsesores = new Set();
  const allModelos = new Set();
  data.forEach(v => {
    allAsesores.add(v.asesor); allModelos.add(v.modelo);
    const key = v.asesor + '|' + v.modelo;
    pivotData[key] = (pivotData[key] || 0) + 1;
  });
  const pivotFact = {};
  if (currentMode === 'ambas') {
    facturadas.forEach(v => {
      const key = v.asesor + '|' + v.modelo;
      pivotFact[key] = (pivotFact[key] || 0) + 1;
    });
  }

  const asesores = [...allAsesores].sort();
  const modelos = [...allModelos].sort();

  let h = '<table><tr><th style="text-align:left;min-width:200px">Asesor \\ Modelo</th>';
  modelos.forEach(m => { h += '<th>' + m.replace(/IPHONE /gi, 'iP ').replace(/256GB/g, '256').replace(/\\(Black\\)/gi, '(B)').replace(/\\(Blue\\)/gi, '(Bl)') + '</th>'; });
  h += '<th>Total</th></tr>';

  let grandTotal = 0;
  const colTotals = {};
  modelos.forEach(m => colTotals[m] = 0);

  asesores.forEach(ase => {
    h += '<tr><td class="left">' + ase + '</td>';
    let rowTotal = 0;
    modelos.forEach(m => {
      const key = ase + '|' + m;
      let val = pivotData[key] || 0;
      if (currentMode === 'ambas') {
        const fVal = pivotFact[key] || 0;
        h += '<td>' + (val || '') + '<br><span class="status-facturada">' + (fVal || '') + '</span></td>';
      } else { h += '<td>' + (val || '') + '</td>'; }
      rowTotal += val; colTotals[m] += val;
    });
    grandTotal += rowTotal;
    h += '<td class="total-col">' + rowTotal + '</td></tr>';
  });

  h += '<tr class="total-row"><td>Total general</td>';
  modelos.forEach(m => { h += '<td class="total-row">' + colTotals[m] + '</td>'; });
  h += '<td class="total-row" style="background:#FFD700">' + grandTotal + '</td></tr></table>';
  if (currentMode === 'ambas') {
    h += '<p style="margin-top:8px;font-size:11px;color:#666"><span class="badge badge-creacion">CREACION</span> / <span class="badge badge-facturada">FACTURADAS</span></p>';
  }
  document.getElementById('pivotTable').innerHTML = h;

  const supCount = {};
  data.forEach(v => { supCount[v.supervisor] = (supCount[v.supervisor] || 0) + 1; });
  let sh = '<table><tr><th class="left">Supervisor</th><th>Ventas</th><th>%</th></tr>';
  Object.entries(supCount).sort((a,b) => b[1] - a[1]).forEach(([sup, cnt], i) => {
    const cls = i===0?'gold':i===1?'silver':i===2?'bronze':'';
    sh += '<tr class="'+cls+'"><td class="left">'+sup+'</td><td>'+cnt+'</td><td>'+(cnt/data.length*100).toFixed(1)+'%</td></tr>';
  });
  sh += '</table>';
  document.getElementById('resumenSup').innerHTML = sh;

  const modCount = {};
  data.forEach(v => { modCount[v.modelo] = (modCount[v.modelo] || 0) + 1; });
  let mh = '<table><tr><th class="left">Modelo</th><th>Ventas</th><th>%</th></tr>';
  Object.entries(modCount).sort((a,b) => b[1] - a[1]).forEach(([mod, cnt], i) => {
    const cls = i===0?'gold':i===1?'silver':i===2?'bronze':'';
    mh += '<tr class="'+cls+'"><td class="left">'+mod+'</td><td>'+cnt+'</td><td>'+(cnt/data.length*100).toFixed(1)+'%</td></tr>';
  });
  mh += '</table>';
  document.getElementById('resumenMod').innerHTML = mh;

  let dh = '<table><tr><th>#</th><th class="left">Cliente</th><th class="left">Modelo</th><th class="left">Supervisor</th><th class="left">Asesor</th><th>Status</th><th>Dia Creacion</th></tr>';
  data.slice(0, 200).forEach((v, i) => {
    const stClass = isFacturada(v.status) ? 'status-facturada' : 'status-cancelada';
    dh += '<tr><td>'+(i+1)+'</td><td class="left">'+v.cliente+'</td><td class="left">'+v.modelo+'</td><td class="left">'+v.supervisor+'</td><td class="left">'+v.asesor+'</td><td class="'+stClass+'">'+(v.status||'').substring(0,40)+'</td><td>'+(v.fecha_creacion||'').substring(0,10)+'</td></tr>';
  });
  if (data.length > 200) dh += '<tr><td colspan="7" style="color:#999">... '+data.length+' registros total</td></tr>';
  dh += '</table>';
  document.getElementById('detalleTable').innerHTML = dh;
}

['filterSup', 'filterDiaCrea', 'filterDiaMod', 'filterModelo'].forEach(id => {
  document.getElementById(id).addEventListener('change', render);
});
render();
</script>
</body>
</html>'''

with open(OUT_HTML, 'w', encoding='utf-8') as f:
    f.write(html)

print(f"  Dashboard: {OUT_HTML}")
print(f"\n{'=' * 60}")
print(f"  RESUMEN")
print(f"{'=' * 60}")
print(f"  SQLite: {DB_PATH} ({db_size/1024:.1f} KB)")
print(f"  Dashboard: {OUT_HTML}")
print(f"  Registros: {len(all_ventas)} iPhones julio 2026")
print(f"  Comboboxes: Supervisor, Dia Crea, Dia Mod, Modelo, Modo (Creacion/Facturadas/Ambas)")
print(f"{'=' * 60}")
