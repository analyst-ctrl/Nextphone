# -*- coding: utf-8 -*-
"""Verifica si los archivos fuente de Lixi/3ra son realmente de AGOSTO 2026."""
import os
import openpyxl
from datetime import datetime, date, timedelta

FOLDER = os.path.dirname(os.path.abspath(__file__))

def parse_date(v):
    if isinstance(v, (datetime, date)):
        return v if isinstance(v, datetime) else datetime(v.year, v.month, v.day)
    if isinstance(v, (int, float)) and v > 10000:
        try:
            return datetime(1899, 12, 30) + timedelta(days=v)
        except Exception:
            return None
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S.%f",
                    "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
    return None

def analyze(fname, sheet, colname, max_rows=300000):
    path = os.path.join(FOLDER, fname)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet in wb.sheetnames else wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        return "SIN HEADER"
    idx = None
    for i, h in enumerate(header):
        if h and colname.lower() in str(h).lower():
            idx = i
            break
    if idx is None:
        return "columna '%s' no encontrada; header=%s" % (
            colname, [str(h)[:15] for h in header[:15]])
    dates = []
    first_val = None
    n = 0
    for r in rows:
        n += 1
        if n > max_rows:
            break
        v = r[idx] if idx < len(r) else None
        if first_val is None and v is not None:
            first_val = v
        d = parse_date(v)
        if d:
            dates.append(d)
    wb.close()
    if not dates:
        return "%d filas, SIN fechas parseables. ejemplo=%r" % (n, first_val)
    dmin, dmax = min(dates), max(dates)
    months = {}
    for d in dates:
        k = d.strftime("%Y-%m")
        months[k] = months.get(k, 0) + 1
    top = sorted(months.items(), key=lambda x: -x[1])[:6]
    return "%d filas | rango: %s .. %s | por mes: %s" % (
        n, dmin.strftime("%Y-%m-%d"), dmax.strftime("%Y-%m-%d"), top)

print("=" * 90)
print("1) VICI COBROS AGOSTO.xlsx  (call_date)")
print("  ", analyze("VICI COBROS AGOSTO.xlsx", "Hoja1", "call_date"))
print("=" * 90)
print("2) VICI COBROS KREDIYA AGOSTO.xlsx  (call_date)")
print("  ", analyze("VICI COBROS KREDIYA AGOSTO.xlsx", "Hoja1", "call_date"))
print("=" * 90)
print("3) Llamadas salientes AGOSTO (Marca temporal)")
print("  ", analyze("Llamadas salientes cobros AGOSTO 2026 (Respuestas) (16).xlsx",
                    "Respuestas de formulario 1", "Marca temporal"))
print("=" * 90)
print("4) Base de datos de Cobros 31072026.xlsx  (F.Registro)")
print("  ", analyze("Base de datos de Cobros 31072026.xlsx", "2026", "F.Registro"))
print("=" * 90)
print("5) NEXTPANAMA __ INNOVATEK __ LLAMAYA 03.08.26.xlsx  (fecha_venta)")
print("  ", analyze("NEXTPANAMA __ INNOVATEK __ LLAMAYA 03.08.26.xlsx", "Sheet 1", "fecha_venta"))
print("=" * 90)
